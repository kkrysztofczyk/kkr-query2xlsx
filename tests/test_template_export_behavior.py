import importlib.machinery
import importlib.util
import tempfile
import threading
import unittest
from pathlib import Path
from unittest import mock

import openpyxl


def load_app_module():
    repo_root = Path(__file__).resolve().parents[1]
    main_path = repo_root / "main.pyw"
    loader = importlib.machinery.SourceFileLoader(
        "app_main_template_behavior", str(main_path)
    )
    spec = importlib.util.spec_from_loader("app_main_template_behavior", loader)
    module = importlib.util.module_from_spec(spec)
    loader.exec_module(module)
    return module


def _make_template_xlsx(path: Path, sheet_name: str = "Sheet1") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(path)
    wb.close()


class TemplateExportBehaviorTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = load_app_module()

    def test_template_export_include_header_writes_header_and_data(self):
        with tempfile.TemporaryDirectory() as td:
            # perf_counter scale safety for total_duration calculation
            sql_start = self.app.time.perf_counter()
            template_path = Path(td) / "template.xlsx"
            output_path = Path(td) / "out.xlsx"
            _make_template_xlsx(template_path, sheet_name="Sheet1")

            with mock.patch.object(
                self.app,
                "_run_query_to_rows",
                return_value=([(1, 2)], ["col1", "col2"], 0.1, sql_start),
            ):
                sql_dur, export_dur, total_dur, rows_count = self.app.run_export_to_template(
                    engine=object(),
                    sql_query="SELECT 1",
                    template_path=str(template_path),
                    output_file_path=str(output_path),
                    sheet_name="Sheet1",
                    start_cell="A1",
                    include_header=True,
                    cancel_event=threading.Event(),
                )

            self.assertEqual(rows_count, 1)
            self.assertGreaterEqual(sql_dur, 0.0)
            self.assertGreaterEqual(export_dur, 0.0)
            self.assertGreaterEqual(total_dur, 0.0)
            self.assertTrue(output_path.exists())

            wb = openpyxl.load_workbook(output_path)
            try:
                ws = wb["Sheet1"]
                # Header
                self.assertEqual(ws["A1"].value, "col1")
                self.assertEqual(ws["B1"].value, "col2")
                # Data starts in the next row
                self.assertEqual(ws["A2"].value, 1)
                self.assertEqual(ws["B2"].value, 2)
            finally:
                wb.close()

    def test_template_export_missing_sheet_raises_and_removes_output(self):
        with tempfile.TemporaryDirectory() as td:
            template_path = Path(td) / "template.xlsx"
            output_path = Path(td) / "out.xlsx"
            _make_template_xlsx(template_path, sheet_name="Sheet1")

            with mock.patch.object(
                self.app,
                "_run_query_to_rows",
                return_value=([(1,)], ["col"], 0.1, 10.0),
            ):
                with self.assertRaises(ValueError):
                    self.app.run_export_to_template(
                        engine=object(),
                        sql_query="SELECT 1",
                        template_path=str(template_path),
                        output_file_path=str(output_path),
                        sheet_name="MissingSheet",
                        start_cell="A1",
                        include_header=False,
                        cancel_event=threading.Event(),
                    )

            # On any exception, output must be cleaned up.
            self.assertFalse(output_path.exists())

    def test_template_export_no_rows_does_not_load_workbook_and_is_byte_copy(self):
        with tempfile.TemporaryDirectory() as td:
            template_path = Path(td) / "template.xlsx"
            output_path = Path(td) / "out.xlsx"

            # Intentionally not a real XLSX. With rows_count == 0 the code must
            # only copy bytes and never call load_workbook().
            template_bytes = b"NOT_AN_XLSX_BUT_OK_FOR_COPY"
            template_path.write_bytes(template_bytes)

            sql_start = self.app.time.perf_counter()
            with mock.patch.object(
                self.app,
                "_run_query_to_rows",
                return_value=([], ["col1"], 0.01, sql_start),
            ), mock.patch.object(
                self.app,
                "load_workbook",
                side_effect=AssertionError(
                    "load_workbook() should not be called when rows_count == 0"
                ),
            ):
                sql_dur, export_dur, total_dur, rows_count = self.app.run_export_to_template(
                    engine=object(),
                    sql_query="SELECT 1",
                    template_path=str(template_path),
                    output_file_path=str(output_path),
                    sheet_name="AnySheetNameIsFine",
                    start_cell="A1",
                    include_header=True,
                    cancel_event=threading.Event(),
                )

            self.assertGreaterEqual(sql_dur, 0.0)
            self.assertGreaterEqual(export_dur, 0.0)
            self.assertGreaterEqual(total_dur, 0.0)
            self.assertEqual(rows_count, 0)
            self.assertTrue(output_path.exists())
            self.assertEqual(output_path.read_bytes(), template_bytes)

    def test_template_export_no_rows_does_not_validate_sheet_name_and_is_byte_copy(self):
        with tempfile.TemporaryDirectory() as td:
            template_path = Path(td) / "template.xlsx"
            output_path = Path(td) / "out.xlsx"

            # Real XLSX (so this test is explicitly about sheet_name not being validated
            # when rows_count == 0, not about workbook format).
            _make_template_xlsx(template_path, sheet_name="Sheet1")
            template_bytes = template_path.read_bytes()

            sql_start = self.app.time.perf_counter()
            with mock.patch.object(
                self.app,
                "_run_query_to_rows",
                return_value=([], ["col1"], 0.01, sql_start),
            ), mock.patch.object(
                self.app,
                "load_workbook",
                side_effect=AssertionError(
                    "load_workbook() should not be called when rows_count == 0"
                ),
            ):
                sql_dur, export_dur, total_dur, rows_count = self.app.run_export_to_template(
                    engine=object(),
                    sql_query="SELECT 1",
                    template_path=str(template_path),
                    output_file_path=str(output_path),
                    sheet_name="MissingSheet",
                    start_cell="A1",
                    include_header=True,
                    cancel_event=threading.Event(),
                )

            self.assertGreaterEqual(sql_dur, 0.0)
            self.assertGreaterEqual(export_dur, 0.0)
            self.assertGreaterEqual(total_dur, 0.0)
            self.assertEqual(rows_count, 0)
            self.assertTrue(output_path.exists())
            self.assertEqual(output_path.read_bytes(), template_bytes)

    def test_template_export_cancel_before_copyfile_does_not_create_output(self):
        cancel_evt = threading.Event()
        cancel_evt.set()
        with tempfile.TemporaryDirectory() as td:
            template_path = Path(td) / "template.xlsx"
            template_path.write_bytes(b"dummy")
            output_path = Path(td) / "out.xlsx"

            with mock.patch.object(
                self.app,
                "_run_query_to_rows",
                return_value=([], ["col"], 0.1, 10.0),
            ), mock.patch.object(self.app.shutil, "copyfile") as mock_copyfile:
                with self.assertRaises(self.app.UserCancelledError):
                    self.app.run_export_to_template(
                        engine=object(),
                        sql_query="SELECT 1",
                        template_path=str(template_path),
                        output_file_path=str(output_path),
                        sheet_name="Sheet1",
                        start_cell="A1",
                        include_header=False,
                        cancel_event=cancel_evt,
                    )

            mock_copyfile.assert_not_called()
            self.assertFalse(output_path.exists())

    def test_template_export_cancel_after_copyfile_raises_before_load_workbook(self):
        cancel_evt = threading.Event()
        with tempfile.TemporaryDirectory() as td:
            template_path = Path(td) / "template.xlsx"
            # The template doesn't have to be a valid XLSX for this test
            # because we expect to cancel before load_workbook().
            template_path.write_bytes(b"dummy")
            output_path = Path(td) / "out.xlsx"

            orig_copyfile = self.app.shutil.copyfile

            def _copy_then_cancel(src, dst):
                orig_copyfile(src, dst)
                cancel_evt.set()

            with mock.patch.object(
                self.app,
                "_run_query_to_rows",
                return_value=([(1,)], ["col"], 0.1, 10.0),
            ), mock.patch.object(
                self.app.shutil, "copyfile", side_effect=_copy_then_cancel
            ), mock.patch.object(
                self.app,
                "load_workbook",
                side_effect=AssertionError(
                    "load_workbook() should not be called after cancel"
                ),
            ):
                with self.assertRaises(self.app.UserCancelledError):
                    self.app.run_export_to_template(
                        engine=object(),
                        sql_query="SELECT 1",
                        template_path=str(template_path),
                        output_file_path=str(output_path),
                        sheet_name="Sheet1",
                        start_cell="A1",
                        include_header=False,
                        cancel_event=cancel_evt,
                    )

            self.assertFalse(output_path.exists())

    def test_template_export_timeout_after_copyfile_removes_output(self):
        with tempfile.TemporaryDirectory() as td:
            template_path = Path(td) / "template.xlsx"
            template_path.write_bytes(b"")
            output_path = Path(td) / "out.xlsx"

            calls = {"n": 0}

            def _check_deadline(deadline, exc_type, message: str):
                calls["n"] += 1
                if calls["n"] == 2:
                    raise exc_type(message)

            with mock.patch.object(
                self.app,
                "_run_query_to_rows",
                return_value=([], ["col"], 0.1, 10.0),
            ), mock.patch.object(
                self.app, "_check_deadline", side_effect=_check_deadline
            ), mock.patch.object(self.app, "_deadline", return_value=object()):
                with self.assertRaises(self.app.ExportTimeoutError):
                    self.app.run_export_to_template(
                        engine=object(),
                        sql_query="SELECT 1",
                        template_path=str(template_path),
                        output_file_path=str(output_path),
                        sheet_name="Sheet1",
                        start_cell="A1",
                        include_header=False,
                        export_timeout_seconds=60,
                        cancel_event=threading.Event(),
                    )

            self.assertFalse(output_path.exists())


if __name__ == "__main__":
    unittest.main()
