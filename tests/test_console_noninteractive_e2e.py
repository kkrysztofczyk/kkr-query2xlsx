import importlib.machinery
import importlib.util
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.pool import NullPool


def load_app_module():
    repo_root = Path(__file__).resolve().parents[1]
    main_path = repo_root / "main.pyw"
    loader = importlib.machinery.SourceFileLoader("app_main_console_e2e", str(main_path))
    spec = importlib.util.spec_from_loader("app_main_console_e2e", loader)
    module = importlib.util.module_from_spec(spec)
    loader.exec_module(module)
    return module


class ConsoleNonInteractiveE2ETests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = load_app_module()

    def setUp(self):
        # Stabilize console messages; restore original language after each test
        self._old_lang = getattr(self.app, "_CURRENT_LANG", "en")
        self.app.set_lang("en")

    def tearDown(self):
        self.app.set_lang(self._old_lang)

    def test_noninteractive_xlsx_export_end_to_end_sqlite_memory(self):
        engine = create_engine("sqlite:///:memory:", poolclass=NullPool)
        try:
            with tempfile.TemporaryDirectory() as td:
                td_path = Path(td)
                sql_path = td_path / "e2e.sql"
                # Works on SQLite and doesn't require any tables
                sql_path.write_text("SELECT 1 AS one, 'a' AS two;", encoding="utf-8")

                out = io.StringIO()
                with redirect_stdout(out):
                    exit_code = self.app.run_console_noninteractive(
                        engine,
                        output_directory=str(td_path),
                        selected_connection={"name": "E2E", "type": "sqlite"},
                        sql_path=str(sql_path),
                        output_format="xlsx",
                        output_override=None,
                        archive_sql=False,
                    )

                self.assertEqual(exit_code, 0)

                xlsx_files = list(td_path.glob("*.xlsx"))
                self.assertTrue(xlsx_files, "Expected at least one .xlsx output file")
                # Prefer strictness: exactly one XLSX in the temp folder
                self.assertEqual(len(xlsx_files), 1)
                expected_xlsx = xlsx_files[0]

                wb = openpyxl.load_workbook(expected_xlsx)
                try:
                    ws = wb.active
                    # Header row
                    self.assertEqual(ws["A1"].value, "one")
                    self.assertEqual(ws["B1"].value, "two")
                    # First data row
                    self.assertEqual(ws["A2"].value, 1)
                    self.assertEqual(ws["B2"].value, "a")
                finally:
                    wb.close()
        finally:
            engine.dispose()

    def test_noninteractive_archive_sql_writes_json_and_sql(self):
        engine = create_engine("sqlite:///:memory:", poolclass=NullPool)
        try:
            with tempfile.TemporaryDirectory() as td:
                td_path = Path(td)

                # Keep archive output isolated from the real DATA_DIR
                # IMPORTANT (Windows): restore DATA_DIR before TemporaryDirectory cleanup,
                # so RotatingFileHandler releases the temp log file lock.
                old_data_dir = self.app.DATA_DIR
                self.app._set_data_dir(str(td_path))
                try:
                    sql_path = td_path / "arch.sql"
                    sql_path.write_text("SELECT 1 AS one;", encoding="utf-8")

                    out = io.StringIO()
                    with redirect_stdout(out):
                        exit_code = self.app.run_console_noninteractive(
                            engine,
                            output_directory=str(td_path),
                            selected_connection={"name": "E2E", "type": "sqlite"},
                            sql_path=str(sql_path),
                            output_format="xlsx",
                            output_override=None,
                            archive_sql=True,
                        )

                    self.assertEqual(exit_code, 0)

                    archive_dir = td_path / "sql_archive"
                    self.assertTrue(archive_dir.exists())

                    json_files = list(archive_dir.glob("*.json"))
                    sql_files = list(archive_dir.glob("*.sql"))
                    self.assertTrue(json_files, "Expected at least one .json archive file")
                    self.assertTrue(sql_files, "Expected at least one .sql archive file")

                    meta = json.loads(json_files[0].read_text(encoding="utf-8"))
                    # Minimal contract checks - keep it stable across refactors
                    self.assertIn("rows_count", meta)
                    self.assertIn("output_file_path", meta)
                    self.assertEqual(meta.get("output_format"), "xlsx")
                finally:
                    self.app._set_data_dir(old_data_dir)
        finally:
            engine.dispose()

    def test_noninteractive_zero_rows_logs_empty_file_or_nothing_saved(self):
        engine = create_engine("sqlite:///:memory:", poolclass=NullPool)
        try:
            with tempfile.TemporaryDirectory() as td:
                td_path = Path(td)
                sql_path = td_path / "zero.sql"
                out_path = td_path / "zero.xlsx"
                sql_path.write_text("SELECT 1 AS one WHERE 1 = 0;", encoding="utf-8")

                out = io.StringIO()
                with redirect_stdout(out):
                    exit_code = self.app.run_console_noninteractive(
                        engine,
                        output_directory=str(td_path),
                        selected_connection={"name": "E2E", "type": "sqlite"},
                        sql_path=str(sql_path),
                        output_format="xlsx",
                        output_override=str(out_path),
                        archive_sql=False,
                    )

                self.assertEqual(exit_code, 0)
                stdout = out.getvalue()
                self.assertTrue(
                    (
                        "No rows, nothing saved." in stdout
                        or "No rows, saved empty file to:" in stdout
                    ),
                    stdout,
                )
                if "No rows, saved empty file to:" in stdout:
                    self.assertTrue(out_path.exists(), "Expected empty output file to be present")
                if "No rows, nothing saved." in stdout:
                    self.assertFalse(out_path.exists(), "Expected no output file to be present")
        finally:
            engine.dispose()

    def test_noninteractive_zero_rows_when_output_already_exists_prints_nothing_saved(self):
        engine = create_engine("sqlite:///:memory:", poolclass=NullPool)
        try:
            with tempfile.TemporaryDirectory() as td:
                td_path = Path(td)
                sql_path = td_path / "zero.sql"
                out_path = td_path / "zero.xlsx"
                sql_path.write_text("SELECT 1 AS one WHERE 1 = 0;", encoding="utf-8")

                # pre-create output file to simulate previous run
                original_payload = b"old"
                out_path.write_bytes(original_payload)

                out = io.StringIO()
                with redirect_stdout(out):
                    exit_code = self.app.run_console_noninteractive(
                        engine,
                        output_directory=str(td_path),
                        selected_connection={"name": "E2E", "type": "sqlite"},
                        sql_path=str(sql_path),
                        output_format="xlsx",
                        output_override=str(out_path),
                        archive_sql=False,
                    )

                self.assertEqual(exit_code, 0)
                self.assertIn("No rows, nothing saved.", out.getvalue())
                self.assertTrue(out_path.exists())
                self.assertEqual(out_path.read_bytes(), original_payload)
        finally:
            engine.dispose()


class CliUxContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.repo_root = Path(__file__).resolve().parents[1]
        cls.main_path = cls.repo_root / "main.pyw"

    def _run_cli(self, args, *, data_home: Path, app_home: Path):
        env = os.environ.copy()
        env["XDG_DATA_HOME"] = str(data_home)
        # Force deterministic language/locale in CI/local environments.
        env["KKR_LANG"] = "en"
        env["LC_ALL"] = "C"
        env["LANG"] = "C"
        env["PYTHONUTF8"] = "1"
        run_main = app_home / "main.pyw"
        shutil.copy2(self.main_path, run_main)
        cli_args = [*args, "--lang", "en"]
        proc = subprocess.run(
            [sys.executable, str(run_main), *cli_args],
            cwd=str(app_home),
            capture_output=True,
            text=True,
            env=env,
        )
        return proc

    def test_list_connections_prints_names_and_exits_zero(self):
        with tempfile.TemporaryDirectory() as td:
            data_home = Path(td)
            app_home = data_home / "app"
            app_home.mkdir(parents=True, exist_ok=True)
            (app_home / "secure.txt").write_text(
                json.dumps(
                    {
                        "connections": [
                            {"name": "Default MSSQL", "type": "mssql_odbc", "details": {}},
                            {"name": "Demo SQLite", "type": "sqlite", "details": {"path": "demo.db"}},
                        ],
                        "last_selected": "Default MSSQL",
                    }
                ),
                encoding="utf-8",
            )

            proc = self._run_cli(["--list-connections"], data_home=data_home, app_home=app_home)

            self.assertEqual(proc.returncode, 0, proc.stderr)
            self.assertEqual(proc.stdout.strip().splitlines(), ["Default MSSQL", "Demo SQLite"])

    def test_connection_not_found_exits_two_and_shows_hint(self):
        with tempfile.TemporaryDirectory() as td:
            data_home = Path(td)
            app_home = data_home / "app"
            app_home.mkdir(parents=True, exist_ok=True)
            (app_home / "secure.txt").write_text(
                json.dumps(
                    {
                        "connections": [
                            {"name": "Default MSSQL", "type": "mssql_odbc", "details": {}},
                        ],
                        "last_selected": "Default MSSQL",
                    }
                ),
                encoding="utf-8",
            )

            proc = self._run_cli(
                ["--console", "--connection", "NOPE"],
                data_home=data_home,
                app_home=app_home,
            )

            self.assertEqual(proc.returncode, 2, proc.stderr)
            self.assertIn("NOPE", proc.stdout)
            self.assertIn("--list-connections", proc.stdout)
            self.assertIn("Default MSSQL", proc.stdout)


if __name__ == "__main__":
    unittest.main()
