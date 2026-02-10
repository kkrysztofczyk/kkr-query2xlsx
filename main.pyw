from __future__ import annotations

import argparse
import csv
import decimal
import json
import getpass
import importlib.metadata
import logging
import math
import platform
import textwrap
import queue
import re
import socket
import threading
import traceback
import os
import sqlite3
import shutil
import subprocess
import sys
import tempfile
import time
import webbrowser
import io
import contextlib
from datetime import datetime
from pathlib import Path
from typing import Literal, Optional, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import quote_plus, unquote_plus
from urllib.request import Request, urlopen

from logging.handlers import RotatingFileHandler

import openpyxl
import sqlalchemy
from sqlalchemy import create_engine, event, text
from sqlalchemy.exc import DBAPIError, NoSuchModuleError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import coordinate_to_tuple

try:
    import tkinter as tk
    import tkinter.font as tkfont
    from tkinter import filedialog, messagebox, simpledialog
    from tkinter import ttk
    from tkinter.scrolledtext import ScrolledText

    _TK_AVAILABLE = True
except Exception:  # noqa: BLE001
    tk = None  # type: ignore[assignment]
    tkfont = None  # type: ignore[assignment]
    filedialog = messagebox = simpledialog = None  # type: ignore[assignment]
    ttk = None  # type: ignore[assignment]
    ScrolledText = None  # type: ignore[assignment]
    _TK_AVAILABLE = False


def _require_tk() -> None:
    if not _TK_AVAILABLE:
        raise RuntimeError(
            "Tkinter is not available. Install python3-tk (Linux) / ensure Tk is bundled."
        )


def _load_optional_sql_highlighter():
    try:
        from chlorophyll import CodeView
        import tklinenums  # noqa: F401
        from pygments.lexers.sql import SqlLexer, MySqlLexer, PostgresLexer, TransactSqlLexer
    except Exception:
        return None, {}

    return CodeView, {
        "SQLite": SqlLexer,
        "MySQL": MySqlLexer,
        "Postgres": PostgresLexer,
        "MSSQL": TransactSqlLexer,
    }


CodeView, SQL_LEXER_CLASSES = _load_optional_sql_highlighter()


def _xlsx_get_sheetnames(xlsx_path: str) -> list[str]:
    """Returns sheetnames or [] on error."""
    wb = None
    try:
        wb = openpyxl.load_workbook(
            xlsx_path,
            read_only=True,
            data_only=True,
            keep_vba=True,
        )
        return list(wb.sheetnames or [])
    except Exception:
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _pick_default_sheet(sheetnames: list[str]) -> str:
    """Prefer 'data' (case-insensitive), else first."""
    for sheet in sheetnames:
        if (sheet or "").strip().lower() == "data":
            return sheet
    return sheetnames[0] if sheetnames else ""


def _show_template_naming_hint_once(
    parent: tk.Misc,
    app_config: dict,
    save_config_fn,
    *,
    sql_path: str | None,
    template_path: str | None,
) -> None:
    """
    Shows one-time hint about naming convention:
      Report.sql, Report.xlsx, Report_template.xlsx
    Has 'Don't show again' persisted in config.
    """
    ui_cfg = app_config.setdefault("ui", {})
    if ui_cfg.get("hide_template_naming_hint", False):
        return

    if not sql_path or not template_path:
        return

    sql_stem = Path(sql_path).stem
    tpl = Path(template_path)
    recommended_tpl_stem = f"{sql_stem}_template"
    if tpl.stem.lower() == recommended_tpl_stem.lower() and tpl.suffix.lower() == ".xlsx":
        return

    win = tk.Toplevel(parent)
    win.title("Tip")
    win.resizable(False, False)
    win.transient(parent)
    win.grab_set()
    try:
        win.lift()
        win.focus_set()
    except Exception:
        pass

    frm = ttk.Frame(win, padding=12)
    frm.grid(row=0, column=0, sticky="nsew")

    msg = (
        "To avoid mix-ups (wrong file / wrong worksheet), we recommend keeping this convention:\n\n"
        "  Report.sql\n"
        "  Report.xlsx\n"
        "  Report_template.xlsx\n\n"
        "It helps prevent mistakes and accidental writes into the wrong sheet (e.g. 'data')."
    )
    lbl = ttk.Label(frm, text=msg, justify="left", wraplength=520)
    lbl.grid(row=0, column=0, columnspan=2, sticky="w")

    dont_var = tk.BooleanVar(value=False)
    chk = ttk.Checkbutton(frm, text="Don't show again", variable=dont_var)
    chk.grid(row=1, column=0, sticky="w", pady=(10, 0))

    def _close() -> None:
        if dont_var.get():
            ui_cfg["hide_template_naming_hint"] = True
            try:
                save_config_fn(app_config)
            except Exception:
                pass
        try:
            win.grab_release()
        except tk.TclError:
            pass
        win.destroy()

    btn = ttk.Button(frm, text="OK", command=_close)
    btn.grid(row=1, column=1, sticky="e", padx=(12, 0), pady=(10, 0))

    win.bind("<Return>", lambda _e: _close())
    win.bind("<Escape>", lambda _e: _close())
    win.protocol("WM_DELETE_WINDOW", _close)

    win.update_idletasks()
    try:
        _center_window(win, parent)
    except Exception:
        pass


def _show_data_dir_notice_once(
    parent: tk.Misc,
    app_config: dict,
    save_config_fn,
    *,
    data_dir: str,
    base_dir: str,
    lang: str | None = None,
) -> None:
    """One-time info popup: where app stores its data (with 'No' option)."""
    notice_lang = _normalize_ui_lang(lang) or _CURRENT_LANG

    def _notice_t(key: str, **kwargs) -> str:
        s = I18N.get(notice_lang, {}).get(key) or I18N["en"].get(key) or key
        return s.format(**kwargs) if kwargs else s

    try:
        dd = os.path.abspath(data_dir or "")
        bd = os.path.abspath(base_dir or "")
    except Exception:
        dd = str(data_dir or "")
        bd = str(base_dir or "")

    # show only when data dir differs from app dir
    if not dd or dd == bd:
        return

    ui_cfg = app_config.setdefault("ui", {})
    if ui_cfg.get("hide_data_dir_notice", False):
        return

    win = tk.Toplevel(parent)
    win.title(_notice_t("TITLE_DATA_DIR_NOTICE"))
    win.resizable(False, False)
    win.transient(parent)
    win.grab_set()

    # Best-effort: keep notice above main window and focused (esp. Windows).
    try:
        win.lift()
    except Exception:
        pass
    try:
        win.focus_force()
    except Exception:
        pass
    try:
        win.attributes("-topmost", True)
        win.attributes("-topmost", False)
    except Exception:
        pass

    frm = ttk.Frame(win, padding=12)
    frm.grid(row=0, column=0, sticky="nsew")

    msg = _notice_t("MSG_DATA_DIR_NOTICE", data_dir=dd)
    lbl = ttk.Label(frm, text=msg, justify="left", wraplength=520)
    lbl.grid(row=0, column=0, columnspan=2, sticky="w")

    def _close() -> None:
        try:
            win.grab_release()
        except tk.TclError:
            pass
        win.destroy()

    def _dont_show_again() -> None:
        ui_cfg["hide_data_dir_notice"] = True
        try:
            save_config_fn(app_config)
        except Exception:
            pass
        _close()

    btn_no = ttk.Button(
        frm,
        text=_notice_t("BTN_DONT_SHOW_AGAIN"),
        command=_dont_show_again,
    )
    btn_no.grid(row=1, column=0, sticky="w", pady=(10, 0))

    btn_ok = ttk.Button(frm, text=_notice_t("BTN_OK"), command=_close)
    btn_ok.grid(row=1, column=1, sticky="e", padx=(12, 0), pady=(10, 0))

    win.bind("<Return>", lambda _e: _close())
    win.bind("<Escape>", lambda _e: _close())
    win.protocol("WM_DELETE_WINDOW", _close)

    win.update_idletasks()
    try:
        _center_window(win, parent)
    except Exception:
        pass

def _get_base_dir() -> str:
    """Return the directory that should be treated as the app "home".

    - source run: folder with main.pyw
    - PyInstaller (onedir/onefile): folder with the .exe
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = _get_base_dir()
DATA_DIR = BASE_DIR  # default: tak jak dziś; zmienimy tylko gdy brak praw zapisu


def _set_data_dir(new_dir: str) -> None:
    """Przełącz katalog danych użytkownika i przelicz globalne ścieżki."""
    global DATA_DIR
    global SECURE_PATH, QUERIES_PATH, APP_CONFIG_PATH, LEGACY_CSV_PROFILES_PATH, SQL_ARCHIVE_DIR

    DATA_DIR = os.path.abspath(new_dir)

    # przelicz ścieżki zależne od _build_path()
    SECURE_PATH = _build_path("secure.txt")
    QUERIES_PATH = _build_path("queries.txt")
    APP_CONFIG_PATH = _build_path("kkr-query2xlsx.json")
    LEGACY_CSV_PROFILES_PATH = _build_path("csv_profiles.json")
    SQL_ARCHIVE_DIR = _build_path("sql_archive")
    _attach_logger_file_handler(_build_path("logs"))


def get_default_user_data_dir() -> str:
    app_name = "kkr-query2xlsx"
    if sys.platform == "win32":
        base = os.environ.get("LOCALAPPDATA") or str(Path.home())
        return os.path.join(base, app_name)
    if sys.platform == "darwin":
        return os.path.join(str(Path.home()), "Library", "Application Support", app_name)
    base = os.environ.get("XDG_DATA_HOME") or os.path.join(str(Path.home()), ".local", "share")
    return os.path.join(base, app_name)


def has_data_markers(path: str | Path) -> bool:
    base = Path(path)
    file_markers = ("secure.txt", "queries.txt", "kkr-query2xlsx.json")
    dir_markers = ("generated_reports", "logs", "sql_archive")

    return any((base / name).exists() for name in file_markers + dir_markers)


def select_startup_data_dir(base_dir: str | Path, user_dir: str | Path) -> str:
    base_raw = str(base_dir)
    user_raw = str(user_dir)

    if has_data_markers(base_raw):
        return base_raw
    if has_data_markers(user_raw):
        return user_raw
    return base_raw


def _path_cmp_key(path: str | Path) -> str:
    """Return a normalized key for path comparisons on the current OS."""
    try:
        return os.path.normcase(os.path.abspath(str(path)))
    except Exception:
        return str(path)


def _same_path(a: str | Path, b: str | Path) -> bool:
    """Best-effort path equality check without mutating caller-provided values."""
    return _path_cmp_key(a) == _path_cmp_key(b)


_APP_ICON_PHOTO = None  # keep a reference to avoid GC in Tk


def apply_app_icon(win) -> None:
    """Best-effort: set window icon from docs/ (no crash if missing)."""
    global _APP_ICON_PHOTO
    docs_dir = Path(BASE_DIR) / "docs"

    ico_candidates = [
        docs_dir / "kkr-query2xlsx.ico",
        docs_dir / "icon.ico",
        docs_dir / "app.ico",
    ]
    png_candidates = [
        docs_dir / "kkr-query2xlsx-icon-256.png",
        docs_dir / "kkr-query2xlsx-icon-512.png",
        docs_dir / "icon.png",
        docs_dir / "app.png",
    ]

    ico_path = next((p for p in ico_candidates if p.exists()), None)
    png_path = next((p for p in png_candidates if p.exists()), None)

    # Windows: .ico is the most reliable
    try:
        if sys.platform == "win32" and ico_path:
            win.iconbitmap(str(ico_path))
    except Exception:
        pass

    # Cross-platform fallback: .png via iconphoto
    try:
        if png_path:
            if _APP_ICON_PHOTO is None:
                _APP_ICON_PHOTO = tk.PhotoImage(file=str(png_path))
            win.iconphoto(True, _APP_ICON_PHOTO)
    except Exception:
        pass


def _fmt_elapsed(seconds: float) -> str:
    s = int(seconds)
    if s < 60:
        return f"{s}s"
    m, s = divmod(s, 60)
    if m < 60:
        return f"{m}m {s}s"
    h, m = divmod(m, 60)
    if h < 24:
        return f"{h}h {m}m"
    d, h = divmod(h, 24)
    return f"{d}d {h}h"


def _coerce_seconds(seconds: Optional[float]) -> float:
    try:
        value = float(seconds)
    except (TypeError, ValueError):
        return 0.0
    if not math.isfinite(value) or value < 0:
        return 0.0
    return value


def _fmt_hms(seconds: Optional[float]) -> str:
    total_seconds = int(_coerce_seconds(seconds))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


class ExportProgressWindow:
    def __init__(self, parent, title: str | None = None, on_stop=None):
        self.parent = parent
        self.win = tk.Toplevel(parent)
        self.win.title(title or t("PROGRESS_WORKING_TITLE"))
        apply_app_icon(self.win)
        self.win.transient(parent)
        self.win.resizable(False, False)
        self._on_stop_callback = on_stop

        self._step_var = tk.StringVar(value=t("PROGRESS_STARTING"))
        self._elapsed_var = tk.StringVar(value="0s")
        self._status_var = tk.StringVar(value="")
        self._t0 = time.time()
        self._running = True
        self._stop_btn = None
        self._tick_job = None

        frm = ttk.Frame(self.win, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text=t("PROGRESS_STEP")).grid(row=0, column=0, sticky="w")
        ttk.Label(frm, textvariable=self._step_var).grid(row=0, column=1, sticky="w")

        ttk.Label(frm, text=t("PROGRESS_ELAPSED")).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Label(frm, textvariable=self._elapsed_var).grid(
            row=1, column=1, sticky="w", pady=(6, 0)
        )

        ttk.Label(frm, textvariable=self._status_var).grid(
            row=2, column=0, columnspan=2, sticky="w", pady=(8, 0)
        )

        self._stop_btn = ttk.Button(
            frm,
            text=f"{t('BTN_CANCEL_RUN')} (Esc)",
            command=self._on_stop_clicked,
        )
        self._stop_btn.grid(row=3, column=0, columnspan=2, sticky="e", pady=(10, 0))

        frm.columnconfigure(1, weight=1)

        self._tick()

        self.win.bind("<Escape>", lambda _event: self._on_stop_clicked())
        self.win.protocol("WM_DELETE_WINDOW", self._on_stop_clicked)

    def _on_stop_clicked(self):
        if self._stop_btn is None:
            return
        if "disabled" in self._stop_btn.state():
            return
        self._status_var.set(t("MSG_CANCEL_REQUESTED"))
        self._stop_btn.state(["disabled"])
        if self._on_stop_callback is not None:
            self._on_stop_callback()

    def _tick(self):
        if not self._running:
            return
        elapsed = time.time() - self._t0
        self._elapsed_var.set(_fmt_elapsed(elapsed))
        self._tick_job = self.win.after(250, self._tick)

    def set_step(self, text: str):
        self._step_var.set(text)

    def close(self):
        self._running = False
        try:
            if self._tick_job is not None:
                self.win.after_cancel(self._tick_job)
                self._tick_job = None
        except Exception:
            pass
        try:
            self.win.destroy()
        except tk.TclError:
            pass

# --- App version -------------------------------------------------------------

APP_VERSION = "0.4.0"  # bump manually for releases

MSSQL_SAFE_SET_SQL = """\
SET NOCOUNT ON;
SET ANSI_NULLS ON;
SET QUOTED_IDENTIFIER ON;
SET ANSI_PADDING ON;
SET ANSI_WARNINGS ON;
SET CONCAT_NULL_YIELDS_NULL ON;
SET ARITHABORT ON;
SET NUMERIC_ROUNDABORT OFF;
"""


def _apply_mssql_safe_set(cur) -> bool:  # noqa: ANN001
    try:
        cur.execute(MSSQL_SAFE_SET_SQL)
        return True
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning(
            "MSSQL session prolog batch failed, retrying per-statement: %s",
            exc,
            exc_info=True,
        )

    ok = True
    for stmt in MSSQL_SAFE_SET_SQL.split(";"):
        s = stmt.strip()
        if not s:
            continue
        try:
            cur.execute(s)
        except Exception as exc:  # noqa: BLE001
            ok = False
            LOGGER.warning(
                "MSSQL session prolog statement failed: %s (stmt=%r)",
                exc,
                s,
                exc_info=True,
            )

    return ok


def _ensure_engine_mssql_set_hook(engine) -> None:
    if getattr(engine, "_kkr_mssql_set_hook_installed", False):
        return

    @event.listens_for(engine, "connect")
    def _kkr_mssql_connect(dbapi_conn, conn_record):  # noqa: ANN001
        # pyodbc.Connection nie ma __dict__ -> nie można doklejać atrybutów setattr().
        # Stan trzymamy w SQLAlchemy ConnectionRecord.info (per-connection w puli).
        info = None
        try:
            info = getattr(conn_record, "info", None)
        except Exception:
            info = None
        if isinstance(info, dict) and info.get("kkr_mssql_safe_set_done"):
            return
        cur = None
        try:
            cur = dbapi_conn.cursor()
        except Exception as exc:  # noqa: BLE001
            # Nie blokuj całego połączenia, jeśli nie uda się utworzyć kursora.
            LOGGER.warning("MSSQL session prolog: cannot create cursor: %s", exc, exc_info=True)
            return
        try:
            if _apply_mssql_safe_set(cur) and isinstance(info, dict):
                info["kkr_mssql_safe_set_done"] = True
        except Exception as exc:  # noqa: BLE001
            # _apply_mssql_safe_set() loguje standardowe ścieżki failover.
            # Tu logujemy tylko naprawdę niespodziewane wyjątki.
            LOGGER.warning("MSSQL session prolog unexpected error: %s", exc, exc_info=True)
        finally:
            try:
                if cur is not None:
                    cur.close()
            except Exception:
                pass

    setattr(engine, "_kkr_mssql_set_hook_installed", True)

GITHUB_ISSUE_CHOOSER_URL = (
    "https://github.com/kkrysztofczyk/kkr-query2xlsx/issues/new/choose"
)
GITHUB_REPO_OWNER = "kkrysztofczyk"
GITHUB_REPO_NAME = "kkr-query2xlsx"
GITHUB_RELEASES_LATEST_URL = (
    f"https://api.github.com/repos/{GITHUB_REPO_OWNER}/{GITHUB_REPO_NAME}/releases/latest"
)
UPDATER_EXE_NAME = "kkr-query2xlsx-updater.exe"
UPDATER_STAGED_EXE_NAME = "kkr-query2xlsx-updater.new.exe"
GITHUB_COMMITS_MAIN_URL = (
    f"https://api.github.com/repos/{GITHUB_REPO_OWNER}/{GITHUB_REPO_NAME}/commits/main"
)


def parse_version(tag: str) -> tuple[int, ...]:
    parts = re.findall(r"\d+", (tag or "").lstrip("v"))
    return tuple(int(part) for part in parts)


def _find_git_root(start: Path) -> Path | None:
    for candidate in [start, *start.parents]:
        if (candidate / ".git").exists():
            return candidate
    return None


def detect_install_mode() -> str:
    if getattr(sys, "frozen", False):
        return "exe"
    if _find_git_root(Path(__file__).resolve().parent):
        return "git"
    return "source"


def _read_json_request(req: Request, timeout_s: int) -> object:
    with urlopen(req, timeout=timeout_s) as resp:  # noqa: S310
        payload = resp.read().decode("utf-8")
    return json.loads(payload)


def _fetch_latest_release() -> dict:
    data = _read_json_request(
        Request(
            GITHUB_RELEASES_LATEST_URL,
            headers={
                "Accept": "application/vnd.github+json",
                "User-Agent": "kkr-query2xlsx-updater",
            },
        ),
        timeout_s=10,
    )
    if not isinstance(data, dict):
        raise ValueError("GitHub release payload is not a JSON object")
    return data


def _fetch_remote_main_sha() -> str | None:
    data = _read_json_request(
        Request(
            GITHUB_COMMITS_MAIN_URL,
            headers={
                "Accept": "application/vnd.github+json",
                "User-Agent": "kkr-query2xlsx-updater",
            },
        ),
        timeout_s=10,
    )
    if not isinstance(data, dict):
        raise ValueError("GitHub commit payload is not a JSON object")
    sha = (data or {}).get("sha") or ""
    return sha.strip() or None


_UPD_MAX_RETRY_AFTER_SECONDS = 7 * 24 * 60 * 60
_UPD_MAX_RESET_FUTURE_SECONDS = 365 * 24 * 60 * 60


def _format_local_ts(ts: float) -> str | None:
    try:
        return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))
    except (OverflowError, OSError, ValueError):
        return None


def _parse_retry_hint(headers) -> str | None:  # noqa: ANN001
    if not headers:
        return None

    retry_after = headers.get("retry-after")
    if retry_after is not None:
        s = str(retry_after).strip()
        if not s:
            return None
        if s.isdigit():
            try:
                seconds = int(s)
            except (OverflowError, ValueError):
                return None
            if 0 <= seconds <= _UPD_MAX_RETRY_AFTER_SECONDS:
                return _format_local_ts(time.time() + seconds)
            return None
        return s[:64]

    reset_raw = headers.get("x-ratelimit-reset")
    if reset_raw is not None:
        s = str(reset_raw).strip()
        if s.isdigit():
            try:
                reset_ts = int(s)
            except (OverflowError, ValueError):
                return None
            now = int(time.time())
            if reset_ts < 0 or reset_ts > now + _UPD_MAX_RESET_FUTURE_SECONDS:
                return None
            try:
                dt = datetime.fromtimestamp(reset_ts)
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except (OverflowError, OSError, ValueError):
                return None
    return None


def _classify_update_check_error(exc: Exception) -> tuple[str, dict]:
    if isinstance(exc, TimeoutError | socket.timeout):
        return "UPD_ERR_TIMEOUT", {}
    if isinstance(exc, HTTPError):
        status = exc.code
        headers = exc.headers or {}
        remaining = str(headers.get("x-ratelimit-remaining", "")).strip()
        retry_at = _parse_retry_hint(headers)
        if status == 429:
            return (
                "UPD_ERR_RATE_LIMITED",
                {"status": status, "retry_at": retry_at or "unknown"},
            )
        if status == 403 and (retry_at or remaining == "0"):
            return (
                "UPD_ERR_RATE_LIMITED",
                {"status": status, "retry_at": retry_at or "unknown"},
            )
        return "UPD_ERR_HTTP", {"status": status}
    if isinstance(exc, URLError):
        reason = getattr(exc, "reason", None)
        if isinstance(reason, TimeoutError | socket.timeout):
            return "UPD_ERR_TIMEOUT", {}
        return "UPD_ERR_NETWORK", {}
    if isinstance(exc, json.JSONDecodeError | UnicodeDecodeError | ValueError):
        return "UPD_ERR_JSON", {}
    return "UPD_CHECK_FAILED", {}


def _build_update_check_message(exc: Exception) -> str:
    try:
        key, params = _classify_update_check_error(exc)
        return t(key, **params)
    except Exception:
        return t("UPD_CHECK_FAILED")


def _fetch_github_compare_status(base_ref: str, head_ref: str) -> str | None:
    """Return GitHub compare status between base and head (best-effort).

    GitHub compare status is one of: ahead, behind, diverged, identical.
    This prevents mislabeling any local!=remote situation as "remote is newer".
    """
    try:
        url = (
            f"https://api.github.com/repos/{GITHUB_REPO_OWNER}/{GITHUB_REPO_NAME}"
            f"/compare/{quote_plus(base_ref)}...{quote_plus(head_ref)}"
        )
        req = Request(
            url,
            headers={
                "Accept": "application/vnd.github+json",
                "User-Agent": "kkr-query2xlsx-updater",
            },
        )
        with urlopen(req, timeout=10) as resp:  # noqa: S310
            payload = resp.read().decode("utf-8")
        data = json.loads(payload) if payload else {}
        status = (data or {}).get("status") or ""
        return status.strip().lower() or None
    except Exception:
        return None


def _classify_git_relation(local_sha: str | None, remote_sha: str | None) -> str:
    """Classify local HEAD relative to remote main SHA (best-effort).

    Returns one of: match, remote_ahead, local_ahead, diverged, different,
    different_unverified
    """
    if not local_sha or not remote_sha:
        return "different"
    if local_sha == remote_sha:
        return "match"
    status = _fetch_github_compare_status(local_sha, remote_sha)
    if status is None:
        return "different_unverified"
    if status == "identical":
        return "match"
    if status == "ahead":
        # GitHub: head is ahead of base. For base=local, head=remote => remote ahead.
        return "remote_ahead"
    if status == "behind":
        return "local_ahead"
    if status == "diverged":
        return "diverged"
    return "different"


def _select_windows_asset(assets: list[dict]) -> dict | None:
    for asset in assets:
        name = (asset.get("name") or "").lower()
        if "windows" in name and name.endswith(".zip"):
            return asset
    for asset in assets:
        name = (asset.get("name") or "").lower()
        if name.endswith(".zip"):
            return asset
    return None


def get_update_info() -> dict:
    data = _fetch_latest_release()
    latest_tag = data.get("tag_name") or ""
    assets = data.get("assets") or []
    asset = _select_windows_asset(assets)
    current_tag = f"v{APP_VERSION}"
    current_version = parse_version(current_tag)
    latest_version = parse_version(latest_tag)
    return {
        "current_tag": current_tag,
        "latest_tag": latest_tag,
        "update_available": latest_version > current_version,
        "asset": asset,
    }


def _get_updater_command() -> list[str] | None:
    updater_path = Path(BASE_DIR) / UPDATER_EXE_NAME
    if updater_path.exists():
        return [str(updater_path)]
    if not getattr(sys, "frozen", False):
        py_updater = Path(__file__).with_name("updater.py")
        if py_updater.exists():
            return [sys.executable, str(py_updater)]
    return None


def launch_updater(wait_pid: int | None = None) -> bool:
    cmd = _get_updater_command()
    if not cmd:
        return False
    if wait_pid:
        cmd += ["--wait-pid", str(wait_pid)]
    subprocess.Popen(cmd, cwd=BASE_DIR)  # noqa: S603
    return True


def _is_windows_image_running(image_name: str) -> bool:
    if sys.platform != "win32":
        return False
    try:
        proc = subprocess.run(
            ["tasklist", "/FI", f"IMAGENAME eq {image_name}"],
            check=False,
            capture_output=True,
            text=True,
        )
        return image_name.lower() in proc.stdout.lower()
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("Process check failed: %s", exc, exc_info=True)
        return False


def _wait_for_windows_image_exit(image_name: str, timeout_s: float = 5.0) -> bool:
    if sys.platform != "win32":
        return True
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        if not _is_windows_image_running(image_name):
            return True
        time.sleep(0.5)
    return False


def _apply_pending_updater_update() -> None:
    if not getattr(sys, "frozen", False):
        return

    staged_path = Path(BASE_DIR) / UPDATER_STAGED_EXE_NAME
    cfg = None
    updates = None
    cfg_loaded_for_staged_lookup = False

    if not staged_path.exists():
        try:
            cfg = load_app_config()
            cfg_loaded_for_staged_lookup = True
            updates = cfg.get("_updates") if isinstance(cfg, dict) else None
            pending = updates.get("pending_updater") if isinstance(updates, dict) else None
            staged_name = pending.get("file") if isinstance(pending, dict) else None
            if isinstance(staged_name, str) and staged_name.strip():
                staged_path = Path(BASE_DIR) / staged_name.strip()
        except Exception:
            return

        if not staged_path.exists():
            return

    if _is_windows_image_running(UPDATER_EXE_NAME):
        if not _wait_for_windows_image_exit(UPDATER_EXE_NAME, timeout_s=5.0):
            LOGGER.info(
                "Updater still running, deferring staged update: %s", staged_path
            )
            return

    target_path = Path(BASE_DIR) / UPDATER_EXE_NAME
    try:
        os.replace(staged_path, target_path)
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("Failed to replace updater: %s", exc, exc_info=True)
        return

    # Sprzątanie configu tylko best-effort.
    if not cfg_loaded_for_staged_lookup:
        try:
            cfg = load_app_config()
            updates = cfg.get("_updates") if isinstance(cfg, dict) else None
        except Exception:
            cfg = None
            updates = None

    if isinstance(cfg, dict) and isinstance(updates, dict) and "pending_updater" in updates:
        updates.pop("pending_updater", None)
        if updates:
            cfg["_updates"] = updates
        else:
            cfg.pop("_updates", None)
        try:
            save_app_config(cfg)
        except Exception:
            pass

    LOGGER.info("Updater updated successfully")


def apply_native_ttk_theme(root: tk.Tk) -> None:
    """Apply a safe, native-looking ttk theme when available."""
    try:
        style = ttk.Style(root)
        available = set(style.theme_names())
        preferred = []
        if sys.platform.startswith("win"):
            preferred = ["vista", "xpnative"]
        elif sys.platform == "darwin":
            preferred = ["aqua"]
        preferred.append("clam")
        for theme in preferred:
            if theme in available:
                try:
                    style.theme_use(theme)
                    break
                except tk.TclError:
                    continue
    except Exception:
        pass

def _get_git_short_sha() -> str | None:
    """Best-effort git short SHA for local/dev runs. Returns None when unavailable."""
    try:
        here = Path(__file__).resolve().parent
        p = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=str(here),
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
            timeout=0.8,
            check=False,
        )
        sha = (p.stdout or "").strip()
        return sha if sha else None
    except Exception:
        return None


def _get_git_full_sha() -> str | None:
    try:
        here = Path(__file__).resolve().parent
        p = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=str(here),
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
            timeout=0.8,
            check=False,
        )
        sha = (p.stdout or "").strip()
        return sha if sha else None
    except Exception:
        return None


def get_app_version_label() -> str:
    sha = _get_git_short_sha()
    if sha:
        return f"v{APP_VERSION} / {sha}"
    return f"v{APP_VERSION}"


def odbc_diagnostics_text() -> str:
    bits = "64-bit" if sys.maxsize > 2**32 else "32-bit"
    lines = [
        f"exe={sys.executable}",
        f"python={sys.version.split()[0]} ({bits})",
        f"platform={sys.platform}",
    ]
    try:
        import pyodbc  # type: ignore

        lines.append("pyodbc=OK")
        try:
            drivers = pyodbc.drivers()
            lines.append(
                "pyodbc.drivers()=" + (", ".join(drivers) if drivers else "<none>")
            )
        except Exception as exc:  # noqa: BLE001
            lines.append(f"pyodbc.drivers()=FAILED ({type(exc).__name__}: {exc})")
        try:
            dsns = dict(pyodbc.dataSources())
            lines.append(
                "pyodbc.dataSources()="
                + (", ".join(sorted(dsns.keys())) if dsns else "<empty>")
            )
        except Exception as exc:  # noqa: BLE001
            lines.append(f"pyodbc.dataSources()=FAILED ({type(exc).__name__}: {exc})")
    except Exception as exc:  # noqa: BLE001
        lines.append(f"pyodbc=FAILED ({type(exc).__name__}: {exc})")
    return "\n".join(lines)


def _is_frozen_exe() -> bool:
    return bool(getattr(sys, "frozen", False))


def _extract_driver_from_conn_str(conn_str: str) -> Optional[str]:
    """
    Extracts DRIVER={...} from ODBC connection string.
    Returns driver name without braces, or None if not present.
    """
    m = re.search(r"(?i)\bDRIVER\s*=\s*\{([^}]+)\}", conn_str or "")
    return m.group(1).strip() if m else None


def _pyodbc_sqlstate_and_msg(exc: BaseException) -> Tuple[Optional[str], str]:
    """
    pyodbc errors often look like:
      e.args == ('08001', '[08001] [Microsoft][ODBC Driver 18 for SQL Server] ...')
    but sometimes it's only one string. We normalize.
    """
    args = getattr(exc, "args", None) or ()
    if len(args) >= 2 and isinstance(args[0], str) and isinstance(args[1], str):
        sqlstate = args[0].strip() or None
        msg = args[1].strip()
        return sqlstate, msg
    if len(args) == 1 and isinstance(args[0], str):
        return None, args[0].strip()
    return None, str(exc)


def _classify_mssql_conn_error(
    *,
    exc: BaseException,
    conn_str: str,
    pyodbc_ok: bool,
    drivers: list[str],
) -> tuple[str, list[str]]:
    """
    Returns: (title, hints[])
    """
    sqlstate, msg = _pyodbc_sqlstate_and_msg(exc)
    msg_l = (msg or "").lower()

    # Missing pyodbc / cannot load module
    if not pyodbc_ok:
        if _is_frozen_exe():
            return (
                "Missing pyodbc in packaged EXE",
                ["pyodbc is missing or failed to load inside the packaged EXE."],
            )
        return (
            "Missing pyodbc",
            ["pyodbc is not installed or failed to load. Try: python -m pip install pyodbc"],
        )

    # Missing / wrong ODBC driver (only if it actually matches reality)
    requested_driver = _extract_driver_from_conn_str(conn_str)
    if requested_driver:
        # connection string explicitly requests a driver - verify it's installed
        if all(requested_driver.lower() != d.lower() for d in (drivers or [])):
            return (
                "Missing ODBC driver",
                [
                    f"Requested ODBC driver is not installed: {requested_driver}",
                    "Install Microsoft 'ODBC Driver 17/18 for SQL Server' and try again.",
                    "Check 32/64-bit mismatch between Python/EXE and the ODBC driver.",
                ],
            )

    # SQLSTATE-based classification (more accurate than a fixed “missing driver”)
    if sqlstate in ("IM002", "IM003"):
        return (
            "ODBC configuration error",
            [
                "ODBC driver/DSN configuration is invalid or driver cannot be loaded.",
                "Verify DRIVER={...} in connection string and installed ODBC drivers.",
            ],
        )

    if sqlstate in ("28000",):
        return (
            "Authentication failed",
            [
                "Check username/password or authentication method (Windows/SQL).",
                "If using Windows auth, verify Trusted_Connection / Integrated Security.",
            ],
        )

    if sqlstate in ("HYT00", "HYT01"):
        return (
            "Connection timeout",
            [
                "Server is not reachable or responds too slowly.",
                "Check network/VPN, host, port, firewall rules.",
            ],
        )

    if sqlstate and sqlstate.startswith("08"):
        return (
            "Cannot connect to SQL Server",
            [
                "Check server/instance name, port, and network/VPN connectivity.",
                "If using ODBC Driver 18, TLS settings may matter (Encrypt/TrustServerCertificate).",
            ],
        )

    # Fallback: use message content
    if "login failed" in msg_l:
        return ("Authentication failed", ["Check credentials and authentication method."])

    return (
        "Cannot connect to SQL Server",
        [
            "Verify connection string (server/instance/database).",
            "Check network/VPN and firewall.",
        ],
    )


def _build_mssql_error_message(
    *,
    exc: BaseException,
    hints: list[str],
    pyodbc_ok: bool,
    drivers: list[str],
) -> str:
    sqlstate, msg = _pyodbc_sqlstate_and_msg(exc)
    msg = _redact_conn_secrets(msg or "")

    diagnostics_lines = [
        f"exe={sys.executable}",
        f"python={sys.version.split()[0]} ({'frozen' if _is_frozen_exe() else 'source'})",
        f"platform={sys.platform}",
        f"pyodbc={'OK' if pyodbc_ok else 'MISSING/FAILED'}",
        f"pyodbc.drivers()={', '.join(drivers) if drivers else '(none)'}",
    ]

    text = []
    text.append(msg if msg else "Connection failed.")
    text.append("")
    if sqlstate:
        text.append(t("CONN_SQLSTATE_LINE", sqlstate=sqlstate))
        text.append("")
    if hints:
        text.append(t("CONN_MOST_COMMON_CAUSES"))
        for hint in hints:
            text.append(f"- {hint}")
        text.append("")
    text.append(t("CONN_DIAGNOSTICS"))
    for line in diagnostics_lines:
        text.append(f"- {line}")
    return "\n".join(text)


def _unwrap_dbapi_original(exc: BaseException) -> BaseException:
    current = exc
    seen = {id(current)}
    while True:
        orig = getattr(current, "orig", None)
        if orig is None:
            return current
        if id(orig) in seen:
            return current
        seen.add(id(orig))
        current = orig


def _best_exception_message(exc: BaseException) -> str:
    args = getattr(exc, "args", None) or ()
    candidates = []
    if len(args) >= 2 and isinstance(args[1], str) and args[1].strip():
        candidates.append(args[1].strip())
    for arg in args:
        if isinstance(arg, str) and arg.strip():
            candidates.append(arg.strip())
    msg = str(exc).strip()
    if msg:
        candidates.append(msg)
    if not candidates:
        return ""
    return min(candidates, key=len)


def _package_version(package_name: str) -> Optional[str]:
    try:
        return importlib.metadata.version(package_name)
    except importlib.metadata.PackageNotFoundError:
        return None
    except Exception:
        return None


def _classify_postgresql_conn_error(exc: BaseException) -> tuple[str, list[str]]:
    msg = _best_exception_message(exc)
    msg_l = msg.lower()
    sqlstate = getattr(exc, "pgcode", None) or getattr(exc, "sqlstate", None)

    if sqlstate in ("28P01", "28000") or "password authentication failed" in msg_l:
        return (
            "Authentication failed",
            [
                "Check username/password.",
                "Verify the authentication method configured on the server.",
            ],
        )

    if sqlstate == "3D000" or ("does not exist" in msg_l and "database" in msg_l):
        return (
            "Database does not exist",
            [
                "Check database name.",
                "Ensure the database exists on the target server.",
            ],
        )

    if sqlstate and sqlstate.startswith("08"):
        return (
            "Cannot connect to PostgreSQL",
            [
                "Check host/port and network/VPN connectivity.",
                "Verify the server is running and accepting TCP connections.",
            ],
        )

    if any(
        phrase in msg_l
        for phrase in (
            "could not translate host name",
            "could not connect to server",
            "connection refused",
            "connection timed out",
            "no route to host",
            "network is unreachable",
            "timeout expired",
            "name or service not known",
            "timeout",
        )
    ):
        return (
            "Cannot connect to PostgreSQL",
            [
                "Check host/port and network/VPN connectivity.",
                "Verify the server is running and accepting TCP connections.",
            ],
        )

    return (
        "Cannot connect to PostgreSQL",
        [
            "Check host/port, database name, and network connectivity.",
            "Verify the server is running and accepting TCP connections.",
        ],
    )


def _classify_mysql_conn_error(exc: BaseException) -> tuple[str, list[str]]:
    msg = _best_exception_message(exc)
    msg_l = msg.lower()
    args = getattr(exc, "args", None) or ()
    errcode = args[0] if args and isinstance(args[0], int) else None

    if errcode == 1045 or "access denied" in msg_l:
        return (
            "Authentication failed",
            [
                "Check username/password.",
                "Verify the authentication plugin and permissions.",
            ],
        )

    if errcode == 1049 or "unknown database" in msg_l:
        return (
            "Database does not exist",
            [
                "Check database name.",
                "Ensure the database exists on the target server.",
            ],
        )

    if errcode in (2002, 2003, 2005) or "can't connect to mysql server" in msg_l or "unknown mysql server host" in msg_l:
        return (
            "Cannot connect to MySQL",
            [
                "Check host/port and network/VPN connectivity.",
                "Verify the server is running and accepting TCP connections.",
            ],
        )

    if errcode in (2006, 2013):
        return (
            "MySQL connection lost",
            [
                "Server closed the connection unexpectedly.",
                "Check network stability and server logs.",
            ],
        )

    if errcode == 2059:
        return (
            "Authentication plugin error",
            [
                "Server requires a different authentication plugin.",
                "Update client driver or change server user plugin.",
            ],
        )

    return (
        "Cannot connect to MySQL",
        [
            "Check host/port, database name, and network connectivity.",
            "Verify the server is running and accepting TCP connections.",
        ],
    )


def _classify_sqlite_conn_error(exc: BaseException) -> tuple[str, list[str]]:
    msg = _best_exception_message(exc)
    msg_l = msg.lower()

    if "unable to open database file" in msg_l:
        return (
            "Cannot open SQLite database file",
            [
                "Check file path and permissions.",
                "Ensure the directory exists and is writable.",
            ],
        )

    if "database is locked" in msg_l:
        return (
            "SQLite database is locked",
            [
                "Close other applications using the file.",
                "Retry after the lock is released.",
            ],
        )

    if "file is not a database" in msg_l:
        return (
            "Invalid SQLite database file",
            [
                "Verify the selected file is a valid SQLite database.",
                "Ensure the file is not corrupted.",
            ],
        )

    if "disk i/o error" in msg_l:
        return (
            "SQLite disk I/O error",
            [
                "Check disk health and permissions.",
                "Verify the storage device is available.",
            ],
        )

    return (
        "Cannot open SQLite database",
        [
            "Check file path and permissions.",
            "Ensure the file exists and is accessible.",
        ],
    )


def _build_connection_error_message(
    *,
    exc: BaseException,
    hints: list[str],
    diagnostics_lines: list[str],
    sqlstate: Optional[str] = None,
) -> str:
    msg = _best_exception_message(exc)
    msg = _redact_conn_secrets(msg or "")
    text = []
    text.append(msg if msg else "Connection failed.")
    text.append("")

    if sqlstate:
        text.append(t("CONN_SQLSTATE_LINE", sqlstate=sqlstate))
        text.append("")

    if hints:
        text.append(t("CONN_MOST_COMMON_CAUSES"))
        for hint in hints:
            text.append(f"- {hint}")
        text.append("")
    text.append(t("CONN_DIAGNOSTICS"))
    for line in diagnostics_lines:
        text.append(f"- {line}")
    return "\n".join(text)


def _format_connection_error(
    *,
    conn_type: str,
    exc: BaseException,
    details: dict,
) -> tuple[str, str]:
    if isinstance(exc, PasswordRequiredError):
        name = getattr(exc, "name", "")
        return (
            t("ERR_PASSWORD_REQUIRED_TITLE"),
            t("ERR_PASSWORD_REQUIRED_BODY", name=name),
        )
    base_diagnostics = [
        f"exe={sys.executable}",
        f"python={sys.version.split()[0]} ({'frozen' if _is_frozen_exe() else 'source'})",
        f"platform={sys.platform}",
        f"sqlalchemy={sqlalchemy.__version__}",
    ]
    conn_type = (conn_type or "").strip()
    details = details or {}

    if conn_type == "mssql_odbc":
        orig = _unwrap_dbapi_original(exc)
        pyodbc_ok = False
        drivers = []
        try:
            import pyodbc  # type: ignore

            pyodbc_ok = True
            drivers = list(pyodbc.drivers())
        except Exception:
            pyodbc_ok = False
            drivers = []

        driver = str(details.get("driver") or "").strip()
        server = str(details.get("server") or "").strip()
        database = str(details.get("database") or "").strip()
        username = str(details.get("username") or "").strip()
        trusted = bool(details.get("trusted"))
        encrypt = bool(details.get("encrypt", True))
        trust_cert = bool(details.get("trust_server_certificate", True))

        conn_parts = []
        if driver:
            conn_parts.append(f"DRIVER={{{driver}}}")
        if server:
            conn_parts.append(f"SERVER={server}")
        if database:
            conn_parts.append(f"DATABASE={database}")
        if trusted:
            conn_parts.append("Trusted_Connection=yes")
        elif username:
            conn_parts.append(f"UID={username}")
        if encrypt:
            conn_parts.append("Encrypt=yes")
        if trust_cert:
            conn_parts.append("TrustServerCertificate=yes")

        conn_str = ";".join(conn_parts)

        title, hints = _classify_mssql_conn_error(
            exc=orig,
            conn_str=conn_str,
            pyodbc_ok=pyodbc_ok,
            drivers=drivers,
        )
        msg = _build_mssql_error_message(
            exc=orig,
            hints=hints,
            pyodbc_ok=pyodbc_ok,
            drivers=drivers,
        )
        return title, msg

    if conn_type == "postgresql":
        orig = _unwrap_dbapi_original(exc)
        title, hints = _classify_postgresql_conn_error(orig)
        pg_sqlstate = getattr(orig, "pgcode", None) or getattr(orig, "sqlstate", None)
        host = str(details.get("host") or "").strip()
        port = str(details.get("port") or "5432").strip() or "5432"
        database = str(details.get("database") or "").strip()
        target = f"{host}:{port}/{database}".strip(":")
        driver_pkg = None
        driver_version = None
        for candidate in ("psycopg", "psycopg2", "psycopg2-binary"):
            version = _package_version(candidate)
            if version:
                driver_pkg = candidate
                driver_version = version
                break
        diagnostics = base_diagnostics + [
            f"driver={driver_pkg or '(unknown)'} {driver_version or ''}".strip(),
            f"target={target}" if target else "target=(not specified)",
        ]
        msg = _build_connection_error_message(
            exc=orig,
            hints=hints,
            diagnostics_lines=diagnostics,
            sqlstate=pg_sqlstate,
        )
        return title, msg

    if conn_type == "mysql":
        orig = _unwrap_dbapi_original(exc)
        title, hints = _classify_mysql_conn_error(orig)
        host = str(details.get("host") or "").strip()
        port = str(details.get("port") or "3306").strip() or "3306"
        database = str(details.get("database") or "").strip()
        target = f"{host}:{port}/{database}".strip(":")
        driver_pkg = None
        driver_version = None
        for candidate in ("pymysql", "mysqlclient", "mysql-connector-python"):
            version = _package_version(candidate)
            if version:
                driver_pkg = candidate
                driver_version = version
                break
        diagnostics = base_diagnostics + [
            f"driver={driver_pkg or '(unknown)'} {driver_version or ''}".strip(),
            f"target={target}" if target else "target=(not specified)",
        ]
        msg = _build_connection_error_message(
            exc=orig,
            hints=hints,
            diagnostics_lines=diagnostics,
        )
        return title, msg

    if conn_type == "sqlite":
        orig = _unwrap_dbapi_original(exc)
        title, hints = _classify_sqlite_conn_error(orig)
        path = str(details.get("path") or "").strip()
        abs_path = os.path.abspath(path) if path else ""
        diagnostics = base_diagnostics + [
            f"sqlite3={sqlite3.sqlite_version}",
            f"path={abs_path}" if abs_path else "path=(not specified)",
        ]
        msg = _build_connection_error_message(
            exc=orig,
            hints=hints,
            diagnostics_lines=diagnostics,
        )
        return title, msg

    title = "Connection failed"
    msg = _build_connection_error_message(
        exc=exc,
        hints=[
            "Verify connection details and network connectivity.",
            "Check logs for more details.",
        ],
        diagnostics_lines=base_diagnostics,
    )
    return title, msg


def _redact_conn_secrets(text: str) -> str:
    if not text:
        return text
    patterns = [
        r"(?i)(pwd|password)\s*=\s*[^;]*",
        r"(?i)(user id|uid)\s*=\s*[^;]*",
        r"(?i)(access_token)\s*=\s*[^;]*",
    ]
    out = text
    for pattern in patterns:
        out = re.sub(pattern, lambda m: m.group(0).split("=")[0] + "=***", out)
    return out


# =========================
# I18N (EN as source)
# =========================
I18N: dict[str, dict[str, str]] = {
    "en": {
        # GUI
        "APP_TITLE": "KKr Query Runner",
        "BTN_RUN": "Run",
        "BTN_EXPORT": "Export",
        "BTN_BROWSE": "Browse...",
        "BTN_CANCEL_RUN": "Stop",
        "BTN_ODBC_DIAGNOSTICS": "ODBC diagnostics",
        "LBL_SQL_FILE": "SQL file",
        "LBL_DB": "Database",
        "LBL_OUTPUT": "Output",
        "LBL_SAVE_AS": "Save as (optional):",
        "BTN_SAVE_AS": "Save as...",
        "BTN_CLEAR": "Clear",
        "LBL_SAVE_AS_HINT": "Leave empty to save into {dir} with the default file name.",
        "INFO_SAVE_AS_EXT_FIXED_TITLE": "Adjusted extension",
        "INFO_SAVE_AS_EXT_FIXED_BODY": "The chosen name had a different extension. It will be saved as:\n{path}",
        "LBL_LANGUAGE": "Language",
        "CHK_ARCHIVE_SQL": "Archive SQL (save query + metadata)",
        "LBL_DB_TIMEOUT_MIN": "DB timeout (minutes) - execution + fetch",
        "LBL_EXPORT_TIMEOUT_MIN": "Export timeout (minutes) - XLSX/CSV generation",
        "LBL_TIMEOUT_NOTE": "0 = no limit. Defaults: 3 minutes each.",
        "CHK_SQL_HIGHLIGHT": "Syntax highlighting (Chlorophyll)",
        "LBL_SQL_DIALECT": "SQL dialect",
        "MSG_DONE": "Done.",
        "MSG_CANCEL_REQUESTED": "Stopping...",
        "MSG_CONFIRM_CANCEL_AND_EXIT": "Export is running. Cancel and exit?",
        "ERR_TITLE": "Error",
        "WARN_TITLE": "Warning",
        "APP_TITLE_FULL": "KKr SQL to XLSX/CSV",
        "BROWSER_OPEN_FAIL_TITLE": "Unable to open browser",
        "BROWSER_OPEN_FAIL_BODY": (
            "Could not automatically open the link.\n"
            "Copy and open it manually:\n\n{url}"
        ),
        "BROWSER_OPEN_FAIL_ERROR_BODY": (
            "Could not open the browser.\n\n{error}\n\nLink:\n{url}"
        ),
        "ODBC_DIAGNOSTICS_TITLE": "ODBC diagnostics",
        "ODBC_DIAGNOSTICS_LABEL": "ODBC diagnostics:",
        "ERR_PG_MISSING_TITLE": "Missing PostgreSQL library",
        "ERR_PG_MISSING_BODY": (
            "Cannot connect to PostgreSQL. Required Python library (e.g. psycopg2) "
            "is not installed. Install the missing library and try again."
        ),
        "ERR_QUERY_TIMEOUT": "DB timeout exceeded (>{minutes} min). Query was interrupted.",
        "ERR_EXPORT_TIMEOUT": "Export timeout exceeded (>{minutes} min). Export was interrupted.",
        "ERR_CANCELLED_BY_USER": "Cancelled by user.",
        "ERR_MYSQL_MISSING_TITLE": "Missing MySQL library",
        "ERR_MYSQL_MISSING_BODY": (
            "Cannot connect to MySQL. Required Python library (e.g. pymysql) "
            "is not installed. Install the missing library and try again."
        ),
        "ERR_ORACLE_MISSING_TITLE": "Missing Oracle library",
        "ERR_ORACLE_MISSING_BODY": (
            "Cannot connect to Oracle. Required Python library (e.g. cx_Oracle) "
            "is not installed. Install the missing library and try again."
        ),
        "INFO_SQL_HIGHLIGHT_TITLE": "Syntax highlighting",
        "INFO_SQL_HIGHLIGHT_BODY": (
            "Chlorophyll is not available. Install it to enable SQL highlighting."
        ),
        "MSG_UI_TRUNCATED": (
            "...\n(Trimmed in UI, full details in kkr-query2xlsx.log)"
        ),
        "CONN_SQLSTATE_LINE": "SQLSTATE: {sqlstate}",
        "CONN_MOST_COMMON_CAUSES": "Most common causes:",
        "CONN_DIAGNOSTICS": "Diagnostics:",
        "CONSOLE_AVAILABLE_FILES": "Available SQL query files:",
        "CONSOLE_CUSTOM_PATH": "0: [Custom path]",
        "CONSOLE_PROMPT_SELECT": (
            "Please enter the number of the SQL query file to execute "
            "(0 for custom path, 1-{max_idx}): "
        ),
        "CONSOLE_PROMPT_CUSTOM_PATH": "Please enter full path to the .sql file: ",
        "CONSOLE_FILE_NOT_FOUND": "File does not exist. Please try again.",
        "CONSOLE_SELECT_RANGE": "Please enter a number between 0 and {max_idx}.",
        "CONSOLE_INVALID_INPUT": "Invalid input. Please enter a number.",
        "CONSOLE_NO_QUERIES": "No SQL query file paths found in queries.txt",
        "CONSOLE_PROMPT_FORMAT": "Please enter the desired output format (xlsx or csv): ",
        "CONSOLE_INVALID_FORMAT": "Invalid input. Please enter 'xlsx' or 'csv'.",
        "CONSOLE_AVAILABLE_CSV_PROFILES": "Available CSV profiles:",
        "CONSOLE_DEFAULT_MARKER": " (default)",
        "CONSOLE_PROMPT_CSV_PROFILE": (
            "Enter CSV profile number to use or press Enter to use the default: "
        ),
        "CONSOLE_INVALID_SELECTION": "Invalid selection. Please try again.",
        "CONSOLE_SAVED_PATH": "Query results have been saved to: {path}",
        "CONSOLE_NO_ROWS": "The query did not return any rows.",
        "CONSOLE_SQL_TIME": "Data fetch time (SQL): {seconds:.2f} seconds",
        "CONSOLE_EXPORT_TIME": "Export time ({fmt}): {seconds:.2f} seconds",
        "CONSOLE_TOTAL_TIME": "Total time: {seconds:.2f} seconds",
        "CONSOLE_PROMPT_PASSWORD": "Password for connection '{name}': ",
        "CLI_DIAG_ODBC_HELP": "Print ODBC diagnostics and exit.",
        "CLI_SELF_TEST_HELP": "Run internal smoke tests and exit.",
        "DEFAULT_MSSQL_NAME": "Default MSSQL",
        "FRAME_MSSQL": "MSSQL (ODBC)",
        "LBL_ODBC_DRIVER": "ODBC driver",
        "LBL_SERVER": "Server",
        "LBL_DATABASE_NAME": "Database name",
        "LBL_LOGIN": "Login",
        "LBL_PASSWORD": "Password",
        "CHK_REMEMBER_PASSWORD": "Remember password",
        "CHK_TRUSTED": "Windows authentication (Trusted_Connection)",
        "CHK_ENCRYPT": "Encrypt",
        "CHK_TRUST_CERT": "TrustServerCertificate",
        "FRAME_POSTGRES": "PostgreSQL",
        "LBL_HOST": "Host",
        "LBL_PORT": "Port",
        "LBL_DATABASE": "Database",
        "LBL_USER": "User",
        "FRAME_MYSQL": "MySQL",
        "FRAME_SQLITE": "SQLite",
        "FILETYPE_SQLITE": "SQLite",
        "FILETYPE_ALL": "All files",
        "TITLE_SELECT_SQLITE": "Select existing SQLite database",
        "TITLE_CREATE_SQLITE": "Create new SQLite database",
        "ASK_CREATE_SQLITE": "Create a new database?",
        "ASK_CREATE_SQLITE_BODY": (
            "No existing database selected. Do you want to create a new one?"
        ),
        "LBL_SQLITE_PATH": "File path",
        "BTN_SELECT": "Select",
        "ERR_DATA_TITLE": "Data error",
        "ERR_FILL_ODBC": "Fill in: driver, server, and database name.",
        "ERR_LOGIN_REQUIRED": (
            "Provide login and password or select Windows authentication "
            "(Trusted_Connection)."
        ),
        "ERR_FILL_PG": "Fill in: host, database name, and user.",
        "ERR_FILL_MYSQL": "Fill in: host, database name, and user.",
        "ERR_FILL_SQLITE": "Provide the SQLite database file path.",
        "PROMPT_PASSWORD_TITLE": "Password required",
        "PROMPT_PASSWORD_BODY": "Enter password for connection:\n{name}",
        "ERR_PASSWORD_REQUIRED_TITLE": "Password required",
        "ERR_PASSWORD_REQUIRED_BODY": (
            "Password was not provided, so the connection was not attempted.\n\n"
            "What to do:\n"
            "- Enter the password when prompted.\n"
            "- Or click 'Manage connections...' and set the password for: {name}\n"
            "- If this is Windows authentication, enable 'Windows authentication (Trusted_Connection)' "
            "and leave login/password empty."
        ),
        "WARN_REMEMBER_PASSWORD_TITLE": "Store password?",
        "WARN_REMEMBER_PASSWORD_BODY": (
            "The password will be saved in plain text in:\n{path}\n\n"
            "Keep this file very safe.\n\nContinue?"
        ),
        "LBL_CONN_NAME": "Connection name",
        "LBL_DB_TYPE": "Database type",
        "ERR_INVALID_CONN_TYPE": "Invalid connection type.",
        "ERR_CONN_NAME_REQUIRED": "Connection name cannot be empty.",
        "ERR_CONN_NAME_EXISTS": "Connection name already exists. Choose another name.",
        "INFO_CONN_SAVED_TITLE": "Saved",
        "INFO_CONN_TEST_OK_TITLE": "Connection works",
        "INFO_CONN_TEST_OK_BODY": "Connection test succeeded.",
        "INFO_CONN_SAVED_BODY": "Connection has been saved.",
        "BTN_SAVE": "Save",
        "BTN_SAVE_NO_TEST": "Save without test",
        "BTN_CANCEL": "Cancel",
        "CSV_PROFILE_TITLE": "CSV profile details",
        "CSV_PROFILE_NAME": "Profile name:",
        "CSV_PROFILE_ENCODING": "Encoding:",
        "CSV_PROFILE_DELIMITER": "Field delimiter:",
        "CSV_PROFILE_DELIM_REPLACE": "Replace delimiter in values:",
        "CSV_PROFILE_DECIMAL": "Decimal separator:",
        "CSV_PROFILE_LINE_END": "Line terminator:",
        "CSV_PROFILE_QUOTECHAR": "Quote character:",
        "CSV_PROFILE_QUOTING": "Quoting:",
        "CSV_PROFILE_ESCAPECHAR": "Escape character:",
        "CSV_PROFILE_ESCAPE_HINT": "(escape char; empty = quoting)",
        "CSV_PROFILE_DOUBLEQUOTE": "Double quote in fields",
        "CSV_PROFILE_DATE_FORMAT": "Date format:",
        "CSV_PROFILE_FIELD_SEPARATOR": "Field separator:",
        "CSV_PROFILE_WARNING_EMPTY": "Profile name cannot be empty.",
        "CSV_PROFILE_WARNING_EXISTS": "A profile with this name already exists.",
        "CSV_PROFILE_WARNING_SELECT": "Select a profile from the list.",
        "CSV_PROFILE_INFO_BUILTIN": "Built-in profiles cannot be deleted.",
        "CSV_PROFILE_WARNING_MIN_ONE": "There must be at least one CSV profile.",
        "CSV_PROFILE_INFO_SAVED_TITLE": "Information",
        "CSV_PROFILE_INFO_SAVED_BODY": "CSV profiles saved.",
        "CSV_PROFILE_CONFIRM_DELETE": (
            "Delete selected profile?\n\nThis cannot be undone."
        ),
        "BTN_SAVE_AS_NEW": "Save as new",
        "BTN_UPDATE_PROFILE": "Update profile",
        "BTN_DELETE": "Delete",
        "BTN_SET_DEFAULT": "Set as default",
        "BTN_CLOSE_SAVE": "Close and save",
        "ERR_QUERY_TITLE": "Query error",
        "BTN_COPY": "Copy",
        "BTN_CLOSE": "Close",
        "BTN_COPY_ALL": "Copy all",
        "BTN_COPY_SUMMARY": "Copy summary",
        "BTN_COPY_SQL": "Copy SQL",
        "BTN_OPEN_SQL": "Open SQL",
        "BTN_OPEN_LOG": "Open log",
        "ERR_TAB_SUMMARY": "Summary",
        "ERR_TAB_SQL": "SQL",
        "ERR_TAB_DETAILS": "Details",
        "PROGRESS_WORKING_TITLE": "Working...",
        "PROGRESS_STARTING": "Starting...",
        "PROGRESS_STEP": "Step:",
        "PROGRESS_ELAPSED": "Elapsed:",
        "PROGRESS_TITLE": "Export in progress",
        "PROGRESS_GETTING_DATA": "Getting data from database...",
        "PROGRESS_EXPORTING_XLSX_TEMPLATE": "Exporting to XLSX (template)...",
        "PROGRESS_EXPORTING_XLSX": "Exporting to XLSX...",
        "PROGRESS_EXPORTING_CSV": "Exporting to CSV...",
        "ERR_NO_CONNECTION_TITLE": "No connection",
        "ERR_NO_CONNECTION_BODY": "No saved connections. Create and save a new connection.",
        "ERR_NO_CONNECTION_DELETE": "No connection to delete.",
        "ASK_DELETE_CONNECTION_TITLE": "Delete connection",
        "ASK_DELETE_CONNECTION_BODY": "Are you sure you want to delete connection {name}?",
        "TITLE_EDIT_SECURE": "Edit secure.txt",
        "INFO_SECURE_SAVED_TITLE": "Saved",
        "INFO_SECURE_SAVED_BODY": "Updated secure.txt content.",
        "ERR_SECURE_SAVE_TITLE": "Save error",
        "ERR_SECURE_SAVE_BODY": (
            "Failed to save secure.txt.\n\nTechnical details:\n{error}"
        ),
        "TITLE_SELECT_SQL": "Select SQL file",
        "FILETYPE_SQL": "SQL files",
        "TITLE_SELECT_TEMPLATE": "Select XLSX template file",
        "FILETYPE_EXCEL": "Excel files",
        "FILETYPE_CSV": "CSV files",
        "ERR_TEMPLATE_TITLE": "Template error",
        "ERR_TEMPLATE_SHEETS": (
            "Cannot read sheets from the template file.\n\nTechnical details:\n{error}"
        ),
        "CSV_DEFAULT_PROFILE_LABEL": "Default CSV profile: {name}",
        "TITLE_EDIT_QUERIES": "Edit queries.txt",
        "TITLE_ADD_SQL_FILES": "Add SQL files",
        "WARN_SKIPPED_FILES_TITLE": "Skipped files",
        "WARN_SKIPPED_FILES_BODY": (
            "Some selected files do not have the .sql extension and were skipped:\n\n"
            "{files}{more}"
        ),
        "INFO_ALREADY_LISTED": "Selected files are already on the list.",
        "TITLE_EDIT_QUERY_PATH": "Edit query path",
        "PROMPT_EDIT_QUERY_PATH": "Edit query path:",
        "WARN_INVALID_SQL_FILE": "The entry must point to a .sql file.",
        "WARN_FILE_MISSING_TITLE": "Warning",
        "WARN_FILE_MISSING_BODY": (
            "File does not exist (or is temporarily unavailable).\n"
            "Saving the path anyway, but make sure it is correct:\n\n{path}"
        ),
        "INFO_SELECT_ENTRY_DELETE": "Select an entry to delete.",
        "ERR_QUERIES_SAVE_TITLE": "Save error",
        "ERR_QUERIES_SAVE_BODY": (
            "Cannot save queries.txt.\n\nTechnical details:\n{error}"
        ),
        "BTN_ADD_FILES": "Add files...",
        "BTN_REMOVE_SELECTED": "Remove selected",
        "TITLE_SELECT_REPORT": "Select report from list",
        "ERR_NO_REPORTS": "No reports in queries.txt",
        "WARN_NO_REPORT_SELECTED": "No report selected.",
        "ERR_NO_SQL_SELECTED": "No SQL file selected.",
        "ERR_SQL_NOT_FOUND": "Selected SQL file does not exist.",
        "ERR_NEED_CONNECTION": "Create a database connection before running the report.",
        "ERR_TEMPLATE_ONLY_XLSX": "Template can only be used for XLSX.",
        "ERR_TEMPLATE_NOT_SELECTED": "No template file selected.",
        "ERR_TEMPLATE_SHEET_NOT_SELECTED": "No template worksheet selected.",
        "MSG_RUNNING": "Running query and export. Please wait...",
        "MSG_SAVED_DETAILS": (
            "Saved: {path}\n"
            "Rows: {rows}\n"
            "SQL time: {sql_time_hms} ({sql_time:.2f} s)\n"
            "Export time: {export_time_hms} ({export_time:.2f} s)\n"
            "Total time: {total_time_hms} ({total_time:.2f} s)"
        ),
        "MSG_SAVED_DETAILS_CSV": "CSV profile: {profile}",
        "MSG_NO_ROWS": (
            "Query returned no rows.\nSQL time: {sql_time_hms} ({sql_time:.2f} s)"
        ),
        "ERR_XLSX_TOO_LARGE": (
            "XLSX export skipped because the result would exceed Excel limits. "
            "Result: {rows} rows, {cols} columns. "
            "Sheet would require: {sheet_rows} rows, {sheet_cols} columns "
            "(max {max_rows} rows, {max_cols} columns). "
            "Please export to CSV instead."
        ),
        "ERR_EXPORT": "Export error. Full details in log.",
        "FRAME_DB_CONNECTION": "Database connection",
        "FRAME_ADVANCED": "Advanced",
        "LBL_CONNECTION": "Connection:",
        "BTN_MANAGE_CONNECTIONS": "Manage connections...",
        "BTN_EDIT_CONNECTION": "Edit connection",
        "BTN_NEW_CONNECTION": "New connection",
        "BTN_DUPLICATE_CONNECTION": "Duplicate",
        "BTN_TEST_CONNECTION": "Test connection",
        "BTN_DELETE_CONNECTION": "Delete connection",
        "BTN_EDIT_SECURE": "Edit secure.txt",
        "BTN_SET_CURRENT_CONNECTION": "Set as current",
        "CONNECTIONS_MANAGER_TITLE": "Connections",
        "WARN_SELECT_CONNECTION": "Select a connection from the list.",
        "FRAME_SQL_SOURCE": "SQL query source",
        "FRAME_OUTPUT_FORMAT": "Output format",
        "FRAME_TEMPLATE_OPTIONS": "XLSX template options (GUI)",
        "FRAME_RESULTS": "Result and actions",
        "LBL_SELECTED_SQL": "Selected SQL file:",
        "BTN_SELECT_SQL": "Select SQL file",
        "BTN_SELECT_FROM_LIST": "Select from report list",
        "BTN_EDIT_QUERIES": "Edit queries.txt",
        "BTN_PASTE_SQL": "Paste SQL",
        "TITLE_PASTE_SQL": "Paste SQL",
        "LBL_REPORT_NAME": "Report name (required)",
        "LBL_PASTE_SQL": "SQL (paste here)",
        "BTN_USE_SQL": "Use SQL",
        "ERR_INVALID_REPORT_NAME_TITLE": "Invalid report name",
        "ERR_REPORT_NAME_REQUIRED": "Report name is required (cannot be empty).",
        "ERR_REPORT_NAME_NO_FOLDERS": "Report name must be a filename only (no folders).",
        "ERR_REPORT_NAME_INVALID_CHARS": "Report name contains invalid characters: {chars}",
        "ERR_REPORT_NAME_CONTROL_CHARS": "Report name contains control characters.",
        "ERR_REPORT_NAME_TRAILING_DOT_SPACE": "Report name cannot end with a space or a dot.",
        "ERR_REPORT_NAME_RESERVED": "Report name '{name}' is reserved on Windows.",
        "ERR_REPORT_NAME_TOO_LONG": "Report name is too long (max {max_len} characters).",
        "ERR_EMPTY_SQL_TITLE": "Empty SQL",
        "ERR_EMPTY_SQL_BODY": "SQL cannot be empty.",
        "ERR_INVALID_SQL_FILE_TITLE": "Invalid SQL file",
        "ERR_NO_FILE_SELECTED": "No file selected.",
        "ERR_SQLFILE_IS_SQLITE_EXT": (
            "Selected a SQLite database file. This is not a .sql file with queries."
        ),
        "ERR_SQLFILE_IS_SPREADSHEET": (
            "This looks like a spreadsheet/data file ({ext}). "
            "Please select a .sql text file containing SQL statements."
        ),
        "ERR_SQLFILE_IS_ZIP": (
            "This file is a ZIP container (starts with 'PK'). "
            "It is not a plain .sql text file."
        ),
        "ERR_SQLFILE_IS_OLD_OFFICE": (
            "This looks like an old Microsoft Office binary file (e.g. .xls). "
            "Please select a .sql text file."
        ),
        "ERR_SQLFILE_IS_BINARY": (
            "This file looks like a binary file, not plain text SQL. "
            "Please select a .sql text file."
        ),
        "ERR_CANNOT_OPEN_FILE": "Cannot open file: {error}",
        "LBL_SQL_PASTED": "Pasted SQL:",
        "ERR_NO_SQL_SOURCE": "Select a SQL file or paste SQL first.",
        "LBL_CSV_PROFILE": "CSV profile:",
        "BTN_MANAGE_CSV_PROFILES": "Manage CSV profiles",
        "CHK_USE_TEMPLATE": "Use template file (XLSX only, GUI only)",
        "LBL_TEMPLATE_FILE": "Template file:",
        "BTN_SELECT_TEMPLATE": "Select template",
        "LBL_TEMPLATE_SHEET": "Worksheet:",
        "LBL_TEMPLATE_START_CELL": "Start cell:",
        "CHK_INCLUDE_HEADERS": "Write headers (column names) to worksheet",
        "BTN_START": "Start",
        "BTN_REPORT_ISSUE": "Report issue / suggestion",
        "BTN_CHECK_UPDATES": "Check updates",
        "BTN_HELP": "Help / README",
        "BTN_OPEN_LOGS": "Open logs folder",
        "LBL_PROJECT_PAGE": "Project page",
        "LBL_EXPORT_INFO": "Export info:",
        "BTN_OPEN_FILE": "Open file",
        "BTN_OPEN_FOLDER": "Open folder",
        "LBL_ERRORS_SHORT": "Errors (summary):",
        "README_WINDOW_TITLE": "kkr-query2xlsx — README",
        "UPD_TITLE": "Updates",
        "UPD_CHECKING": "Checking for updates...",
        "UPD_CHECK_FAILED": "Could not check updates.",
        "UPD_ERR_NETWORK": "Could not reach GitHub (network/DNS problem). Check your connection and try again.",
        "UPD_ERR_TIMEOUT": "GitHub did not respond in time (timeout). Try again in a moment.",
        "UPD_ERR_HTTP": "GitHub returned HTTP {status} while checking updates.",
        "UPD_ERR_RATE_LIMITED": "GitHub returned HTTP {status}. It looks like rate limiting. Try again after: {retry_at}.",
        "UPD_ERR_JSON": "Received an invalid response from GitHub (JSON parse error).",
        "UPD_NO_UPDATE": (
            "No Release updates available. Current: {current}, latest Release: {latest}."
        ),
        "UPD_UPDATE_AVAILABLE": "Release update available: {current} → {latest}.",
        "UPD_PROMPT_INSTALL": "Install the update now?",
        "UPD_GIT_MODE": (
            "This run comes from a Git checkout. This check looks only for official Releases. "
            "To update the code, run:\n\n{command}"
        ),
        "UPD_GIT_STATUS_MATCH": "Git status: local {local}, remote main {remote}.",
        "UPD_GIT_STATUS_AHEAD": "Git status: local {local}, remote main {remote} (update available).",
        "UPD_GIT_STATUS_LOCAL_AHEAD": "Git status: local {local} is ahead of remote main {remote}.",
        "UPD_GIT_STATUS_DIVERGED": "Git status: local {local} and remote main {remote} have diverged.",
        "UPD_GIT_STATUS_DIFFERENT": "Git status: local {local}, remote main {remote} (different).",
        "UPD_GIT_STATUS_DIFFERENT_UNVERIFIED": (
            "Git status: local {local}, remote main {remote} (different) "
            "(could not compare — no connection or the local commit does not exist on GitHub)."
        ),
        "UPD_SOURCE_MODE": (
            "This run comes from source files without Git. This check looks only for official "
            "Releases. To update, download a new ZIP from Releases or clone the repository."
        ),
        "UPD_UPDATER_MISSING": "Updater is missing. Download the latest ZIP manually.",
        "UPD_START_FAILED": "Failed to start updater.",
        "STATUS_NO_CONNECTION": "No connection. Create a new connection.",
        "STATUS_CONNECTION_ERROR": "Connection error. Create a new connection.",
        "STATUS_PASSWORD_REQUIRED": "Password required.",
        "ERR_CONNECTION_TITLE": "Connection error",
        "ERR_CONNECTION_BODY": (
            "Failed to establish a connection.\n\nTechnical details:\n{error}"
        ),
        "STATUS_CONNECTED": "Connected to {name} ({type}).",
        "INFO_CONNECTION_OK_TITLE": "Connection works",
        "INFO_CONNECTION_OK_BODY": "Connection {name} succeeded.",
        "ERR_NO_SECURE_CONFIG": "No connection configuration.",
        "ERR_SELECT_TEMPLATE": "Select template file",
        "ERR_GENERIC": "Error",
        "ERR_TEMPLATE_MISSING_SHEET": "Worksheet '{sheet}' does not exist in template file.",
        "CLI_DESC": "Run SQL files and export results to XLSX/CSV.",
        "CLI_LANG_HELP": "UI language (en/pl).",
        "CLI_CONSOLE_HELP": "Run in console mode.",
        "CLI_OUTPUT_HELP": (
            "Optional output file or directory. A directory may be existing or a new path "
            "without an extension. Use a trailing slash/backslash to force directory "
            "interpretation."
        ),
        "CLI_NO_CONNECTIONS": (
            "No saved connections. Create a connection in GUI mode to run console."
        ),
        "CLI_CONNECTION_FAIL": "Failed to create connection. Full details in log.",
        "MENU_LANGUAGE": "Language",
        "ERR_FILE_PATH": "Path: {path}",
        "ERR_FILE_LOCKED": (
            "The output file already exists and may be open in another app (e.g. Excel). "
            "Close it and try again.{path}"
        ),
        "ERR_NO_WRITE_PERMISSION": (
            "No permission to write the output file or the path is unavailable. "
            "Check the file location."
        ),
        "ERR_SQL_SOURCE": "SQL source:",
        "ERR_DB_MESSAGE": "Database message (excerpt):",
        "ERR_SQL_PREVIEW": "SQL (start):",
        "ERR_FULL_LOG": "Full error saved in kkr-query2xlsx.log",
        "ERR_HINT_LABEL": "Hint:",
        "DB_TYPE_MSSQL": "SQL Server (ODBC)",
        "DB_TYPE_PG": "PostgreSQL",
        "DB_TYPE_MYSQL": "MySQL",
        "DB_TYPE_SQLITE": "SQLite (.db file)",
        "TITLE_CONN_DIALOG_EDIT": "Editing: {name}",
        "TITLE_CONN_DIALOG_NEW": "New connection",
        "TITLE_CONN_DIALOG_DUPLICATE": "Duplicating: {name}",
        "CONN_DIALOG_HINT_EDIT": "Editing connection: {name}",
        "CONN_DIALOG_HINT_NEW": "Create a new connection by entering its details.",
        "CONN_DIALOG_HINT_DUPLICATE": "Duplicating connection: {name}",
        "CONN_DUPLICATE_SUFFIX": "(copy)",
        "INFO_CONN_SAVED_NO_TEST": (
            "Connection saved without testing.\n"
            "Use the \"Test connection\" button to verify it."
        ),
        "INFO_ICON": "i",
        "BTN_OK": "OK",
        "TITLE_DATA_DIR_NOTICE": "Data folder",
        "MSG_DATA_DIR_NOTICE": (
            "This app stores its data in:\n\n"
            "{data_dir}\n\n"
            "This includes config files (secure.txt, queries.txt, kkr-query2xlsx.json) "
            "and folders like generated_reports/ and sql_archive/."
        ),
        "BTN_DONT_SHOW_AGAIN": "Don't show again",
        "FORMAT_XLSX": "XLSX",
        "FORMAT_CSV": "CSV",
        "CSV_PROFILE_DIALOG_TITLE": "CSV profiles",
        "CSV_PROFILE_DEFAULT_SUFFIX": "(default)",
        "CSV_PROFILE_BUILTIN_SUFFIX": "[built-in]",
        "CSV_PROFILE_INVALID_DATE": (
            "The provided date format is invalid. Use strftime syntax."
        ),
        "CSV_PROFILE_DATE_DEFAULT": "Pandas default format (example: {example})",
        "CSV_PROFILE_DATE_INVALID": (
            "Invalid date format (use strftime syntax, e.g. %Y-%m-%d)."
        ),
        "CSV_PROFILE_DATE_PREVIEW": "Current time in this format: {example}",
        "CSV_PROFILE_BUILTIN_NOTICE": (
            "Built-in profile: changes cannot be saved or deleted. "
            "Use Save as new to create your own variant."
        ),
        "CSV_PROFILE_NAME_RESERVED": (
            "This name is reserved for a built-in profile. Choose another name."
        ),
        "CSV_PROFILE_NO_SELECTION_TITLE": "No profile",
        "CSV_PROFILE_BUILTIN_OVERWRITE": (
            "You cannot overwrite a built-in profile. Change the name and save as a new profile."
        ),
        "CSV_PROFILE_UNSAVED_TITLE": "Unsaved changes",
        "CSV_PROFILE_UNSAVED_BODY": (
            "You have unsaved CSV profile changes. Save before closing?"
        ),
        "WARN_SKIPPED_FILES_MORE": "\n\n(+ {count} more)",
        "CSV_HELP_NAME_TITLE": "Profile name",
        "CSV_HELP_NAME_BODY": (
            "Any unique name that helps select the profile, e.g. "
            "\"UTF-8 (comma)\" or \"Windows-1250 (semicolon)\"."
        ),
        "CSV_HELP_ENCODING_TITLE": "Encoding",
        "CSV_HELP_ENCODING_BODY": (
            "Character encoding used in the CSV file. Default is UTF-8; for older "
            "Excel sheets you can use windows-1250."
        ),
        "CSV_HELP_DELIMITER_TITLE": "Field delimiter",
        "CSV_HELP_DELIMITER_BODY": (
            "Character separating columns. Usually comma (,) or semicolon (;), "
            "depending on regional settings."
        ),
        "CSV_HELP_DELIM_REPLACE_TITLE": "Replace delimiter in values",
        "CSV_HELP_DELIM_REPLACE_BODY": (
            "Optionally replace the delimiter character inside values (e.g. semicolon to comma). "
            "Useful when the import system does not handle escaping. Note: replacement is global "
            "for all text fields (including JSON/IDs)."
        ),
        "CSV_HELP_DECIMAL_TITLE": "Decimal separator",
        "CSV_HELP_DECIMAL_BODY": (
            "Character separating integer and fractional parts. Dot (.) for English format, "
            "comma (,) for Polish."
        ),
        "CSV_HELP_LINETERM_TITLE": "Line terminator",
        "CSV_HELP_LINETERM_BODY": (
            "Default is \\n. For full Windows compatibility you can use \\r\\n. "
            "Change only if required by import."
        ),
        "CSV_HELP_QUOTECHAR_TITLE": "Quote character",
        "CSV_HELP_QUOTECHAR_BODY": (
            "Usually \". Used to wrap fields that require quoting (e.g. containing the delimiter)."
        ),
        "CSV_HELP_QUOTING_TITLE": "Quoting strategy",
        "CSV_HELP_QUOTING_BODY": (
            "minimal – only when needed (recommended), all – always, nonnumeric – for text, "
            "none – no quoting (requires escapechar)."
        ),
        "CSV_HELP_ESCAPECHAR_TITLE": "Escape character",
        "CSV_HELP_ESCAPECHAR_BODY": (
            "Escape character used when quoting=none or fields may include the delimiter. "
            "Leave empty when using standard quoting."
        ),
        "CSV_HELP_DOUBLEQUOTE_TITLE": "Double quote",
        "CSV_HELP_DOUBLEQUOTE_BODY": (
            "When enabled, internal \" in a field becomes \"\". Keep enabled unless the import "
            "system requires otherwise."
        ),
        "CSV_HELP_DATE_FORMAT_TITLE": "Date format",
        "CSV_HELP_DATE_FORMAT_BODY": (
            "Optional strftime pattern, e.g. %Y-%m-%d or %d.%m.%Y. Leave empty to use Pandas defaults."
        ),
    },
    "pl": {
        # GUI
        "APP_TITLE": "KKr Runner Zapytań",
        "BTN_RUN": "Uruchom",
        "BTN_EXPORT": "Eksportuj",
        "BTN_BROWSE": "Wybierz...",
        "BTN_CANCEL_RUN": "Wstrzymaj wykonywanie",
        "BTN_ODBC_DIAGNOSTICS": "Diagnostyka ODBC",
        "LBL_SQL_FILE": "Plik SQL",
        "LBL_DB": "Baza danych",
        "LBL_OUTPUT": "Wyjście",
        "LBL_SAVE_AS": "Zapisz jako (opcjonalnie):",
        "BTN_SAVE_AS": "Zapisz jako...",
        "BTN_CLEAR": "Wyczyść",
        "LBL_SAVE_AS_HINT": "Zostaw puste, aby zapisać w {dir} pod domyślną nazwą.",
        "INFO_SAVE_AS_EXT_FIXED_TITLE": "Poprawiono rozszerzenie",
        "INFO_SAVE_AS_EXT_FIXED_BODY": "Wybrana nazwa miała inne rozszerzenie. Zapiszę jako:\n{path}",
        "LBL_LANGUAGE": "Język",
        "CHK_ARCHIVE_SQL": "Archiwizuj SQL (zapisz zapytanie + metadane)",
        "LBL_DB_TIMEOUT_MIN": "Limit czasu DB (minuty) - wykonanie + pobieranie",
        "LBL_EXPORT_TIMEOUT_MIN": "Limit czasu eksportu (minuty) - generowanie XLSX/CSV",
        "LBL_TIMEOUT_NOTE": "0 = brak limitu. Domyślnie: po 3 minuty.",
        "CHK_SQL_HIGHLIGHT": "Podświetlanie składni (Chlorophyll)",
        "LBL_SQL_DIALECT": "Dialekt SQL",
        "MSG_DONE": "Gotowe.",
        "MSG_CANCEL_REQUESTED": "Zatrzymywanie...",
        "MSG_CONFIRM_CANCEL_AND_EXIT": "Eksport trwa. Przerwać i zamknąć?",
        "ERR_TITLE": "Błąd",
        "WARN_TITLE": "Uwaga",
        "APP_TITLE_FULL": "KKr SQL to XLSX/CSV",
        "BROWSER_OPEN_FAIL_TITLE": "Nie mogę otworzyć przeglądarki",
        "BROWSER_OPEN_FAIL_BODY": (
            "Nie udało się automatycznie otworzyć linku.\n"
            "Skopiuj i otwórz ręcznie:\n\n{url}"
        ),
        "BROWSER_OPEN_FAIL_ERROR_BODY": (
            "Nie udało się otworzyć przeglądarki.\n\n{error}\n\nLink:\n{url}"
        ),
        "ODBC_DIAGNOSTICS_TITLE": "Diagnostyka ODBC",
        "ODBC_DIAGNOSTICS_LABEL": "Diagnostyka ODBC:",
        "ERR_PG_MISSING_TITLE": "Brak biblioteki PostgreSQL",
        "ERR_PG_MISSING_BODY": (
            "Nie można połączyć z PostgreSQL. Wymagana biblioteka Pythona (np. psycopg2) "
            "nie jest zainstalowana. Zainstaluj brakującą bibliotekę i spróbuj ponownie."
        ),
        "ERR_QUERY_TIMEOUT": "Przekroczono limit czasu DB (>{minutes} min). Zapytanie przerwano.",
        "ERR_EXPORT_TIMEOUT": "Przekroczono limit czasu eksportu (>{minutes} min). Eksport przerwano.",
        "ERR_CANCELLED_BY_USER": "Przerwano na żądanie użytkownika.",
        "ERR_MYSQL_MISSING_TITLE": "Brak biblioteki MySQL",
        "ERR_MYSQL_MISSING_BODY": (
            "Nie można połączyć z MySQL. Wymagana biblioteka Pythona (np. pymysql) "
            "nie jest zainstalowana. Zainstaluj brakującą bibliotekę i spróbuj ponownie."
        ),
        "ERR_ORACLE_MISSING_TITLE": "Brak biblioteki Oracle",
        "ERR_ORACLE_MISSING_BODY": (
            "Nie można połączyć z Oracle. Wymagana biblioteka Pythona (np. cx_Oracle) "
            "nie jest zainstalowana. Zainstaluj brakującą bibliotekę i spróbuj ponownie."
        ),
        "INFO_SQL_HIGHLIGHT_TITLE": "Podświetlanie składni",
        "INFO_SQL_HIGHLIGHT_BODY": (
            "Chlorophyll nie jest dostępny. Zainstaluj go, aby włączyć "
            "podświetlanie SQL."
        ),
        "MSG_UI_TRUNCATED": (
            "...\n(Przycięto w UI, pełna treść w kkr-query2xlsx.log)"
        ),
        "CONN_SQLSTATE_LINE": "SQLSTATE: {sqlstate}",
        "CONN_MOST_COMMON_CAUSES": "Najczęstsze przyczyny:",
        "CONN_DIAGNOSTICS": "Diagnostyka:",
        "CONSOLE_AVAILABLE_FILES": "Dostępne pliki zapytań SQL:",
        "CONSOLE_CUSTOM_PATH": "0: [Własna ścieżka]",
        "CONSOLE_PROMPT_SELECT": (
            "Podaj numer pliku SQL do uruchomienia "
            "(0 = własna ścieżka, 1-{max_idx}): "
        ),
        "CONSOLE_PROMPT_CUSTOM_PATH": "Podaj pełną ścieżkę do pliku .sql: ",
        "CONSOLE_FILE_NOT_FOUND": "Plik nie istnieje. Spróbuj ponownie.",
        "CONSOLE_SELECT_RANGE": "Wpisz liczbę z zakresu 0-{max_idx}.",
        "CONSOLE_INVALID_INPUT": "Nieprawidłowe dane. Wpisz liczbę.",
        "CONSOLE_NO_QUERIES": "Brak ścieżek do plików SQL w queries.txt",
        "CONSOLE_PROMPT_FORMAT": "Podaj format wyjściowy (xlsx lub csv): ",
        "CONSOLE_INVALID_FORMAT": "Nieprawidłowe dane. Wpisz 'xlsx' lub 'csv'.",
        "CONSOLE_AVAILABLE_CSV_PROFILES": "Dostępne profile CSV:",
        "CONSOLE_DEFAULT_MARKER": " (domyślny)",
        "CONSOLE_PROMPT_CSV_PROFILE": (
            "Podaj numer profilu CSV lub naciśnij Enter, aby użyć domyślnego: "
        ),
        "CONSOLE_INVALID_SELECTION": "Nieprawidłowy wybór. Spróbuj ponownie.",
        "CONSOLE_SAVED_PATH": "Wyniki zapytania zapisano w: {path}",
        "CONSOLE_NO_ROWS": "Zapytanie nie zwróciło żadnych wierszy.",
        "CONSOLE_SQL_TIME": "Czas pobrania danych (SQL): {seconds:.2f} s",
        "CONSOLE_EXPORT_TIME": "Czas eksportu ({fmt}): {seconds:.2f} s",
        "CONSOLE_TOTAL_TIME": "Czas łączny: {seconds:.2f} s",
        "CONSOLE_PROMPT_PASSWORD": "Hasło dla połączenia '{name}': ",
        "CLI_DIAG_ODBC_HELP": "Wypisz diagnostykę ODBC i zakończ.",
        "CLI_SELF_TEST_HELP": "Uruchom wewnętrzne testy (smoke tests) i zakończ.",
        "DEFAULT_MSSQL_NAME": "Domyślne MSSQL",
        "FRAME_MSSQL": "MSSQL (ODBC)",
        "LBL_ODBC_DRIVER": "Sterownik ODBC",
        "LBL_SERVER": "Serwer",
        "LBL_DATABASE_NAME": "Nazwa bazy",
        "LBL_LOGIN": "Login",
        "LBL_PASSWORD": "Hasło",
        "CHK_REMEMBER_PASSWORD": "Zapamiętaj hasło",
        "CHK_TRUSTED": "Logowanie Windows (Trusted_Connection)",
        "CHK_ENCRYPT": "Encrypt",
        "CHK_TRUST_CERT": "TrustServerCertificate",
        "FRAME_POSTGRES": "PostgreSQL",
        "LBL_HOST": "Host",
        "LBL_PORT": "Port",
        "LBL_DATABASE": "Baza danych",
        "LBL_USER": "Użytkownik",
        "FRAME_MYSQL": "MySQL",
        "FRAME_SQLITE": "SQLite",
        "FILETYPE_SQLITE": "SQLite",
        "FILETYPE_ALL": "Wszystkie pliki",
        "TITLE_SELECT_SQLITE": "Wybierz istniejącą bazę SQLite",
        "TITLE_CREATE_SQLITE": "Utwórz nową bazę SQLite",
        "ASK_CREATE_SQLITE": "Utworzyć nową bazę?",
        "ASK_CREATE_SQLITE_BODY": (
            "Nie wybrano istniejącej bazy. Czy chcesz utworzyć nową?"
        ),
        "LBL_SQLITE_PATH": "Ścieżka do pliku",
        "BTN_SELECT": "Wybierz",
        "ERR_DATA_TITLE": "Błąd danych",
        "ERR_FILL_ODBC": "Wypełnij: sterownik, serwer i nazwę bazy danych.",
        "ERR_LOGIN_REQUIRED": (
            "Podaj login i hasło lub zaznacz logowanie Windows (Trusted_Connection)."
        ),
        "ERR_FILL_PG": "Wypełnij: host, nazwę bazy i użytkownika.",
        "ERR_FILL_MYSQL": "Wypełnij: host, nazwę bazy i użytkownika.",
        "ERR_FILL_SQLITE": "Podaj ścieżkę do pliku bazy SQLite.",
        "PROMPT_PASSWORD_TITLE": "Wymagane hasło",
        "PROMPT_PASSWORD_BODY": "Podaj hasło dla połączenia:\n{name}",
        "ERR_PASSWORD_REQUIRED_TITLE": "Wymagane hasło",
        "ERR_PASSWORD_REQUIRED_BODY": (
            "Nie podano hasła, więc połączenie nie zostało wykonane.\n\n"
            "Co zrobić:\n"
            "- Podaj hasło w oknie, które się pojawia.\n"
            "- Albo kliknij 'Zarządzaj połączeniami...' i ustaw hasło dla: {name}\n"
            "- Jeśli to logowanie Windows, zaznacz 'Logowanie Windows (Trusted_Connection)' "
            "i zostaw login/hasło puste."
        ),
        "WARN_REMEMBER_PASSWORD_TITLE": "Zapisać hasło?",
        "WARN_REMEMBER_PASSWORD_BODY": (
            "Hasło zostanie zapisane jawnym tekstem w pliku:\n{path}\n\n"
            "Trzymaj ten plik bardzo ostrożnie.\n\nKontynuować?"
        ),
        "LBL_CONN_NAME": "Nazwa połączenia",
        "LBL_DB_TYPE": "Typ bazy",
        "ERR_INVALID_CONN_TYPE": "Nieprawidłowy typ połączenia.",
        "ERR_CONN_NAME_REQUIRED": "Nazwa połączenia nie może być pusta.",
        "ERR_CONN_NAME_EXISTS": "Nazwa połączenia już istnieje. Wybierz inną nazwę.",
        "INFO_CONN_SAVED_TITLE": "Zapisano",
        "INFO_CONN_TEST_OK_TITLE": "Połączenie działa",
        "INFO_CONN_TEST_OK_BODY": "Test połączenia zakończony sukcesem.",
        "INFO_CONN_SAVED_BODY": "Połączenie zostało zapisane.",
        "BTN_SAVE": "Zapisz",
        "BTN_SAVE_NO_TEST": "Zapisz bez testu",
        "BTN_CANCEL": "Anuluj",
        "CSV_PROFILE_TITLE": "Szczegóły profilu",
        "CSV_PROFILE_NAME": "Nazwa profilu:",
        "CSV_PROFILE_ENCODING": "Kodowanie:",
        "CSV_PROFILE_DELIMITER": "Separator pól:",
        "CSV_PROFILE_DELIM_REPLACE": "Zastąp separator w wartościach:",
        "CSV_PROFILE_DECIMAL": "Separator dziesiętny:",
        "CSV_PROFILE_LINE_END": "Znak końca linii:",
        "CSV_PROFILE_QUOTECHAR": "Znak cudzysłowu:",
        "CSV_PROFILE_QUOTING": "Tryb cytowania:",
        "CSV_PROFILE_ESCAPECHAR": "Znak ucieczki:",
        "CSV_PROFILE_ESCAPE_HINT": "(znak ucieczki; puste = cytowanie)",
        "CSV_PROFILE_DOUBLEQUOTE": "Podwajaj cudzysłowy w polach",
        "CSV_PROFILE_DATE_FORMAT": "Format daty:",
        "CSV_PROFILE_FIELD_SEPARATOR": "Separator pól:",
        "CSV_PROFILE_WARNING_EMPTY": "Nazwa profilu nie może być pusta.",
        "CSV_PROFILE_WARNING_EXISTS": "Profil o podanej nazwie już istnieje.",
        "CSV_PROFILE_WARNING_SELECT": "Zaznacz profil na liście.",
        "CSV_PROFILE_INFO_BUILTIN": "Wbudowanych profili nie można usuwać.",
        "CSV_PROFILE_WARNING_MIN_ONE": "Musi istnieć co najmniej jeden profil CSV.",
        "CSV_PROFILE_INFO_SAVED_TITLE": "Informacja",
        "CSV_PROFILE_INFO_SAVED_BODY": "Zapisano profile CSV.",
        "CSV_PROFILE_CONFIRM_DELETE": "Usunąć wybrany profil?\n\nNie można cofnąć.",
        "BTN_SAVE_AS_NEW": "Zapisz jako nowy",
        "BTN_UPDATE_PROFILE": "Zaktualizuj profil",
        "BTN_DELETE": "Usuń",
        "BTN_SET_DEFAULT": "Ustaw jako domyślny",
        "BTN_CLOSE_SAVE": "Zamknij i zapisz",
        "ERR_QUERY_TITLE": "Błąd zapytania",
        "BTN_COPY": "Kopiuj",
        "BTN_CLOSE": "Zamknij",
        "BTN_COPY_ALL": "Kopiuj wszystko",
        "BTN_COPY_SUMMARY": "Kopiuj podsumowanie",
        "BTN_COPY_SQL": "Kopiuj SQL",
        "BTN_OPEN_SQL": "Otwórz SQL",
        "BTN_OPEN_LOG": "Otwórz log",
        "ERR_TAB_SUMMARY": "Podsumowanie",
        "ERR_TAB_SQL": "SQL",
        "ERR_TAB_DETAILS": "Szczegóły",
        "PROGRESS_WORKING_TITLE": "Trwa praca...",
        "PROGRESS_STARTING": "Startowanie...",
        "PROGRESS_STEP": "Krok:",
        "PROGRESS_ELAPSED": "Upłynęło:",
        "PROGRESS_TITLE": "Trwa eksport",
        "PROGRESS_GETTING_DATA": "Pobieranie danych z bazy...",
        "PROGRESS_EXPORTING_XLSX_TEMPLATE": "Eksport do XLSX (template)...",
        "PROGRESS_EXPORTING_XLSX": "Eksport do XLSX...",
        "PROGRESS_EXPORTING_CSV": "Eksport do CSV...",
        "ERR_NO_CONNECTION_TITLE": "Brak połączenia",
        "ERR_NO_CONNECTION_BODY": "Brak zapisanych połączeń. Utwórz i zapisz nowe połączenie.",
        "ERR_NO_CONNECTION_DELETE": "Brak połączenia do usunięcia.",
        "ASK_DELETE_CONNECTION_TITLE": "Usuń połączenie",
        "ASK_DELETE_CONNECTION_BODY": "Czy na pewno chcesz usunąć połączenie {name}?",
        "TITLE_EDIT_SECURE": "Edytuj secure.txt",
        "INFO_SECURE_SAVED_TITLE": "Zapisano",
        "INFO_SECURE_SAVED_BODY": "Zaktualizowano zawartość pliku secure.txt.",
        "ERR_SECURE_SAVE_TITLE": "Błąd zapisu",
        "ERR_SECURE_SAVE_BODY": (
            "Nie udało się zapisać pliku secure.txt.\n\nSzczegóły techniczne:\n{error}"
        ),
        "TITLE_SELECT_SQL": "Wybierz plik SQL",
        "FILETYPE_SQL": "Pliki SQL",
        "TITLE_SELECT_TEMPLATE": "Wybierz plik template XLSX",
        "FILETYPE_EXCEL": "Pliki Excel",
        "FILETYPE_CSV": "Pliki CSV",
        "ERR_TEMPLATE_TITLE": "Błąd template",
        "ERR_TEMPLATE_SHEETS": (
            "Nie można odczytać arkuszy z pliku template.\n\n"
            "Szczegóły techniczne:\n{error}"
        ),
        "CSV_DEFAULT_PROFILE_LABEL": "Domyślny profil CSV: {name}",
        "TITLE_EDIT_QUERIES": "Edycja queries.txt",
        "TITLE_ADD_SQL_FILES": "Dodaj pliki SQL",
        "WARN_SKIPPED_FILES_TITLE": "Pominięto pliki",
        "WARN_SKIPPED_FILES_BODY": (
            "Niektóre wybrane pliki nie mają rozszerzenia .sql i zostały pominięte:\n\n"
            "{files}{more}"
        ),
        "INFO_ALREADY_LISTED": "Wybrane pliki są już na liście.",
        "TITLE_EDIT_QUERY_PATH": "Edytuj ścieżkę zapytania",
        "PROMPT_EDIT_QUERY_PATH": "Edytuj ścieżkę zapytania:",
        "WARN_INVALID_SQL_FILE": "Wpis musi wskazywać plik z rozszerzeniem .sql.",
        "WARN_FILE_MISSING_TITLE": "Uwaga",
        "WARN_FILE_MISSING_BODY": (
            "Plik nie istnieje (lub jest chwilowo niedostępny).\n"
            "Zapisuję ścieżkę, ale upewnij się, że jest poprawna:\n\n{path}"
        ),
        "INFO_SELECT_ENTRY_DELETE": "Zaznacz wpis do usunięcia.",
        "ERR_QUERIES_SAVE_TITLE": "Błąd zapisu",
        "ERR_QUERIES_SAVE_BODY": (
            "Nie można zapisać queries.txt.\n\nSzczegóły techniczne:\n{error}"
        ),
        "BTN_ADD_FILES": "Dodaj pliki...",
        "BTN_REMOVE_SELECTED": "Usuń zaznaczone",
        "TITLE_SELECT_REPORT": "Wybierz raport z listy",
        "ERR_NO_REPORTS": "Brak raportów w queries.txt",
        "WARN_NO_REPORT_SELECTED": "Nie wybrano żadnego raportu.",
        "ERR_NO_SQL_SELECTED": "Nie wybrano pliku SQL.",
        "ERR_SQL_NOT_FOUND": "Wybrany plik SQL nie istnieje.",
        "ERR_SQLFILE_IS_SQLITE_EXT": (
            "Wybrano bazę SQLite. To nie jest plik .sql z zapytaniami."
        ),
        "ERR_SQLFILE_IS_SPREADSHEET": (
            "To wygląda jak arkusz danych ({ext}). "
            "Wybierz plik .sql zawierający zapytania."
        ),
        "ERR_SQLFILE_IS_ZIP": (
            "To plik ZIP (zaczyna się od 'PK'). Nie jest to zwykły plik .sql."
        ),
        "ERR_SQLFILE_IS_OLD_OFFICE": (
            "To wygląda na stary binarny plik Microsoft Office (np. .xls). "
            "Wybierz plik .sql."
        ),
        "ERR_SQLFILE_IS_BINARY": (
            "Plik wygląda na binarny, a nie tekstowy SQL. Wybierz plik .sql."
        ),
        "ERR_CANNOT_OPEN_FILE": "Nie można otworzyć pliku: {error}",
        "ERR_NEED_CONNECTION": "Utwórz połączenie z bazą danych przed uruchomieniem raportu.",
        "ERR_TEMPLATE_ONLY_XLSX": "Template można użyć tylko dla formatu XLSX.",
        "ERR_TEMPLATE_NOT_SELECTED": "Nie wybrano pliku template.",
        "ERR_TEMPLATE_SHEET_NOT_SELECTED": "Nie wybrano arkusza z pliku template.",
        "MSG_RUNNING": "Trwa wykonywanie zapytania i eksport. Proszę czekać...",
        "MSG_SAVED_DETAILS": (
            "Zapisano: {path}\n"
            "Wiersze: {rows}\n"
            "Czas SQL: {sql_time_hms} ({sql_time:.2f} s)\n"
            "Czas eksportu: {export_time_hms} ({export_time:.2f} s)\n"
            "Czas łączny: {total_time_hms} ({total_time:.2f} s)"
        ),
        "MSG_SAVED_DETAILS_CSV": "Profil CSV: {profile}",
        "MSG_NO_ROWS": (
            "Zapytanie nie zwróciło wierszy.\n"
            "Czas SQL: {sql_time_hms} ({sql_time:.2f} s)"
        ),
        "ERR_XLSX_TOO_LARGE": (
            "Eksport XLSX pominięty, ponieważ wynik przekracza limity Excela. "
            "Wynik: {rows} wierszy, {cols} kolumn. "
            "Arkusz wymaga: {sheet_rows} wierszy, {sheet_cols} kolumn "
            "(maks. {max_rows} wierszy, {max_cols} kolumn). "
            "Zamiast tego użyj CSV."
        ),
        "ERR_EXPORT": "Błąd eksportu. Pełne szczegóły w logu.",
        "FRAME_DB_CONNECTION": "Połączenie z bazą danych",
        "FRAME_ADVANCED": "Zaawansowane",
        "LBL_CONNECTION": "Połączenie:",
        "BTN_MANAGE_CONNECTIONS": "Zarządzaj połączeniami...",
        "BTN_EDIT_CONNECTION": "Edytuj połączenie",
        "BTN_NEW_CONNECTION": "Nowe połączenie",
        "BTN_DUPLICATE_CONNECTION": "Duplikuj",
        "BTN_TEST_CONNECTION": "Testuj połączenie",
        "BTN_DELETE_CONNECTION": "Usuń połączenie",
        "BTN_EDIT_SECURE": "Edytuj secure.txt",
        "BTN_SET_CURRENT_CONNECTION": "Ustaw jako bieżące",
        "CONNECTIONS_MANAGER_TITLE": "Połączenia",
        "WARN_SELECT_CONNECTION": "Wybierz połączenie z listy.",
        "FRAME_SQL_SOURCE": "Źródło zapytania SQL",
        "FRAME_OUTPUT_FORMAT": "Format wyjściowy",
        "FRAME_TEMPLATE_OPTIONS": "Opcje template XLSX (GUI)",
        "FRAME_RESULTS": "Wynik i akcje",
        "LBL_SELECTED_SQL": "Wybrany plik SQL:",
        "BTN_SELECT_SQL": "Wybierz plik SQL",
        "BTN_SELECT_FROM_LIST": "Wybierz z listy raportów",
        "BTN_EDIT_QUERIES": "Edytuj queries.txt",
        "BTN_PASTE_SQL": "Wklej SQL",
        "TITLE_PASTE_SQL": "Wklej SQL",
        "LBL_REPORT_NAME": "Nazwa raportu (wymagana)",
        "LBL_PASTE_SQL": "SQL (wklej tutaj)",
        "BTN_USE_SQL": "Użyj SQL",
        "ERR_INVALID_REPORT_NAME_TITLE": "Nieprawidłowa nazwa raportu",
        "ERR_REPORT_NAME_REQUIRED": "Nazwa raportu jest wymagana (nie może być pusta).",
        "ERR_REPORT_NAME_NO_FOLDERS": "Nazwa raportu musi być samą nazwą pliku (bez folderów).",
        "ERR_REPORT_NAME_INVALID_CHARS": "Nazwa raportu zawiera niedozwolone znaki: {chars}",
        "ERR_REPORT_NAME_CONTROL_CHARS": "Nazwa raportu zawiera znaki sterujące.",
        "ERR_REPORT_NAME_TRAILING_DOT_SPACE": "Nazwa raportu nie może kończyć się spacją ani kropką.",
        "ERR_REPORT_NAME_RESERVED": "Nazwa raportu '{name}' jest zarezerwowana w Windows.",
        "ERR_REPORT_NAME_TOO_LONG": "Nazwa raportu jest za długa (max {max_len} znaków).",
        "ERR_EMPTY_SQL_TITLE": "Pusty SQL",
        "ERR_EMPTY_SQL_BODY": "SQL nie może być pusty.",
        "ERR_INVALID_SQL_FILE_TITLE": "Nieprawidłowy plik SQL",
        "ERR_NO_FILE_SELECTED": "Nie wybrano pliku.",
        "LBL_SQL_PASTED": "Wklejony SQL:",
        "ERR_NO_SQL_SOURCE": "Wybierz plik SQL albo wklej SQL.",
        "LBL_CSV_PROFILE": "Profil CSV:",
        "BTN_MANAGE_CSV_PROFILES": "Zarządzaj profilami CSV",
        "CHK_USE_TEMPLATE": "Użyj pliku template (tylko dla XLSX, tylko w GUI)",
        "LBL_TEMPLATE_FILE": "Plik template:",
        "BTN_SELECT_TEMPLATE": "Wybierz template",
        "LBL_TEMPLATE_SHEET": "Arkusz:",
        "LBL_TEMPLATE_START_CELL": "Startowa komórka:",
        "CHK_INCLUDE_HEADERS": "Zapisz nagłówki (nazwy kolumn) w arkuszu",
        "BTN_START": "Start",
        "BTN_REPORT_ISSUE": "Zgłoś problem / sugestię",
        "BTN_CHECK_UPDATES": "Sprawdź aktualizacje",
        "BTN_HELP": "Pomoc / README",
        "BTN_OPEN_LOGS": "Otwórz katalog logów",
        "LBL_PROJECT_PAGE": "Strona projektu",
        "LBL_EXPORT_INFO": "Informacje o eksporcie:",
        "BTN_OPEN_FILE": "Otwórz plik",
        "BTN_OPEN_FOLDER": "Otwórz katalog",
        "LBL_ERRORS_SHORT": "Błędy (skrót):",
        "README_WINDOW_TITLE": "kkr-query2xlsx — README",
        "UPD_TITLE": "Aktualizacje",
        "UPD_CHECKING": "Sprawdzanie aktualizacji...",
        "UPD_CHECK_FAILED": "Nie udało się sprawdzić aktualizacji.",
        "UPD_ERR_NETWORK": "Brak połączenia z GitHub (problem sieci/DNS). Sprawdź połączenie i spróbuj ponownie.",
        "UPD_ERR_TIMEOUT": "GitHub nie odpowiedział na czas (timeout). Spróbuj ponownie za chwilę.",
        "UPD_ERR_HTTP": "GitHub zwrócił HTTP {status} podczas sprawdzania aktualizacji.",
        "UPD_ERR_RATE_LIMITED": "GitHub zwrócił HTTP {status}. Wygląda na limit zapytań. Spróbuj ponownie po: {retry_at}.",
        "UPD_ERR_JSON": "Otrzymano nieprawidłową odpowiedź z GitHub (błąd parsowania JSON).",
        "UPD_NO_UPDATE": (
            "Brak aktualizacji wydań (Release). Obecna: {current}, najnowsze Release: {latest}."
        ),
        "UPD_UPDATE_AVAILABLE": "Dostępna aktualizacja wydania (Release): {current} → {latest}.",
        "UPD_PROMPT_INSTALL": "Zainstalować aktualizację teraz?",
        "UPD_GIT_MODE": (
            "Uruchomiono z repozytorium Git. Ta funkcja sprawdza tylko oficjalne wydania "
            "(Release). Aby zaktualizować kod, uruchom:\n\n{command}"
        ),
        "UPD_GIT_STATUS_MATCH": "Status Git: lokalnie {local}, zdalnie main {remote}.",
        "UPD_GIT_STATUS_AHEAD": "Status Git: lokalnie {local}, zdalnie main {remote} (dostępna aktualizacja).",
        "UPD_GIT_STATUS_LOCAL_AHEAD": "Status Git: lokalnie {local} jest przed zdalnym main {remote}.",
        "UPD_GIT_STATUS_DIVERGED": "Status Git: lokalnie {local} i zdalnie main {remote} są rozbieżne.",
        "UPD_GIT_STATUS_DIFFERENT": "Status Git: lokalnie {local}, zdalnie main {remote} (różne).",
        "UPD_GIT_STATUS_DIFFERENT_UNVERIFIED": (
            "Status Git: lokalnie {local}, zdalnie main {remote} (różne) "
            "(nie udało się porównać — brak połączenia lub lokalny commit nie istnieje na GitHub)."
        ),
        "UPD_SOURCE_MODE": (
            "Uruchomiono ze źródeł bez repo Git. Ta funkcja sprawdza tylko oficjalne wydania "
            "(Release). Aby zaktualizować, pobierz nowy ZIP z Releases albo sklonuj repozytorium."
        ),
        "UPD_UPDATER_MISSING": "Brak updatera. Pobierz najnowszy ZIP ręcznie.",
        "UPD_START_FAILED": "Nie udało się uruchomić updatera.",
        "STATUS_NO_CONNECTION": "Brak połączenia. Utwórz nowe połączenie.",
        "STATUS_CONNECTION_ERROR": "Błąd połączenia. Utwórz nowe połączenie.",
        "STATUS_PASSWORD_REQUIRED": "Wymagane hasło.",
        "ERR_CONNECTION_TITLE": "Błąd połączenia",
        "ERR_CONNECTION_BODY": (
            "Nie udało się nawiązać połączenia.\n\nSzczegóły techniczne:\n{error}"
        ),
        "STATUS_CONNECTED": "Połączono z {name} ({type}).",
        "INFO_CONNECTION_OK_TITLE": "Połączenie działa",
        "INFO_CONNECTION_OK_BODY": "Połączenie {name} powiodło się.",
        "ERR_NO_SECURE_CONFIG": "Brak konfiguracji połączenia",
        "ERR_SELECT_TEMPLATE": "Wybierz plik template",
        "ERR_GENERIC": "Błąd",
        "ERR_TEMPLATE_MISSING_SHEET": "Arkusz '{sheet}' nie istnieje w pliku template.",
        "CLI_DESC": "Uruchamiaj pliki SQL i eksportuj wyniki do XLSX/CSV.",
        "CLI_LANG_HELP": "Język interfejsu (en/pl).",
        "CLI_CONSOLE_HELP": "Uruchom w trybie konsolowym.",
        "CLI_OUTPUT_HELP": (
            "Opcjonalny plik lub katalog docelowy. Katalog może być istniejący lub nowy "
            "(ścieżka bez rozszerzenia). Aby wymusić katalog, dodaj na końcu / lub \\."
        ),
        "CLI_NO_CONNECTIONS": (
            "Brak zapisanych połączeń. Utwórz połączenie w trybie GUI, aby uruchomić konsolę."
        ),
        "CLI_CONNECTION_FAIL": "Nie udało się utworzyć połączenia. Pełne szczegóły w logu.",
        "MENU_LANGUAGE": "Język",
        "ERR_FILE_PATH": "Ścieżka: {path}",
        "ERR_FILE_LOCKED": (
            "Plik docelowy już istnieje i może być otwarty w innej aplikacji "
            "(np. Excel). Zamknij go i spróbuj ponownie.{path}"
        ),
        "ERR_NO_WRITE_PERMISSION": (
            "Brak uprawnień do zapisu pliku docelowego lub ścieżka jest niedostępna. "
            "Sprawdź lokalizację pliku."
        ),
        "ERR_SQL_SOURCE": "Źródło SQL:",
        "ERR_DB_MESSAGE": "Komunikat bazy (fragment):",
        "ERR_SQL_PREVIEW": "SQL (początek):",
        "ERR_FULL_LOG": "Pełny błąd zapisany w pliku kkr-query2xlsx.log",
        "ERR_HINT_LABEL": "Podpowiedź:",
        "DB_TYPE_MSSQL": "SQL Server (ODBC)",
        "DB_TYPE_PG": "PostgreSQL",
        "DB_TYPE_MYSQL": "MySQL",
        "DB_TYPE_SQLITE": "SQLite (plik .db)",
        "TITLE_CONN_DIALOG_EDIT": "Edycja: {name}",
        "TITLE_CONN_DIALOG_NEW": "Nowe połączenie",
        "TITLE_CONN_DIALOG_DUPLICATE": "Duplikowanie: {name}",
        "CONN_DIALOG_HINT_EDIT": "Edycja połączenia: {name}",
        "CONN_DIALOG_HINT_NEW": "Utwórz nowe połączenie, wpisując jego szczegóły.",
        "CONN_DIALOG_HINT_DUPLICATE": "Duplikowanie połączenia: {name}",
        "CONN_DUPLICATE_SUFFIX": "(kopia)",
        "INFO_CONN_SAVED_NO_TEST": (
            "Połączenie zapisane bez testu.\n"
            "Użyj przycisku „Testuj połączenie”, aby je sprawdzić."
        ),
        "INFO_ICON": "i",
        "BTN_OK": "OK",
        "TITLE_DATA_DIR_NOTICE": "Katalog danych",
        "MSG_DATA_DIR_NOTICE": (
            "Aplikacja używa katalogu danych:\n\n"
            "{data_dir}\n\n"
            "Tu trzymane są m.in. pliki konfiguracyjne (secure.txt, queries.txt, kkr-query2xlsx.json) "
            "oraz foldery typu generated_reports/ i sql_archive/."
        ),
        "BTN_DONT_SHOW_AGAIN": "Nie pokazuj ponownie",
        "FORMAT_XLSX": "XLSX",
        "FORMAT_CSV": "CSV",
        "CSV_PROFILE_DIALOG_TITLE": "Profile CSV",
        "CSV_PROFILE_DEFAULT_SUFFIX": "(domyślny)",
        "CSV_PROFILE_BUILTIN_SUFFIX": "[wbudowany]",
        "CSV_PROFILE_INVALID_DATE": (
            "Podany format daty jest nieprawidłowy. Skorzystaj ze składni strftime."
        ),
        "CSV_PROFILE_DATE_DEFAULT": "Domyślny format Pandas (przykład: {example})",
        "CSV_PROFILE_DATE_INVALID": (
            "Nieprawidłowy wzorzec daty (użyj składni strftime, np. %Y-%m-%d)."
        ),
        "CSV_PROFILE_DATE_PREVIEW": "Bieżący czas w tym formacie: {example}",
        "CSV_PROFILE_BUILTIN_NOTICE": (
            "Profil wbudowany: nie można zapisać zmian ani usuwać. "
            "Użyj Zapisz jako nowy, aby stworzyć własny wariant."
        ),
        "CSV_PROFILE_NAME_RESERVED": (
            "Ta nazwa jest zarezerwowana dla wbudowanego profilu. Wybierz inną nazwę."
        ),
        "CSV_PROFILE_NO_SELECTION_TITLE": "Brak profilu",
        "CSV_PROFILE_BUILTIN_OVERWRITE": (
            "Nie możesz nadpisać wbudowanego profilu. Zmień nazwę i zapisz jako nowy profil."
        ),
        "CSV_PROFILE_UNSAVED_TITLE": "Niezapisane zmiany",
        "CSV_PROFILE_UNSAVED_BODY": (
            "Masz niezapisane zmiany profili CSV. Zapisać przed zamknięciem?"
        ),
        "WARN_SKIPPED_FILES_MORE": "\n\n(+ {count} kolejnych)",
        "CSV_HELP_NAME_TITLE": "Nazwa profilu",
        "CSV_HELP_NAME_BODY": (
            "Dowolna, unikalna nazwa ułatwiająca wybór profilu, np. "
            "\"UTF-8 (przecinek)\" lub \"Windows-1250 (średnik)\"."
        ),
        "CSV_HELP_ENCODING_TITLE": "Kodowanie",
        "CSV_HELP_ENCODING_BODY": (
            "Sposób kodowania znaków w pliku CSV. Domyślnie UTF-8; dla "
            "starszych arkuszy Excel można użyć windows-1250."
        ),
        "CSV_HELP_DELIMITER_TITLE": "Separator pól",
        "CSV_HELP_DELIMITER_BODY": (
            "Znak oddzielający kolumny. Najczęściej przecinek (,) lub "
            "średnik (;), zgodnie z ustawieniami regionalnymi arkusza."
        ),
        "CSV_HELP_DELIM_REPLACE_TITLE": "Zastąp separator w wartościach",
        "CSV_HELP_DELIM_REPLACE_BODY": (
            "Opcjonalnie zamienia znak separatora w wartościach na inny (np. "
            "średnik na przecinek). Przydatne, gdy system importujący nie "
            "obsługuje poprawnego eskapowania separatorów w polach. "
            "Uwaga: zamiana jest globalna dla wszystkich pól tekstowych "
            "(również JSON/ID w formie tekstu)."
        ),
        "CSV_HELP_DECIMAL_TITLE": "Separator dziesiętny",
        "CSV_HELP_DECIMAL_BODY": (
            "Znak rozdzielający część całkowitą od ułamkowej. Kropka (.) "
            "dla układu angielskiego, przecinek (,) dla polskiego."
        ),
        "CSV_HELP_LINETERM_TITLE": "Znak końca linii",
        "CSV_HELP_LINETERM_BODY": (
            "Domyślnie \\n. Dla pełnej zgodności z Windows można użyć "
            "\\r\\n. Zmień tylko gdy import wymaga konkretnego formatu."
        ),
        "CSV_HELP_QUOTECHAR_TITLE": "Znak cudzysłowu",
        "CSV_HELP_QUOTECHAR_BODY": (
            "Najczęściej \". Używany do otaczania pól wymagających "
            "cytowania (np. zawierających separator)."
        ),
        "CSV_HELP_QUOTING_TITLE": "Strategia cudzysłowów",
        "CSV_HELP_QUOTING_BODY": (
            "minimal – tylko gdy potrzebne (zalecane), all – zawsze, "
            "nonnumeric – dla tekstu, none – bez cytowania (wymaga "
            "escapechar)."
        ),
        "CSV_HELP_ESCAPECHAR_TITLE": "Znak ucieczki",
        "CSV_HELP_ESCAPECHAR_BODY": (
            "Znak ucieczki używany, gdy quoting=none lub pola mogą "
            "zawierać separator. Zostaw pusty, jeżeli stosujesz "
            "standardowe cytowanie."
        ),
        "CSV_HELP_DOUBLEQUOTE_TITLE": "Podwajanie cudzysłowów",
        "CSV_HELP_DOUBLEQUOTE_BODY": (
            "Gdy zaznaczone, wewnętrzny \" w polu staje się \"\". "
            "Zostaw włączone, chyba że system importujący wymaga inaczej."
        ),
        "CSV_HELP_DATE_FORMAT_TITLE": "Format daty",
        "CSV_HELP_DATE_FORMAT_BODY": (
            "Opcjonalny wzorzec strftime, np. %Y-%m-%d lub %d.%m.%Y. "
            "Pozostaw puste, aby użyć domyślnego formatowania Pandas."
        ),
    },
}


def _detect_lang() -> str:
    # Default language is English.
    return "en"


_CURRENT_LANG = _detect_lang()


def _normalize_ui_lang(lang: str | None) -> str | None:
    normalized = (lang or "").lower()
    return normalized if normalized in I18N else None


def set_lang(lang: str) -> None:
    global _CURRENT_LANG
    normalized = _normalize_ui_lang(lang)
    _CURRENT_LANG = normalized or "en"


def t(key: str, **kwargs) -> str:
    # fallback: en -> key
    s = I18N.get(_CURRENT_LANG, {}).get(key) or I18N["en"].get(key) or key
    return s.format(**kwargs) if kwargs else s


def _build_path(name: str) -> str:
    return os.path.join(DATA_DIR, name)


def _center_window(win, parent=None):
    """Center the window on the parent or screen, clamping to screen bounds."""
    win.update()
    w = win.winfo_width() or win.winfo_reqwidth()
    h = win.winfo_height() or win.winfo_reqheight()
    if parent and parent.winfo_ismapped():
        parent.update()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        pw = parent.winfo_width() or parent.winfo_reqwidth()
        ph = parent.winfo_height() or parent.winfo_reqheight()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
    else:
        vroot_x = win.winfo_vrootx()
        vroot_y = win.winfo_vrooty()
        vroot_w = win.winfo_vrootwidth()
        vroot_h = win.winfo_vrootheight()
        x = vroot_x + (vroot_w - w) // 2
        y = vroot_y + (vroot_h - h) // 2
    vroot_x = win.winfo_vrootx()
    vroot_y = win.winfo_vrooty()
    vroot_w = win.winfo_vrootwidth()
    vroot_h = win.winfo_vrootheight()
    if x + w > vroot_x + vroot_w:
        x = vroot_x + vroot_w - w
    if y + h > vroot_y + vroot_h:
        y = vroot_y + vroot_h - h
    x = max(x, vroot_x)
    y = max(y, vroot_y)
    win.geometry(f"+{x}+{y}")


def _center_toplevel_on_parent(win: "tk.Toplevel", parent: "tk.Misc") -> None:
    """Center `win` on `parent` (or screen) and clamp to visible bounds."""
    win.update_idletasks()

    # IMPORTANT: before map, winfo_width/height can be 1 -> use req sizes as fallback
    ww = win.winfo_width() or win.winfo_reqwidth()
    wh = win.winfo_height() or win.winfo_reqheight()

    # Visible desktop bounds (multi-monitor safe-ish)
    vroot_x = win.winfo_vrootx()
    vroot_y = win.winfo_vrooty()
    vroot_w = win.winfo_vrootwidth()
    vroot_h = win.winfo_vrootheight()

    x = vroot_x + max((vroot_w - ww) // 2, 0)
    y = vroot_y + max((vroot_h - wh) // 2, 0)

    try:
        parent.update_idletasks()
        if parent.winfo_viewable():
            pw = parent.winfo_width() or parent.winfo_reqwidth()
            ph = parent.winfo_height() or parent.winfo_reqheight()
            px = parent.winfo_rootx()
            py = parent.winfo_rooty()
            x = px + (pw - ww) // 2
            y = py + (ph - wh) // 2
    except Exception:
        pass

    # Clamp to visible area
    if x + ww > vroot_x + vroot_w:
        x = vroot_x + vroot_w - ww
    if y + wh > vroot_y + vroot_h:
        y = vroot_y + vroot_h - wh
    x = max(x, vroot_x)
    y = max(y, vroot_y)

    win.geometry(f"+{x}+{y}")


def open_github_issue_chooser(parent=None) -> None:
    version = get_app_version_label()
    # GitHub Issue Forms: przekazanie wartości do pola "App version"
    url = f"{GITHUB_ISSUE_CHOOSER_URL}?app_version={quote_plus(version)}"
    try:
        ok = webbrowser.open_new_tab(url)
        if not ok and parent is not None:
            messagebox.showwarning(
                t("BROWSER_OPEN_FAIL_TITLE"),
                t(
                    "BROWSER_OPEN_FAIL_BODY",
                    url=url,
                ),
                parent=parent,
            )
    except Exception as exc:  # noqa: BLE001
        if parent is not None:
            messagebox.showerror(
                t("ERR_TITLE"),
                t(
                    "BROWSER_OPEN_FAIL_ERROR_BODY",
                    error=exc,
                    url=url,
                ),
                parent=parent,
            )


# =========================
# README / DOCS VIEWER (GUI)
# =========================

REPO_URL = "https://github.com/kkrysztofczyk/kkr-query2xlsx"


def _resource_path(rel_path: str) -> Path:
    """Resolve paths for source run and PyInstaller (_MEIPASS)."""
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return (base / rel_path).resolve()


def _read_text_if_exists(path: Path) -> str | None:
    try:
        if path.exists():
            return path.read_text(encoding="utf-8")
    except Exception:
        return None
    return None


def _cleanup_md_for_viewer(md: str) -> str:
    """
    Removes HTML blocks like <p align="center">...</p> and Markdown image lines,
    because we show images separately inside the app.
    """
    lines: list[str] = []
    in_html_block = False

    for raw in md.splitlines():
        s = raw.strip()

        if s.startswith("<p") and "align" in s:
            in_html_block = True
            continue
        if in_html_block:
            if "</p>" in s:
                in_html_block = False
            continue

        if s.startswith("<img") or s.startswith("</") or s.startswith("<br"):
            continue

        if re.match(r"^!\[.*?\]\(.*?\)\s*$", s):
            continue

        lines.append(raw)

    return "\n".join(lines).strip()


def _maybe_insert_image(
    txt: "tkinter.Text", img_refs: list, path: Path, max_width: int = 880
) -> None:
    try:
        if not path.exists():
            return

        img = __import__("tkinter").PhotoImage(file=str(path))

        w = img.width()
        if w > max_width:
            factor = (w + max_width - 1) // max_width
            img = img.subsample(factor, factor)

        img_refs.append(img)  # keep reference alive
        txt.image_create("end", image=img)
        txt.insert("end", "\n")
    except Exception:
        return


def _insert_md_simple(txt: "tkinter.Text", md: str) -> None:
    in_code = False

    for raw in md.splitlines():
        line = raw.rstrip("\n")
        s = line.lstrip()

        if s.startswith("```"):
            in_code = not in_code
            txt.insert("end", "\n")
            continue

        if in_code:
            txt.insert("end", line + "\n", ("codeblock",))
            continue

        if s.startswith("# "):
            txt.insert("end", s[2:] + "\n", ("h1",))
            continue
        if s.startswith("## "):
            txt.insert("end", s[3:] + "\n", ("h2",))
            continue
        if s.startswith("### "):
            txt.insert("end", s[4:] + "\n", ("h3",))
            continue

        if re.match(r"^(\* \* \*|---+)\s*$", s):
            txt.insert("end", "\n")
            continue

        if s.startswith(">"):
            txt.insert("end", s.lstrip("> ").rstrip() + "\n", ("quote",))
            continue

        if s.startswith(("- ", "* ")):
            txt.insert("end", "• " + s[2:] + "\n", ("bullet",))
            continue

        if re.match(r"^\d+\.\s+", s):
            txt.insert("end", s + "\n", ("bullet",))
            continue

        txt.insert("end", line + "\n")


def show_readme_window(parent) -> None:
    import tkinter as tk
    from tkinter import ttk

    win = tk.Toplevel(parent)
    apply_app_icon(win)
    win.title(t("README_WINDOW_TITLE"))
    win.geometry("920x700")
    win.minsize(720, 480)
    win.transient(parent)
    try:
        win.lift()
        win.focus_set()
    except Exception:
        pass

    top = ttk.Frame(win)
    top.pack(fill="x", padx=10, pady=(10, 6))

    ttk.Label(top, text="README.md", font=("TkDefaultFont", 12, "bold")).pack(
        side="left"
    )

    txt = ScrolledText(win, wrap="word")
    txt.pack(fill="both", expand=True, padx=10, pady=(0, 10))
    txt.configure(state="normal")

    txt.tag_configure("h1", font=("TkDefaultFont", 16, "bold"), spacing3=8)
    txt.tag_configure("h2", font=("TkDefaultFont", 14, "bold"), spacing3=6)
    txt.tag_configure("h3", font=("TkDefaultFont", 12, "bold"), spacing3=4)
    txt.tag_configure(
        "codeblock", font=("TkFixedFont", 10), lmargin1=12, lmargin2=12
    )
    txt.tag_configure("quote", foreground="gray40", lmargin1=12, lmargin2=12)
    txt.tag_configure("bullet", lmargin1=12, lmargin2=24)

    win._img_refs = []

    _maybe_insert_image(txt, win._img_refs, _resource_path("docs/logo.png"), max_width=880)

    readme_path = _resource_path("README.md")
    md = _read_text_if_exists(readme_path)
    if not md:
        md = (
            "# README.md not found\n\n"
            f"Missing file: {readme_path}\n\n"
            "Project page:\n"
            f"- {REPO_URL}\n"
        )
    md = _cleanup_md_for_viewer(md)
    _insert_md_simple(txt, md)

    txt.insert("end", "\n")
    _maybe_insert_image(txt, win._img_refs, _resource_path("docs/gui.png"), max_width=880)

    txt.configure(state="disabled")
    win.update_idletasks()
    _center_window(win, parent)


SECURE_PATH = _build_path("secure.txt")
QUERIES_PATH = _build_path("queries.txt")
APP_CONFIG_PATH = _build_path("kkr-query2xlsx.json")
UI_CONFIG_FILENAME = "kkr-query2xlsx.user.json"
LEGACY_CSV_PROFILES_PATH = _build_path("csv_profiles.json")
SQL_ARCHIVE_DIR = _build_path("sql_archive")

SECURE_SAMPLE_PATH = os.path.join(BASE_DIR, "secure.sample.json")
QUERIES_SAMPLE_PATH = os.path.join(BASE_DIR, "queries.sample.txt")


BUILTIN_CSV_PROFILES = [
    {
        "name": "CSV standard (comma, dot)",
        "encoding": "utf-8-sig",
        "delimiter": ",",
        "delimiter_replacement": "",
        "decimal": ".",
        "lineterminator": "\n",
        "quotechar": '"',
        "quoting": "minimal",
        "escapechar": "",
        "doublequote": True,
        "date_format": "",
    },
    {
        "name": "CSV Excel Europe (semicolon, comma)",
        "encoding": "utf-8-sig",
        "delimiter": ";",
        "delimiter_replacement": "",
        "decimal": ",",
        "lineterminator": "\n",
        "quotechar": '"',
        "quoting": "minimal",
        "escapechar": "",
        "doublequote": True,
        "date_format": "",
    },
]

BUILTIN_CSV_PROFILE_NAMES = {p["name"] for p in BUILTIN_CSV_PROFILES}

PREFERRED_DEFAULT_CSV_PROFILE_BY_LANG = {
    "en": "CSV standard (comma, dot)",
    "pl": "CSV Excel Europe (semicolon, comma)",
}


def _default_ui_config() -> dict:
    return {
        "ui": {
            "sql_highlight_enabled": False,
            "hide_template_naming_hint": False,
            "hide_data_dir_notice": False,
        }
    }


def load_ui_config(app_dir: Path) -> dict:
    path = Path(app_dir) / UI_CONFIG_FILENAME
    if not path.exists():
        return _default_ui_config()
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return _default_ui_config()

    if not isinstance(raw, dict):
        return _default_ui_config()

    cfg = _default_ui_config()
    ui_section = raw.get("ui")
    if isinstance(ui_section, dict):
        raw_val = ui_section.get("sql_highlight_enabled")
        if isinstance(raw_val, bool):
            cfg["ui"]["sql_highlight_enabled"] = raw_val
        elif isinstance(raw_val, int) and raw_val in (0, 1):
            cfg["ui"]["sql_highlight_enabled"] = bool(raw_val)
        raw_hint = ui_section.get("hide_template_naming_hint")
        if isinstance(raw_hint, bool):
            cfg["ui"]["hide_template_naming_hint"] = raw_hint
        elif isinstance(raw_hint, int) and raw_hint in (0, 1):
            cfg["ui"]["hide_template_naming_hint"] = bool(raw_hint)
        raw_data_dir = ui_section.get("hide_data_dir_notice")
        if isinstance(raw_data_dir, bool):
            cfg["ui"]["hide_data_dir_notice"] = raw_data_dir
        elif isinstance(raw_data_dir, int) and raw_data_dir in (0, 1):
            cfg["ui"]["hide_data_dir_notice"] = bool(raw_data_dir)
    return cfg


def save_ui_config(app_dir: Path, data: dict) -> None:
    tmp_path: Path | None = None
    try:
        if not isinstance(data, dict):
            data = _default_ui_config()

        ui_section = data.get("ui") if isinstance(data.get("ui"), dict) else {}
        normalized = _default_ui_config()
        raw_enabled = ui_section.get("sql_highlight_enabled")
        if isinstance(raw_enabled, bool):
            normalized["ui"]["sql_highlight_enabled"] = raw_enabled
        elif isinstance(raw_enabled, int) and raw_enabled in (0, 1):
            normalized["ui"]["sql_highlight_enabled"] = bool(raw_enabled)
        raw_hint = ui_section.get("hide_template_naming_hint")
        if isinstance(raw_hint, bool):
            normalized["ui"]["hide_template_naming_hint"] = raw_hint
        elif isinstance(raw_hint, int) and raw_hint in (0, 1):
            normalized["ui"]["hide_template_naming_hint"] = bool(raw_hint)
        raw_data_dir = ui_section.get("hide_data_dir_notice")
        if isinstance(raw_data_dir, bool):
            normalized["ui"]["hide_data_dir_notice"] = raw_data_dir
        elif isinstance(raw_data_dir, int) and raw_data_dir in (0, 1):
            normalized["ui"]["hide_data_dir_notice"] = bool(raw_data_dir)

        path = Path(app_dir) / UI_CONFIG_FILENAME
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_suffix(path.suffix + ".tmp")
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(normalized, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
    except Exception:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def is_builtin_csv_profile(name: str) -> bool:
    return name in BUILTIN_CSV_PROFILE_NAMES


def load_app_config():
    def _load_legacy_csv_profiles():
        if not os.path.exists(LEGACY_CSV_PROFILES_PATH):
            return None

        try:
            with open(LEGACY_CSV_PROFILES_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data or None
            if isinstance(data, list):
                return {"profiles": data} if data else None
        except Exception:  # noqa: BLE001
            logging.exception("Nie udało się odczytać csv_profiles.json")
        return None

    if os.path.exists(APP_CONFIG_PATH):
        try:
            with open(APP_CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            cfg = data if isinstance(data, dict) else {}
        except Exception:  # noqa: BLE001
            logging.exception("Nie udało się odczytać kkr-query2xlsx.json")
            return {}

        csv_section = cfg.get("csv")
        if not isinstance(csv_section, dict) or not csv_section:
            legacy_csv = _load_legacy_csv_profiles()
            if legacy_csv:
                cfg["csv"] = legacy_csv
        return cfg

    legacy_csv = _load_legacy_csv_profiles()
    if legacy_csv:
        return {"csv": legacy_csv}

    return {}


def save_app_config(cfg: dict) -> None:
    if not isinstance(cfg, dict):
        cfg = {}

    path = APP_CONFIG_PATH
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    tmp_path = f"{path}.tmp"
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def load_persisted_ui_lang() -> str | None:
    app_config = load_app_config()
    ui_lang = _normalize_ui_lang(app_config.get("ui_lang"))
    if ui_lang:
        return ui_lang

    legacy_lang = None
    if os.path.exists(SECURE_PATH):
        legacy_lang = _normalize_ui_lang(load_connections().get("ui_lang"))

    if legacy_lang:
        app_config["ui_lang"] = legacy_lang
        save_app_config(app_config)
        return legacy_lang

    return None


def persist_ui_lang(ui_lang: str) -> None:
    normalized = _normalize_ui_lang(ui_lang)
    if not normalized:
        return
    app_config = load_app_config()
    app_config["ui_lang"] = normalized
    save_app_config(app_config)


def load_persisted_archive_sql() -> bool:
    app_config = load_app_config()
    raw = app_config.get("archive_sql")
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, int) and raw in (0, 1):
        return bool(raw)
    if isinstance(raw, str):
        value = raw.strip().lower()
        if value in ("1", "true", "yes", "y", "on"):
            return True
        if value in ("0", "false", "no", "n", "off"):
            return False
    return False


def persist_archive_sql(enabled: bool) -> None:
    app_config = load_app_config()
    app_config["archive_sql"] = bool(enabled)
    save_app_config(app_config)


# --- Timeouts (persisted, GUI + CLI) -----------------------------------------

# Defaults: 3 minutes for DB execution/fetch + 3 minutes for export generation.
DEFAULT_DB_TIMEOUT_SECONDS = 3 * 60
DEFAULT_EXPORT_TIMEOUT_SECONDS = 3 * 60


def _normalize_timeout_seconds(value, default_seconds: int) -> int:
    """Parse timeout from config. Accepts int/float/str, clamps to >=0."""
    if value is None:
        return int(default_seconds)
    try:
        if isinstance(value, bool):
            return int(default_seconds)
        if isinstance(value, (int, float)):
            seconds = int(value)
        else:
            s = str(value).strip()
            if not s:
                return int(default_seconds)
            seconds = int(float(s))
        return max(0, seconds)
    except Exception:  # noqa: BLE001
        return int(default_seconds)


def load_persisted_db_timeout_seconds() -> int:
    app_config = load_app_config()
    timeouts = app_config.get("timeouts", {})
    if not isinstance(timeouts, dict):
        return DEFAULT_DB_TIMEOUT_SECONDS
    raw = timeouts.get("db_seconds")
    return _normalize_timeout_seconds(raw, DEFAULT_DB_TIMEOUT_SECONDS)


def persist_db_timeout_seconds(seconds: int) -> None:
    seconds = max(0, int(seconds))
    app_config = load_app_config()
    timeouts = app_config.get("timeouts")
    if not isinstance(timeouts, dict):
        timeouts = {}
    timeouts["db_seconds"] = seconds
    app_config["timeouts"] = timeouts
    save_app_config(app_config)


def load_persisted_export_timeout_seconds() -> int:
    app_config = load_app_config()
    timeouts = app_config.get("timeouts", {})
    if not isinstance(timeouts, dict):
        return DEFAULT_EXPORT_TIMEOUT_SECONDS
    raw = timeouts.get("export_seconds")
    return _normalize_timeout_seconds(raw, DEFAULT_EXPORT_TIMEOUT_SECONDS)


def persist_export_timeout_seconds(seconds: int) -> None:
    seconds = max(0, int(seconds))
    app_config = load_app_config()
    timeouts = app_config.get("timeouts")
    if not isinstance(timeouts, dict):
        timeouts = {}
    timeouts["export_seconds"] = seconds
    app_config["timeouts"] = timeouts
    save_app_config(app_config)


def _dialect_from_db_kind(db_kind: str) -> str:
    kind = (db_kind or "").strip().lower()
    if "sqlite" in kind:
        return "SQLite"
    if "postgres" in kind:
        return "Postgres"
    if "mysql" in kind:
        return "MySQL"
    if any(token in kind for token in ("mssql", "sqlserver", "sql server", "odbc")):
        return "MSSQL"
    return "SQLite"


def _apply_codeview_pygments_style(w: tk.Text, style_name: str = "vs") -> None:
    """
    Apply a Pygments style (e.g. 'vs') to Chlorophyll CodeView by configuring Text tags.
    This gives a much more coherent look (keywords/types/punctuation) than ad-hoc coloring.
    """
    try:
        from pygments.styles import get_style_by_name
        from pygments.token import Token, string_to_tokentype
    except Exception:
        return

    def _parse_style(s: str) -> dict:
        # Pygments style strings look like: "bold #RRGGBB" or "#RRGGBB"
        if not s:
            return {}
        fg = None
        for part in str(s).split():
            if part.startswith("#") and len(part) == 7:
                fg = part
        return {"foreground": fg} if fg else {}

    def _lookup_style(pyg_style, ttype):
        # climb parents until we find a defined style
        cur = ttype
        while cur not in pyg_style.styles and cur is not Token:
            cur = cur.parent
        return pyg_style.styles.get(cur, "")

    try:
        pyg_style = get_style_by_name(style_name)
    except Exception:
        return

    # keep light background (light UI)
    bg = getattr(pyg_style, "background_color", None) or "white"
    w.configure(
        bg=bg,
        fg="black",
        insertbackground="black",
        selectbackground="#cce8ff",
        selectforeground="black",
        inactiveselectbackground="#cce8ff",
    )

    # Apply per-tag colors from Pygments
    for tag in w.tag_names():
        tag_norm = tag.replace("_", ".")
        if not tag_norm.startswith("Token"):
            continue
        try:
            ttype = string_to_tokentype(tag_norm)
        except Exception:
            continue

        style_str = _lookup_style(pyg_style, ttype)
        cfg = _parse_style(style_str)
        if cfg:
            try:
                w.tag_configure(tag, **cfg)
            except Exception:
                pass

    # Light theme tweaks (readability on white background):
    # - operators (incl. MSSQL [ ]) forced to black for consistency
    # - punctuation slightly darker so quotes like '' are visible
    # - functions use a subtle accent (not pure black)
    # - types (VARCHAR/NVARCHAR/INT...) and keywords strong blue
    # - strings in readable red
    # - numbers dark for contrast
    for tag in w.tag_names():
        lt = tag.lower().replace(".", "_").replace("-", "_")

        if "punctuation" in lt:
            try:
                w.tag_configure(tag, foreground="#404040")
            except Exception:
                pass

        if "operator" in lt:
            try:
                w.tag_configure(tag, foreground="#000000")
            except Exception:
                pass

        if "name_function" in lt:
            try:
                w.tag_configure(tag, foreground="#2f5597")
            except Exception:
                pass

        if "keyword" in lt or "keyword_type" in lt:
            try:
                w.tag_configure(tag, foreground="#0000ff")
            except Exception:
                pass

        if "name_builtin" in lt:
            try:
                w.tag_configure(tag, foreground="#0000ff")
            except Exception:
                pass

        if "string" in lt:
            try:
                w.tag_configure(tag, foreground="#a31515")
            except Exception:
                pass

        if "number" in lt:
            try:
                w.tag_configure(tag, foreground="#000000")
            except Exception:
                pass


def _apply_sql_light_theme_simple(w: tk.Text) -> None:
    # base
    w.configure(
        bg="white",
        fg="black",
        insertbackground="black",
        selectbackground="#cce8ff",
        inactiveselectbackground="#cce8ff",
    )

    def _set(tag: str, **cfg):
        try:
            w.tag_configure(tag, **cfg)
        except Exception:
            pass

    for tag in w.tag_names():
        lt = tag.lower()
        norm = lt.replace(".", "_").replace("-", "_")

        # comments
        if "comment" in norm:
            _set(tag, foreground="#008000")
            continue

        # strings
        if "string" in norm:
            _set(tag, foreground="#a31515")
            continue

        # keywords (SELECT/CREATE/INSERT/...)
        if "keyword" in norm:
            _set(tag, foreground="#0000ff")
            continue

        # datatypes often come as Name.Builtin / Keyword.Type
        if "name_builtin" in norm or "keyword_type" in norm or "type" in norm:
            _set(tag, foreground="#0000ff")
            continue

        # numbers (opcjonalnie)
        if "number" in norm:
            _set(tag, foreground="#098658")
            continue


def shorten_path(path, max_len=80):
    if not path:
        return ""
    if len(path) <= max_len:
        return path
    head, tail = os.path.split(path)
    short = f"...{os.sep}{tail}"
    if len(short) > max_len:
        short = short[-max_len:]
    return short


def resolve_path(path: str) -> str:
    path = (path or "").strip()
    path = os.path.expandvars(os.path.expanduser(path))
    if path and not os.path.isabs(path):
        path = os.path.join(BASE_DIR, path)
    return os.path.abspath(os.path.normpath(path))


def _normalize_missing_path(path: str) -> str:
    normalized = os.path.abspath(os.path.normpath(path))
    if sys.platform == "win32":
        return os.path.normcase(normalized)
    return normalized


def query_path_key(path: str) -> tuple:
    resolved = resolve_path(path)
    try:
        stat_result = os.stat(resolved)
    except (FileNotFoundError, OSError):
        return ("path", _normalize_missing_path(resolved))

    inode = getattr(stat_result, "st_ino", None)
    if not inode:
        return ("path", _normalize_missing_path(resolved))

    return ("stat", stat_result.st_dev, stat_result.st_ino)


def to_storage_path(path: str) -> str:
    resolved = resolve_path(path)
    base_dir = os.path.abspath(BASE_DIR)
    if sys.platform == "win32":
        resolved_cmp = os.path.normcase(resolved)
        base_cmp = os.path.normcase(base_dir)
    else:
        resolved_cmp = resolved
        base_cmp = base_dir
    try:
        if os.path.commonpath([resolved_cmp, base_cmp]) == base_cmp:
            return os.path.relpath(resolved, base_dir)
    except ValueError:
        pass
    return resolved


def is_sql_path(path: str) -> bool:
    # Filtr w oknie wyboru to nie walidacja -> walidujemy w kodzie
    return (path or "").strip().lower().endswith(".sql")


_WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}
_INVALID_FILENAME_CHARS = r'<>:"/\|?*'


def validate_report_basename(name: str) -> tuple[bool, str, str]:
    normalized = (name or "").strip()
    if not normalized:
        return False, t("ERR_REPORT_NAME_REQUIRED"), ""

    separators = {"/", "\\"}
    if os.sep:
        separators.add(os.sep)
    if os.altsep:
        separators.add(os.altsep)
    if any(sep in normalized for sep in separators):
        return False, t("ERR_REPORT_NAME_NO_FOLDERS"), ""

    invalid_match = re.search(f"[{re.escape(_INVALID_FILENAME_CHARS)}]", normalized)
    if invalid_match:
        return False, t("ERR_REPORT_NAME_INVALID_CHARS", chars=_INVALID_FILENAME_CHARS), ""

    if any(ord(ch) < 32 for ch in normalized):
        return False, t("ERR_REPORT_NAME_CONTROL_CHARS"), ""

    if normalized[-1] in {" ", "."}:
        return False, t("ERR_REPORT_NAME_TRAILING_DOT_SPACE"), ""

    reserved_check = normalized.split(".", 1)[0].upper()
    if reserved_check in _WINDOWS_RESERVED_NAMES:
        return False, t("ERR_REPORT_NAME_RESERVED", name=normalized), ""

    if len(normalized) > 120:
        return False, t("ERR_REPORT_NAME_TOO_LONG", max_len=120), ""

    return True, "", normalized


def _looks_binary(data: bytes) -> bool:
    if not data:
        return False

    # BOM -> tekst (nawet jesli potem beda NUL-e)
    if data.startswith(b"\xef\xbb\xbf"):  # UTF-8 BOM
        return False
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):  # UTF-16 BOM
        return False
    if data.startswith((b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")):  # UTF-32 BOM
        return False

    # NUL zwykle oznacza binarke, ale dopusc UTF-16 bez BOM (NUL glownie na parzystych albo nieparzystych)
    if b"\x00" in data:
        nul_positions = [i for i, b in enumerate(data) if b == 0]
        if nul_positions:
            even = sum(1 for i in nul_positions if i % 2 == 0)
            odd = len(nul_positions) - even
            if max(even, odd) / len(nul_positions) >= 0.90:
                return False
        return True

    # Zbyt duzo znakow kontrolnych (poza whitespace) -> binarka
    allowed_controls = {0x09, 0x0A, 0x0D, 0x0C}  # \t \n \r \f
    control_bytes = sum(
        1
        for b in data
        if (b < 32 and b not in allowed_controls) or b == 0x7F
    )

    # Lagodny prog (benefit of doubt)
    return (control_bytes / len(data)) > 0.10


def validate_sql_text_file(path: str) -> tuple[bool, str]:
    """
    Very shallow validation:
    - blocks obvious mistakes: SQLite db, Excel, zip-like, binary blobs
    - allows normal text files (even if not actual SQL)
    Returns: (ok, message_if_not_ok)
    """
    if not path:
        return False, t("ERR_NO_FILE_SELECTED")

    ext = os.path.splitext(path)[1].lower()

    # Oczywiste pomylki po rozszerzeniu
    if ext in {".db", ".sqlite", ".sqlite3"}:
        return False, t("ERR_SQLFILE_IS_SQLITE_EXT")
    if ext in {".xlsx", ".xls"}:
        return False, t("ERR_SQLFILE_IS_SPREADSHEET", ext=ext)

    try:
        with open(path, "rb") as f:
            head = f.read(8192)
    except OSError as e:
        return False, t("ERR_CANNOT_OPEN_FILE", error=e)

    # Oczywiste binarki po magic bytes
    if head.startswith(b"SQLite format 3\x00"):
        return False, t("ERR_SQLFILE_IS_SQLITE_EXT")
    if head.startswith(b"PK\x03\x04"):
        return False, t("ERR_SQLFILE_IS_ZIP")
    if head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return False, t("ERR_SQLFILE_IS_OLD_OFFICE")

    if _looks_binary(head):
        return False, t("ERR_SQLFILE_IS_BINARY")

    return True, ""


def remove_bom(content: bytes) -> str:
    """
    Decode text from bytes, handling UTF-8/16/32 BOM if present.

    Also detects UTF-16 without BOM using a light heuristic:
    NUL bytes mostly on even or odd positions -> likely UTF-16-BE/LE.

    Falls back to UTF-8 when no BOM is detected and attempts legacy
    codepages when UTF-8 decoding fails (useful for queries saved with
    Windows encodings).
    """
    if not content:
        return ""

    # UTF-8 with BOM
    if content.startswith(b"\xef\xbb\xbf"):
        return content[3:].decode("utf-8")

    # UTF-32 LE / BE with BOM
    if content.startswith((b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")):
        return content.decode("utf-32")

    # UTF-16 LE / BE with BOM
    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        return content.decode("utf-16")

    # UTF-16 without BOM (heuristic): aligned NUL bytes suggest LE/BE text
    head = content[:8192]
    if b"\x00" in head:
        nul_positions = [i for i, current_byte in enumerate(head) if current_byte == 0]
        if nul_positions:
            even = sum(1 for i in nul_positions if i % 2 == 0)
            odd = len(nul_positions) - even
            if max(even, odd) / len(nul_positions) >= 0.90:
                encoding = "utf-16-le" if odd > even else "utf-16-be"
                try:
                    return content.decode(encoding)
                except UnicodeDecodeError:
                    pass

    # Default: attempt UTF-8 without BOM, then try common Windows encodings
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        for fallback_encoding in ("cp1250", "cp1252"):
            try:
                LOGGER.warning(
                    "Failed to decode bytes as UTF-8. Falling back to %s.",
                    fallback_encoding,
                )
                return content.decode(fallback_encoding)
            except UnicodeDecodeError:
                continue

        LOGGER.error(
            "Failed to decode bytes with UTF-8 and Windows fallbacks. "
            "Replacing invalid bytes.",
            exc_info=True,
        )
        return content.decode("utf-8", errors="replace")


def _normalize_connections(data):
    """Return normalized structure for saved connections with legacy support."""

    def _default_store():
        return {"connections": [], "last_selected": None, "ui_lang": None}

    if isinstance(data, str):
        legacy_str = data.strip()
        if not legacy_str:
            return _default_store()
        name = t("DEFAULT_MSSQL_NAME")
        return {
            "connections": [
                {
                    "name": name,
                    "type": "mssql_odbc",
                    "url": f"mssql+pyodbc:///?odbc_connect={quote_plus(legacy_str)}",
                    "details": {"odbc_connect": legacy_str},
                }
            ],
            "last_selected": name,
        }

    connections = []
    last_selected = None
    ui_lang = None

    if isinstance(data, dict):
        last_selected = data.get("last_selected")
        ui_lang = data.get("ui_lang")
        for item in data.get("connections", []):
            name = str(item.get("name") or "").strip()
            # url may contain secrets in legacy files - ignore it going forward
            raw_url = str(item.get("url") or "").strip()
            url = ""
            conn_type = str(item.get("type") or "custom").strip() or "custom"
            details = item.get("details")
            details = details if isinstance(details, dict) else {}

            # Legacy support: old entries may have only URL with odbc_connect
            if (not details) and raw_url.startswith("mssql+pyodbc://") and "odbc_connect=" in raw_url:
                try:
                    odbc_enc = raw_url.split("odbc_connect=", 1)[1]
                    # cut off any extra params (rare)
                    odbc_enc = odbc_enc.split("&", 1)[0]
                    odbc_raw = unquote_plus(odbc_enc).strip()
                    parsed = _parse_odbc_connect_string(odbc_raw)
                    if parsed:
                        details = parsed
                        # keep type as mssql_odbc if missing/wrong
                        if conn_type in ("", "custom"):
                            conn_type = "mssql_odbc"
                except Exception:
                    pass
            if not name:
                continue
            connections.append(
                {
                    "name": name,
                    "type": conn_type,
                    "url": url,  # may be empty for new-format entries
                    "details": details,
                }
            )

    if not connections:
        return _default_store()

    names = {c["name"] for c in connections}
    if last_selected not in names:
        last_selected = connections[0]["name"]

    return {
        "connections": connections,
        "last_selected": last_selected,
        "ui_lang": ui_lang,
    }


def load_connections(path: str | None = None):
    path = path or SECURE_PATH
    if not os.path.exists(path):
        return _normalize_connections({})

    with open(path, "r", encoding="utf-8") as file:
        raw = file.read().strip()

    if not raw:
        return _normalize_connections({})

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        data = raw

    return _normalize_connections(data)


def save_connections(store, path: str | None = None):
    path = path or SECURE_PATH
    normalized = _normalize_connections(store)
    normalized.pop("ui_lang", None)
    with open(path, "w", encoding="utf-8") as file:
        # Persist only safe structure: name/type/details (+ last_selected).
        # url may contain secrets (legacy), so we intentionally do not persist it anymore.
        safe = {
            "connections": [
                {"name": c.get("name"), "type": c.get("type"), "details": c.get("details") or {}}
                for c in (normalized.get("connections") or [])
                if (c.get("name") or "").strip()
            ],
            "last_selected": normalized.get("last_selected"),
        }
        json.dump(safe, file, ensure_ascii=False, indent=2)


def load_query_paths(queries_file: str | None = None):
    queries_file = queries_file or QUERIES_PATH
    paths = []
    if not os.path.exists(queries_file):
        return paths

    with open(queries_file, "r", encoding="utf-8") as f:
        for line in f:
            path = line.strip()
            if path:
                paths.append(path)
    return paths


def save_query_paths(paths, queries_file: str | None = None):
    queries_file = queries_file or QUERIES_PATH
    with open(queries_file, "w", encoding="utf-8") as f:
        for path in paths:
            f.write(f"{path}\n")


def bootstrap_local_files():
    """Create local config files from *.sample.* on first run.

    The app will copy (only if missing):
    - secure.sample.json  -> secure.txt
    - queries.sample.txt  -> queries.txt

    Existing files are never overwritten.
    """

    created = []

    if not os.path.exists(SECURE_PATH) and os.path.exists(SECURE_SAMPLE_PATH):
        try:
            shutil.copyfile(SECURE_SAMPLE_PATH, SECURE_PATH)
            created.append(os.path.basename(SECURE_PATH))
        except Exception:  # noqa: BLE001
            LOGGER.exception(
                "Failed to create secure.txt from secure.sample.json",
                exc_info=True,
            )

    if not os.path.exists(QUERIES_PATH) and os.path.exists(QUERIES_SAMPLE_PATH):
        try:
            shutil.copyfile(QUERIES_SAMPLE_PATH, QUERIES_PATH)
            created.append(os.path.basename(QUERIES_PATH))
        except Exception:  # noqa: BLE001
            LOGGER.exception(
                "Failed to create queries.txt from queries.sample.txt",
                exc_info=True,
            )

    return created


DEFAULT_CSV_PROFILE = {
    "name": "UTF-8 (comma)",
    "encoding": "utf-8",
    "delimiter": ",",
    "delimiter_replacement": "",
    "decimal": ".",
    "lineterminator": "\n",
    "quotechar": '"',
    "quoting": "minimal",
    "escapechar": "",
    "doublequote": True,
    "date_format": "",
}
XLSX_MAX_ROWS = 1_048_576
XLSX_MAX_COLS = 16_384


# --- Logging setup ----------------------------------------------------------

LOG_DIR: str | None = None
LOG_FILE_PATH: str | None = None
LOG_FORMATTER = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s - %(message)s")


def _attach_logger_file_handler(log_dir: str) -> bool:
    global LOG_DIR, LOG_FILE_PATH
    logger = logging.getLogger("kkr-query2xlsx")

    try:
        os.makedirs(log_dir, exist_ok=True)
        log_path = os.path.join(log_dir, "kkr-query2xlsx.log")
        handler = RotatingFileHandler(
            log_path,
            maxBytes=1_000_000,
            backupCount=3,
            encoding="utf-8",
        )
        handler.setFormatter(LOG_FORMATTER)
    except Exception:
        return False

    for existing in list(logger.handlers):
        if isinstance(existing, RotatingFileHandler):
            logger.removeHandler(existing)
            try:
                existing.close()
            except Exception:
                pass
            continue

        is_fallback_stream = bool(getattr(existing, "_kkr_fallback", False))
        if is_fallback_stream:
            logger.removeHandler(existing)
            try:
                existing.close()
            except Exception:
                pass

    logger.addHandler(handler)
    LOG_DIR = log_dir
    LOG_FILE_PATH = log_path
    return True


def _setup_logger():
    """
    Prosty globalny logger z rotacją (ok. 1 MB na plik, 3 backupy).
    Domyślnie loguje do DATA_DIR/logs.
    """
    logger = logging.getLogger("kkr-query2xlsx")
    logger.setLevel(logging.INFO)
    logger.propagate = False

    # Nie dodawaj handlerów ponownie przy imporcie
    if not logger.handlers:
        global LOG_DIR, LOG_FILE_PATH
        candidates: list[str] = [os.path.join(DATA_DIR, "logs")]

        # Ostateczny fallback: temp.
        try:
            candidates.append(os.path.join(tempfile.gettempdir(), "kkr-query2xlsx", "logs"))
        except Exception:
            pass

        attached = False
        for log_dir in candidates:
            if _attach_logger_file_handler(log_dir):
                attached = True
                break

        if not attached:
            LOG_DIR = None
            LOG_FILE_PATH = None
            if sys.stderr is not None:
                fallback_handler = logging.StreamHandler(sys.stderr)
                fallback_handler.setFormatter(LOG_FORMATTER)
                setattr(fallback_handler, "_kkr_fallback", True)
                logger.addHandler(fallback_handler)
            else:
                logger.addHandler(logging.NullHandler())

    return logger


def _startup_ask_yes_no(
    title: str,
    message: str,
    *,
    mode: Literal["gui", "console", "auto"] = "auto",
    prefer_gui: bool = False,
) -> bool:
    def _console_prompt() -> bool:
        stdin = getattr(sys, "stdin", None)
        try:
            if stdin is None or not getattr(stdin, "isatty", lambda: False)():
                return False
        except Exception:
            return False

        try:
            print(f"{title}\n{message}", file=sys.stderr)
            resp = input("Use user folder? [y/N]: ").strip().lower()
            return resp in ("y", "yes")
        except Exception:
            return False

    def _gui_prompt() -> bool:
        _require_tk()
        try:
            import tkinter as _tk
            from tkinter import messagebox as _mb

            r = _tk.Tk()
            r.withdraw()
            try:
                return bool(_mb.askyesno(title, message, parent=r))
            finally:
                r.destroy()
        except Exception:
            return _console_prompt()

    if mode == "console":
        return _console_prompt()
    if mode == "gui":
        return _gui_prompt()
    if mode == "auto":
        return _gui_prompt() if prefer_gui else _console_prompt()

    return False


def _suggest_user_data_dir() -> str:
    return get_default_user_data_dir()


def _startup_show_error(title: str, message: str) -> None:
    """Best-effort popup (dla .exe bez konsoli)."""
    try:
        import tkinter as _tk
        from tkinter import messagebox as _mb

        r = _tk.Tk()
        r.withdraw()
        _mb.showerror(title, message, parent=r)
        r.destroy()
    except Exception:
        try:
            if sys.stderr is not None:
                print(f"{title}: {message}", file=sys.stderr)
        except Exception:
            pass


LOGGER = _setup_logger()
logger = logging.getLogger(__name__)
logger.setLevel(LOGGER.level)
if not logger.handlers:
    logger.handlers = LOGGER.handlers
    logger.propagate = False


def _log_unhandled_exception(exc_type, exc_value, exc_traceback):
    """Wszystkie nieobsłużone wyjątki lądują w logu.

    Ctrl+C zostawiamy w spokoju.
    """
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    LOGGER.critical(
        "Unhandled exception",
        exc_info=(exc_type, exc_value, exc_traceback),
    )

    # Jeśli uruchomione z konsoli (python.exe) - pokaż traceback na stderr,
    # bo inaczej wygląda jak "normalne zamknięcie".
    try:
        if sys.stderr is not None:
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
    except Exception:
        pass


sys.excepthook = _log_unhandled_exception

_ORIG_UNRAISABLEHOOK = getattr(sys, "unraisablehook", None)
_KKR_UNRAISABLE_INSTALLED = False
_PKG_VER_CACHE: dict[str, str] = {}


def _pkg_version(name: str) -> str:
    if name in _PKG_VER_CACHE:
        return _PKG_VER_CACHE[name]
    try:
        ver = importlib.metadata.version(name) or ""
    except Exception:
        ver = ""
    _PKG_VER_CACHE[name] = ver
    return ver


def _install_unraisablehook() -> None:
    global _KKR_UNRAISABLE_INSTALLED
    if _KKR_UNRAISABLE_INSTALLED or _ORIG_UNRAISABLEHOOK is None:
        return

    def _debug_unraisable_enabled() -> bool:
        return os.environ.get("KKR_DEBUG_UNRAISABLE") == "1"

    def _call_orig_unraisablehook(unraisable) -> None:  # noqa: ANN001
        try:
            _ORIG_UNRAISABLEHOOK(unraisable)
        except Exception:
            pass

    def _hook(unraisable):  # noqa: ANN001
        try:
            debug_mode = _debug_unraisable_enabled()

            exc_type = getattr(unraisable, "exc_type", None)
            exc = getattr(unraisable, "exc_value", None)
            tb = getattr(unraisable, "exc_traceback", None)
            obj = getattr(unraisable, "object", None)
            err_msg = getattr(unraisable, "err_msg", None) or ""

            exc_name = getattr(exc_type, "__name__", "") if exc_type is not None else ""
            try:
                obj_repr = repr(obj) if obj is not None else ""
            except Exception:
                obj_repr = "<repr unavailable>"

            is_openpyxl_lxml_noise = (
                exc_name in {"LxmlSyntaxError", "XMLSyntaxError"}
                and (
                    "WorksheetWriter.get_stream" in obj_repr
                    or "WriteOnlyWorksheet._write_rows" in obj_repr
                )
            )

            if is_openpyxl_lxml_noise:
                (LOGGER.debug if debug_mode else LOGGER.warning)(
                    "%s noisy openpyxl/lxml unraisable exception during cleanup "
                    "(usually after failed/cancelled XLSX save; check earlier logs for root cause). "
                    "Python=%s openpyxl=%s lxml=%s; %s",
                    ("Not suppressed" if debug_mode else "Suppressed"),
                    (sys.version.split()[0] if sys.version else ""),
                    _pkg_version("openpyxl"),
                    _pkg_version("lxml"),
                    str(exc) if exc is not None else "",
                )
                if debug_mode:
                    _call_orig_unraisablehook(unraisable)
                return

            LOGGER.error(
                "Unraisable exception: %s",
                err_msg,
                exc_info=(exc_type, exc, tb),
            )

            if debug_mode:
                _call_orig_unraisablehook(unraisable)
        except Exception:
            if _debug_unraisable_enabled():
                _call_orig_unraisablehook(unraisable)

    sys.unraisablehook = _hook
    _KKR_UNRAISABLE_INSTALLED = True


_install_unraisablehook()

_ORIG_SYS_EXIT = sys.exit


def _install_debug_sys_exit_hook() -> None:
    """DEBUG: pokaż stacktrace dla sys.exit (SystemExit nie przechodzi przez sys.excepthook)."""
    def _exit(code=0):  # noqa: ANN001
        import traceback
        try:
            print(f"[DEBUG] sys.exit({code}) called", file=sys.stderr)
            traceback.print_stack()
        except Exception:
            pass
        _ORIG_SYS_EXIT(code)

    sys.exit = _exit


def _debug_enabled() -> bool:
    return os.environ.get("KKR_DEBUG_EXIT") == "1"


def _dbg(msg: str) -> None:
    if not _debug_enabled():
        return
    try:
        print(f"[DEBUG] {msg}", file=sys.stderr, flush=True)
    except Exception:
        pass


def handle_db_driver_error(exc, db_type, profile_name=None, show_message=None):
    """Show user-friendly info for missing DB drivers and log details.

    Returns True when the error was recognized and presented to the user.
    """

    # Log the *passed* exception reliably (even if this function is called outside an except block)
    LOGGER.error(
        "Database driver or library issue (type=%s, profile=%s): %s",
        db_type,
        profile_name,
        exc,
        exc_info=(type(exc), exc, exc.__traceback__),
    )

    show = show_message or messagebox.showerror
    exc_text = str(exc).lower()

    def _notify(title, msg):
        try:
            show(title, msg)
        except Exception:
            # In console mode, show may be a simple print function
            try:
                show(msg)
            except Exception:
                pass

    if db_type == "mssql_odbc":
        missing_pyodbc = isinstance(exc, (ImportError, ModuleNotFoundError)) and (
            getattr(exc, "name", "") == "pyodbc" or "pyodbc" in exc_text
        )
        if isinstance(exc, NoSuchModuleError) and "pyodbc" in exc_text:
            missing_pyodbc = True

        if missing_pyodbc:
            title, hints = _classify_mssql_conn_error(
                exc=exc,
                conn_str="",
                pyodbc_ok=False,
                drivers=[],
            )
            msg = _build_mssql_error_message(
                exc=exc,
                hints=hints,
                pyodbc_ok=False,
                drivers=[],
            )
            _notify(title, msg)
            return True

    if db_type == "postgresql":
        missing_psycopg2 = isinstance(exc, (ImportError, ModuleNotFoundError)) and (
            getattr(exc, "name", "") == "psycopg2" or "psycopg2" in exc_text
        )
        if isinstance(exc, NoSuchModuleError) and "psycopg2" in exc_text:
            missing_psycopg2 = True

        if missing_psycopg2:
            msg = t("ERR_PG_MISSING_BODY")
            _notify(t("ERR_PG_MISSING_TITLE"), msg)
            return True

    if db_type == "mysql":
        missing_pymysql = isinstance(exc, (ImportError, ModuleNotFoundError)) and (
            getattr(exc, "name", "") == "pymysql" or "pymysql" in exc_text
        )
        if isinstance(exc, NoSuchModuleError) and "pymysql" in exc_text:
            missing_pymysql = True

        if missing_pymysql:
            msg = t("ERR_MYSQL_MISSING_BODY")
            _notify(t("ERR_MYSQL_MISSING_TITLE"), msg)
            return True

    return False


def _escape_visible(text):
    return (
        text.replace("\\", "\\\\")
        .replace("\n", "\\n")
        .replace("\r", "\\r")
        .replace("\t", "\\t")
    )


def _unescape_visible(text):
    return (
        text.replace("\\n", "\n")
        .replace("\\r", "\r")
        .replace("\\t", "\t")
        .replace("\\\\", "\\")
    )


def _normalize_user_csv_profiles(raw_profiles):
    if not isinstance(raw_profiles, list):
        raw_profiles = []

    normalized_profiles = []
    seen = set()
    for raw in raw_profiles:
        name = str(raw.get("name") or "").strip()
        if not name or name in seen or is_builtin_csv_profile(name):
            continue
        seen.add(name)
        normalized_profiles.append(
            {
                "name": name,
                "encoding": raw.get("encoding") or DEFAULT_CSV_PROFILE["encoding"],
                "delimiter": raw.get("delimiter") or DEFAULT_CSV_PROFILE["delimiter"],
                "delimiter_replacement": raw.get("delimiter_replacement", ""),
                "decimal": raw.get("decimal") or DEFAULT_CSV_PROFILE["decimal"],
                "lineterminator": raw.get("lineterminator")
                or DEFAULT_CSV_PROFILE["lineterminator"],
                "quotechar": raw.get("quotechar") or DEFAULT_CSV_PROFILE["quotechar"],
                "quoting": (raw.get("quoting") or DEFAULT_CSV_PROFILE["quoting"]).lower(),
                "escapechar": raw.get("escapechar", ""),
                "doublequote": bool(
                    raw.get("doublequote")
                    if raw.get("doublequote") is not None
                    else DEFAULT_CSV_PROFILE["doublequote"]
                ),
                "date_format": raw.get("date_format", ""),
            }
        )

    return normalized_profiles


def _merge_builtin_and_user_profiles(user_profiles):
    profiles_by_name = {p["name"]: dict(p) for p in BUILTIN_CSV_PROFILES}
    for prof in user_profiles:
        name = prof.get("name")
        if not name or is_builtin_csv_profile(name):
            continue
        profiles_by_name[name] = prof

    profiles = list(profiles_by_name.values())
    _sort_csv_profiles_in_place(profiles)
    return profiles


def _sort_csv_profiles_in_place(profiles):
    profiles.sort(
        key=lambda p: (0 if is_builtin_csv_profile(p["name"]) else 1, p["name"].lower())
    )


def _read_csv_profiles_from_data(data):
    if isinstance(data, dict) and "profiles" in data:
        profiles = data.get("profiles") or []
    else:
        profiles = data if isinstance(data, list) else []

    return [p for p in profiles if p.get("name") not in BUILTIN_CSV_PROFILE_NAMES]


def get_all_csv_profiles(data):
    user_profiles = _normalize_user_csv_profiles(_read_csv_profiles_from_data(data))
    return _merge_builtin_and_user_profiles(user_profiles)


def _normalize_csv_config(data):
    profiles = []
    default_profile = None
    if isinstance(data, dict):
        default_profile = data.get("default_profile")
        profiles = data.get("profiles") or []
    elif isinstance(data, list):
        profiles = data

    user_profiles = _normalize_user_csv_profiles(profiles)
    merged_profiles = _merge_builtin_and_user_profiles(user_profiles)

    if not merged_profiles:
        merged_profiles = [DEFAULT_CSV_PROFILE.copy()]

    names = {p["name"] for p in merged_profiles}
    if default_profile not in names:
        preferred_name = PREFERRED_DEFAULT_CSV_PROFILE_BY_LANG.get(_CURRENT_LANG)
        if preferred_name in names:
            default_profile = preferred_name
        else:
            default_profile = merged_profiles[0]["name"]

    return {"default_profile": default_profile, "profiles": merged_profiles}


def load_csv_profiles():
    app_config = load_app_config()
    csv_config = app_config.get("csv", {})
    config = _normalize_csv_config(csv_config)
    config["profiles"] = get_all_csv_profiles(csv_config)
    if config["default_profile"] not in {p["name"] for p in config["profiles"]}:
        config["default_profile"] = config["profiles"][0]["name"]

    return config


def save_csv_profiles(config):
    normalized = _normalize_csv_config(config)
    data_to_save = {
        "default_profile": normalized.get("default_profile"),
        "profiles": [
            p for p in normalized.get("profiles", []) if not is_builtin_csv_profile(p.get("name", ""))
        ],
    }
    app_config = load_app_config()
    app_config["csv"] = data_to_save
    save_app_config(app_config)


def remember_last_used_csv_profile(
    profile_name: str,
    current_config: dict,
) -> dict:
    """
    Ustawia profile_name jako default_profile i zapisuje do kkr-query2xlsx.json.
    Zwraca odświeżoną konfigurację (load_csv_profiles).
    Ma być bezpieczna: jeśli zapis się nie uda, nie przerywa działania eksportu.
    """
    if not profile_name:
        return current_config or {}

    config = current_config or {}
    profiles = config.get("profiles") or []
    names = {p.get("name") for p in profiles if p.get("name")}

    if profile_name not in names:
        return config

    if config.get("default_profile") == profile_name:
        return config

    new_config = dict(config)
    new_config["default_profile"] = profile_name

    try:
        save_csv_profiles(new_config)
    except OSError as exc:
        LOGGER.warning(
            "Nie udało się zapisać domyślnego profilu CSV (%s): %s",
            profile_name,
            exc,
        )
        return config

    try:
        return load_csv_profiles()
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning(
            "Nie udało się ponownie wczytać kkr-query2xlsx.json: %s",
            exc,
        )
        return new_config


def ensure_directories(paths):
    for path in paths:
        os.makedirs(path, exist_ok=True)


def _required_work_dirs(output_dir: str) -> list[str]:
    """Single source of truth for required working directories.

    NOTE: compute subpaths via ``_build_path`` to avoid stale import-time globals
    after a runtime DATA_DIR switch.
    """
    return [
        output_dir,
        _build_path("sql_archive"),
        _build_path("templates"),
        _build_path("queries"),
    ]


def _ensure_required_work_dirs(output_dir: str) -> None:
    """Create all required working directories for the current DATA_DIR."""
    ensure_directories(_required_work_dirs(output_dir))


def bootstrap_data_dir_and_workdirs_or_exit(*, prefer_gui_prompt: bool = False) -> str:
    """Ensure working directories are ready, fallback to user dir, or exit with code 1."""
    user_dir = _suggest_user_data_dir()
    selected_data_dir = select_startup_data_dir(BASE_DIR, user_dir)
    _set_data_dir(selected_data_dir)

    output_dir = _build_path("generated_reports")
    try:
        _ensure_required_work_dirs(output_dir)
        return output_dir
    except OSError as exc:
        base_dir_raw = str(BASE_DIR)
        user_dir_raw = str(user_dir)
        current_dir_raw = str(DATA_DIR)

        if _same_path(current_dir_raw, user_dir_raw) and not _same_path(base_dir_raw, user_dir_raw):
            LOGGER.warning(
                "DATA_DIR bootstrap failed in user data dir (%s). Trying app dir (%s). Error: %s",
                user_dir_raw,
                base_dir_raw,
                exc,
                exc_info=True,
            )

            _set_data_dir(base_dir_raw)
            output_dir = _build_path("generated_reports")
            try:
                _ensure_required_work_dirs(output_dir)
                return output_dir
            except OSError as exc_base:
                _startup_show_error(
                    "No write permission",
                    "App cannot create its working folders in either location:\n\n"
                    f"- User data folder: {user_dir_raw}\n"
                    f"  Error: {exc}\n\n"
                    f"- App folder: {base_dir_raw}\n"
                    f"  Error: {exc_base}\n\n"
                    "Try:\n"
                    "- Ensure your user profile storage is available and writable.\n"
                    "- Or move the app to a writable folder (e.g. Desktop/Documents) and try again.\n",
                )
                sys.exit(1)

        msg = (
            "App cannot create its working folders under:\n"
            f"{DATA_DIR}\n\n"
            "Do you want to store app data in your user folder instead?\n"
            f"{user_dir_raw}\n\n"
            f"Error: {exc}"
        )

        already_user_dir = _same_path(DATA_DIR, user_dir_raw)
        if (not already_user_dir) and _startup_ask_yes_no(
            "No write permission",
            msg,
            mode="auto",
            prefer_gui=prefer_gui_prompt,
        ):
            _set_data_dir(user_dir_raw)
            output_dir = _build_path("generated_reports")
            try:
                _ensure_required_work_dirs(output_dir)
                return output_dir
            except OSError as exc2:
                _startup_show_error(
                    "No write permission",
                    "Still cannot create working folders in:\n"
                    f"{DATA_DIR}\n\n"
                    f"Error: {exc2}",
                )
                sys.exit(1)

        _startup_show_error(
            "No write permission",
            "App cannot create its working folders under:\n"
            f"{DATA_DIR}\n\n"
            "Move the app to a writable folder (e.g. Desktop/Documents) and try again.\n\n"
            f"Error: {exc}",
        )
        sys.exit(1)


def _expected_output_extension(output_format: str) -> str:
    return ".xlsx" if (output_format or "").lower() == "xlsx" else ".csv"


def _looks_like_directory(path: str) -> bool:
    if not path:
        return False
    if path.endswith(os.sep) or (os.altsep and path.endswith(os.altsep)):
        return True
    resolved = resolve_path(path)
    return os.path.isdir(resolved)


def _looks_like_new_directory_override(path: str) -> bool:
    if not path:
        return False
    if path.endswith(os.sep) or (os.altsep and path.endswith(os.altsep)):
        return False
    resolved = resolve_path(path)
    if os.path.exists(resolved):
        return False
    _, ext = os.path.splitext(path)
    if ext:
        return False
    return True


def normalize_output_file_path(
    *,
    output_directory: str,
    default_file_name: str,
    output_format: str,
    override_path: str | None = None,
    prefer_dir_for_extensionless_nonexistent: bool = False,
) -> tuple[str, bool]:
    """Normalize output file path and detect extension mismatch.

    Creates output_dir if missing.
    """
    expected_ext = _expected_output_extension(output_format)
    override = (override_path or "").strip()
    ext_mismatch = False

    if override:
        resolved = resolve_path(override)
        if _looks_like_directory(override) or (
            prefer_dir_for_extensionless_nonexistent
            and _looks_like_new_directory_override(override)
        ):
            output_dir = resolved
            output_file_path = os.path.join(output_dir, default_file_name)
        else:
            root, ext = os.path.splitext(resolved)
            ext_mismatch = bool(ext) and ext.lower() != expected_ext
            output_file_path = (
                root + expected_ext if ext.lower() != expected_ext else resolved
            )
            output_dir = os.path.dirname(output_file_path) or output_directory
    else:
        output_dir = output_directory
        output_file_path = os.path.join(output_directory, default_file_name)

    os.makedirs(output_dir, exist_ok=True)
    return output_file_path, ext_mismatch


def _sanitize_filename_part(value: str, max_len: int = 80) -> str:
    """
    Unicode-friendly, cross-platform-ish:
    - keeps letters (incl. Polish), digits, spaces, dot, dash, underscore
    - replaces the rest with "_"
    - collapses repeated "_" and trims ends
    """
    cleaned = (value or "").strip()
    if not cleaned:
        return "unknown"

    # normalize whitespace -> underscores
    cleaned = re.sub(r"\s+", "_", cleaned, flags=re.UNICODE)

    # keep unicode word chars + . - _
    # \w in Python (re.UNICODE) includes many unicode letters/digits/underscore
    cleaned = re.sub(r"[^\w\.\-]+", "_", cleaned, flags=re.UNICODE)

    # collapse multiple underscores and trim
    cleaned = re.sub(r"_+", "_", cleaned)
    cleaned = cleaned.strip("._- ")

    if not cleaned:
        return "unknown"

    return cleaned[:max_len]


def write_sql_archive_entry(
    sql_query: str,
    report_label: str,
    sql_source_path: str,
    output_file_path: str,
    output_format: str,
    rows_count: int,
    sql_duration: float,
    export_duration: float,
    total_duration: float,
    connection_name: str,
    connection_type: str,
) -> None:
    rows_count = max(0, int(rows_count))
    timestamp = datetime.now()
    base_label = os.path.splitext(os.path.basename(report_label or ""))[0]
    safe_label = _sanitize_filename_part(base_label)
    stamp = timestamp.strftime("%Y%m%d_%H%M%S")
    base_name = f"{stamp}__{safe_label}__{rows_count}rows"

    os.makedirs(SQL_ARCHIVE_DIR, exist_ok=True)
    sql_path = os.path.join(SQL_ARCHIVE_DIR, f"{base_name}.sql")
    metadata_path = os.path.join(SQL_ARCHIVE_DIR, f"{base_name}.json")

    with open(sql_path, "w", encoding="utf-8") as sql_file:
        sql_file.write((sql_query or "").rstrip() + "\n")

    metadata = {
        "timestamp": timestamp.isoformat(timespec="seconds"),
        "rows_count": rows_count,
        "report_label": report_label,
        "sql_source_path": sql_source_path,
        "output_file_path": output_file_path,
        "output_format": output_format,
        "sql_duration": sql_duration,
        "export_duration": export_duration,
        "total_duration": total_duration,
        "connection_name": connection_name,
        "connection_type": connection_type,
        "sql_file": os.path.basename(sql_path),
        "metadata_file": os.path.basename(metadata_path),
    }

    with open(metadata_path, "w", encoding="utf-8") as meta_file:
        json.dump(metadata, meta_file, ensure_ascii=False, indent=2)


def get_csv_profile(config, name):
    for profile in config.get("profiles", []):
        if profile.get("name") == name:
            return profile
    return None


class XlsxSizeError(ValueError):
    pass


class UserCancelledError(RuntimeError):
    pass


class PasswordRequiredError(UserCancelledError):
    def __init__(self, *, name: str, conn_type: str):
        super().__init__("Password required.")
        self.name = str(name or "")
        self.conn_type = str(conn_type or "")



def _ensure_xlsx_limits(
    rows_count: int,
    columns_count: int,
    *,
    header_rows: int = 0,
    start_row: int = 1,
    start_col: int = 1,
) -> None:
    """Ensure writing data to XLSX will not exceed Excel sheet limits.

    rows_count/columns_count refer to DATA size.
    header_rows accounts for extra header rows written above the data.
    start_row/start_col are 1-based coordinates of the top-left cell
    where header/data begins.
    """
    rows_count = max(0, int(rows_count))
    columns_count = max(0, int(columns_count))
    header_rows = max(0, int(header_rows))
    start_row = max(1, int(start_row))
    start_col = max(1, int(start_col))

    total_rows = rows_count + header_rows
    if total_rows <= 0 or columns_count <= 0:
        return

    last_row = start_row + total_rows - 1
    last_col = start_col + columns_count - 1

    if last_row > XLSX_MAX_ROWS or last_col > XLSX_MAX_COLS:
        raise XlsxSizeError(
            t(
                "ERR_XLSX_TOO_LARGE",
                # Keep message semantics: {rows}/{cols} = query result (data),
                # and show effective worksheet bounds separately.
                rows=rows_count,
                cols=columns_count,
                sheet_rows=last_row,
                sheet_cols=last_col,
                max_rows=XLSX_MAX_ROWS,
                max_cols=XLSX_MAX_COLS,
            )
        )


class QueryTimeoutError(RuntimeError):
    pass


class ExportTimeoutError(RuntimeError):
    pass


def _engine_backend_name(engine) -> str:
    try:
        return str(engine.url.get_backend_name() or "").lower()
    except Exception:  # noqa: BLE001
        return ""


def _ensure_engine_timeout_hook(engine) -> None:
    """Install a per-engine hook that can set driver-level cursor timeouts (best-effort)."""
    if getattr(engine, "_kkr_timeout_hook_installed", False):
        return
    try:

        @event.listens_for(engine, "before_cursor_execute")
        def _kkr_before_cursor_execute(
            conn, cursor, statement, parameters, context, executemany
        ):  # noqa: ANN001
            seconds = int(getattr(engine, "_kkr_db_timeout_seconds", 0) or 0)
            try:
                if hasattr(cursor, "timeout"):
                    # Important: set to 0 as well, so pooled/reused cursors don't keep old timeout.
                    cursor.timeout = max(0, seconds)
            except Exception:
                pass

    except Exception:
        pass

    setattr(engine, "_kkr_timeout_hook_installed", True)


def set_engine_db_timeout(engine, timeout_seconds: int) -> None:
    """Set current DB timeout on engine (used by the cursor hook)."""
    if engine is None:
        return
    _ensure_engine_timeout_hook(engine)
    setattr(engine, "_kkr_db_timeout_seconds", max(0, int(timeout_seconds)))


def _extract_dbapi_connection(sqlalchemy_connection):
    """Best-effort: extract the underlying DBAPI connection from a SQLAlchemy Connection."""
    if sqlalchemy_connection is None:
        return None
    try:
        if hasattr(sqlalchemy_connection, "driver_connection"):
            dbapi_conn = sqlalchemy_connection.driver_connection
        else:
            dbapi_conn = getattr(sqlalchemy_connection, "connection", None)
        if dbapi_conn is not None:
            if hasattr(dbapi_conn, "driver_connection"):
                dbapi_conn = dbapi_conn.driver_connection
            elif hasattr(dbapi_conn, "connection"):
                dbapi_conn = dbapi_conn.connection
        return dbapi_conn
    except Exception:  # noqa: BLE001
        return None


def _cancel_db_operation(dbapi_conn) -> None:
    """Best-effort cancel across DBAPIs."""
    if dbapi_conn is None:
        return
    try:
        if hasattr(dbapi_conn, "cancel"):
            dbapi_conn.cancel()
            return
    except Exception:
        pass
    try:
        if hasattr(dbapi_conn, "interrupt"):
            dbapi_conn.interrupt()
            return
    except Exception:
        pass
    try:
        dbapi_conn.close()
    except Exception:
        pass


def _apply_server_side_timeout_if_possible(
    backend: str, sqlalchemy_connection, timeout_seconds: int
) -> None:
    """Apply server-enforced timeouts when they exist (best-effort, per session)."""
    backend = (backend or "").lower()

    # Important: when timeout_seconds == 0 ("no limit"), explicitly reset session values.
    # Otherwise pooled connections can keep a previous non-zero value.

    if backend in ("postgresql", "postgres"):
        ms = int(timeout_seconds * 1000) if timeout_seconds and timeout_seconds > 0 else 0
        try:
            sqlalchemy_connection.exec_driver_sql(f"SET statement_timeout = {ms}")
        except Exception:
            pass
    if backend in ("mysql", "mariadb"):
        ms = int(timeout_seconds * 1000) if timeout_seconds and timeout_seconds > 0 else 0
        try:
            sqlalchemy_connection.exec_driver_sql(
                f"SET SESSION max_execution_time = {ms}"
            )
        except Exception:
            pass


def _deadline(timeout_seconds: int):
    if timeout_seconds and int(timeout_seconds) > 0:
        return time.monotonic() + int(timeout_seconds)
    return None


def _check_deadline(deadline, exc_type, message: str) -> None:
    if deadline is not None and time.monotonic() > deadline:
        raise exc_type(message)

def _sql_excerpt_preserve_lines(
    sql: str,
    *,
    max_chars: int = 1200,
    max_lines: int = 60,
) -> str:
    """
    UI-friendly excerpt that preserves original newlines/indentation.
    Avoids turning SQL into one line (which is unreadable and hard to debug).
    """
    s = (sql or "").rstrip()
    if not s:
        return ""

    lines = s.splitlines()
    trimmed = False
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        trimmed = True

    out = "\n".join(lines).rstrip()
    if len(out) > max_chars:
        out = out[:max_chars].rstrip()
        trimmed = True

    if trimmed:
        out += "\n...\n(Trimmed in UI, full query in kkr-query2xlsx.log)"
    return out

def _limit_text_for_widget(text: str, max_chars: int = 50000) -> str:
    s = text or ""
    if len(s) <= max_chars:
        return s
    return s[:max_chars] + "\n...\n(Trimmed in UI)"


def _dbapi_conn_from_raw_connection(raw_conn):
    """
    engine.raw_connection() returns a SQLAlchemy wrapper (e.g. ConnectionFairy).
    For cancel()/close fallback we want the underlying DBAPI connection object.
    """
    if raw_conn is None:
        return None
    for attr in ("driver_connection", "connection"):
        try:
            value = getattr(raw_conn, attr, None)
            if value is not None:
                return value
        except Exception:
            pass
    return raw_conn


def _mssql_dbapi_from_raw_connection(raw_conn):
    """Backward-compatible alias for legacy call sites."""
    return _dbapi_conn_from_raw_connection(raw_conn)


def _split_sql_statements(sql: str, backend: str) -> list[str]:
    """
    Split SQL script into statements by ';' while respecting:
    - single/double quotes
    - MySQL backticks
    - -- line comments, /* */ block comments, MySQL # comments
    - PostgreSQL dollar-quoted blocks ($$...$$ or $tag$...$tag$)
    """
    s = sql or ""
    if not s.strip():
        return []

    backend = (backend or "").lower()
    out: list[str] = []
    buf: list[str] = []

    in_sq = False
    in_dq = False
    in_bt = False
    in_line_comment = False
    in_block_comment = False
    dollar_tag: str | None = None

    i = 0
    n = len(s)

    while i < n:
        if in_line_comment:
            ch = s[i]
            buf.append(ch)
            i += 1
            if ch == "\n":
                in_line_comment = False
            continue

        if in_block_comment:
            if i + 1 < n and s[i] == "*" and s[i + 1] == "/":
                buf.append("*/")
                i += 2
                in_block_comment = False
            else:
                buf.append(s[i])
                i += 1
            continue

        if dollar_tag is not None:
            if s.startswith(dollar_tag, i):
                buf.append(dollar_tag)
                i += len(dollar_tag)
                dollar_tag = None
            else:
                buf.append(s[i])
                i += 1
            continue

        ch = s[i]
        nxt = s[i + 1] if i + 1 < n else ""

        if not (in_sq or in_dq or in_bt):
            if ch == "-" and nxt == "-":
                buf.append("--")
                i += 2
                in_line_comment = True
                continue
            if ch == "/" and nxt == "*":
                buf.append("/*")
                i += 2
                in_block_comment = True
                continue
            if backend in ("mysql", "mariadb") and ch == "#":
                buf.append("#")
                i += 1
                in_line_comment = True
                continue

            if backend in ("postgresql", "postgres") and ch == "$":
                m = re.match(r"^\$[A-Za-z_][A-Za-z0-9_]*\$", s[i:]) or re.match(
                    r"^\$\$", s[i:]
                )
                if m:
                    tag = m.group(0)
                    buf.append(tag)
                    i += len(tag)
                    dollar_tag = tag
                    continue

        if in_sq:
            if backend in ("mysql", "mariadb") and ch == "\\" and nxt:
                buf.append(ch)
                buf.append(nxt)
                i += 2
                continue
            if ch == "'" and nxt == "'":
                buf.append("''")
                i += 2
                continue
            if ch == "'":
                in_sq = False
            buf.append(ch)
            i += 1
            continue

        if in_dq:
            if ch == '"' and nxt == '"':
                buf.append('""')
                i += 2
                continue
            if ch == '"':
                in_dq = False
            buf.append(ch)
            i += 1
            continue

        if in_bt:
            if ch == "`":
                in_bt = False
            buf.append(ch)
            i += 1
            continue

        if ch == "'":
            in_sq = True
            buf.append(ch)
            i += 1
            continue
        if ch == '"':
            in_dq = True
            buf.append(ch)
            i += 1
            continue
        if backend in ("mysql", "mariadb") and ch == "`":
            in_bt = True
            buf.append(ch)
            i += 1
            continue

        if ch == ";":
            stmt = "".join(buf).strip()
            if stmt:
                out.append(stmt)
            buf = []
            i += 1
            continue

        buf.append(ch)
        i += 1

    tail = "".join(buf).strip()
    if tail:
        out.append(tail)

    return out


def has_multiple_statements(sql_query: str, backend: str) -> bool:
    """Return True only when SQL contains more than one non-empty statement."""
    return len(_split_sql_statements(sql_query, backend)) > 1


def _apply_server_side_timeout_dbapi(backend: str, cur, timeout_seconds: int) -> None:
    """Best-effort server-side timeouts for DBAPI cursor sessions."""
    backend = (backend or "").lower()
    ms = int(timeout_seconds * 1000) if timeout_seconds and timeout_seconds > 0 else 0
    try:
        if backend in ("postgresql", "postgres"):
            cur.execute(f"SET statement_timeout = {ms}")
        elif backend in ("mysql", "mariadb"):
            cur.execute(f"SET SESSION max_execution_time = {ms}")
    except Exception:
        pass


def _run_dbapi_batch_fetch_last_select(
    engine,
    sql_query: str,
    *,
    backend: str,
    timeout_seconds: int,
    cancel_event: threading.Event | None,
    timed_out_flag: dict,
    dbapi_conn_holder: dict,
) -> tuple[list, list]:
    """
    Generic batch runner for Postgres/MySQL/SQLite:
    - split script into statements
    - execute sequentially on one DBAPI cursor/session
    - return rows/cols from the LAST statement that returns rows
    """
    raw = engine.raw_connection()
    cur = None
    try:
        cur = raw.cursor()
        # Keep a cancellable object: prefer cursor.cancel() when available,
        # otherwise use the underlying DBAPI connection.
        dbapi_conn_holder["conn"] = (
            cur if hasattr(cur, "cancel") else _dbapi_conn_from_raw_connection(raw)
        )

        try:
            if hasattr(cur, "timeout"):
                cur.timeout = max(0, int(timeout_seconds or 0))
        except Exception:
            pass

        _apply_server_side_timeout_dbapi(backend, cur, int(timeout_seconds or 0))

        stmts = _split_sql_statements(sql_query or "", backend)
        if not stmts:
            return [], []

        last_rows: list = []
        last_cols: list = []
        fetch_size = 2000

        def _raise_timeout():
            raise QueryTimeoutError(
                t(
                    "ERR_QUERY_TIMEOUT",
                    minutes=max(1, int(math.ceil(max(0, int(timeout_seconds or 0)) / 60))),
                )
            )

        for stmt in stmts:
            if cancel_event is not None and cancel_event.is_set():
                raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
            if timed_out_flag.get("flag"):
                _raise_timeout()

            try:
                cur.execute(stmt)
            except (UserCancelledError, QueryTimeoutError):
                raise
            except Exception as exc:  # noqa: BLE001
                raise DBAPIError(
                    statement=stmt,
                    params=None,
                    orig=exc,
                    connection_invalidated=False,
                ) from exc

            if cur.description is not None:
                cols = [str(c[0]) for c in (cur.description or [])]
                rows: list = []
                while True:
                    if cancel_event is not None and cancel_event.is_set():
                        raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                    if timed_out_flag.get("flag"):
                        _raise_timeout()
                    try:
                        chunk = cur.fetchmany(fetch_size)
                    except Exception as exc:  # noqa: BLE001
                        raise DBAPIError(
                            statement=stmt,
                            params=None,
                            orig=exc,
                            connection_invalidated=False,
                        ) from exc
                    if not chunk:
                        break
                    rows.extend(chunk)

                last_cols = cols
                last_rows = rows

        try:
            raw.commit()
        except Exception:
            pass

        return last_rows, last_cols

    finally:
        try:
            if cur is not None:
                cur.close()
        except Exception:
            pass
        try:
            raw.close()
        except Exception:
            pass


def _run_mssql_batch_fetch_last_select(
    engine,
    sql_query: str,
    *,
    timeout_seconds: int,
    cancel_event: threading.Event | None,
    timed_out_flag: dict,
    dbapi_conn_holder: dict,
) -> tuple[list, list]:
    """
    MSSQL: execute a multi-statement batch on one DBAPI cursor/session,
    walk result sets with nextset() and return rows/columns from the LAST SELECT.

    Returns: (rows, columns)
    """
    raw = engine.raw_connection()
    cur = None
    try:
        cur = raw.cursor()
        # Keep a cancellable object: prefer cursor.cancel() when available,
        # otherwise use the underlying DBAPI connection.
        dbapi_conn_holder["conn"] = (
            cur if hasattr(cur, "cancel") else _dbapi_conn_from_raw_connection(raw)
        )

        try:
            if hasattr(cur, "timeout"):
                cur.timeout = max(0, int(timeout_seconds or 0))
        except Exception:
            pass

        batch = "SET NOCOUNT ON;\n" + (sql_query or "")

        if cancel_event is not None and cancel_event.is_set():
            raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))

        try:
            cur.execute(batch)
        except (UserCancelledError, QueryTimeoutError):
            raise
        except Exception as exc:  # noqa: BLE001
            # Wrap raw DBAPI errors so _run_query_to_rows keeps retry/timeout logic
            raise DBAPIError(
                statement=batch,
                params=None,
                orig=exc,
                connection_invalidated=False,
            ) from exc

        last_rows: list = []
        last_cols: list = []

        while True:
            if cancel_event is not None and cancel_event.is_set():
                raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
            if timed_out_flag.get("flag"):
                raise QueryTimeoutError(
                    t(
                        "ERR_QUERY_TIMEOUT",
                        minutes=max(
                            1,
                            int(
                                math.ceil(
                                    max(0, int(timeout_seconds or 0)) / 60
                                )
                            ),
                        ),
                    )
                )

            if cur.description is not None:
                cols = [str(c[0]) for c in (cur.description or [])]
                rows: list = []
                fetch_size = 2000

                while True:
                    if cancel_event is not None and cancel_event.is_set():
                        raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                    if timed_out_flag.get("flag"):
                        raise QueryTimeoutError(
                            t(
                                "ERR_QUERY_TIMEOUT",
                                minutes=max(
                                    1,
                                    int(
                                        math.ceil(
                                            max(0, int(timeout_seconds or 0)) / 60
                                        )
                                    ),
                                ),
                            )
                        )
                    try:
                        chunk = cur.fetchmany(fetch_size)
                    except (UserCancelledError, QueryTimeoutError):
                        raise
                    except Exception as exc:  # noqa: BLE001
                        raise DBAPIError(
                            statement=batch,
                            params=None,
                            orig=exc,
                            connection_invalidated=False,
                        ) from exc
                    if not chunk:
                        break
                    rows.extend(chunk)

                last_cols = cols
                last_rows = rows

            try:
                has_next = cur.nextset()
            except (UserCancelledError, QueryTimeoutError):
                raise
            except Exception as exc:  # noqa: BLE001
                raise DBAPIError(
                    statement=batch,
                    params=None,
                    orig=exc,
                    connection_invalidated=False,
                ) from exc

            if not has_next:
                break

        try:
            raw.commit()
        except Exception:
            pass

        return last_rows, last_cols

    finally:
        try:
            if cur is not None:
                cur.close()
        except Exception:
            pass
        try:
            raw.close()
        except Exception:
            pass


def _run_query_to_rows(
    engine,
    sql_query,
    timeout_seconds: int = 0,
    cancel_event: threading.Event | None = None,
):
    """
    Execute SQL with retry/deadlock handling and return:
    rows, columns, sql_duration, sql_start.

    timeout_seconds applies to the full DB phase (execute + fetch).
    0 disables the timeout.
    """
    max_retries = 3
    last_exception = None
    timeout_seconds = max(0, int(timeout_seconds or 0))
    backend = _engine_backend_name(engine)

    for attempt in range(1, max_retries + 1):
        done = None
        timed_out = {"flag": False}
        cancelled = {"flag": False}
        dbapi_conn_holder = {"conn": None}
        try:

            def watchdog():
                if timeout_seconds <= 0:
                    return
                if not done.wait(timeout_seconds):  # type: ignore[union-attr]
                    timed_out["flag"] = True
                    _cancel_db_operation(dbapi_conn_holder.get("conn"))

            def canceller():
                if cancel_event is None:
                    return
                cancel_event.wait()
                cancelled["flag"] = True

                while True:
                    if done.wait(0.05):  # type: ignore[union-attr]
                        return
                    conn = dbapi_conn_holder.get("conn")
                    if conn is not None:
                        _cancel_db_operation(conn)
                        return

            done = threading.Event()
            t_watchdog = threading.Thread(target=watchdog, daemon=True)
            t_watchdog.start()
            t_canceller = threading.Thread(target=canceller, daemon=True)
            t_canceller.start()

            sql_start = time.perf_counter()

            if backend == "mssql":
                rows, columns = _run_mssql_batch_fetch_last_select(
                    engine,
                    sql_query,
                    timeout_seconds=timeout_seconds,
                    cancel_event=cancel_event,
                    timed_out_flag=timed_out,
                    dbapi_conn_holder=dbapi_conn_holder,
                )
            elif backend in ("postgresql", "postgres", "mysql", "mariadb", "sqlite") and has_multiple_statements(
                sql_query, backend
            ):
                rows, columns = _run_dbapi_batch_fetch_last_select(
                    engine,
                    sql_query,
                    backend=backend,
                    timeout_seconds=timeout_seconds,
                    cancel_event=cancel_event,
                    timed_out_flag=timed_out,
                    dbapi_conn_holder=dbapi_conn_holder,
                )
            else:
                with engine.connect() as connection:
                    if cancel_event is not None and cancel_event.is_set():
                        raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                    set_engine_db_timeout(engine, timeout_seconds)
                    _apply_server_side_timeout_if_possible(
                        backend, connection, timeout_seconds
                    )
                    dbapi_conn_holder["conn"] = _extract_dbapi_connection(connection)
                    result = connection.execution_options(stream_results=True).execute(
                        text(sql_query)
                    )

                    if result.returns_rows:
                        rows = []
                        columns = list(result.keys())
                        fetch_size = 2000
                        while True:
                            if cancel_event is not None and cancel_event.is_set():
                                raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                            chunk = result.fetchmany(fetch_size)
                            if not chunk:
                                break
                            rows.extend(chunk)
                            if timed_out["flag"]:
                                raise QueryTimeoutError(
                                    t(
                                        "ERR_QUERY_TIMEOUT",
                                        minutes=max(
                                            1, int(math.ceil(timeout_seconds / 60))
                                        ),
                                    )
                                )
                    else:
                        rows = []
                        columns = []
            sql_end = time.perf_counter()
            sql_duration = sql_end - sql_start

            return rows, columns, sql_duration, sql_start

        except DBAPIError as e:
            msg = str(getattr(e, "orig", e))
            msg_lower = msg.lower()
            last_exception = e

            if cancelled.get("flag"):
                raise UserCancelledError(t("ERR_CANCELLED_BY_USER")) from e

            if timed_out.get("flag"):
                raise QueryTimeoutError(
                    t(
                        "ERR_QUERY_TIMEOUT",
                        minutes=max(1, int(math.ceil(timeout_seconds / 60))),
                    )
                ) from e

            if timeout_seconds > 0:
                timeout_error_signatures = (
                    # PostgreSQL
                    "statement timeout",
                    "canceling statement due to statement timeout",
                    "query canceled",
                    "query cancelled",
                    "57014",
                    # MySQL / MariaDB
                    "max_execution_time",
                    "maximum statement execution time exceeded",
                    "query execution was interrupted",
                    "execution time exceeded",
                    # ODBC / SQL Server
                    "query timeout",
                    "timeout expired",
                    "operation timed out",
                    "hyt00",
                    "hyt01",
                )
                if any(
                    signature in msg_lower for signature in timeout_error_signatures
                ):
                    raise QueryTimeoutError(
                        t(
                            "ERR_QUERY_TIMEOUT",
                            minutes=max(1, int(math.ceil(timeout_seconds / 60))),
                        )
                    ) from e

            # Keep this as a safe superset (avoid regressions across backends).
            retryable_error_signatures = (
                # SQL Server
                "1205",
                "deadlock victim",
                "lock request time out period exceeded",
                # Generic / SQLSTATE-ish
                "40001",
                "serialization failure",
                # PostgreSQL (common messages)
                "deadlock detected",
                "could not serialize access due to",
                "could not serialize access",
                # MySQL / MariaDB (common messages / codes)
                "1213",
                "deadlock found when trying to get lock",
                "lock wait timeout exceeded",
                # Older SQL Server phrasing
                "deadlocked on lock",
            )

            if any(
                signature in msg_lower for signature in retryable_error_signatures
            ) and attempt < max_retries:
                wait_seconds = 2 ** attempt
                LOGGER.warning(
                    "Deadlock-like DBAPIError while executing SQL "
                    "(attempt %s/%s, waiting %s s). Query:\n%s",
                    attempt,
                    max_retries,
                    wait_seconds,
                    sql_query,
                )
                time.sleep(wait_seconds)
                continue

            LOGGER.exception("DBAPIError while executing SQL. Query:\n%s", sql_query)
            raise

        except (UserCancelledError, QueryTimeoutError) as e:
            # Expected control-flow (user stop / timeout) -> do not log as "Unexpected error".
            last_exception = e
            if isinstance(e, UserCancelledError):
                LOGGER.info(
                    "SQL execution cancelled by user (attempt %s/%s).",
                    attempt,
                    max_retries,
                )
            else:
                LOGGER.warning(
                    "SQL execution timed out (timeout=%ss, attempt %s/%s).",
                    int(timeout_seconds or 0),
                    attempt,
                    max_retries,
                )
            raise

        except Exception:  # noqa: BLE001
            last_exception = sys.exc_info()[1]
            LOGGER.exception(
                "Unexpected error while executing SQL. Query:\n%s",
                sql_query,
            )
            raise
        finally:
            try:
                if done is not None:
                    done.set()
            except Exception:
                pass

    if last_exception:
        raise last_exception
    raise RuntimeError("Unknown SQL execution error")


def format_error_for_ui(
    exc: Exception,
    sql_query: str,
    sql_source_path: str | None = None,
    max_chars: int = 2000,
    context: Literal["sql", "export"] = "sql",
) -> str:
    """Log full error and return a shortened message for UI display."""
    # pełny traceback + SQL tylko do loga
    if context == "export":
        LOGGER.exception("Export failed. Query:\n%s", sql_query)
    else:
        LOGGER.exception("Query failed. Query:\n%s", sql_query)

    orig = getattr(exc, "orig", exc)
    orig_s = str(orig).strip() if orig is not None else ""
    first_line = (
        f"{type(orig).__name__}: {orig_s.splitlines()[0]}"
        if orig_s
        else type(exc).__name__
    )

    # pierwsza linia komunikatu z bazy (preferuj exc.orig)
    db_msg_first_line = orig_s.splitlines()[0] if orig_s else ""

    # SQL: preserve original formatting (no forced one-line flattening)
    sql_preview = _sql_excerpt_preserve_lines(sql_query, max_chars=1200, max_lines=60)

    src_line = ""
    if sql_source_path:
        src_line = f"{t('ERR_SQL_SOURCE')}\n{shorten_path(sql_source_path, max_len=220)}\n\n"

    hints: list[str] = []
    if isinstance(exc, PermissionError):
        blocked_file = getattr(exc, "filename", "")
        shortened = shorten_path(blocked_file, max_len=200) if blocked_file else ""
        exists = os.path.exists(blocked_file) if blocked_file else False

        if exists:
            path_note = (
                f"\n{t('ERR_FILE_PATH', path=shortened)}" if shortened else ""
            )
            hints.append(
                t("ERR_FILE_LOCKED", path=path_note)
            )
        else:
            hints.append(t("ERR_NO_WRITE_PERMISSION"))

    msg = (
        f"{first_line}\n\n"
        f"{t('ERR_DB_MESSAGE')}\n{db_msg_first_line}\n\n"
        f"{src_line}"
        f"{t('ERR_SQL_PREVIEW')}\n{sql_preview}\n\n"
        f"{t('ERR_FULL_LOG')}"
    )

    if hints:
        msg += "\n\n" + t("ERR_HINT_LABEL") + "\n" + "\n".join(hints)

    if len(msg) > max_chars:
        msg = msg[:max_chars] + "\n" + t("MSG_UI_TRUNCATED")

    return msg


def _safe_remove(path: str) -> None:
    try:
        if path and os.path.exists(path):
            os.remove(path)
    except Exception:
        pass


def _csv_quoting_value(quoting_name: str):
    mapping = {
        "minimal": csv.QUOTE_MINIMAL,
        "all": csv.QUOTE_ALL,
        "nonnumeric": csv.QUOTE_NONNUMERIC,
        "none": csv.QUOTE_NONE,
    }
    return mapping.get((quoting_name or "minimal").lower(), csv.QUOTE_MINIMAL)


def _coerce_csv_value(value, *, decimal_sep: str, date_format: Optional[str]):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(date_format) if date_format else value.isoformat(sep=" ")
    try:
        import datetime as _dt

        if isinstance(value, _dt.date) and not isinstance(value, _dt.datetime):
            return value.strftime(date_format) if date_format else value.isoformat()
    except Exception:
        pass

    if isinstance(value, (float, decimal.Decimal)):
        s = str(value)
        if decimal_sep and decimal_sep != ".":
            s = s.replace(".", decimal_sep)
        return s

    return str(value)


def _export_rows_to_csv(
    output_file_path: str,
    columns: list,
    rows: list,
    profile: dict,
    timeout_seconds: int,
    cancel_event: threading.Event | None = None,
) -> None:
    deadline = _deadline(timeout_seconds)
    profile = profile or DEFAULT_CSV_PROFILE

    encoding = profile.get("encoding") or DEFAULT_CSV_PROFILE["encoding"]
    delimiter = profile.get("delimiter") or DEFAULT_CSV_PROFILE["delimiter"]
    delimiter_replacement = profile.get("delimiter_replacement", "")
    quotechar = profile.get("quotechar") or DEFAULT_CSV_PROFILE["quotechar"]
    quoting = _csv_quoting_value(profile.get("quoting") or DEFAULT_CSV_PROFILE["quoting"])
    escapechar = profile.get("escapechar") or None
    doublequote = bool(profile.get("doublequote", DEFAULT_CSV_PROFILE["doublequote"]))
    lineterminator = (
        profile.get("line_terminator")
        or profile.get("lineterminator")
        or DEFAULT_CSV_PROFILE["lineterminator"]
    )
    decimal_sep = profile.get("decimal") or DEFAULT_CSV_PROFILE["decimal"]
    date_format = profile.get("date_format") or None

    with open(output_file_path, "w", encoding=encoding, newline="") as f:
        writer = csv.writer(
            f,
            delimiter=delimiter,
            quotechar=quotechar,
            quoting=quoting,
            escapechar=escapechar,
            doublequote=doublequote,
            lineterminator=lineterminator,
        )

        writer.writerow([str(c) for c in (columns or [])])

        for idx, row in enumerate(rows or [], start=1):
            if idx % 100 == 0:
                if cancel_event is not None and cancel_event.is_set():
                    raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                _check_deadline(
                    deadline,
                    ExportTimeoutError,
                    t(
                        "ERR_EXPORT_TIMEOUT",
                        minutes=max(1, int(math.ceil(timeout_seconds / 60))),
                    ),
                )

            values = list(row)
            if delimiter and delimiter_replacement:
                values = [
                    (v.replace(delimiter, delimiter_replacement) if isinstance(v, str) else v)
                    for v in values
                ]

            writer.writerow(
                [
                    _coerce_csv_value(v, decimal_sep=decimal_sep, date_format=date_format)
                    for v in values
                ]
            )

    _check_deadline(
        deadline,
        ExportTimeoutError,
        t("ERR_EXPORT_TIMEOUT", minutes=max(1, int(math.ceil(timeout_seconds / 60)))),
    )


def _export_rows_to_xlsx(
    output_file_path: str,
    columns: list,
    rows: list,
    timeout_seconds: int,
    cancel_event: threading.Event | None = None,
) -> None:
    deadline = _deadline(timeout_seconds)

    wb = Workbook(write_only=True)
    try:
        ws = wb.create_sheet(title="Results")
        ws.append([str(c) for c in (columns or [])])

        for idx, row in enumerate(rows or [], start=1):
            if idx % 100 == 0:
                if cancel_event is not None and cancel_event.is_set():
                    raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                _check_deadline(
                    deadline,
                    ExportTimeoutError,
                    t(
                        "ERR_EXPORT_TIMEOUT",
                        minutes=max(1, int(math.ceil(timeout_seconds / 60))),
                    ),
                )
            ws.append(list(row))

        wb.save(output_file_path)
    finally:
        # Avoid leaving open zip/file handles (notably visible on Python 3.13 at shutdown).
        try:
            wb.close()
        except Exception:
            pass

    _check_deadline(
        deadline,
        ExportTimeoutError,
        t("ERR_EXPORT_TIMEOUT", minutes=max(1, int(math.ceil(timeout_seconds / 60)))),
    )


def run_export(
    engine,
    sql_query,
    output_file_path,
    output_format,
    csv_profile=None,
    *,
    db_timeout_seconds: int = 0,
    export_timeout_seconds: int = 0,
):
    """Execute SQL, export the result, and return timing + row count details."""
    rows, columns, sql_duration, sql_start = _run_query_to_rows(
        engine, sql_query, timeout_seconds=db_timeout_seconds
    )
    rows_count = len(rows)
    columns_count = len(columns)
    if output_format == "xlsx":
        # pandas.DataFrame.to_excel writes a header row by default.
        # We only write the file when rows_count > 0, so reserve header row only then.
        _ensure_xlsx_limits(
            rows_count,
            columns_count,
            header_rows=(1 if rows_count else 0),
            start_row=1,
            start_col=1,
        )

    export_duration = 0.0
    if rows_count:
        export_start = time.perf_counter()
        try:
            if output_format == "xlsx":
                _export_rows_to_xlsx(
                    output_file_path,
                    columns=list(columns),
                    rows=rows,
                    timeout_seconds=int(export_timeout_seconds or 0),
                )
            else:
                profile = csv_profile or DEFAULT_CSV_PROFILE
                _export_rows_to_csv(
                    output_file_path,
                    columns=list(columns),
                    rows=rows,
                    profile=profile,
                    timeout_seconds=int(export_timeout_seconds or 0),
                )
        except Exception:
            _safe_remove(output_file_path)
            raise
        export_end = time.perf_counter()
        export_duration = export_end - export_start
        total_duration = export_end - sql_start
    else:
        total_duration = sql_duration

    return sql_duration, export_duration, total_duration, rows_count


def run_export_to_template(
    engine,
    sql_query,
    template_path,
    output_file_path,
    sheet_name,
    start_cell,
    include_header,
    *,
    db_timeout_seconds: int = 0,
    export_timeout_seconds: int = 0,
    cancel_event: Optional[threading.Event] = None,
):
    """
    Execute SQL, copy XLSX template and paste data into given sheet starting at start_cell.

    Returns the same tuple as run_export:
    (sql_duration, export_duration, total_duration, rows_count).
    """
    rows, columns, sql_duration, sql_start = _run_query_to_rows(
        engine,
        sql_query,
        timeout_seconds=db_timeout_seconds,
        cancel_event=cancel_event,
    )
    rows_count = len(rows)
    columns_count = len(columns)

    # Validate worksheet bounds before copying the template.
    # Note: we only write header/data when rows_count > 0.
    start_row, start_col = coordinate_to_tuple(start_cell)
    _ensure_xlsx_limits(
        rows_count,
        columns_count,
        header_rows=(1 if (include_header and rows_count) else 0),
        start_row=start_row,
        start_col=start_col,
    )

    export_start = time.perf_counter()
    try:
        # Zawsze kopiujemy template, nawet jeśli nie ma wierszy z SQL
        shutil.copyfile(template_path, output_file_path)

        wb = None
        try:
            if rows_count:
                deadline = _deadline(int(export_timeout_seconds or 0))
                wb = load_workbook(output_file_path)
                if sheet_name not in wb.sheetnames:
                    raise ValueError(t("ERR_TEMPLATE_MISSING_SHEET", sheet=sheet_name))

                ws = wb[sheet_name]

                data_start_row = start_row
                if include_header:
                    if cancel_event is not None and cancel_event.is_set():
                        raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                    for c_offset, col_name in enumerate(list(columns)):
                        if c_offset % 50 == 0:
                            if cancel_event is not None and cancel_event.is_set():
                                raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                            _check_deadline(
                                deadline,
                                ExportTimeoutError,
                                t(
                                    "ERR_EXPORT_TIMEOUT",
                                    minutes=max(
                                        1,
                                        int(
                                            math.ceil(
                                                int(export_timeout_seconds or 0) / 60
                                            )
                                        ),
                                    ),
                                ),
                            )
                        cell = ws.cell(row=start_row, column=start_col + c_offset)
                        cell.value = col_name
                    data_start_row = start_row + 1

                for r_offset, row in enumerate(rows):
                    if r_offset % 100 == 0:
                        if cancel_event is not None and cancel_event.is_set():
                            raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                        _check_deadline(
                            deadline,
                            ExportTimeoutError,
                            t(
                                "ERR_EXPORT_TIMEOUT",
                                minutes=max(
                                    1,
                                    int(
                                        math.ceil(
                                            int(export_timeout_seconds or 0) / 60
                                        )
                                    ),
                                ),
                            ),
                        )
                    for c_offset, value in enumerate(list(row)):
                        cell = ws.cell(
                            row=data_start_row + r_offset,
                            column=start_col + c_offset,
                        )
                        cell.value = value
                if cancel_event is not None and cancel_event.is_set():
                    raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                wb.save(output_file_path)
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
    except Exception:
        _safe_remove(output_file_path)
        raise

    export_end = time.perf_counter()
    export_duration = export_end - export_start
    total_duration = export_end - sql_start

    return sql_duration, export_duration, total_duration, rows_count


def run_console(
    engine,
    output_directory,
    selected_connection,
    archive_sql: bool,
    output_override: str | None = None,
):
    sql_query_file_paths = load_query_paths()
    csv_config = load_csv_profiles()

    db_timeout_seconds = load_persisted_db_timeout_seconds()
    export_timeout_seconds = load_persisted_export_timeout_seconds()

    sql_query_file_path = None
    if sql_query_file_paths:
        print(t("CONSOLE_AVAILABLE_FILES"))
        print(t("CONSOLE_CUSTOM_PATH"))
        for idx, path in enumerate(sql_query_file_paths, start=1):
            print(f"{idx}: {path}")

        while True:
            try:
                selection = int(
                    input(
                        t(
                            "CONSOLE_PROMPT_SELECT",
                            max_idx=len(sql_query_file_paths),
                        )
                    )
                )
                if selection == 0:
                    custom_path = input(t("CONSOLE_PROMPT_CUSTOM_PATH")).strip()
                    resolved = resolve_path(custom_path)
                    if not os.path.isfile(resolved):
                        print(t("CONSOLE_FILE_NOT_FOUND"))
                        continue
                    ok, msg = validate_sql_text_file(resolved)
                    if not ok:
                        print(f"ERROR: {msg}")
                        continue
                    sql_query_file_path = resolved
                    break
                if 1 <= selection <= len(sql_query_file_paths):
                    resolved = resolve_path(sql_query_file_paths[selection - 1])
                    if not os.path.isfile(resolved):
                        print(t("CONSOLE_FILE_NOT_FOUND"))
                        continue
                    ok, msg = validate_sql_text_file(resolved)
                    if not ok:
                        print(f"ERROR: {msg}")
                        continue
                    sql_query_file_path = resolved
                    break
                print(
                    t(
                        "CONSOLE_SELECT_RANGE",
                        max_idx=len(sql_query_file_paths),
                    )
                )
            except ValueError:
                print(t("CONSOLE_INVALID_INPUT"))
    else:
        print(t("CONSOLE_NO_QUERIES"))
        while True:
            custom_path = input(t("CONSOLE_PROMPT_CUSTOM_PATH")).strip()
            resolved = resolve_path(custom_path)
            if os.path.isfile(resolved):
                ok, msg = validate_sql_text_file(resolved)
                if not ok:
                    print(f"ERROR: {msg}")
                    continue
                sql_query_file_path = resolved
                break
            print(t("CONSOLE_FILE_NOT_FOUND"))

    while True:
        output_format = (
            input(t("CONSOLE_PROMPT_FORMAT")).strip().lower()
        )
        if output_format in ["xlsx", "csv"]:
            break
        print(t("CONSOLE_INVALID_FORMAT"))

    selected_csv_profile = get_csv_profile(csv_config, csv_config.get("default_profile"))
    if output_format == "csv":
        profiles = csv_config.get("profiles", [])
        profile_names = [p.get("name") for p in profiles]
        default_profile_name = csv_config.get("default_profile") or profile_names[0]

        print(t("CONSOLE_AVAILABLE_CSV_PROFILES"))
        for idx, name in enumerate(profile_names, start=1):
            default_marker = (
                t("CONSOLE_DEFAULT_MARKER") if name == default_profile_name else ""
            )
            print(f"{idx}: {name}{default_marker}")

        while True:
            selection = input(
                t("CONSOLE_PROMPT_CSV_PROFILE")
            ).strip()
            if not selection:
                break
            if selection.isdigit():
                idx = int(selection)
                if 1 <= idx <= len(profile_names):
                    selected_csv_profile = profiles[idx - 1]
                    break
            print(t("CONSOLE_INVALID_SELECTION"))

    if output_format == "csv" and selected_csv_profile:
        prof_name = (selected_csv_profile.get("name") or "").strip()
        if prof_name:
            csv_config = remember_last_used_csv_profile(prof_name, csv_config)

    with open(sql_query_file_path, "rb") as file:
        content = file.read()

    sql_query = remove_bom(content).strip()

    base_name = os.path.basename(sql_query_file_path)
    output_file_name = os.path.splitext(base_name)[0] + (".xlsx" if output_format == "xlsx" else ".csv")
    output_file_path, _ = normalize_output_file_path(
        output_directory=output_directory,
        default_file_name=output_file_name,
        output_format=output_format,
        override_path=(output_override.strip() if output_override else None),
        prefer_dir_for_extensionless_nonexistent=True,
    )

    try:
        sql_dur, export_dur, total_dur, rows_count = run_export(
            engine,
            sql_query,
            output_file_path,
            output_format,
            csv_profile=selected_csv_profile,
            db_timeout_seconds=db_timeout_seconds,
            export_timeout_seconds=export_timeout_seconds,
        )
    except XlsxSizeError as exc:
        print(str(exc))
        return
    except QueryTimeoutError as exc:
        print(str(exc))
        return
    except ExportTimeoutError as exc:
        print(str(exc))
        return

    if archive_sql:
        try:
            write_sql_archive_entry(
                sql_query=sql_query,
                report_label=os.path.basename(sql_query_file_path),
                sql_source_path=sql_query_file_path,
                output_file_path=output_file_path,
                output_format=output_format,
                rows_count=rows_count,
                sql_duration=sql_dur,
                export_duration=export_dur,
                total_duration=total_dur,
                connection_name=selected_connection.get("name", ""),
                connection_type=selected_connection.get("type", ""),
            )
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("SQL archive failed: %s", exc, exc_info=True)

    if rows_count > 0:
        print(t("CONSOLE_SAVED_PATH", path=output_file_path))
    else:
        print(t("CONSOLE_NO_ROWS"))
    print(t("CONSOLE_SQL_TIME", seconds=sql_dur))
    if rows_count > 0:
        print(t("CONSOLE_EXPORT_TIME", fmt=output_format, seconds=export_dur))
        print(t("CONSOLE_TOTAL_TIME", seconds=total_dur))


def _create_mssql_frame(parent):
    frame = tk.LabelFrame(parent, text=t("FRAME_MSSQL"))
    frame.grid(row=3, column=0, columnspan=4, sticky="we", padx=10, pady=(5, 0))
    for idx in range(4):
        frame.columnconfigure(idx, weight=1)

    driver_var = tk.StringVar(value="ODBC Driver 17 for SQL Server")
    server_var = tk.StringVar()
    database_var = tk.StringVar()
    username_var = tk.StringVar()
    password_var = tk.StringVar()
    remember_password_var = tk.BooleanVar(value=False)
    trusted_var = tk.BooleanVar(value=False)
    encrypt_var = tk.BooleanVar(value=True)
    trust_cert_var = tk.BooleanVar(value=True)

    tk.Label(frame, text=t("LBL_ODBC_DRIVER")).grid(
        row=0, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=driver_var, width=30).grid(
        row=0, column=1, columnspan=3, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_SERVER")).grid(
        row=1, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=server_var, width=30).grid(
        row=1, column=1, columnspan=3, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_DATABASE_NAME")).grid(
        row=2, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=database_var, width=30).grid(
        row=2, column=1, columnspan=3, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_LOGIN")).grid(
        row=3, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=username_var, width=25).grid(
        row=3, column=1, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_PASSWORD")).grid(
        row=3, column=2, sticky="w", padx=5, pady=(5, 0)
    )
    password_entry = tk.Entry(frame, textvariable=password_var, show="*", width=25)
    password_entry.grid(row=3, column=3, sticky="we", padx=5, pady=(5, 0))

    def _sync_password_state():  # OFF -> disabled + clear
        if trusted_var.get():
            # Trusted auth: password not used
            remember_password_var.set(False)
            password_var.set("")
            password_entry.config(state=tk.DISABLED)
            return

        if remember_password_var.get():
            password_entry.config(state=tk.NORMAL)
        else:
            password_var.set("")
            password_entry.config(state=tk.DISABLED)

    tk.Checkbutton(
        frame, text=t("CHK_TRUSTED"), variable=trusted_var
    ).grid(
        row=4, column=1, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Checkbutton(frame, text=t("CHK_ENCRYPT"), variable=encrypt_var).grid(
        row=4, column=2, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Checkbutton(frame, text=t("CHK_TRUST_CERT"), variable=trust_cert_var).grid(
        row=4, column=3, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Checkbutton(
        frame,
        text=t("CHK_REMEMBER_PASSWORD"),
        variable=remember_password_var,
        command=_sync_password_state,
    ).grid(row=5, column=2, columnspan=2, sticky="w", padx=5, pady=(5, 0))

    def _on_trusted_toggle(*_):  # noqa: ANN001
        _sync_password_state()

    trusted_var.trace_add("write", _on_trusted_toggle)
    _sync_password_state()

    return frame, {
        "driver": driver_var,
        "server": server_var,
        "database": database_var,
        "username": username_var,
        "password": password_var,
        "remember_password": remember_password_var,
        "trusted": trusted_var,
        "encrypt": encrypt_var,
        "trust_cert": trust_cert_var,
    }


def _create_pg_frame(parent):
    frame = tk.LabelFrame(parent, text=t("FRAME_POSTGRES"))
    frame.grid(row=4, column=0, columnspan=4, sticky="we", padx=10, pady=(5, 0))
    for idx in range(4):
        frame.columnconfigure(idx, weight=1)

    host_var = tk.StringVar(value="localhost")
    port_var = tk.StringVar(value="5432")
    db_var = tk.StringVar()
    user_var = tk.StringVar()
    password_var = tk.StringVar()
    remember_password_var = tk.BooleanVar(value=False)

    tk.Label(frame, text=t("LBL_HOST")).grid(
        row=0, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=host_var).grid(
        row=0, column=1, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_PORT")).grid(
        row=0, column=2, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=port_var, width=8).grid(
        row=0, column=3, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_DATABASE")).grid(
        row=1, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=db_var).grid(
        row=1, column=1, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_USER")).grid(
        row=1, column=2, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=user_var).grid(
        row=1, column=3, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_PASSWORD")).grid(
        row=2, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    password_entry = tk.Entry(frame, textvariable=password_var, show="*")
    password_entry.grid(row=2, column=1, sticky="we", padx=5, pady=(5, 0))

    tk.Checkbutton(
        frame,
        text=t("CHK_REMEMBER_PASSWORD"),
        variable=remember_password_var,
        command=lambda: (
            password_entry.config(state=tk.NORMAL)
            if remember_password_var.get()
            else (password_var.set(""), password_entry.config(state=tk.DISABLED))
        ),
    ).grid(row=3, column=1, columnspan=3, sticky="w", padx=5, pady=(5, 0))

    # init OFF -> disabled
    password_var.set("")
    password_entry.config(state=tk.DISABLED)

    return frame, {
        "host": host_var,
        "port": port_var,
        "database": db_var,
        "user": user_var,
        "password": password_var,
        "remember_password": remember_password_var,
    }


def _create_mysql_frame(parent):
    frame = tk.LabelFrame(parent, text=t("FRAME_MYSQL"))
    frame.grid(row=6, column=0, columnspan=4, sticky="we", padx=10, pady=(5, 0))
    for idx in range(4):
        frame.columnconfigure(idx, weight=1)

    host_var = tk.StringVar(value="localhost")
    port_var = tk.StringVar(value="3306")
    db_var = tk.StringVar()
    user_var = tk.StringVar()
    password_var = tk.StringVar()
    remember_password_var = tk.BooleanVar(value=False)

    tk.Label(frame, text=t("LBL_HOST")).grid(
        row=0, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=host_var).grid(
        row=0, column=1, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_PORT")).grid(
        row=0, column=2, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=port_var, width=8).grid(
        row=0, column=3, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_DATABASE")).grid(
        row=1, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=db_var).grid(
        row=1, column=1, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_USER")).grid(
        row=1, column=2, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=user_var).grid(
        row=1, column=3, sticky="we", padx=5, pady=(5, 0)
    )

    tk.Label(frame, text=t("LBL_PASSWORD")).grid(
        row=2, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    password_entry = tk.Entry(frame, textvariable=password_var, show="*")
    password_entry.grid(row=2, column=1, sticky="we", padx=5, pady=(5, 0))

    tk.Checkbutton(
        frame,
        text=t("CHK_REMEMBER_PASSWORD"),
        variable=remember_password_var,
        command=lambda: (
            password_entry.config(state=tk.NORMAL)
            if remember_password_var.get()
            else (password_var.set(""), password_entry.config(state=tk.DISABLED))
        ),
    ).grid(row=3, column=1, columnspan=3, sticky="w", padx=5, pady=(5, 0))

    # init OFF -> disabled
    password_var.set("")
    password_entry.config(state=tk.DISABLED)

    return frame, {
        "host": host_var,
        "port": port_var,
        "database": db_var,
        "user": user_var,
        "password": password_var,
        "remember_password": remember_password_var,
    }


def _create_sqlite_frame(parent):
    frame = tk.LabelFrame(parent, text=t("FRAME_SQLITE"))
    frame.grid(row=5, column=0, columnspan=4, sticky="we", padx=10, pady=(5, 0))
    frame.columnconfigure(1, weight=1)

    path_var = tk.StringVar()
    sqlite_filetypes = [
        (t("FILETYPE_SQLITE"), "*.db *.sqlite *.sqlite3"),
        (t("FILETYPE_ALL"), "*.*"),
    ]

    def _choose_sqlite_path():
        chosen_path = filedialog.askopenfilename(
            title=t("TITLE_SELECT_SQLITE"),
            filetypes=sqlite_filetypes,
        )
        if not chosen_path:
            if not messagebox.askyesno(
                t("ASK_CREATE_SQLITE"),
                t("ASK_CREATE_SQLITE_BODY"),
            ):
                return
            chosen_path = filedialog.asksaveasfilename(
                title=t("TITLE_CREATE_SQLITE"),
                defaultextension=".sqlite",
                filetypes=sqlite_filetypes,
            )
        if chosen_path:
            path_var.set(os.path.abspath(chosen_path))

    tk.Label(frame, text=t("LBL_SQLITE_PATH")).grid(
        row=0, column=0, sticky="w", padx=5, pady=(5, 0)
    )
    tk.Entry(frame, textvariable=path_var).grid(
        row=0, column=1, sticky="we", padx=5, pady=(5, 0)
    )
    tk.Button(
        frame,
        text=t("BTN_SELECT"),
        command=_choose_sqlite_path,
    ).grid(row=0, column=2, padx=5, pady=(5, 0))

    return frame, {"path": path_var}

def _parse_odbc_connect_string(conn_str: str) -> dict:
    """
    Best-effort parser for ODBC connection strings like:
    DRIVER={ODBC Driver 17 for SQL Server};SERVER=...;DATABASE=...;UID=...;PWD=...;Encrypt=yes;TrustServerCertificate=yes;Trusted_Connection=yes
    Returns normalized details keys used by the GUI.
    """
    raw = (conn_str or "").strip()
    if not raw:
        return {}

    parts: dict[str, str] = {}
    for chunk in raw.split(";"):
        chunk = chunk.strip()
        if not chunk or "=" not in chunk:
            continue
        k, v = chunk.split("=", 1)
        k = k.strip().lower()
        v = v.strip()
        parts[k] = v

    def _strip_braces(v: str) -> str:
        v = (v or "").strip()
        if len(v) >= 2 and v.startswith("{") and v.endswith("}"):
            return v[1:-1].strip()
        return v

    def _as_bool(v: str, default: bool) -> bool:
        if v is None:
            return default
        s = str(v).strip().lower()
        if s in ("yes", "true", "1", "y"):
            return True
        if s in ("no", "false", "0", "n"):
            return False
        return default

    driver = _strip_braces(parts.get("driver", ""))
    server = parts.get("server", "") or parts.get("data source", "")
    database = parts.get("database", "") or parts.get("initial catalog", "")
    username = parts.get("uid", "") or parts.get("user id", "")
    password = parts.get("pwd", "") or parts.get("password", "")

    trusted = _as_bool(parts.get("trusted_connection"), False) or _as_bool(
        parts.get("trusted_connection".replace("_", "")), False
    )
    # Some users may have "Integrated Security=SSPI"
    trusted = trusted or (
        str(parts.get("integrated security", "")).strip().lower()
        in ("sspi", "true", "yes", "1")
    )

    encrypt = _as_bool(parts.get("encrypt"), True)
    trust_cert = _as_bool(parts.get("trustservercertificate"), True)

    return {
        "driver": driver,
        "server": server,
        "database": database,
        "username": username,
        "password": password,
        "trusted": trusted,
        "encrypt": encrypt,
        "trust_server_certificate": trust_cert,
        "odbc_connect": raw,
    }


def _load_connection_details(conn_type, vars_by_type, details):
    details = details or {}
    if conn_type == "mssql_odbc":
        # Legacy support: some saved connections only have {"odbc_connect": "..."}
        if (
            "odbc_connect" in details
            and not any(
                details.get(k)
                for k in ("driver", "server", "database", "username", "password")
            )
        ):
            details = _parse_odbc_connect_string(details.get("odbc_connect", "")) or details

        vars_by_type["driver"].set(details.get("driver", ""))
        vars_by_type["server"].set(details.get("server", ""))
        vars_by_type["database"].set(details.get("database", ""))
        vars_by_type["username"].set(details.get("username", ""))
        vars_by_type["password"].set(details.get("password", ""))
        remember_var = vars_by_type.get("remember_password")
        if remember_var is not None:
            remember_var.set(bool(details.get("password")))
        vars_by_type["trusted"].set(bool(details.get("trusted", False)))
        vars_by_type["encrypt"].set(bool(details.get("encrypt", True)))
        vars_by_type["trust_cert"].set(bool(details.get("trust_server_certificate", True)))
    elif conn_type == "postgresql":
        vars_by_type["host"].set(details.get("host", ""))
        vars_by_type["port"].set(str(details.get("port", "5432")))
        vars_by_type["database"].set(details.get("database", ""))
        vars_by_type["user"].set(details.get("user", ""))
        vars_by_type["password"].set(details.get("password", ""))
        remember_var = vars_by_type.get("remember_password")
        if remember_var is not None:
            remember_var.set(bool(details.get("password")))
    elif conn_type == "mysql":
        vars_by_type["host"].set(details.get("host", "localhost"))
        vars_by_type["port"].set(str(details.get("port", "3306")))
        vars_by_type["database"].set(details.get("database", ""))
        vars_by_type["user"].set(details.get("user", ""))
        vars_by_type["password"].set(details.get("password", ""))
        remember_var = vars_by_type.get("remember_password")
        if remember_var is not None:
            remember_var.set(bool(details.get("password")))
    elif conn_type == "sqlite":
        vars_by_type["path"].set(details.get("path", ""))


def _build_runtime_url(conn_type: str, details: dict, runtime_password: str | None) -> str:
    """
    Build SQLAlchemy URL from stored details + runtime password (not persisted).
    IMPORTANT: runtime_password may be None/"" for trusted auth.
    """
    details = details or {}
    pwd = runtime_password or ""

    if conn_type == "mssql_odbc":
        driver = str(details.get("driver") or "").strip()
        server = str(details.get("server") or "").strip()
        database = str(details.get("database") or "").strip()
        username = str(details.get("username") or "").strip()
        trusted = bool(details.get("trusted"))
        encrypt = bool(details.get("encrypt", True))
        trust_cert = bool(details.get("trust_server_certificate", True))

        parts = [
            f"DRIVER={{{driver}}}",
            f"SERVER={server}",
            f"DATABASE={database}",
        ]
        if trusted:
            parts.append("Trusted_Connection=yes")
        else:
            parts.append(f"UID={username}")
            parts.append(f"PWD={pwd}")
        if encrypt:
            parts.append("Encrypt=yes")
        if trust_cert:
            parts.append("TrustServerCertificate=yes")
        odbc = ";".join(parts)
        return f"mssql+pyodbc:///?odbc_connect={quote_plus(odbc)}"

    if conn_type == "postgresql":
        host = str(details.get("host") or "").strip()
        port = str(details.get("port") or "5432").strip() or "5432"
        database = str(details.get("database") or "").strip()
        user = str(details.get("user") or "").strip()
        return (
            f"postgresql+psycopg2://{quote_plus(user)}:{quote_plus(pwd)}@"
            f"{host}:{port}/{database}"
        )

    if conn_type == "mysql":
        host = str(details.get("host") or "").strip()
        port = str(details.get("port") or "3306").strip() or "3306"
        database = str(details.get("database") or "").strip()
        user = str(details.get("user") or "").strip()
        return (
            f"mysql+pymysql://{quote_plus(user)}:{quote_plus(pwd)}@"
            f"{host}:{port}/{database}"
        )

    # sqlite
    db_path = str(details.get("path") or "").strip()
    return f"sqlite:///{os.path.abspath(db_path)}"


def _resolve_password_for_storage(vars_by_type, password, *, confirm_store: bool):
    remember_var = vars_by_type.get("remember_password")
    if remember_var is None:
        return password
    if not remember_var.get():
        return ""
    if not password:
        return ""
    if not confirm_store:
        # e.g. "Test connection" - never ask, never store (caller decides)
        return password
    if messagebox.askyesno(
        t("WARN_REMEMBER_PASSWORD_TITLE"),
        t("WARN_REMEMBER_PASSWORD_BODY", path=SECURE_PATH),
    ):
        return password
    remember_var.set(False)  # user refused storing -> turn off remember
    return ""


def _build_connection_entry(conn_type, vars_by_type, name, *, confirm_store: bool):
    base_entry = {"name": name, "type": conn_type, "url": "", "details": {}}
    if conn_type == "mssql_odbc":
        driver = vars_by_type["driver"].get().strip()
        server = vars_by_type["server"].get().strip()
        database = vars_by_type["database"].get().strip()
        username = vars_by_type["username"].get().strip()
        password = vars_by_type["password"].get()  # may be empty when remember=OFF (disabled entry)
        remember_password = vars_by_type.get("remember_password")
        password_for_storage = _resolve_password_for_storage(
            vars_by_type, password, confirm_store=confirm_store
        )

        if not driver or not server or not database:
            messagebox.showerror(
                t("ERR_DATA_TITLE"),
                t("ERR_FILL_ODBC"),
            )
            return None

        parts = [
            f"DRIVER={{{driver}}}",
            f"SERVER={server}",
            f"DATABASE={database}",
        ]

        if vars_by_type["trusted"].get():
            parts.append("Trusted_Connection=yes")
        else:
            if not username:
                messagebox.showerror(
                    t("ERR_DATA_TITLE"),
                    t("ERR_LOGIN_REQUIRED"),
                )
                return None
            if remember_password is not None and remember_password.get() and not password:
                messagebox.showerror(
                    t("ERR_DATA_TITLE"),
                    t("ERR_LOGIN_REQUIRED"),
                )
                return None
            # URL/test uses the runtime password; storage uses password_for_storage
            parts.append(f"UID={username}")
            if password:
                parts.append(f"PWD={password}")

        if vars_by_type["encrypt"].get():
            parts.append("Encrypt=yes")
        if vars_by_type["trust_cert"].get():
            parts.append("TrustServerCertificate=yes")

        # Do NOT persist URL anymore (may contain secrets). Built at runtime.
        base_entry["url"] = ""
        base_entry["details"] = {
            "driver": driver,
            "server": server,
            "database": database,
            "username": username,
            "password": password_for_storage,
            "trusted": vars_by_type["trusted"].get(),
            "encrypt": vars_by_type["encrypt"].get(),
            "trust_server_certificate": vars_by_type["trust_cert"].get(),
        }
        return base_entry

    if conn_type == "postgresql":
        host = vars_by_type["host"].get().strip()
        port = vars_by_type["port"].get().strip() or "5432"
        database = vars_by_type["database"].get().strip()
        user = vars_by_type["user"].get().strip()
        password = vars_by_type["password"].get()
        remember_password = vars_by_type.get("remember_password")
        password_for_storage = _resolve_password_for_storage(
            vars_by_type, password, confirm_store=confirm_store
        )

        if not host or not database or not user:
            messagebox.showerror(
                t("ERR_DATA_TITLE"),
                t("ERR_FILL_PG"),
            )
            return None
        if remember_password is not None and remember_password.get() and not password:
            messagebox.showerror(
                t("ERR_DATA_TITLE"),
                t("ERR_LOGIN_REQUIRED"),
            )
            return None

        # Do NOT persist URL anymore (may contain secrets). Built at runtime.
        base_entry["url"] = ""
        base_entry["details"] = {
            "host": host,
            "port": port,
            "database": database,
            "user": user,
            "password": password_for_storage,
        }
        return base_entry

    if conn_type == "mysql":
        host = vars_by_type["host"].get().strip()
        port = vars_by_type["port"].get().strip() or "3306"
        database = vars_by_type["database"].get().strip()
        user = vars_by_type["user"].get().strip()
        password = vars_by_type["password"].get()
        remember_password = vars_by_type.get("remember_password")
        password_for_storage = _resolve_password_for_storage(
            vars_by_type, password, confirm_store=confirm_store
        )

        if not host or not database or not user:
            messagebox.showerror(
                t("ERR_DATA_TITLE"),
                t("ERR_FILL_MYSQL"),
            )
            return None
        if remember_password is not None and remember_password.get() and not password:
            messagebox.showerror(
                t("ERR_DATA_TITLE"),
                t("ERR_LOGIN_REQUIRED"),
            )
            return None

        # Do NOT persist URL anymore (may contain secrets). Built at runtime.
        base_entry["url"] = ""
        base_entry["details"] = {
            "host": host,
            "port": port,
            "database": database,
            "user": user,
            "password": password_for_storage,
        }
        return base_entry

    db_path = vars_by_type["path"].get().strip()
    if not db_path:
        messagebox.showerror(
            t("ERR_DATA_TITLE"), t("ERR_FILL_SQLITE")
        )
        return None
    # Safe even if persisted, but keep consistent: runtime build
    base_entry["url"] = ""
    base_entry["details"] = {"path": os.path.abspath(db_path)}
    return base_entry


# Map internal DB types to user-friendly labels (UI-only)
def _db_type_labels():
    return {
        "mssql_odbc": t("DB_TYPE_MSSQL"),
        "postgresql": t("DB_TYPE_PG"),
        "mysql": t("DB_TYPE_MYSQL"),
        "sqlite": t("DB_TYPE_SQLITE"),
    }


def _db_type_by_label():
    labels = _db_type_labels()
    return {label: key for key, label in labels.items()}


def connection_name_exists(store, name):
    return any(conn.get("name") == name for conn in store.get("connections", []))


def generate_unique_connection_name(store, base):
    if not base:
        return ""
    if not connection_name_exists(store, base):
        return base
    counter = 2
    while True:
        candidate = f"{base} {counter}"
        if not connection_name_exists(store, candidate):
            return candidate
        counter += 1


def _reset_connection_details(conn_type, vars_by_type):
    if conn_type == "mssql_odbc":
        vars_by_type["driver"].set("ODBC Driver 17 for SQL Server")
        vars_by_type["server"].set("")
        vars_by_type["database"].set("")
        vars_by_type["username"].set("")
        vars_by_type["password"].set("")
        remember_var = vars_by_type.get("remember_password")
        if remember_var is not None:
            remember_var.set(False)
        vars_by_type["trusted"].set(False)
        vars_by_type["encrypt"].set(True)
        vars_by_type["trust_cert"].set(True)
    elif conn_type == "postgresql":
        vars_by_type["host"].set("localhost")
        vars_by_type["port"].set("5432")
        vars_by_type["database"].set("")
        vars_by_type["user"].set("")
        vars_by_type["password"].set("")
        remember_var = vars_by_type.get("remember_password")
        if remember_var is not None:
            remember_var.set(False)
    elif conn_type == "mysql":
        vars_by_type["host"].set("localhost")
        vars_by_type["port"].set("3306")
        vars_by_type["database"].set("")
        vars_by_type["user"].set("")
        vars_by_type["password"].set("")
        remember_var = vars_by_type.get("remember_password")
        if remember_var is not None:
            remember_var.set(False)
    elif conn_type == "sqlite":
        vars_by_type["path"].set("")


def _build_connection_dialog_ui(root):
    dlg = tk.Toplevel(root)
    apply_app_icon(dlg)
    dlg.title("")
    dlg.transient(root)
    dlg.grab_set()

    for idx in range(4):
        dlg.columnconfigure(idx, weight=1)

    hint_label = tk.Label(
        dlg,
        text="",
        justify="left",
    )
    hint_label.grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 0))

    tk.Label(dlg, text=t("LBL_CONN_NAME")).grid(
        row=1, column=0, sticky="w", padx=10, pady=(10, 0)
    )
    name_var = tk.StringVar(value="")
    tk.Entry(dlg, textvariable=name_var, width=40).grid(
        row=1, column=1, columnspan=3, sticky="we", padx=10, pady=(10, 0)
    )

    tk.Label(dlg, text=t("LBL_DB_TYPE")).grid(
        row=2, column=0, sticky="w", padx=10, pady=(5, 0)
    )
    type_var = tk.StringVar(value="mssql_odbc")
    db_type_labels = _db_type_labels()
    type_label_var = tk.StringVar(value=db_type_labels["mssql_odbc"])
    type_combo = ttk.Combobox(
        dlg,
        textvariable=type_label_var,
        values=list(db_type_labels.values()),
        state="readonly",
    )
    type_combo.grid(row=2, column=1, sticky="w", padx=10, pady=(5, 0))

    mssql_frame, mssql_vars = _create_mssql_frame(dlg)
    pg_frame, pg_vars = _create_pg_frame(dlg)
    mysql_frame, mysql_vars = _create_mysql_frame(dlg)
    sqlite_frame, sqlite_vars = _create_sqlite_frame(dlg)

    type_sections = {
        "mssql_odbc": (mssql_frame, mssql_vars),
        "postgresql": (pg_frame, pg_vars),
        "mysql": (mysql_frame, mysql_vars),
        "sqlite": (sqlite_frame, sqlite_vars),
    }

    def show_type_frame(*_):  # noqa: ANN001
        for frame, _ in type_sections.values():
            frame.grid_remove()
        frame, _ = type_sections.get(type_var.get(), (mssql_frame, mssql_vars))
        frame.grid()

    def update_type_from_label(*_):  # noqa: ANN001
        type_var.set(_db_type_by_label().get(type_label_var.get(), "mssql_odbc"))
        show_type_frame()

    show_type_frame()
    type_var.trace_add("write", show_type_frame)
    type_label_var.trace_add("write", update_type_from_label)

    return dlg, name_var, type_var, type_label_var, type_sections, show_type_frame, hint_label


def _load_existing_connection(
    name_var,
    type_var,
    type_label_var,
    type_sections,
    show_type_frame,
    connection_name,
    get_connection_by_name,
):
    existing = get_connection_by_name(connection_name)
    if not existing:
        return
    name_var.set(existing.get("name", ""))
    conn_type = existing.get("type", "mssql_odbc")
    type_var.set(conn_type)
    db_type_labels = _db_type_labels()
    type_label_var.set(db_type_labels.get(conn_type, db_type_labels["mssql_odbc"]))
    section = type_sections.get(conn_type)
    if section:
        show_type_frame()
        _load_connection_details(conn_type, section[1], existing.get("details"))


def _build_and_test_connection_entry(
    name, conn_type, type_sections, create_engine_from_entry, handle_db_driver_error, parent=None
):
    section = type_sections.get(conn_type)
    if not section:
        messagebox.showerror(
            t("ERR_DATA_TITLE"),
            t("ERR_INVALID_CONN_TYPE"),
            parent=parent,
        )
        return None

    new_entry = _build_connection_entry(conn_type, section[1], name, confirm_store=False)
    if not new_entry:
        return None

    try:
        engine = create_engine_from_entry(new_entry, parent=parent)
        with engine.connect() as connection:
            connection.execute(text("SELECT 1"))
    except Exception as exc:  # noqa: BLE001
        if isinstance(exc, PasswordRequiredError):
            messagebox.showwarning(
                t("ERR_PASSWORD_REQUIRED_TITLE"),
                t("ERR_PASSWORD_REQUIRED_BODY", name=exc.name or name),
                parent=parent,
            )
            return None
        if handle_db_driver_error(exc, conn_type, name):
            return None
        LOGGER.exception(
            "Connection creation failed for %s (%s)",
            name,
            conn_type,
            exc_info=True,
        )
        title, msg = _format_connection_error(
            conn_type=conn_type,
            exc=exc,
            details=new_entry.get("details") or {},
        )
        messagebox.showerror(title, msg, parent=parent)
        return None

    return new_entry


def _validate_connection_name(name, mode, store, original_name=None):
    if not name:
        messagebox.showerror(t("ERR_DATA_TITLE"), t("ERR_CONN_NAME_REQUIRED"))
        return False
    if mode == "edit":
        if name != (original_name or "") and connection_name_exists(store, name):
            messagebox.showerror(t("ERR_DATA_TITLE"), t("ERR_CONN_NAME_EXISTS"))
            return False
    elif connection_name_exists(store, name):
        messagebox.showerror(t("ERR_DATA_TITLE"), t("ERR_CONN_NAME_EXISTS"))
        return False
    return True


def _replace_or_append_connection(connections_state, new_entry, original_name=None, allow_replace=True):
    replaced = False
    connections = connections_state["store"].setdefault("connections", [])
    if allow_replace:
        target_name = original_name or new_entry.get("name")
        for idx, c in enumerate(connections):
            if c.get("name") == target_name:
                connections[idx] = new_entry
                replaced = True
                break

    if allow_replace and not replaced:
        for idx, c in enumerate(connections):
            if c.get("name") == new_entry.get("name"):
                connections[idx] = new_entry
                replaced = True
                break

    if not replaced:
        connections.append(new_entry)


def _save_connection_without_test(
    name_var,
    type_var,
    type_sections,
    connections_state,
    mode,
    original_name,
    set_selected_connection,
    persist_connections,
    refresh_connection_combobox,
):
    name = name_var.get().strip()
    if not _validate_connection_name(name, mode, connections_state["store"], original_name):
        return False

    conn_type = type_var.get()
    section = type_sections.get(conn_type)
    if not section:
        messagebox.showerror(t("ERR_DATA_TITLE"), t("ERR_INVALID_CONN_TYPE"))
        return False

    new_entry = _build_connection_entry(conn_type, section[1], name, confirm_store=True)
    if not new_entry:
        return False

    _replace_or_append_connection(
        connections_state,
        new_entry,
        original_name=original_name,
        allow_replace=(mode == "edit"),
    )

    set_selected_connection(name)
    persist_connections()
    refresh_connection_combobox()

    messagebox.showinfo(
        t("INFO_CONN_SAVED_TITLE"),
        t("INFO_CONN_SAVED_NO_TEST"),
    )
    return True


def _save_connection_from_dialog(
    name_var,
    type_var,
    type_sections,
    connections_state,
    mode,
    original_name,
    set_selected_connection,
    persist_connections,
    refresh_connection_combobox,
    apply_selected_connection,
    handle_db_driver_error,
    create_engine_from_entry,
    parent=None,
):
    name = name_var.get().strip()
    if not _validate_connection_name(name, mode, connections_state["store"], original_name):
        return False

    # Build+test WITHOUT storing prompt, then rebuild for storage with prompt (if needed)
    tested_entry = _build_and_test_connection_entry(
        name,
        type_var.get(),
        type_sections,
        create_engine_from_entry,
        handle_db_driver_error,
        parent=parent,
    )
    if not tested_entry:
        return False

    conn_type = type_var.get()
    section = type_sections.get(conn_type)
    if not section:
        messagebox.showerror(t("ERR_DATA_TITLE"), t("ERR_INVALID_CONN_TYPE"))
        return False
    new_entry = _build_connection_entry(conn_type, section[1], name, confirm_store=True)
    if not new_entry:
        return False

    _replace_or_append_connection(
        connections_state,
        new_entry,
        original_name=original_name,
        allow_replace=(mode == "edit"),
    )

    set_selected_connection(name)
    persist_connections()
    refresh_connection_combobox()
    apply_selected_connection(show_success=True)
    return True


def open_connection_dialog_gui(
    root,
    connections_state,
    selected_connection_var,
    get_connection_by_name,
    set_selected_connection,
    persist_connections,
    refresh_connection_combobox,
    apply_selected_connection,
    handle_db_driver_error,
    create_engine_from_entry,
    mode="edit",
    source_name=None,
):
    if mode in ("edit", "duplicate") and not (source_name or selected_connection_var.get()):
        messagebox.showerror(t("ERR_NO_CONNECTION_TITLE"), t("ERR_NO_CONNECTION_BODY"))
        return

    (
        dlg,
        name_var,
        type_var,
        type_label_var,
        type_sections,
        show_type_frame,
        hint_label,
    ) = _build_connection_dialog_ui(root)

    original_name = None
    active_name = source_name or selected_connection_var.get()
    if mode in ("edit", "duplicate") and active_name:
        _load_existing_connection(
            name_var,
            type_var,
            type_label_var,
            type_sections,
            show_type_frame,
            active_name,
            get_connection_by_name,
        )
        original_name = active_name

    name_edit_tracking = {"enabled": True}
    name_edited = {"value": False}

    def set_name_value(value):
        name_edit_tracking["enabled"] = False
        name_var.set(value)
        name_edit_tracking["enabled"] = True

    def update_dialog_copy():
        if mode == "edit":
            current_name = name_var.get().strip() or (original_name or "")
            dlg.title(t("TITLE_CONN_DIALOG_EDIT", name=current_name))
            hint_label.config(text=t("CONN_DIALOG_HINT_EDIT", name=current_name))
        elif mode == "duplicate":
            dlg.title(t("TITLE_CONN_DIALOG_DUPLICATE", name=original_name or ""))
            hint_label.config(text=t("CONN_DIALOG_HINT_DUPLICATE", name=original_name or ""))
        else:
            dlg.title(t("TITLE_CONN_DIALOG_NEW"))
            hint_label.config(text=t("CONN_DIALOG_HINT_NEW"))

    def on_name_change(*_):  # noqa: ANN001
        if name_edit_tracking["enabled"]:
            name_edited["value"] = True
            if mode == "edit":
                update_dialog_copy()

    name_var.trace_add("write", on_name_change)

    if mode == "duplicate" and original_name:
        duplicate_base = f"{original_name} {t('CONN_DUPLICATE_SUFFIX')}"
        set_name_value(generate_unique_connection_name(connections_state["store"], duplicate_base))

    if mode == "new":
        name_edited["value"] = False

    update_dialog_copy()

    if mode == "new":
        def on_type_change(*_):  # noqa: ANN001
            conn_type = type_var.get()
            section = type_sections.get(conn_type)
            if section:
                _reset_connection_details(conn_type, section[1])
            if not name_edited["value"]:
                set_name_value("")

        type_var.trace_add("write", on_type_change)
        on_type_change()

    def on_save(*_):
        saved = _save_connection_from_dialog(
            name_var,
            type_var,
            type_sections,
            connections_state,
            mode,
            original_name,
            set_selected_connection,
            persist_connections,
            refresh_connection_combobox,
            apply_selected_connection,
            handle_db_driver_error,
            create_engine_from_entry,
            parent=dlg,
        )
        if saved:
            dlg.destroy()

    def on_save_without_test(*_):
        saved = _save_connection_without_test(
            name_var,
            type_var,
            type_sections,
            connections_state,
            mode,
            original_name,
            set_selected_connection,
            persist_connections,
            refresh_connection_combobox,
        )
        if saved:
            dlg.destroy()

    def on_cancel(*_):
        dlg.destroy()

    def on_test(*_):
        name = name_var.get().strip()
        if not name:
            messagebox.showerror(t("ERR_DATA_TITLE"), t("ERR_CONN_NAME_REQUIRED"))
            return
        conn_type = type_var.get()
        tested = _build_and_test_connection_entry(
            name,
            conn_type,
            type_sections,
            create_engine_from_entry,
            handle_db_driver_error,
            parent=dlg,
        )
        if tested:
            messagebox.showinfo(t("INFO_CONN_TEST_OK_TITLE"), t("INFO_CONN_TEST_OK_BODY"))

    button_frame = tk.Frame(dlg)
    button_frame.grid(row=7, column=0, columnspan=4, pady=10)

    if mode == "edit":
        def on_duplicate(*_):  # noqa: ANN001
            dlg.destroy()
            open_connection_dialog_gui(
                root,
                connections_state,
                selected_connection_var,
                get_connection_by_name,
                set_selected_connection,
                persist_connections,
                refresh_connection_combobox,
                apply_selected_connection,
                handle_db_driver_error,
                create_engine_from_entry,
                mode="duplicate",
                source_name=original_name,
            )

        tk.Button(
            button_frame,
            text=t("BTN_DUPLICATE_CONNECTION"),
            command=on_duplicate,
            width=14,
        ).pack(side="left", padx=(0, 5))

    tk.Button(
        button_frame,
        text=t("BTN_TEST_CONNECTION"),
        command=on_test,
        width=14,
    ).pack(side="left", padx=(0, 5))

    tk.Button(button_frame, text=t("BTN_SAVE"), command=on_save, width=14).pack(
        side="left", padx=(0, 5)
    )
    tk.Button(
        button_frame,
        text=t("BTN_SAVE_NO_TEST"),
        command=on_save_without_test,
        width=18,
    ).pack(side="left", padx=(0, 5))
    tk.Button(
        button_frame, text=t("BTN_CANCEL"), command=on_cancel, width=12
    ).pack(side="left")

    dlg.bind("<Return>", on_save)
    dlg.bind("<Escape>", on_cancel)

    _center_window(dlg, root)


def _init_csv_profile_vars():
    return {
        "name": tk.StringVar(value=""),
        "encoding": tk.StringVar(value="utf-8"),
        "delimiter": tk.StringVar(value=","),
        "delimiter_replacement": tk.StringVar(value=""),
        "decimal": tk.StringVar(value="."),
        "lineterminator": tk.StringVar(value="\\n"),
        "quotechar": tk.StringVar(value='"'),
        "quoting": tk.StringVar(value="minimal"),
        "escapechar": tk.StringVar(value=""),
        "doublequote": tk.BooleanVar(value=True),
        "date_format": tk.StringVar(value=""),
        "date_preview": tk.StringVar(value=""),
    }


def _validate_date_format(raw):
    if not raw:
        example = datetime.now().isoformat(sep=" ", timespec="seconds")
        return (
            True,
            t("CSV_PROFILE_DATE_DEFAULT", example=example),
        )
    try:
        example = datetime.now().strftime(raw)
    except (ValueError, TypeError):
        return (
            False,
            t("CSV_PROFILE_DATE_INVALID"),
        )
    return True, t("CSV_PROFILE_DATE_PREVIEW", example=example)


def _load_csv_profile(vars_dict, profile):
    vars_dict["name"].set(profile.get("name", ""))
    vars_dict["encoding"].set(profile.get("encoding", ""))
    vars_dict["delimiter"].set(profile.get("delimiter", ","))
    vars_dict["delimiter_replacement"].set(profile.get("delimiter_replacement", ""))
    vars_dict["decimal"].set(profile.get("decimal", "."))
    vars_dict["lineterminator"].set(_escape_visible(profile.get("lineterminator", "\\n")))
    vars_dict["quotechar"].set(profile.get("quotechar", '"'))
    vars_dict["quoting"].set((profile.get("quoting") or "minimal").lower())
    vars_dict["escapechar"].set(profile.get("escapechar", ""))
    vars_dict["doublequote"].set(bool(profile.get("doublequote", True)))
    vars_dict["date_format"].set(profile.get("date_format", ""))


def _read_csv_profile(vars_dict):
    valid_format, preview = _validate_date_format(vars_dict["date_format"].get())
    vars_dict["date_preview"].set(preview)
    if not valid_format:
        return None
    return {
        "name": vars_dict["name"].get().strip(),
        "encoding": vars_dict["encoding"].get().strip()
        or DEFAULT_CSV_PROFILE["encoding"],
        "delimiter": vars_dict["delimiter"].get() or DEFAULT_CSV_PROFILE["delimiter"],
        "delimiter_replacement": vars_dict["delimiter_replacement"].get(),
        "decimal": vars_dict["decimal"].get() or DEFAULT_CSV_PROFILE["decimal"],
        "lineterminator": _unescape_visible(
            vars_dict["lineterminator"].get() or DEFAULT_CSV_PROFILE["lineterminator"]
        ),
        "quotechar": vars_dict["quotechar"].get()
        or DEFAULT_CSV_PROFILE["quotechar"],
        "quoting": vars_dict["quoting"].get() or DEFAULT_CSV_PROFILE["quoting"],
        "escapechar": vars_dict["escapechar"].get(),
        "doublequote": bool(vars_dict["doublequote"].get()),
        "date_format": vars_dict["date_format"].get(),
    }


def _csv_field_help():
    return {
        "name": (
            t("CSV_HELP_NAME_TITLE"),
            t("CSV_HELP_NAME_BODY"),
        ),
        "encoding": (
            t("CSV_HELP_ENCODING_TITLE"),
            t("CSV_HELP_ENCODING_BODY"),
        ),
        "delimiter": (
            t("CSV_HELP_DELIMITER_TITLE"),
            t("CSV_HELP_DELIMITER_BODY"),
        ),
        "delimiter_replacement": (
            t("CSV_HELP_DELIM_REPLACE_TITLE"),
            t("CSV_HELP_DELIM_REPLACE_BODY"),
        ),
        "decimal": (
            t("CSV_HELP_DECIMAL_TITLE"),
            t("CSV_HELP_DECIMAL_BODY"),
        ),
        "lineterminator": (
            t("CSV_HELP_LINETERM_TITLE"),
            t("CSV_HELP_LINETERM_BODY"),
        ),
        "quotechar": (
            t("CSV_HELP_QUOTECHAR_TITLE"),
            t("CSV_HELP_QUOTECHAR_BODY"),
        ),
        "quoting": (
            t("CSV_HELP_QUOTING_TITLE"),
            t("CSV_HELP_QUOTING_BODY"),
        ),
        "escapechar": (
            t("CSV_HELP_ESCAPECHAR_TITLE"),
            t("CSV_HELP_ESCAPECHAR_BODY"),
        ),
        "doublequote": (
            t("CSV_HELP_DOUBLEQUOTE_TITLE"),
            t("CSV_HELP_DOUBLEQUOTE_BODY"),
        ),
        "date_format": (
            t("CSV_HELP_DATE_FORMAT_TITLE"),
            t("CSV_HELP_DATE_FORMAT_BODY"),
        ),
    }


def _build_csv_profile_list_ui(dlg):
    list_var = tk.StringVar(value=[])
    listbox = tk.Listbox(
        dlg,
        listvariable=list_var,
        width=40,
        height=8,
        activestyle="dotbox",
        exportselection=False,
    )
    listbox.grid(row=1, column=0, sticky="nsew", padx=(10, 5), pady=10)

    normal_font = tkfont.nametofont(listbox.cget("font"))
    bold_font = normal_font.copy()
    bold_font.configure(weight="bold")
    listbox._fonts = {"normal": normal_font, "bold": bold_font}

    scrollbar = tk.Scrollbar(dlg, orient="vertical", command=listbox.yview)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.grid(row=1, column=1, sticky="ns", pady=10)
    return listbox, list_var


def _build_csv_profile_form_ui(form_frame, form_vars, field_help):
    def show_field_help(key):
        title, message = field_help.get(key, (t("CSV_PROFILE_INFO_SAVED_TITLE"), ""))
        messagebox.showinfo(title, message)

    def add_info_button(row, key):
        tk.Button(
            form_frame, text=t("INFO_ICON"), width=2, command=lambda k=key: show_field_help(k)
        ).grid(row=row, column=2, sticky="w", padx=(5, 0))

    date_preview_var = form_vars["date_preview"]

    def update_date_preview(*_args):  # noqa: ANN001
        valid, preview = _validate_date_format(form_vars["date_format"].get())
        date_preview_var.set(preview)
        return valid

    widgets = []

    tk.Label(form_frame, text=t("CSV_PROFILE_NAME")).grid(row=0, column=0, sticky="w")
    name_entry = tk.Entry(form_frame, textvariable=form_vars["name"])
    name_entry.grid(
        row=0, column=1, sticky="we"
    )
    widgets.append((name_entry, "normal"))
    add_info_button(0, "name")

    tk.Label(form_frame, text=t("CSV_PROFILE_ENCODING")).grid(row=1, column=0, sticky="w")
    encoding_entry = tk.Entry(form_frame, textvariable=form_vars["encoding"])
    encoding_entry.grid(
        row=1, column=1, sticky="we"
    )
    widgets.append((encoding_entry, "normal"))
    add_info_button(1, "encoding")

    tk.Label(form_frame, text=t("CSV_PROFILE_DELIMITER")).grid(row=2, column=0, sticky="w")
    delimiter_entry = tk.Entry(form_frame, textvariable=form_vars["delimiter"], width=5)
    delimiter_entry.grid(
        row=2, column=1, sticky="w"
    )
    widgets.append((delimiter_entry, "normal"))
    add_info_button(2, "delimiter")

    tk.Label(form_frame, text=t("CSV_PROFILE_DELIM_REPLACE")).grid(
        row=3, column=0, sticky="w"
    )
    delimiter_replacement_entry = tk.Entry(
        form_frame, textvariable=form_vars["delimiter_replacement"], width=5
    )
    delimiter_replacement_entry.grid(
        row=3, column=1, sticky="w"
    )
    widgets.append((delimiter_replacement_entry, "normal"))
    add_info_button(3, "delimiter_replacement")

    tk.Label(form_frame, text=t("CSV_PROFILE_DECIMAL")).grid(
        row=4, column=0, sticky="w"
    )
    decimal_entry = tk.Entry(form_frame, textvariable=form_vars["decimal"], width=5)
    decimal_entry.grid(
        row=4, column=1, sticky="w"
    )
    widgets.append((decimal_entry, "normal"))
    add_info_button(4, "decimal")

    tk.Label(form_frame, text=t("CSV_PROFILE_LINE_END")).grid(
        row=5, column=0, sticky="w"
    )
    lineterminator_entry = tk.Entry(
        form_frame, textvariable=form_vars["lineterminator"], width=10
    )
    lineterminator_entry.grid(
        row=5, column=1, sticky="w"
    )
    widgets.append((lineterminator_entry, "normal"))
    add_info_button(5, "lineterminator")

    tk.Label(form_frame, text=t("CSV_PROFILE_QUOTECHAR")).grid(
        row=6, column=0, sticky="w"
    )
    quotechar_entry = tk.Entry(form_frame, textvariable=form_vars["quotechar"], width=5)
    quotechar_entry.grid(
        row=6, column=1, sticky="w"
    )
    widgets.append((quotechar_entry, "normal"))
    add_info_button(6, "quotechar")

    tk.Label(form_frame, text=t("CSV_PROFILE_QUOTING")).grid(row=7, column=0, sticky="w")
    quoting_combo = ttk.Combobox(
        form_frame,
        textvariable=form_vars["quoting"],
        values=["minimal", "all", "nonnumeric", "none"],
        state="readonly",
        width=15,
    )
    quoting_combo.grid(row=7, column=1, sticky="w")
    widgets.append((quoting_combo, "readonly"))
    add_info_button(7, "quoting")

    tk.Label(
        form_frame,
        text=t("CSV_PROFILE_ESCAPECHAR"),
    ).grid(row=8, column=0, sticky="w")
    escapechar_entry = tk.Entry(form_frame, textvariable=form_vars["escapechar"], width=5)
    escapechar_entry.grid(
        row=8, column=1, sticky="w"
    )
    widgets.append((escapechar_entry, "normal"))
    add_info_button(8, "escapechar")
    tk.Label(
        form_frame,
        text=t("CSV_PROFILE_ESCAPE_HINT"),
        fg="gray",
    ).grid(row=8, column=3, sticky="w", padx=(5, 0))

    doublequote_check = tk.Checkbutton(
        form_frame,
        text=t("CSV_PROFILE_DOUBLEQUOTE"),
        variable=form_vars["doublequote"],
    )
    doublequote_check.grid(row=9, column=0, columnspan=2, sticky="w")
    widgets.append((doublequote_check, "normal"))
    add_info_button(9, "doublequote")

    tk.Label(form_frame, text=t("CSV_PROFILE_DATE_FORMAT")).grid(
        row=10, column=0, sticky="w"
    )
    date_format_entry = tk.Entry(form_frame, textvariable=form_vars["date_format"])
    date_format_entry.grid(
        row=10, column=1, columnspan=1, sticky="we"
    )
    widgets.append((date_format_entry, "normal"))
    add_info_button(10, "date_format")
    tk.Label(
        form_frame,
        textvariable=date_preview_var,
        fg="gray",
    ).grid(row=11, column=0, columnspan=4, sticky="w", pady=(2, 0))
    form_vars["date_format"].trace_add("write", update_date_preview)
    update_date_preview()

    return update_date_preview, widgets


def _create_csv_profiles_dialog(root, csv_profile_state):
    dlg = tk.Toplevel(root)
    apply_app_icon(dlg)
    dlg.title(t("CSV_PROFILE_DIALOG_TITLE"))
    dlg.transient(root)
    dlg.grab_set()
    dlg.resizable(True, True)

    working_profiles = [dict(p) for p in csv_profile_state["config"].get("profiles", [])]
    _sort_csv_profiles_in_place(working_profiles)
    default_profile_var = tk.StringVar(value=csv_profile_state["config"].get("default_profile"))
    display_default_var = tk.StringVar(
        value=t("CSV_DEFAULT_PROFILE_LABEL", name=default_profile_var.get() or "")
    )

    dlg.columnconfigure(1, weight=1)
    dlg.rowconfigure(1, weight=1)

    tk.Label(dlg, textvariable=display_default_var, anchor="w").grid(
        row=0, column=0, columnspan=3, sticky="we", padx=10, pady=(10, 0)
    )

    return dlg, working_profiles, default_profile_var, display_default_var


def _refresh_csv_profile_list(
    listbox, list_var, working_profiles, default_profile_var, display_default_var
):
    display = []
    for prof in working_profiles:
        suffix = (
            f" {t('CSV_PROFILE_DEFAULT_SUFFIX')}"
            if prof["name"] == default_profile_var.get()
            else ""
        )
        builtin_tag = (
            f" {t('CSV_PROFILE_BUILTIN_SUFFIX')}"
            if is_builtin_csv_profile(prof["name"])
            else ""
        )
        display.append(f"{prof['name']}{builtin_tag}{suffix}")
    list_var.set(display)

    fonts = getattr(listbox, "_fonts", {})
    item_font_supported = True
    for idx, prof in enumerate(working_profiles):
        font_to_use = (
            fonts.get("bold") if is_builtin_csv_profile(prof["name"]) else fonts.get("normal")
        )
        if item_font_supported and font_to_use is not None:
            try:
                listbox.itemconfig(idx, font=font_to_use)
            except tk.TclError:
                # Some Tk builds (incl. some Windows bundles) do not support per-item fonts in Listbox.
                item_font_supported = False
                break

    display_default_var.set(
        t("CSV_DEFAULT_PROFILE_LABEL", name=default_profile_var.get() or "")
    )


def _ensure_default_profile(default_profile_var, working_profiles, preferred_name=None):
    names = {p["name"] for p in working_profiles}
    if default_profile_var.get() in names:
        return
    if preferred_name and preferred_name in names:
        default_profile_var.set(preferred_name)
    else:
        default_profile_var.set(working_profiles[0]["name"])


def _read_profile_from_form_or_warn(form_vars):
    profile = _read_csv_profile(form_vars)
    if profile is None:
        messagebox.showwarning(
            t("WARN_TITLE"),
            t("CSV_PROFILE_INVALID_DATE"),
        )
    elif not profile.get("name"):
        messagebox.showwarning(t("WARN_TITLE"), t("CSV_PROFILE_WARNING_EMPTY"))
        profile = None
    return profile


def _save_csv_profile_config(
    csv_profile_state, default_profile_var, working_profiles, refresh_csv_profile_controls
):
    config = {
        "default_profile": default_profile_var.get() or working_profiles[0]["name"],
        "profiles": working_profiles,
    }
    save_csv_profiles(config)
    csv_profile_state["config"] = load_csv_profiles()
    refresh_csv_profile_controls(csv_profile_state["config"].get("default_profile"))


def open_csv_profiles_manager_gui(
    root,
    csv_profile_state,
    selected_csv_profile_var,
    refresh_csv_profile_controls,
):
    current_profile_name = {"name": None}
    unsaved_changes = False

    dlg, working_profiles, default_profile_var, display_default_var = _create_csv_profiles_dialog(
        root, csv_profile_state
    )

    listbox, list_var = _build_csv_profile_list_ui(dlg)

    form_frame = tk.LabelFrame(dlg, text=t("CSV_PROFILE_TITLE"), padx=10, pady=10)
    form_frame.grid(row=1, column=2, sticky="nsew", padx=(5, 10), pady=10)
    form_frame.columnconfigure(1, weight=1)
    form_frame.columnconfigure(3, weight=1)

    form_vars = _init_csv_profile_vars()
    update_date_preview, form_widgets = _build_csv_profile_form_ui(
        form_frame, form_vars, _csv_field_help()
    )
    builtin_notice_var = tk.StringVar(value="")
    tk.Label(
        form_frame,
        textvariable=builtin_notice_var,
        fg="gray",
    ).grid(row=12, column=0, columnspan=4, sticky="w", pady=(2, 0))

    def refresh_list():
        _sort_csv_profiles_in_place(working_profiles)
        _refresh_csv_profile_list(
            listbox,
            list_var,
            working_profiles,
            default_profile_var,
            display_default_var,
        )

        selected_name = current_profile_name.get("name")
        if selected_name:
            for idx, prof in enumerate(working_profiles):
                if prof.get("name") == selected_name:
                    listbox.selection_clear(0, tk.END)
                    listbox.selection_set(idx)
                    listbox.see(idx)
                    break

    def set_editable_state(_is_builtin):
        for widget, normal_state in form_widgets:
            widget.configure(state=tk.DISABLED if _is_builtin else normal_state)

    def update_builtin_indicator(idx=None):
        is_builtin = False
        if idx is not None and 0 <= idx < len(working_profiles):
            is_builtin = is_builtin_csv_profile(working_profiles[idx].get("name", ""))
        builtin_notice_var.set(
            t("CSV_PROFILE_BUILTIN_NOTICE") if is_builtin else ""
        )
        set_editable_state(is_builtin)
        update_button.configure(state=tk.DISABLED if is_builtin else tk.NORMAL)
        delete_button.configure(state=tk.DISABLED if is_builtin else tk.NORMAL)

    def load_profile(idx):
        if idx < 0 or idx >= len(working_profiles):
            return
        _load_csv_profile(form_vars, working_profiles[idx])
        update_date_preview()
        update_builtin_indicator(idx)
        current_profile_name["name"] = working_profiles[idx].get("name")

    def sync_selection(event=None):  # noqa: ANN001
        sel = listbox.curselection()
        if sel:
            load_profile(sel[0])
        else:
            selected_name = current_profile_name.get("name")
            reselect_idx = None
            if selected_name:
                reselect_idx = next(
                    (
                        idx
                        for idx, prof in enumerate(working_profiles)
                        if prof.get("name") == selected_name
                    ),
                    None,
                )

            if reselect_idx is not None:
                listbox.selection_set(reselect_idx)
                listbox.see(reselect_idx)
                load_profile(reselect_idx)
            else:
                update_builtin_indicator()

    listbox.bind("<<ListboxSelect>>", sync_selection)

    def add_profile():
        nonlocal unsaved_changes
        prof = _read_profile_from_form_or_warn(form_vars)
        if not prof:
            return
        if is_builtin_csv_profile(prof["name"]):
            messagebox.showerror(
                t("ERR_TITLE"),
                t("CSV_PROFILE_NAME_RESERVED"),
            )
            return
        if any(p["name"] == prof["name"] for p in working_profiles):
            messagebox.showwarning(t("WARN_TITLE"), t("CSV_PROFILE_WARNING_EXISTS"))
            return
        working_profiles.append(prof)
        unsaved_changes = True
        refresh_list()
        listbox.selection_clear(0, tk.END)
        listbox.selection_set(len(working_profiles) - 1)
        sync_selection()

    def update_profile():
        nonlocal unsaved_changes
        sel = listbox.curselection()
        if not sel:
            messagebox.showwarning(
                t("CSV_PROFILE_NO_SELECTION_TITLE"),
                t("CSV_PROFILE_WARNING_SELECT"),
            )
            return
        prof = _read_profile_from_form_or_warn(form_vars)
        if not prof:
            return
        selected_profile = working_profiles[sel[0]]
        if is_builtin_csv_profile(prof["name"]):
            messagebox.showerror(
                t("ERR_TITLE"),
                t("CSV_PROFILE_BUILTIN_OVERWRITE"),
            )
            return
        for idx, existing in enumerate(working_profiles):
            if idx != sel[0] and existing["name"] == prof["name"]:
                messagebox.showwarning(t("WARN_TITLE"), t("CSV_PROFILE_WARNING_EXISTS"))
                return
        if is_builtin_csv_profile(selected_profile.get("name", "")):
            working_profiles.append(prof)
            unsaved_changes = True
            refresh_list()
            new_idx = next(
                (i for i, p in enumerate(working_profiles) if p.get("name") == prof["name"]),
                None,
            )
            if new_idx is not None:
                listbox.selection_clear(0, tk.END)
                listbox.selection_set(new_idx)
            sync_selection()
            return

        working_profiles[sel[0]] = prof
        _ensure_default_profile(default_profile_var, working_profiles, prof["name"])
        unsaved_changes = True
        refresh_list()
        listbox.selection_set(sel[0])
        sync_selection()

    def delete_profile():
        nonlocal unsaved_changes
        sel = listbox.curselection()
        if not sel:
            messagebox.showwarning(
                t("CSV_PROFILE_NO_SELECTION_TITLE"),
                t("CSV_PROFILE_WARNING_SELECT"),
            )
            return
        idx = sel[0]
        if is_builtin_csv_profile(working_profiles[idx].get("name", "")):
            messagebox.showinfo(
                t("CSV_PROFILE_INFO_SAVED_TITLE"),
                t("CSV_PROFILE_INFO_BUILTIN"),
            )
            return
        working_profiles.pop(idx)
        if not working_profiles:
            working_profiles.append(DEFAULT_CSV_PROFILE.copy())
        _ensure_default_profile(default_profile_var, working_profiles)
        unsaved_changes = True
        refresh_list()
        listbox.selection_set(0)
        sync_selection()

    def set_default_profile():
        nonlocal unsaved_changes
        sel = listbox.curselection()
        if not sel:
            messagebox.showwarning(
                t("CSV_PROFILE_NO_SELECTION_TITLE"),
                t("CSV_PROFILE_WARNING_SELECT"),
            )
            return
        selected_name = working_profiles[sel[0]]["name"]
        if default_profile_var.get() != selected_name:
            default_profile_var.set(selected_name)
            unsaved_changes = True
            refresh_list()

    def save_and_close():
        nonlocal unsaved_changes
        if not working_profiles:
            messagebox.showwarning(t("WARN_TITLE"), t("CSV_PROFILE_WARNING_MIN_ONE"))
            return
        _save_csv_profile_config(
            csv_profile_state, default_profile_var, working_profiles, refresh_csv_profile_controls
        )
        messagebox.showinfo(
            t("CSV_PROFILE_INFO_SAVED_TITLE"),
            t("CSV_PROFILE_INFO_SAVED_BODY"),
        )
        unsaved_changes = False
        dlg.destroy()

    def on_close():
        if not unsaved_changes:
            dlg.destroy()
            return

        resp = messagebox.askyesnocancel(
            t("CSV_PROFILE_UNSAVED_TITLE"),
            t("CSV_PROFILE_UNSAVED_BODY"),
        )
        if resp is True:
            save_and_close()
        elif resp is False:
            dlg.destroy()

    button_frame = tk.Frame(dlg)
    button_frame.grid(row=2, column=0, columnspan=3, pady=(0, 10))

    tk.Button(button_frame, text=t("BTN_SAVE_AS_NEW"), command=add_profile, width=14).pack(
        side="left", padx=(0, 5)
    )
    update_button = tk.Button(
        button_frame, text=t("BTN_UPDATE_PROFILE"), command=update_profile, width=14
    )
    update_button.pack(side="left", padx=(0, 5))
    delete_button = tk.Button(button_frame, text=t("BTN_DELETE"), command=delete_profile, width=10)
    delete_button.pack(side="left", padx=(0, 5))
    tk.Button(
        button_frame, text=t("BTN_SET_DEFAULT"), command=set_default_profile, width=18
    ).pack(
        side="left", padx=(0, 5)
    )
    tk.Button(
        button_frame, text=t("BTN_CLOSE_SAVE"), command=save_and_close, width=14
    ).pack(side="left")

    refresh_list()
    if working_profiles:
        preferred_name = selected_csv_profile_var.get() or default_profile_var.get()
        selected_idx = next(
            (
                idx
                for idx, prof in enumerate(working_profiles)
                if prof.get("name") == preferred_name
            ),
            0,
        )
        listbox.selection_set(selected_idx)
        listbox.see(selected_idx)
        sync_selection()
    else:
        update_builtin_indicator()

    # 1) Set keyboard focus to the list of profiles
    listbox.focus_set()

    # 2) Esc korzysta z on_close (uwzględnia niezapisane zmiany)
    dlg.bind("<Escape>", lambda *_: on_close())

    # 3) Profesjonalne centrowanie względem okna rodzica (aplikacji)
    dlg.update()
    _center_window(dlg, root)

    dlg.protocol("WM_DELETE_WINDOW", on_close)

    dlg.wait_window(dlg)


def run_gui(connection_store, output_directory):
    _require_tk()
    query_paths_state = {"paths": load_query_paths()}
    csv_profile_state = {"config": load_csv_profiles(), "combobox": None}
    connections_state = {
        "store": connection_store or {
            "connections": [],
            "last_selected": None,
        },
        "combobox": None,
    }

    _dbg("run_gui(): start")
    root = tk.Tk()
    _dbg("run_gui(): Tk root created")
    root.title(f"{t('APP_TITLE_FULL')} {get_app_version_label()}")
    apply_app_icon(root)
    apply_native_ttk_theme(root)

    selected_sql_path_full = tk.StringVar(value="")
    sql_label_var = tk.StringVar(value="")
    format_var = tk.StringVar(value="xlsx")
    selected_csv_profile_var = tk.StringVar(value="")
    default_csv_label_var = tk.StringVar(value="")
    save_as_path_var = tk.StringVar(value="")
    result_info_var = tk.StringVar(value="")
    last_output_path = {"path": None}
    engine_holder = {"engine": None}
    # Per-session password cache: { cache_key: password }
    # Exists only in memory for the current app run.
    password_cache: dict[tuple, str] = {}
    connection_status_var = tk.StringVar(value="")
    connection_status_state = {"key": None, "params": {}}
    secure_edit_state = {"button": None}
    start_button_holder = {"widget": None}
    cancel_state = {"event": None}
    close_state = {"after_cancel": False}
    error_display = {"widget": None}
    selected_connection_var = tk.StringVar(
        value=connections_state["store"].get("last_selected") or ""
    )
    pasted_sql_state = {"sql": None, "report_name": None}
    lang_var = tk.StringVar(
        value=_CURRENT_LANG.upper()
    )
    archive_sql_var = tk.BooleanVar(value=load_persisted_archive_sql())
    db_timeout_min_var = tk.IntVar(value=int(load_persisted_db_timeout_seconds() / 60))
    export_timeout_min_var = tk.IntVar(
        value=int(load_persisted_export_timeout_seconds() / 60)
    )
    ui_config = load_ui_config(Path(DATA_DIR))
    _show_data_dir_notice_once(
        root,
        ui_config,
        lambda cfg: save_ui_config(Path(DATA_DIR), cfg),
        data_dir=DATA_DIR,
        base_dir=BASE_DIR,
        lang=(lang_var.get() or "").lower(),
    )
    sql_highlight_enabled = bool(
        (ui_config.get("ui") or {}).get("sql_highlight_enabled", False)
    )
    if sql_highlight_enabled and (CodeView is None or not SQL_LEXER_CLASSES):
        sql_highlight_enabled = False
    sql_highlight_var = tk.BooleanVar(value=sql_highlight_enabled)
    sql_dialect_var = tk.StringVar(value="SQLite")
    sql_editor_state = {"dialog": None, "widget": None, "rebuild": None}

    # Template-related state (GUI only; console mode has no template support)
    use_template_var = tk.BooleanVar(value=False)
    template_path_var = tk.StringVar(value="")
    template_label_var = tk.StringVar(value="")
    sheet_name_var = tk.StringVar(value="")
    start_cell_var = tk.StringVar(value="A2")
    include_header_var = tk.BooleanVar(value=False)
    template_state = {
        "sheet_combobox": None,
        "choose_button": None,
        "start_cell_entry": None,
        "include_header_check": None,
    }
    i18n_widgets = {}

    def on_archive_sql_toggle():  # noqa: ANN001
        try:
            persist_archive_sql(bool(archive_sql_var.get()))
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Persist archive_sql failed: %s", exc, exc_info=True)

    def persist_timeouts_from_ui():  # noqa: ANN001
        """Persist GUI timeout fields into app config (minutes -> seconds)."""
        try:
            db_seconds = max(0, int(db_timeout_min_var.get() or 0)) * 60
            export_seconds = max(0, int(export_timeout_min_var.get() or 0)) * 60
            persist_db_timeout_seconds(db_seconds)
            persist_export_timeout_seconds(export_seconds)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Persist timeouts failed: %s", exc, exc_info=True)

    def _sql_highlight_available() -> bool:
        return CodeView is not None and bool(SQL_LEXER_CLASSES)

    def persist_sql_highlight_setting(enabled: bool) -> None:
        config = load_ui_config(Path(DATA_DIR))
        config.setdefault("ui", {})["sql_highlight_enabled"] = bool(enabled)
        save_ui_config(Path(DATA_DIR), config)

    def reset_sql_dialect_for_connection(conn_type: str) -> None:
        sql_dialect_var.set(_dialect_from_db_kind(conn_type))
        rebuild = sql_editor_state.get("rebuild")
        if rebuild and sql_highlight_var.get():
            rebuild(keep_text=True)

    def _set_sql_path(path):
        pasted_sql_state["sql"] = None
        pasted_sql_state["report_name"] = None
        resolved = resolve_path(path)
        selected_sql_path_full.set(resolved)
        sql_label_var.set(shorten_path(path))

    def set_connection_status(message=None, connected=False, key=None, **params):
        if key:
            connection_status_state["key"] = key
            connection_status_state["params"] = params
            display_params = dict(params)
            if key == "STATUS_CONNECTED" and "type" in display_params:
                type_key = (display_params.get("type") or "").strip()
                display_params["type"] = _db_type_labels().get(type_key, type_key)
            message = t(key, **display_params)
        connection_status_var.set(message or "")
        btn = start_button_holder.get("widget")
        if btn is not None:
            btn_state = tk.NORMAL if connected else tk.DISABLED
            btn.config(state=btn_state)

    def apply_engine(new_engine):
        old_engine = engine_holder.get("engine")
        if old_engine is not None and old_engine is not new_engine:
            old_engine.dispose()
        engine_holder["engine"] = new_engine

    def get_connection_by_name(name):
        for conn in connections_state["store"].get("connections", []):
            if conn.get("name") == name:
                return conn
        return None

    def persist_connections():
        save_connections(connections_state["store"])
        refresh_secure_edit_button()

    def refresh_connection_combobox():
        combo = connections_state.get("combobox")
        if combo is None:
            return
        names = [c.get("name", "") for c in connections_state["store"].get("connections", [])]
        combo["values"] = names
        if selected_connection_var.get() not in names:
            selected_connection_var.set(names[0] if names else "")

    def set_selected_connection(name):
        if name and get_connection_by_name(name):
            connections_state["store"]["last_selected"] = name
            selected_connection_var.set(name)
            persist_connections()

    def on_connection_change(*_):  # noqa: ANN001
        name = selected_connection_var.get()
        if not name:
            set_connection_status(connected=False, key="STATUS_NO_CONNECTION")
            apply_engine(None)
            return
        set_selected_connection(name)
        apply_selected_connection(show_success=False)

    def _password_cache_key(entry):
        conn_type = (entry.get("type") or "").strip()
        name = (entry.get("name") or "").strip()
        details = entry.get("details") or {}
        if conn_type == "mssql_odbc":
            return (
                conn_type,
                name,
                details.get("server"),
                details.get("database"),
                details.get("username"),
                bool(details.get("trusted")),
            )
        if conn_type == "postgresql":
            return (
                conn_type,
                name,
                details.get("host"),
                details.get("port"),
                details.get("database"),
                details.get("user"),
            )
        if conn_type == "mysql":
            return (
                conn_type,
                name,
                details.get("host"),
                details.get("port"),
                details.get("database"),
                details.get("user"),
            )
        if conn_type == "sqlite":
            return (conn_type, name, details.get("path"))
        return (conn_type, name)

    def create_engine_from_entry(entry, parent=None):
        if not entry:
            raise ValueError(t("ERR_NO_SECURE_CONFIG"))
        engine_kwargs = {}
        if entry.get("type") == "mssql_odbc":
            engine_kwargs["isolation_level"] = "AUTOCOMMIT"

        conn_type = str(entry.get("type") or "").strip()
        name = str(entry.get("name") or "").strip()
        details = entry.get("details") or {}

        # SQLite: no password prompt
        if conn_type == "sqlite":
            url = _build_runtime_url(conn_type, details, None)
            return create_engine(url, **engine_kwargs)

        # MSSQL trusted auth: no password prompt
        if conn_type == "mssql_odbc" and bool(details.get("trusted")):
            url = _build_runtime_url(conn_type, details, None)
            engine = create_engine(url, **engine_kwargs)
            _ensure_engine_mssql_set_hook(engine)
            return engine

        stored_pwd = str(details.get("password") or "")
        if stored_pwd:
            url = _build_runtime_url(conn_type, details, stored_pwd)
            engine = create_engine(url, **engine_kwargs)
            if conn_type == "mssql_odbc":
                _ensure_engine_mssql_set_hook(engine)
            return engine

        # Not stored -> use cache or prompt once per session
        cache_key = _password_cache_key(entry)
        cached = password_cache.get(cache_key)
        if not cached:
            pwd = simpledialog.askstring(
                t("PROMPT_PASSWORD_TITLE"),
                t("PROMPT_PASSWORD_BODY", name=name),
                show="*",
                parent=parent or root,
            )
            if pwd is None or not str(pwd):
                raise PasswordRequiredError(name=name, conn_type=conn_type)
            cached = str(pwd)
            password_cache[cache_key] = cached

        url = _build_runtime_url(conn_type, details, cached)
        engine = create_engine(url, **engine_kwargs)
        if conn_type == "mssql_odbc":
            _ensure_engine_mssql_set_hook(engine)
        return engine

    def apply_selected_connection(show_success=False):
        conn = get_connection_by_name(selected_connection_var.get())
        if not conn:
            set_connection_status(connected=False, key="STATUS_NO_CONNECTION")
            apply_engine(None)
            return
        reset_sql_dialect_for_connection(str(conn.get("type") or ""))
        try:
            engine = create_engine_from_entry(conn)
            with engine.connect() as connection:
                connection.execute(text("SELECT 1"))
            apply_engine(engine)
        except Exception as exc:  # noqa: BLE001
            if isinstance(exc, PasswordRequiredError):
                set_connection_status(connected=False, key="STATUS_PASSWORD_REQUIRED")
                messagebox.showwarning(
                    t("ERR_PASSWORD_REQUIRED_TITLE"),
                    t("ERR_PASSWORD_REQUIRED_BODY", name=exc.name or conn.get("name", "")),
                )
                return
            set_connection_status(connected=False, key="STATUS_CONNECTION_ERROR")
            if handle_db_driver_error(exc, conn.get("type"), conn.get("name")):
                return
            LOGGER.exception(
                "Connection test failed for %s (%s)",
                conn.get("name"),
                conn.get("type"),
                exc_info=True,
            )
            title, msg = _format_connection_error(
                conn_type=str(conn.get("type") or ""),
                exc=exc,
                details=conn.get("details") or {},
            )
            messagebox.showerror(title, msg)
            return

        set_connection_status(
            connected=True,
            key="STATUS_CONNECTED",
            name=conn.get("name", ""),
            type=conn.get("type", ""),
        )
        if show_success:
            messagebox.showinfo(
                t("INFO_CONNECTION_OK_TITLE"),
                t("INFO_CONNECTION_OK_BODY", name=conn.get("name", "")),
            )

    def update_error_display(message):
        widget = error_display.get("widget")
        if widget is None:
            return
        widget.config(state="normal")
        widget.delete("1.0", tk.END)
        if message:
            widget.insert("1.0", message)
            widget.see(tk.END)
        widget.config(state="disabled")

    def request_cancel():
        ev = cancel_state.get("event")
        if ev and not ev.is_set():
            ev.set()
            result_info_var.set(t("MSG_CANCEL_REQUESTED"))
            btn_cancel.config(state=tk.DISABLED)

    def on_close():  # noqa: ANN001
        ev = cancel_state.get("event")
        if ev is not None and not ev.is_set():
            if messagebox.askyesno(
                t("WARN_TITLE"),
                t("MSG_CONFIRM_CANCEL_AND_EXIT"),
            ):
                close_state["after_cancel"] = True
                ev.set()
                result_info_var.set(t("MSG_CANCEL_REQUESTED"))
                try:
                    btn_cancel.config(state=tk.DISABLED)
                except Exception:  # noqa: BLE001
                    pass
            return

        # If user confirmed "Cancel+Exit" and cancel is already in progress,
        # ignore subsequent WM_DELETE requests; _reset_run_state() will close.
        if ev is not None and ev.is_set() and close_state.get("after_cancel"):
            return
        root.destroy()

    def show_error_popup(ui_msg, *, sql_query: str | None = None, sql_source_path: str | None = None):
        # --- helper: split the rendered error into nicer sections (best-effort) ---
        def _split_error_ui_msg(msg: str) -> dict:
            db_lbl = t("ERR_DB_MESSAGE")
            sql_lbl = t("ERR_SQL_PREVIEW")
            hint_lbl = t("ERR_HINT_LABEL")
            full_log_lbl = t("ERR_FULL_LOG")

            first_line = (msg.splitlines()[0] if msg else "").strip()

            def _idx(s: str) -> int:
                try:
                    return msg.index(s)
                except ValueError:
                    return -1

            db_i = _idx(db_lbl)
            sql_i = _idx(sql_lbl)
            hint_i = _idx(hint_lbl)
            full_log_i = _idx(full_log_lbl)

            db_part = ""
            sql_part = ""
            hint_part = ""

            if db_i >= 0:
                start = db_i + len(db_lbl)
                end = sql_i if sql_i >= 0 else len(msg)
                db_part = msg[start:end].strip()

            if sql_i >= 0:
                start = sql_i + len(sql_lbl)
                candidates = [i for i in (hint_i, full_log_i) if i >= 0]
                end = min(candidates) if candidates else len(msg)
                sql_part = msg[start:end].strip()
                if full_log_lbl and full_log_lbl in sql_part:
                    sql_part = sql_part.split(full_log_lbl, 1)[0].strip()

            if hint_i >= 0:
                start = hint_i + len(hint_lbl)
                hint_part = msg[start:].strip()

            # Summary tries to be short and scannable
            summary_lines = []
            if first_line:
                summary_lines.append(first_line)
            if db_part:
                summary_lines.append("")
                summary_lines.append(db_lbl)
                summary_lines.append(db_part.strip())
            if hint_part:
                summary_lines.append("")
                summary_lines.append(hint_lbl)
                summary_lines.append(hint_part.strip())

            summary = "\n".join(summary_lines).strip() if summary_lines else msg.strip()

            return {
                "summary": summary,
                "sql": sql_part.strip(),
                "full": msg.strip(),
            }

        parts = _split_error_ui_msg(ui_msg)

        sql_tab_text = parts.get("sql", "") or ""
        if sql_query:
            src = (sql_source_path or "").strip()
            header = ""
            if src:
                header = f"{t('ERR_SQL_SOURCE')}\n{src}\n\n"
            sql_tab_text = header + (sql_query.rstrip() + "\n")
        sql_tab_text = _limit_text_for_widget(sql_tab_text, max_chars=80000)

        popup = tk.Toplevel(root)
        apply_app_icon(popup)
        popup.title(t("ERR_QUERY_TITLE"))
        popup.transient(root)
        popup.grab_set()

        popup.minsize(860, 420)
        popup.geometry("980x520")
        _center_window(popup, root)

        # Layout grid
        popup.columnconfigure(0, weight=1)
        popup.rowconfigure(1, weight=1)

        header = ttk.Frame(popup, padding=(12, 10, 12, 8))
        header.grid(row=0, column=0, sticky="we")
        header.columnconfigure(0, weight=1)

        title_lbl = ttk.Label(
            header,
            text=t("ERR_QUERY_TITLE"),
            font=("Segoe UI", 12, "bold") if sys.platform == "win32" else None,
        )
        title_lbl.grid(row=0, column=0, sticky="w")

        sub_lbl = ttk.Label(
            header,
            text=t("ERR_FULL_LOG"),
            foreground="gray",
        )
        sub_lbl.grid(row=1, column=0, sticky="w", pady=(2, 0))

        body = ttk.Frame(popup, padding=(12, 0, 12, 10))
        body.grid(row=1, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        nb = ttk.Notebook(body)
        nb.grid(row=0, column=0, sticky="nsew")

        def _make_text_tab(name: str, content: str, wrap: str):
            frame = ttk.Frame(nb, padding=0)
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)

            txt = tk.Text(
                frame,
                wrap=wrap,
                font=("Consolas", 9) if sys.platform == "win32" else None,
                borderwidth=1,
                relief="solid",
            )
            y = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)
            txt.configure(yscrollcommand=y.set)

            txt.grid(row=0, column=0, sticky="nsew")
            y.grid(row=0, column=1, sticky="ns")

            x = None
            if wrap == "none":
                x = ttk.Scrollbar(frame, orient="horizontal", command=txt.xview)
                txt.configure(xscrollcommand=x.set)
                x.grid(row=1, column=0, sticky="we")

            txt.insert("1.0", content or "")
            txt.config(state="disabled")

            nb.add(frame, text=name)
            return txt

        summary_txt = _make_text_tab(t("ERR_TAB_SUMMARY"), parts.get("summary", ""), wrap="word")
        sql_txt = _make_text_tab(t("ERR_TAB_SQL"), sql_tab_text, wrap="none")
        full_txt = _make_text_tab(t("ERR_TAB_DETAILS"), parts.get("full", ""), wrap="none")

        # Buttons
        btns = ttk.Frame(popup, padding=(12, 0, 12, 12))
        btns.grid(row=2, column=0, sticky="we")
        btns.columnconfigure(0, weight=1)

        def _copy_to_clipboard(text: str):
            popup.clipboard_clear()
            popup.clipboard_append(text or "")

        def copy_summary():
            _copy_to_clipboard(parts.get("summary", ""))

        def copy_sql():
            if sql_query:
                _copy_to_clipboard(sql_query)
            else:
                _copy_to_clipboard(parts.get("sql", ""))

        def copy_all():
            _copy_to_clipboard(ui_msg)

        def open_sql_file():
            src = (sql_source_path or "").strip()
            if not src or src.startswith("<pasted:"):
                return
            if not os.path.exists(src):
                messagebox.showwarning(t("WARN_TITLE"), f"{t('ERR_FILE_PATH', path=src)}")
                return
            _open_path(src)

        def open_log():
            log_path = LOG_FILE_PATH or ""
            if not os.path.exists(log_path):
                messagebox.showwarning(t("WARN_TITLE"), f"{t('ERR_FILE_PATH', path=log_path)}")
                return
            try:
                if sys.platform.startswith("win"):
                    os.startfile(log_path)  # type: ignore[attr-defined]
                elif sys.platform == "darwin":
                    subprocess.run(["open", log_path], check=False)
                else:
                    subprocess.run(["xdg-open", log_path], check=False)
            except Exception as err:  # noqa: BLE001
                messagebox.showerror(t("ERR_TITLE"), str(err))

        left = ttk.Frame(btns)
        left.grid(row=0, column=0, sticky="w")

        ttk.Button(left, text=t("BTN_COPY_SUMMARY"), command=copy_summary).pack(side="left", padx=(0, 8))
        ttk.Button(left, text=t("BTN_COPY_SQL"), command=copy_sql).pack(side="left", padx=(0, 8))
        ttk.Button(left, text=t("BTN_COPY_ALL"), command=copy_all).pack(side="left", padx=(0, 8))
        btn_open_sql = ttk.Button(left, text=t("BTN_OPEN_SQL"), command=open_sql_file)
        btn_open_sql.pack(side="left", padx=(0, 8))
        ttk.Button(left, text=t("BTN_OPEN_LOG"), command=open_log).pack(side="left")

        src_for_btn = (sql_source_path or "").strip()
        if (not src_for_btn) or src_for_btn.startswith("<pasted:") or (not os.path.exists(src_for_btn)):
            btn_open_sql.state(["disabled"])

        right = ttk.Frame(btns)
        right.grid(row=0, column=1, sticky="e")

        ttk.Button(
            right,
            text=t("BTN_REPORT_ISSUE"),
            command=lambda: open_github_issue_chooser(parent=popup),
        ).pack(side="left", padx=(0, 8))
        ttk.Button(right, text=t("BTN_CLOSE"), command=popup.destroy).pack(side="left")

        popup.bind("<Escape>", lambda *_: popup.destroy())
        popup.focus_set()

    def show_odbc_diagnostics_popup():
        text = odbc_diagnostics_text()

        dlg = tk.Toplevel(root)
        apply_app_icon(dlg)
        dlg.title(t("ODBC_DIAGNOSTICS_TITLE"))
        dlg.transient(root)
        dlg.grab_set()

        dlg.minsize(680, 420)
        dlg.geometry("760x420")
        _center_window(dlg, root)

        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(0, weight=1)

        body = ttk.Frame(dlg, padding=(12, 12, 12, 8))
        body.grid(row=0, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        text_widget = tk.Text(
            body,
            wrap="none",
            font=("Consolas", 9) if sys.platform == "win32" else None,
            borderwidth=1,
            relief="solid",
        )
        y_scroll = ttk.Scrollbar(body, orient="vertical", command=text_widget.yview)
        x_scroll = ttk.Scrollbar(body, orient="horizontal", command=text_widget.xview)
        text_widget.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        text_widget.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="we")

        text_widget.insert("1.0", text)
        text_widget.config(state="disabled")

        btns = ttk.Frame(dlg, padding=(12, 0, 12, 12))
        btns.grid(row=1, column=0, sticky="we")
        btns.columnconfigure(0, weight=1)

        def copy_all():
            dlg.clipboard_clear()
            dlg.clipboard_append(text or "")

        ttk.Button(btns, text=t("BTN_COPY"), command=copy_all).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(btns, text=t("BTN_CLOSE"), command=dlg.destroy).pack(side="left")

        dlg.bind("<Escape>", lambda *_: dlg.destroy())
        dlg.focus_set()

    def open_connections_manager_gui():
        dlg = tk.Toplevel(root)
        apply_app_icon(dlg)
        dlg.title(t("CONNECTIONS_MANAGER_TITLE"))
        dlg.transient(root)
        dlg.grab_set()

        dlg.minsize(560, 320)
        dlg.geometry("680x360")
        dlg.update_idletasks()
        _center_window(dlg, root)

        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(0, weight=1)

        body = ttk.Frame(dlg, padding=(12, 12, 12, 12))
        body.grid(row=0, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        list_frame = ttk.Frame(body)
        list_frame.grid(row=0, column=0, sticky="nsew")
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        tree = ttk.Treeview(list_frame, show="tree", selectmode="browse", height=12)
        scroll = ttk.Scrollbar(list_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")

        button_frame = ttk.Frame(body)
        button_frame.grid(row=0, column=1, sticky="ns", padx=(12, 0))

        def selected_name() -> str:
            sel = tree.selection()
            return sel[0] if sel else ""

        def update_button_state() -> None:
            has_selection = bool(selected_name())
            state = tk.NORMAL if has_selection else tk.DISABLED
            set_current_btn.config(state=state)
            edit_btn.config(state=state)
            delete_btn.config(state=state)

        def refresh_list(select_name: str | None = None) -> None:
            tree.delete(*tree.get_children())
            names = [
                c.get("name", "") for c in connections_state["store"].get("connections", [])
            ]
            for name in names:
                if name:
                    tree.insert("", "end", iid=name, text=name)
            preferred = select_name or selected_connection_var.get()
            if preferred and preferred in names:
                tree.selection_set(preferred)
                tree.see(preferred)
            elif names:
                tree.selection_set(names[0])
                tree.see(names[0])
            update_button_state()

        def warn_no_selection() -> None:
            messagebox.showwarning(t("WARN_TITLE"), t("WARN_SELECT_CONNECTION"))

        def on_set_current() -> None:
            name = selected_name()
            if not name:
                warn_no_selection()
                return
            set_selected_connection(name)
            refresh_connection_combobox()
            apply_selected_connection(show_success=False)
            refresh_list(select_name=name)

        def on_new() -> None:
            open_connection_dialog_gui(
                root,
                connections_state,
                selected_connection_var,
                get_connection_by_name,
                set_selected_connection,
                persist_connections,
                refresh_connection_combobox,
                apply_selected_connection,
                handle_db_driver_error,
                create_engine_from_entry,
                mode="new",
            )
            refresh_list(select_name=selected_connection_var.get() or None)

        def on_edit() -> None:
            name = selected_name()
            if not name:
                warn_no_selection()
                return
            set_selected_connection(name)
            open_connection_dialog_gui(
                root,
                connections_state,
                selected_connection_var,
                get_connection_by_name,
                set_selected_connection,
                persist_connections,
                refresh_connection_combobox,
                apply_selected_connection,
                handle_db_driver_error,
                create_engine_from_entry,
                mode="edit",
            )
            refresh_list(select_name=name)

        def on_delete() -> None:
            name = selected_name()
            if not name:
                warn_no_selection()
                return
            set_selected_connection(name)
            delete_selected_connection()
            refresh_list()

        def on_close() -> None:
            if secure_edit_state.get("button") is security_btn:
                secure_edit_state["button"] = None
            dlg.destroy()

        set_current_btn = ttk.Button(
            button_frame,
            text=t("BTN_SET_CURRENT_CONNECTION"),
            command=on_set_current,
            width=20,
        )
        set_current_btn.pack(fill="x", pady=(0, 6))

        new_btn = ttk.Button(
            button_frame, text=t("BTN_NEW_CONNECTION"), command=on_new, width=20
        )
        new_btn.pack(fill="x", pady=(0, 6))

        edit_btn = ttk.Button(
            button_frame, text=t("BTN_EDIT_CONNECTION"), command=on_edit, width=20
        )
        edit_btn.pack(fill="x", pady=(0, 6))

        delete_btn = ttk.Button(
            button_frame, text=t("BTN_DELETE_CONNECTION"), command=on_delete, width=20
        )
        delete_btn.pack(fill="x", pady=(0, 6))

        security_btn = ttk.Button(
            button_frame,
            text=t("BTN_EDIT_SECURE"),
            command=open_secure_editor,
            width=20,
        )
        security_btn.pack(fill="x", pady=(0, 6))

        odbc_btn = ttk.Button(
            button_frame,
            text=t("BTN_ODBC_DIAGNOSTICS"),
            command=show_odbc_diagnostics_popup,
            width=20,
        )
        odbc_btn.pack(fill="x", pady=(0, 6))

        close_btn = ttk.Button(
            button_frame, text=t("BTN_CLOSE"), command=on_close, width=20
        )
        close_btn.pack(fill="x")

        secure_edit_state["button"] = security_btn
        refresh_secure_edit_button()
        refresh_list()

        tree.bind("<<TreeviewSelect>>", lambda *_: update_button_state())
        tree.bind("<Double-Button-1>", lambda *_: on_set_current())
        dlg.bind("<Escape>", lambda *_: on_close())
        dlg.protocol("WM_DELETE_WINDOW", on_close)
        tree.focus_set()
        dlg.wait_window(dlg)

    def refresh_secure_edit_button():
        btn = secure_edit_state.get("button")
        if btn is None:
            return
        exists = os.path.exists(SECURE_PATH)
        btn_state = tk.NORMAL if exists else tk.DISABLED
        btn.config(state=btn_state)

    def test_connection_only():
        if not connections_state["store"].get("connections"):
            messagebox.showerror(
                t("ERR_NO_CONNECTION_TITLE"),
                t("ERR_NO_CONNECTION_BODY"),
            )
            return
        apply_selected_connection(show_success=True)

    def delete_selected_connection():
        connections = connections_state["store"].get("connections", [])
        name = selected_connection_var.get()
        if not connections or not name:
            messagebox.showerror(
                t("ERR_NO_CONNECTION_TITLE"),
                t("ERR_NO_CONNECTION_DELETE"),
            )
            return

        if not messagebox.askyesno(
            t("ASK_DELETE_CONNECTION_TITLE"),
            t("ASK_DELETE_CONNECTION_BODY", name=name),
        ):
            return

        connections_state["store"]["connections"] = [
            c for c in connections if c.get("name") != name
        ]

        remaining = connections_state["store"].get("connections", [])
        if remaining:
            new_selection = remaining[0].get("name", "")
            set_selected_connection(new_selection)
            refresh_connection_combobox()
            apply_selected_connection(show_success=False)
        else:
            selected_connection_var.set("")
            connections_state["store"]["last_selected"] = None
            persist_connections()
            refresh_connection_combobox()
            apply_engine(None)
            set_connection_status(connected=False, key="STATUS_NO_CONNECTION")

    def open_secure_editor():
        dlg = tk.Toplevel(root)
        apply_app_icon(dlg)
        dlg.title(t("TITLE_EDIT_SECURE"))
        dlg.transient(root)
        dlg.grab_set()

        pretty = json.dumps(connections_state["store"], ensure_ascii=False, indent=2)
        text_widget = tk.Text(dlg, width=80, height=16, wrap="word")
        text_widget.insert("1.0", pretty)
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 5))

        def save_and_close(*_):
            new_content = text_widget.get("1.0", tk.END).strip()
            try:
                parsed = json.loads(new_content) if new_content else {}
                connections_state["store"] = _normalize_connections(parsed)
                persist_connections()
                refresh_connection_combobox()
                dlg.destroy()
                apply_selected_connection(show_success=False)
                messagebox.showinfo(
                    t("INFO_SECURE_SAVED_TITLE"),
                    t("INFO_SECURE_SAVED_BODY"),
                )
            except Exception as exc:  # noqa: BLE001
                messagebox.showerror(
                    t("ERR_SECURE_SAVE_TITLE"),
                    t("ERR_SECURE_SAVE_BODY", error=exc),
                )

        def cancel(*_):
            dlg.destroy()

        btn_frame = tk.Frame(dlg)
        btn_frame.pack(pady=(0, 10))

        tk.Button(btn_frame, text=t("BTN_SAVE"), width=12, command=save_and_close).pack(
            side="left", padx=(0, 5)
        )
        tk.Button(btn_frame, text=t("BTN_CANCEL"), width=12, command=cancel).pack(
            side="left"
        )

        dlg.bind("<Return>", save_and_close)
        dlg.bind("<Escape>", cancel)

        _center_window(dlg, root)

    def choose_sql_file():
        path = filedialog.askopenfilename(
            title=t("TITLE_SELECT_SQL"),
            filetypes=[(t("FILETYPE_SQL"), "*.sql"), (t("FILETYPE_ALL"), "*.*")],
        )
        if not path:
            return
        ok, msg = validate_sql_text_file(path)
        if not ok:
            messagebox.showerror(t("ERR_INVALID_SQL_FILE_TITLE"), msg)
            return
        _set_sql_path(path)

    def open_paste_sql_dialog():
        dlg = tk.Toplevel(root)
        apply_app_icon(dlg)
        dlg.title(t("TITLE_PASTE_SQL"))
        dlg.transient(root)
        dlg.grab_set()
        dlg.minsize(520, 320)
        _center_window(dlg, root)

        report_name_var = tk.StringVar(value=pasted_sql_state["report_name"] or "")

        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(4, weight=1)

        tk.Label(dlg, text=t("LBL_REPORT_NAME")).grid(
            row=0, column=0, sticky="w", padx=12, pady=(12, 4)
        )
        report_entry = tk.Entry(dlg, textvariable=report_name_var, width=40)
        report_entry.grid(row=1, column=0, sticky="we", padx=12)

        tk.Label(dlg, text=t("LBL_PASTE_SQL")).grid(
            row=2, column=0, sticky="nw", padx=12, pady=(12, 4)
        )
        options_frame = tk.Frame(dlg)
        options_frame.grid(row=3, column=0, sticky="we", padx=12)
        options_frame.columnconfigure(2, weight=1)
        chk_sql_highlight = None

        def update_dialect_controls_state():
            combo_state = "readonly" if sql_highlight_var.get() else "disabled"
            dialect_combo.config(state=combo_state)

        def rebuild_sql_editor(keep_text: bool = True) -> None:
            current_text = ""
            existing = sql_editor_state.get("widget")
            if keep_text and existing is not None:
                try:
                    current_text = existing.get("1.0", tk.END).rstrip()
                except Exception:
                    current_text = ""
            elif keep_text and pasted_sql_state.get("sql"):
                current_text = pasted_sql_state.get("sql") or ""

            for child in editor_container.winfo_children():
                child.destroy()

            editor_container.columnconfigure(0, weight=1)
            editor_container.rowconfigure(0, weight=1)

            def _build_plain_editor() -> tk.Text:
                sql_widget = tk.Text(editor_container, wrap="none", height=10)

                y_scroll = ttk.Scrollbar(
                    editor_container, orient="vertical", command=sql_widget.yview
                )
                x_scroll = ttk.Scrollbar(
                    editor_container, orient="horizontal", command=sql_widget.xview
                )
                sql_widget.configure(
                    yscrollcommand=y_scroll.set,
                    xscrollcommand=x_scroll.set,
                )
                sql_widget.grid(row=0, column=0, sticky="nsew")
                y_scroll.grid(row=0, column=1, sticky="ns")
                x_scroll.grid(row=1, column=0, sticky="we")
                return sql_widget

            if sql_highlight_var.get() and _sql_highlight_available():
                try:
                    lexer_cls = (
                        SQL_LEXER_CLASSES.get(sql_dialect_var.get())
                        or SQL_LEXER_CLASSES.get("SQLite")
                    )
                    sql_widget = CodeView(  # type: ignore[call-arg]
                        editor_container,
                        lexer=lexer_cls,
                        wrap="none",
                        height=10,
                    )
                    sql_widget.grid(row=0, column=0, sticky="nsew")
                    _apply_codeview_pygments_style(sql_widget, "vs")
                except Exception as exc:  # noqa: BLE001
                    LOGGER.warning(
                        "SQL highlighting failed; falling back to plain editor: %s",
                        exc,
                        exc_info=True,
                    )
                    messagebox.showinfo(
                        t("INFO_SQL_HIGHLIGHT_TITLE"),
                        t("INFO_SQL_HIGHLIGHT_BODY"),
                        parent=dlg,
                    )
                    sql_highlight_var.set(False)
                    persist_sql_highlight_setting(False)
                    update_dialect_controls_state()
                    if chk_sql_highlight is not None:
                        chk_sql_highlight.config(state=tk.DISABLED)
                    sql_widget = _build_plain_editor()
            else:
                sql_widget = _build_plain_editor()

            if current_text:
                sql_widget.insert("1.0", current_text)

            sql_editor_state["widget"] = sql_widget

        def on_sql_highlight_toggle():  # noqa: ANN001
            if sql_highlight_var.get():
                if not _sql_highlight_available():
                    messagebox.showinfo(
                        t("INFO_SQL_HIGHLIGHT_TITLE"),
                        t("INFO_SQL_HIGHLIGHT_BODY"),
                        parent=dlg,
                    )
                    sql_highlight_var.set(False)
                    persist_sql_highlight_setting(False)
                    update_dialect_controls_state()
                    return
                persist_sql_highlight_setting(True)
            else:
                persist_sql_highlight_setting(False)
            update_dialect_controls_state()
            rebuild_sql_editor(keep_text=True)

        chk_sql_highlight = tk.Checkbutton(
            options_frame,
            text=t("CHK_SQL_HIGHLIGHT"),
            variable=sql_highlight_var,
            command=on_sql_highlight_toggle,
        )
        chk_sql_highlight.grid(row=0, column=0, sticky="w")

        tk.Label(options_frame, text=t("LBL_SQL_DIALECT")).grid(
            row=0, column=1, sticky="w", padx=(12, 4)
        )
        dialect_combo = ttk.Combobox(
            options_frame,
            textvariable=sql_dialect_var,
            values=["SQLite", "MSSQL", "Postgres", "MySQL"],
            state="readonly",
            width=12,
        )
        dialect_combo.grid(row=0, column=2, sticky="w")

        editor_container = tk.Frame(dlg)
        editor_container.grid(row=4, column=0, sticky="nsew", padx=12)

        update_dialect_controls_state()
        rebuild_sql_editor(keep_text=True)
        sql_editor_state["dialog"] = dlg
        sql_editor_state["rebuild"] = rebuild_sql_editor

        def on_dialect_change(*_):  # noqa: ANN001
            if sql_highlight_var.get():
                rebuild_sql_editor(keep_text=True)

        dialect_combo.bind("<<ComboboxSelected>>", on_dialect_change)

        btn_frame = tk.Frame(dlg)
        btn_frame.grid(row=5, column=0, sticky="e", padx=12, pady=(12, 12))

        def use_sql(*_):
            ok, msg, normalized_name = validate_report_basename(report_name_var.get())
            if not ok:
                messagebox.showerror(
                    t("ERR_INVALID_REPORT_NAME_TITLE"),
                    msg,
                    parent=dlg,
                )
                return
            sql_widget = sql_editor_state.get("widget")
            sql = sql_widget.get("1.0", tk.END).strip() if sql_widget else ""
            if not sql:
                messagebox.showerror(
                    t("ERR_EMPTY_SQL_TITLE"),
                    t("ERR_EMPTY_SQL_BODY"),
                    parent=dlg,
                )
                return
            pasted_sql_state["sql"] = sql
            pasted_sql_state["report_name"] = normalized_name
            selected_sql_path_full.set("")
            sql_label_var.set(f"{t('LBL_SQL_PASTED')} {normalized_name}")
            sql_editor_state["dialog"] = None
            sql_editor_state["widget"] = None
            sql_editor_state["rebuild"] = None
            dlg.destroy()

        def cancel(*_):
            sql_editor_state["dialog"] = None
            sql_editor_state["widget"] = None
            sql_editor_state["rebuild"] = None
            dlg.destroy()

        cancel_btn = tk.Button(btn_frame, text=t("BTN_CANCEL"), width=12, command=cancel)
        use_btn = tk.Button(btn_frame, text=t("BTN_USE_SQL"), width=12, command=use_sql)
        cancel_btn.pack(side="left", padx=(0, 6))
        use_btn.pack(side="left")

        dlg.bind("<Escape>", cancel)
        dlg.bind("<Return>", use_sql)
        dlg.protocol("WM_DELETE_WINDOW", cancel)
        report_entry.focus_set()

    def update_template_controls_state():
        enabled = use_template_var.get()
        tpl_path = (template_path_var.get() or "").strip()
        tpl_ok = enabled and tpl_path and Path(tpl_path).exists()

        choose_btn = template_state.get("choose_button")
        if choose_btn is not None:
            choose_btn_state = tk.NORMAL if enabled else tk.DISABLED
            choose_btn.config(state=choose_btn_state)

        sheet_combo = template_state.get("sheet_combobox")
        if sheet_combo is not None:
            sheet_state = "readonly" if tpl_ok else "disabled"
            sheet_combo.config(state=sheet_state)

        start_cell_entry = template_state.get("start_cell_entry")
        if start_cell_entry is not None:
            start_cell_state = tk.NORMAL if enabled else tk.DISABLED
            start_cell_entry.config(state=start_cell_state)

        include_header_check = template_state.get("include_header_check")
        if include_header_check is not None:
            include_header_state = tk.NORMAL if enabled else tk.DISABLED
            include_header_check.config(state=include_header_state)

    def _refresh_template_ui() -> None:
        use_tpl = bool(use_template_var.get())
        tpl_path = (template_path_var.get() or "").strip()
        tpl_ok = use_tpl and tpl_path and Path(tpl_path).exists()

        sheet_combo = template_state.get("sheet_combobox")
        if sheet_combo is None:
            return

        if not tpl_ok:
            sheet_combo.configure(state="disabled", values=[])
            sheet_name_var.set("")
            return

        sheetnames = _xlsx_get_sheetnames(tpl_path)
        sheet_combo.configure(values=sheetnames)

        if not sheetnames:
            sheet_combo.configure(state="disabled")
            sheet_name_var.set("")
            return

        sheet_combo.configure(state="readonly")

        current = (sheet_name_var.get() or "").strip()
        if current and current in sheetnames:
            return

        sheet_name_var.set(_pick_default_sheet(sheetnames))

    def update_csv_profile_controls_state():
        enabled = format_var.get() == "csv"

        combo = csv_profile_state.get("combobox")
        if combo is not None:
            combo_state = "readonly" if enabled else "disabled"
            combo.config(state=combo_state)

        manage_btn = csv_profile_state.get("manage_button")
        if manage_btn is not None:
            manage_state = tk.NORMAL if enabled else tk.DISABLED
            manage_btn.config(state=manage_state)

    def on_toggle_template():
        # Template jest sensowny tylko dla XLSX; jeśli zaznaczono template przy CSV, przełącz na XLSX.
        if use_template_var.get() and format_var.get() != "xlsx":
            format_var.set("xlsx")

        update_template_controls_state()
        update_csv_profile_controls_state()
        _refresh_template_ui()

    def _sync_save_as_extension(show_info: bool = False) -> None:
        raw = (save_as_path_var.get() or "").strip()
        if not raw:
            return
        if _looks_like_directory(raw):
            return
        expected_ext = _expected_output_extension(format_var.get())
        root, ext = os.path.splitext(raw)
        if ext.lower() == expected_ext:
            return
        adjusted_path = root + expected_ext
        save_as_path_var.set(adjusted_path)
        if show_info and ext:
            messagebox.showinfo(
                t("INFO_SAVE_AS_EXT_FIXED_TITLE"),
                t("INFO_SAVE_AS_EXT_FIXED_BODY", path=resolve_path(adjusted_path)),
            )

    def on_format_change(*_):
        """Keep template option consistent with selected output format."""
        if format_var.get() == "csv":
            use_template_var.set(False)

        update_template_controls_state()
        update_csv_profile_controls_state()
        _sync_save_as_extension(show_info=False)

    def choose_template_file():
        path = filedialog.askopenfilename(
            title=t("TITLE_SELECT_TEMPLATE"),
            filetypes=[(t("FILETYPE_EXCEL"), "*.xlsx"), (t("FILETYPE_ALL"), "*.*")],
        )
        if not path:
            return

        template_path_var.set(path)
        template_label_var.set(shorten_path(path))
        use_template_var.set(True)
        on_toggle_template()
        _refresh_template_ui()
        _show_template_naming_hint_once(
            root,
            ui_config,
            lambda cfg: save_ui_config(Path(DATA_DIR), cfg),
            sql_path=selected_sql_path_full.get() or None,
            template_path=template_path_var.get(),
        )

    def choose_save_as_path():
        output_format = format_var.get()
        default_ext = _expected_output_extension(output_format)
        filetypes = (
            [(t("FILETYPE_CSV"), "*.csv"), (t("FILETYPE_ALL"), "*.*")]
            if output_format == "csv"
            else [(t("FILETYPE_EXCEL"), "*.xlsx"), (t("FILETYPE_ALL"), "*.*")]
        )
        path = filedialog.asksaveasfilename(
            title=t("BTN_SAVE_AS"),
            defaultextension=default_ext,
            initialdir=output_directory,
            filetypes=filetypes,
        )
        if not path:
            return
        save_as_path_var.set(path)
        _sync_save_as_extension(show_info=True)

    def clear_save_as_path():
        save_as_path_var.set("")

    def refresh_csv_profile_controls(selected_name=None):
        config = csv_profile_state.get("config", {"profiles": []})
        profiles = config.get("profiles", [])
        names = [p["name"] for p in profiles]

        combo = csv_profile_state.get("combobox")
        if combo is not None:
            combo["values"] = names

        chosen = selected_name or selected_csv_profile_var.get()
        if chosen and chosen in names:
            selected_csv_profile_var.set(chosen)
        elif names:
            selected_csv_profile_var.set(config.get("default_profile") or names[0])
        else:
            selected_csv_profile_var.set("")

        default_name = config.get("default_profile")
        if default_name:
            default_csv_label_var.set(
                t("CSV_DEFAULT_PROFILE_LABEL", name=default_name)
            )
        else:
            default_csv_label_var.set("")

    def open_queries_manager():
        dlg = tk.Toplevel(root)
        apply_app_icon(dlg)
        dlg.title(t("TITLE_EDIT_QUERIES"))
        dlg.transient(root)
        dlg.grab_set()
        dlg.resizable(True, True)

        # Uwaga: nie modyfikuj query_paths_state przed udanym zapisem (bez efektów ubocznych).
        raw_paths = load_query_paths()
        paths = []
        seen_keys = set()
        for raw in raw_paths:
            key = query_path_key(raw)
            if key in seen_keys:
                continue
            paths.append(raw)
            seen_keys.add(key)

        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(0, weight=1)
        dlg.rowconfigure(1, weight=0)

        list_frame = tk.Frame(dlg)
        list_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=(10, 5))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        list_var = tk.StringVar(value=[])
        listbox = tk.Listbox(
            list_frame,
            listvariable=list_var,
            width=80,
            height=10,
            activestyle="dotbox",
            selectmode=tk.EXTENDED,
        )
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        listbox.config(yscrollcommand=scrollbar.set)
        listbox.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        def refresh_list():
            list_var.set(list(paths))
            update_delete_state()

        def update_delete_state(*_):  # noqa: ANN001
            has_selection = bool(listbox.curselection())
            delete_state = tk.NORMAL if has_selection else tk.DISABLED
            delete_btn.config(state=delete_state)

        def add_from_dialog():
            selected = filedialog.askopenfilenames(
                title=t("TITLE_ADD_SQL_FILES"),
                filetypes=[(t("FILETYPE_SQL"), "*.sql"), (t("FILETYPE_ALL"), "*.*")],
            )
            if not selected:
                return
            existing_keys = {query_path_key(p) for p in paths}

            added_indices = []
            skipped_non_sql = []
            skipped_duplicates = 0

            for p in selected:
                if not is_sql_path(p):
                    skipped_non_sql.append(p)
                    continue

                p_key = query_path_key(p)
                if p_key in existing_keys:
                    skipped_duplicates += 1
                    continue

                paths.append(to_storage_path(p))
                existing_keys.add(p_key)
                added_indices.append(len(paths) - 1)

            if skipped_non_sql:
                messagebox.showwarning(
                    t("WARN_SKIPPED_FILES_TITLE"),
                    t(
                        "WARN_SKIPPED_FILES_BODY",
                        files="\n".join(skipped_non_sql[:20]),
                        more=(
                            t("WARN_SKIPPED_FILES_MORE", count=len(skipped_non_sql) - 20)
                            if len(skipped_non_sql) > 20
                            else ""
                        ),
                    ),
                )

            if not added_indices:
                if skipped_duplicates and not skipped_non_sql:
                    messagebox.showinfo(
                        t("CSV_PROFILE_INFO_SAVED_TITLE"),
                        t("INFO_ALREADY_LISTED"),
                    )
                return

            refresh_list()

            listbox.selection_clear(0, tk.END)
            for idx in added_indices:
                listbox.selection_set(idx)

            last_idx = added_indices[-1]
            listbox.activate(last_idx)
            listbox.see(last_idx)
            update_delete_state()

        def edit_selected(event=None):  # noqa: ANN001
            sel = listbox.curselection()
            if not sel:
                return

            idx = sel[0]
            current_path = paths[idx]
            new_path = simpledialog.askstring(
                t("TITLE_EDIT_QUERY_PATH"),
                t("PROMPT_EDIT_QUERY_PATH"),
                initialvalue=current_path,
                parent=dlg,
            )
            if new_path is None:
                return

            new_path = new_path.strip()
            if not new_path:
                return
            if not is_sql_path(new_path):
                messagebox.showwarning(
                    t("WARN_TITLE"),
                    t("WARN_INVALID_SQL_FILE"),
                )
                return

            new_key = query_path_key(new_path)

            # Duplikaty sprawdzamy po kluczu pliku (z pominięciem edytowanego indeksu)
            for j, existing in enumerate(paths):
                if j == idx:
                    continue
                if query_path_key(existing) == new_key:
                    messagebox.showinfo(
                        t("CSV_PROFILE_INFO_SAVED_TITLE"),
                        t("INFO_ALREADY_LISTED"),
                    )
                    return

            # Opcjonalna polerka: ostrzeż, ale nie blokuj (sieciówki / dyski zewnętrzne)
            resolved_new_path = resolve_path(new_path)
            if not os.path.isfile(resolved_new_path):
                messagebox.showwarning(
                    t("WARN_FILE_MISSING_TITLE"),
                    t("WARN_FILE_MISSING_BODY", path=resolved_new_path),
                )

            # Zapis „ładnej” wersji ścieżki (bez normcase)
            paths[idx] = to_storage_path(new_path)

            refresh_list()
            listbox.selection_clear(0, tk.END)
            listbox.selection_set(idx)
            listbox.activate(idx)
            listbox.see(idx)
            update_delete_state()

        def delete_selected(event=None):  # noqa: ANN001
            sel = listbox.curselection()
            if not sel:
                messagebox.showinfo(
                    t("CSV_PROFILE_INFO_SAVED_TITLE"),
                    t("INFO_SELECT_ENTRY_DELETE"),
                )
                return "break" if event is not None else None

            for idx in reversed(sel):
                paths.pop(idx)

            refresh_list()

            if paths:
                next_idx = min(sel[0], len(paths) - 1)
                listbox.selection_set(next_idx)
                listbox.activate(next_idx)
                listbox.see(next_idx)
                update_delete_state()

            return "break" if event is not None else None

        def save_and_close(event=None):  # noqa: ANN001
            try:
                save_query_paths(paths)
            except OSError as exc:
                messagebox.showerror(
                    t("ERR_QUERIES_SAVE_TITLE"),
                    t("ERR_QUERIES_SAVE_BODY", error=exc),
                )
                return

            query_paths_state["paths"] = list(paths)
            dlg.destroy()

        def cancel_dialog(event=None):  # noqa: ANN001
            dlg.destroy()

        button_frame = tk.Frame(dlg)
        button_frame.grid(row=1, column=0, pady=(0, 10), padx=10, sticky="e")

        add_btn = tk.Button(
            button_frame,
            text=t("BTN_ADD_FILES"),
            command=add_from_dialog,
            width=15,
        )
        delete_btn = tk.Button(
            button_frame,
            text=t("BTN_REMOVE_SELECTED"),
            command=delete_selected,
            width=18,
        )
        save_btn = tk.Button(button_frame, text=t("BTN_SAVE"), command=save_and_close, width=12)
        cancel_btn = tk.Button(
            button_frame, text=t("BTN_CANCEL"), command=cancel_dialog, width=12
        )

        add_btn.pack(side="left", padx=(0, 5))
        delete_btn.pack(side="left", padx=(0, 20))
        save_btn.pack(side="left", padx=(0, 5))
        cancel_btn.pack(side="left")

        listbox.bind("<<ListboxSelect>>", update_delete_state)
        listbox.bind("<Double-Button-1>", edit_selected)
        listbox.bind("<Delete>", delete_selected)
        dlg.bind("<Return>", save_and_close)
        dlg.bind("<Escape>", cancel_dialog)
        dlg.protocol("WM_DELETE_WINDOW", cancel_dialog)

        refresh_list()

        if paths:
            listbox.selection_set(0)
            listbox.activate(0)
            listbox.see(0)
        update_delete_state()

        listbox.focus_set()

        _center_window(dlg, root)

    def choose_from_list():
        current_paths = query_paths_state["paths"]
        if not current_paths:
            messagebox.showerror(t("ERR_TITLE"), t("ERR_NO_REPORTS"))
            return

        dlg = tk.Toplevel(root)
        apply_app_icon(dlg)
        dlg.title(t("TITLE_SELECT_REPORT"))

        dlg.transient(root)
        dlg.resizable(True, True)

        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(0, weight=1)

        list_frame = tk.Frame(dlg)
        list_frame.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=10, pady=(10, 5))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        listbox = tk.Listbox(list_frame, width=80, activestyle="dotbox")
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        listbox.config(yscrollcommand=scrollbar.set)
        listbox.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        for p in current_paths:
            listbox.insert(tk.END, os.path.basename(p))

        path_label_var = tk.StringVar(value="")
        path_label = tk.Label(
            dlg,
            textvariable=path_label_var,
            anchor="w",
            justify="left",
            wraplength=600,
        )
        path_label.grid(row=1, column=0, columnspan=2, sticky="we", padx=10, pady=(0, 10))

        button_frame = tk.Frame(dlg)
        button_frame.grid(row=2, column=0, columnspan=2, pady=(0, 10), padx=10, sticky="e")

        def update_path_label(*_):
            sel = listbox.curselection()
            if sel:
                path_label_var.set(current_paths[sel[0]])
            else:
                path_label_var.set("")

        def on_ok(*_):
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning(t("WARN_TITLE"), t("WARN_NO_REPORT_SELECTED"))
                return
            idx = sel[0]
            _set_sql_path(current_paths[idx])
            dlg.destroy()

        def on_cancel(*_):
            dlg.destroy()

        ok_btn = tk.Button(button_frame, text=t("BTN_OK"), width=12, command=on_ok)
        cancel_btn = tk.Button(
            button_frame, text=t("BTN_CANCEL"), width=12, command=on_cancel
        )
        ok_btn.pack(side="left", padx=(0, 5))
        cancel_btn.pack(side="left")

        listbox.bind("<<ListboxSelect>>", update_path_label)
        listbox.bind("<Double-Button-1>", on_ok)
        dlg.bind("<Return>", on_ok)
        dlg.bind("<Escape>", on_cancel)

        if current_paths:
            listbox.selection_set(0)
            listbox.activate(0)
            listbox.see(0)
            update_path_label()

        listbox.focus_set()

        _center_window(dlg, root)

        dlg.grab_set()
        root.wait_window(dlg)

    def _build_export_params():
        use_pasted_sql = bool(
            pasted_sql_state["sql"] and pasted_sql_state["report_name"]
        )
        if use_pasted_sql:
            sql_query = pasted_sql_state["sql"]
            base_name = pasted_sql_state["report_name"]
            sql_source_path = f"<pasted:{base_name}>"
        else:
            path = selected_sql_path_full.get()
            if not path:
                messagebox.showerror(t("ERR_TITLE"), t("ERR_NO_SQL_SOURCE"))
                return None
            if not os.path.isfile(path):
                messagebox.showerror(t("ERR_TITLE"), t("ERR_SQL_NOT_FOUND"))
                return None

            ok, msg = validate_sql_text_file(path)
            if not ok:
                messagebox.showerror(t("ERR_INVALID_SQL_FILE_TITLE"), msg)
                return None

            with open(path, "rb") as f:
                content = f.read()
            sql_query = remove_bom(content).strip()
            base_name = os.path.basename(path)
            sql_source_path = path

        engine = engine_holder.get("engine")
        current_connection = get_connection_by_name(selected_connection_var.get())
        if engine is None or current_connection is None:
            messagebox.showerror(
                t("ERR_NO_CONNECTION_TITLE"),
                t("ERR_NEED_CONNECTION"),
            )
            return None

        output_format = format_var.get()
        use_template = use_template_var.get()

        csv_profile = None
        if output_format == "csv":
            csv_config = csv_profile_state["config"]
            profile_name = selected_csv_profile_var.get() or csv_config.get("default_profile")
            csv_profile = (
                get_csv_profile(csv_config, profile_name)
                or get_csv_profile(csv_config, csv_config.get("default_profile"))
                or csv_config.get("profiles", [DEFAULT_CSV_PROFILE])[0]
            )

        if use_template:
            if output_format != "xlsx":
                messagebox.showerror(
                    t("ERR_TITLE"),
                    t("ERR_TEMPLATE_ONLY_XLSX"),
                )
                return None
            if not template_path_var.get():
                messagebox.showerror(t("ERR_TITLE"), t("ERR_TEMPLATE_NOT_SELECTED"))
                return None
            if not sheet_name_var.get():
                messagebox.showerror(
                    t("ERR_TITLE"), t("ERR_TEMPLATE_SHEET_NOT_SELECTED")
                )
                return None

            if use_pasted_sql:
                output_file_name = f"{base_name}.xlsx"
            else:
                output_file_name = os.path.splitext(base_name)[0] + ".xlsx"
            output_override = (save_as_path_var.get() or "").strip()
            output_file_path, ext_mismatch = normalize_output_file_path(
                output_directory=output_directory,
                default_file_name=output_file_name,
                output_format=output_format,
                override_path=(output_override or None),
            )
            if (
                ext_mismatch
                and output_override
                and not _looks_like_directory(output_override)
            ):
                messagebox.showinfo(
                    t("INFO_SAVE_AS_EXT_FIXED_TITLE"),
                    t("INFO_SAVE_AS_EXT_FIXED_BODY", path=output_file_path),
                )
                save_as_path_var.set(output_file_path)

            return {
                "engine": engine,
                "sql_query": sql_query,
                "report_label": base_name,
                "sql_source_path": sql_source_path,
                "connection_name": current_connection.get("name", ""),
                "connection_type": current_connection.get("type", ""),
                "output_format": output_format,
                "output_file_path": output_file_path,
                "csv_profile": csv_profile,
                "use_template": True,
                "template": {
                    "template_path": template_path_var.get(),
                    "sheet_name": sheet_name_var.get(),
                    "start_cell": (start_cell_var.get() or "A2").strip(),
                    "include_header": include_header_var.get(),
                },
            }

        if use_pasted_sql:
            output_file_name = base_name + (".xlsx" if output_format == "xlsx" else ".csv")
        else:
            output_file_name = os.path.splitext(base_name)[0] + (
                ".xlsx" if output_format == "xlsx" else ".csv"
            )
        output_override = (save_as_path_var.get() or "").strip()
        output_file_path, ext_mismatch = normalize_output_file_path(
            output_directory=output_directory,
            default_file_name=output_file_name,
            output_format=output_format,
            override_path=(output_override or None),
        )
        if (
            ext_mismatch
            and output_override
            and not _looks_like_directory(output_override)
        ):
            messagebox.showinfo(
                t("INFO_SAVE_AS_EXT_FIXED_TITLE"),
                t("INFO_SAVE_AS_EXT_FIXED_BODY", path=output_file_path),
            )
            save_as_path_var.set(output_file_path)

        return {
            "engine": engine,
            "sql_query": sql_query,
            "report_label": base_name,
            "sql_source_path": sql_source_path,
            "connection_name": current_connection.get("name", ""),
            "connection_type": current_connection.get("type", ""),
            "output_format": output_format,
            "output_file_path": output_file_path,
            "csv_profile": csv_profile,
            "use_template": False,
        }

    def _reset_run_state():
        cancel_state["event"] = None
        try:
            btn_cancel.config(state=tk.DISABLED)
        except Exception:  # noqa: BLE001
            pass
        try:
            btn_cancel.grid()
            lbl_export_info.grid()
            lbl_export_info_value.grid()
        except Exception:  # noqa: BLE001
            pass
        try:
            btn_start.config(state=tk.NORMAL)
        except Exception:  # noqa: BLE001
            pass
        if close_state.get("after_cancel"):
            root.after(0, root.destroy)


    def run_export_gui():
        sql_query = ""
        params: dict | None = None
        db_timeout_seconds = 0
        export_timeout_seconds = 0
    
        try:
            params = _build_export_params()
            if not params:
                return
    
            sql_query = params.get("sql_query", "") or ""
            db_timeout_seconds = max(0, int(db_timeout_min_var.get() or 0)) * 60
            export_timeout_seconds = max(0, int(export_timeout_min_var.get() or 0)) * 60
        except Exception as exc:  # noqa: BLE001
            sql_src = (params or {}).get("sql_source_path", "") if isinstance(params, dict) else ""
            ui_msg = format_error_for_ui(
                exc,
                sql_query,
                sql_source_path=sql_src,
                context="export",
            )
            result_info_var.set(t("ERR_EXPORT"))
            update_error_display(ui_msg)
            show_error_popup(
                ui_msg,
                sql_query=sql_query,
                sql_source_path=sql_src,
            )
            try:
                btn_start.config(state=tk.NORMAL)
            except Exception:
                pass
            return
    
        try:
            persist_db_timeout_seconds(db_timeout_seconds)
            persist_export_timeout_seconds(export_timeout_seconds)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Persist timeouts failed: %s", exc, exc_info=True)
    
        result_info_var.set(t("MSG_RUNNING"))
        cancel_state["event"] = threading.Event()
        btn_cancel.config(state=tk.NORMAL)
        try:
            btn_cancel.grid_remove()
            lbl_export_info.grid_remove()
            lbl_export_info_value.grid_remove()
        except Exception:  # noqa: BLE001
            pass
        btn_start.config(state=tk.DISABLED)
        root.update_idletasks()
    
        output_format = params["output_format"]
        csv_profile = params.get("csv_profile")
    
        def _handle_export_error(exc: Exception):
            if isinstance(exc, XlsxSizeError):
                msg = str(exc)
                result_info_var.set(msg)
                update_error_display("")
                messagebox.showwarning(t("WARN_TITLE"), msg)
            elif isinstance(exc, QueryTimeoutError):
                msg = str(exc)
                result_info_var.set(msg)
                update_error_display("")
                messagebox.showwarning(t("WARN_TITLE"), msg)
            elif isinstance(exc, ExportTimeoutError):
                msg = str(exc)
                result_info_var.set(msg)
                update_error_display("")
                messagebox.showwarning(t("WARN_TITLE"), msg)
            elif isinstance(exc, UserCancelledError):
                msg = str(exc)
                result_info_var.set(msg)
                update_error_display("")
                if not close_state.get("after_cancel"):
                    messagebox.showinfo(t("WARN_TITLE"), msg)
            else:
                ui_msg = format_error_for_ui(
                    exc,
                    sql_query,
                    sql_source_path=params.get("sql_source_path", ""),
                    context="export",
                )
                result_info_var.set(t("ERR_EXPORT"))
                update_error_display(ui_msg)
                show_error_popup(
                    ui_msg,
                    sql_query=sql_query,
                    sql_source_path=params.get("sql_source_path", ""),
                )
            _reset_run_state()
    
        def _handle_export_success(result: dict):
            sql_dur = result["sql_duration"]
            export_dur = result["export_duration"]
            total_dur = result["total_duration"]
            rows_count = result["rows_count"]
    
            last_output_path["path"] = params["output_file_path"]
    
            if archive_sql_var.get():
                try:
                    write_sql_archive_entry(
                        sql_query=sql_query,
                        report_label=params.get("report_label", ""),
                        sql_source_path=params.get("sql_source_path", ""),
                        output_file_path=params["output_file_path"],
                        output_format=("xlsx" if params.get("use_template") else output_format),
                        rows_count=rows_count,
                        sql_duration=sql_dur,
                        export_duration=export_dur,
                        total_duration=total_dur,
                        connection_name=params.get("connection_name", ""),
                        connection_type=params.get("connection_type", ""),
                    )
                except Exception as exc:  # noqa: BLE001
                    LOGGER.warning("SQL archive failed: %s", exc, exc_info=True)
    
            if output_format == "csv" and csv_profile:
                prof_name = (csv_profile.get("name") or "").strip()
                if prof_name:
                    csv_profile_state["config"] = remember_last_used_csv_profile(
                        prof_name,
                        csv_profile_state["config"],
                    )
                    refresh_csv_profile_controls(prof_name)
    
            if rows_count > 0:
                safe_sql_dur = _coerce_seconds(sql_dur)
                safe_export_dur = _coerce_seconds(export_dur)
                safe_total_dur = _coerce_seconds(total_dur)
                msg = t(
                    "MSG_SAVED_DETAILS",
                    path=params["output_file_path"],
                    rows=rows_count,
                    sql_time=safe_sql_dur,
                    sql_time_hms=_fmt_hms(safe_sql_dur),
                    export_time=safe_export_dur,
                    export_time_hms=_fmt_hms(safe_export_dur),
                    total_time=safe_total_dur,
                    total_time_hms=_fmt_hms(safe_total_dur),
                )
                if output_format == "csv" and csv_profile:
                    msg += "\n" + t("MSG_SAVED_DETAILS_CSV", profile=csv_profile.get("name", ""))
            else:
                safe_sql_dur = _coerce_seconds(sql_dur)
                msg = t(
                    "MSG_NO_ROWS",
                    sql_time=safe_sql_dur,
                    sql_time_hms=_fmt_hms(safe_sql_dur),
                )
                if output_format == "csv" and csv_profile:
                    msg += "\n" + t("MSG_SAVED_DETAILS_CSV", profile=csv_profile.get("name", ""))
    
            result_info_var.set(msg)
            messagebox.showinfo(t("MSG_DONE"), msg)
            btn_open_file.config(state=tk.NORMAL)
            btn_open_folder.config(state=tk.NORMAL)
            update_error_display("")
            _reset_run_state()
    
        def _run_export_in_background(work_fn):
            q = queue.Queue()
            progress = ExportProgressWindow(
                root,
                title=t("PROGRESS_TITLE"),
                on_stop=request_cancel,
            )
            progress.win.after(0, lambda: _center_toplevel_on_parent(progress.win, root))
            try:
                progress.win.lift()
                progress.win.focus_set()
            except Exception:
                pass
    
            def _worker():
                try:
                    def report_step(msg: str):
                        q.put(("step", msg))
    
                    result = work_fn(report_step)
                    q.put(("done", result))
                except Exception as exc:  # noqa: BLE001
                    q.put(("error", exc))
    
            threading.Thread(target=_worker, daemon=True).start()
    
            def _poll():
                should_stop = False
                try:
                    while True:
                        kind, payload = q.get_nowait()
                        if kind == "step":
                            progress.set_step(str(payload))
                        elif kind == "done":
                            progress.close()
                            _handle_export_success(payload)
                            should_stop = True
                        elif kind == "error":
                            progress.close()
                            _handle_export_error(payload)
                            should_stop = True
                except queue.Empty:
                    pass
                if should_stop:
                    return
                root.after(200, _poll)
    
            root.after(200, _poll)
    
        def _work(report_step):
            engine = params["engine"]
            report_step(t("PROGRESS_GETTING_DATA"))
    
            rows, columns, sql_duration, sql_start = _run_query_to_rows(
                engine,
                sql_query,
                timeout_seconds=db_timeout_seconds,
                cancel_event=cancel_state["event"],
            )
            rows_count = len(rows)
            columns_count = len(columns)
    
            if params.get("use_template"):
                template = params["template"]
                start_row, start_col = coordinate_to_tuple(template["start_cell"])
                _ensure_xlsx_limits(
                    rows_count,
                    columns_count,
                    header_rows=(1 if (template["include_header"] and rows_count) else 0),
                    start_row=start_row,
                    start_col=start_col,
                )
                report_step(t("PROGRESS_EXPORTING_XLSX_TEMPLATE"))
    
                export_start = time.perf_counter()
                try:
                    shutil.copyfile(template["template_path"], params["output_file_path"])
    
                    wb = None
                    try:
                        if rows_count:
                            deadline = _deadline(int(export_timeout_seconds or 0))
                            wb = load_workbook(params["output_file_path"])
                            if template["sheet_name"] not in wb.sheetnames:
                                raise ValueError(
                                    t("ERR_TEMPLATE_MISSING_SHEET", sheet=template["sheet_name"])
                                )
    
                            ws = wb[template["sheet_name"]]
    
                            data_start_row = start_row
                            if template["include_header"]:
                                for c_offset, col_name in enumerate(list(columns)):
                                    if c_offset % 50 == 0:
                                        if cancel_state["event"] is not None and cancel_state["event"].is_set():
                                            raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                                        _check_deadline(
                                            deadline,
                                            ExportTimeoutError,
                                            t(
                                                "ERR_EXPORT_TIMEOUT",
                                                minutes=max(1, int(math.ceil(int(export_timeout_seconds or 0) / 60))),
                                            ),
                                        )
                                    ws.cell(row=start_row, column=start_col + c_offset).value = col_name
                                data_start_row = start_row + 1
    
                            for r_offset, row in enumerate(rows):
                                if r_offset % 100 == 0:
                                    if cancel_state["event"] is not None and cancel_state["event"].is_set():
                                        raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                                    _check_deadline(
                                        deadline,
                                        ExportTimeoutError,
                                        t(
                                            "ERR_EXPORT_TIMEOUT",
                                            minutes=max(1, int(math.ceil(int(export_timeout_seconds or 0) / 60))),
                                        ),
                                    )
                                for c_offset, value in enumerate(list(row)):
                                    ws.cell(
                                        row=data_start_row + r_offset,
                                        column=start_col + c_offset,
                                    ).value = value
                            if cancel_state["event"] is not None and cancel_state["event"].is_set():
                                raise UserCancelledError(t("ERR_CANCELLED_BY_USER"))
                            wb.save(params["output_file_path"])
                    finally:
                        if wb is not None:
                            try:
                                wb.close()
                            except Exception:
                                pass
                except Exception:
                    _safe_remove(params["output_file_path"])
                    raise
    
                export_end = time.perf_counter()
                export_duration = export_end - export_start
                total_duration = export_end - sql_start
            else:
                if output_format == "xlsx":
                    _ensure_xlsx_limits(
                        rows_count,
                        columns_count,
                        header_rows=(1 if rows_count else 0),
                        start_row=1,
                        start_col=1,
                    )
    
                report_step(t("PROGRESS_EXPORTING_CSV") if output_format == "csv" else t("PROGRESS_EXPORTING_XLSX"))
    
                export_duration = 0.0
                if rows_count:
                    export_start = time.perf_counter()
                    try:
                        if output_format == "xlsx":
                            _export_rows_to_xlsx(
                                params["output_file_path"],
                                columns=list(columns),
                                rows=rows,
                                timeout_seconds=int(export_timeout_seconds or 0),
                                cancel_event=cancel_state["event"],
                            )
                        else:
                            profile = csv_profile or DEFAULT_CSV_PROFILE
                            _export_rows_to_csv(
                                params["output_file_path"],
                                columns=list(columns),
                                rows=rows,
                                profile=profile,
                                timeout_seconds=int(export_timeout_seconds or 0),
                                cancel_event=cancel_state["event"],
                            )
                    except Exception:
                        _safe_remove(params["output_file_path"])
                        raise
                    export_end = time.perf_counter()
                    export_duration = export_end - export_start
                    total_duration = export_end - sql_start
                else:
                    total_duration = sql_duration
    
            return {
                "sql_duration": sql_duration,
                "export_duration": export_duration,
                "total_duration": total_duration,
                "rows_count": rows_count,
            }
    
        _run_export_in_background(_work)

    def _open_path(target):
        if not target or not os.path.exists(target):
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(target)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.run(["open", target], check=False)
            else:
                subprocess.run(["xdg-open", target], check=False)
        except Exception as err:  # noqa: BLE001
            messagebox.showerror(t("ERR_TITLE"), str(err))

    def open_file():
        path = last_output_path.get("path")
        _open_path(path)

    def open_folder():
        path = last_output_path.get("path")
        if path and os.path.isfile(path):
            folder = os.path.dirname(path)
            _open_path(folder)

    def open_logs_folder():
        target = LOG_DIR
        if not target and LOG_FILE_PATH:
            try:
                target = os.path.dirname(LOG_FILE_PATH)
            except Exception:
                target = None
        if not target:
            target = _build_path("logs")
        _open_path(target)

    def show_help_window():
        # Help == README viewer (no nested modal windows).
        show_readme_window(root)

    def check_updates_gui():
        prev_status = connection_status_var.get()
        prev_key = connection_status_state.get("key")
        prev_params = connection_status_state.get("params") or {}
        connection_status_var.set(t("UPD_CHECKING"))
        root.update_idletasks()
        mode = detect_install_mode()
        git_status_line = None
        if mode == "git":
            local_sha = _get_git_full_sha()
            remote_sha = None
            try:
                remote_sha = _fetch_remote_main_sha()
            except Exception:
                remote_sha = None
            if local_sha and remote_sha:
                local_short = local_sha[:7]
                remote_short = remote_sha[:7]
                relation = _classify_git_relation(local_sha, remote_sha)
                if relation == "match":
                    key = "UPD_GIT_STATUS_MATCH"
                elif relation == "remote_ahead":
                    key = "UPD_GIT_STATUS_AHEAD"
                elif relation == "local_ahead":
                    key = "UPD_GIT_STATUS_LOCAL_AHEAD"
                elif relation == "diverged":
                    key = "UPD_GIT_STATUS_DIVERGED"
                elif relation == "different_unverified":
                    key = "UPD_GIT_STATUS_DIFFERENT_UNVERIFIED"
                else:
                    key = "UPD_GIT_STATUS_DIFFERENT"
                git_status_line = t(key, local=local_short, remote=remote_short)
        try:
            info = get_update_info()
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Update check failed: %s", exc, exc_info=True)
            messagebox.showerror(t("UPD_TITLE"), _build_update_check_message(exc))
            return
        finally:
            if prev_key:
                set_connection_status(key=prev_key, **prev_params)
            else:
                connection_status_var.set(prev_status)

        latest_tag = info.get("latest_tag") or ""
        current_tag = info.get("current_tag") or f"v{APP_VERSION}"
        if not latest_tag:
            messagebox.showerror(t("UPD_TITLE"), t("UPD_ERR_JSON"))
            return
        if mode == "exe" and not info.get("asset"):
            messagebox.showerror(t("UPD_TITLE"), t("UPD_ERR_JSON"))
            return

        if not info.get("update_available"):
            message_lines = [
                t("UPD_NO_UPDATE", current=current_tag, latest=latest_tag)
            ]
            if mode == "git":
                if git_status_line:
                    message_lines.append(git_status_line)
                message_lines.append(t("UPD_GIT_MODE", command="git pull"))
            elif mode == "source":
                message_lines.append(t("UPD_SOURCE_MODE"))
            messagebox.showinfo(
                t("UPD_TITLE"),
                "\n\n".join(message_lines),
            )
            return
        if mode in {"git", "source"}:
            message_lines = [
                t("UPD_UPDATE_AVAILABLE", current=current_tag, latest=latest_tag)
            ]
            if mode == "git":
                if git_status_line:
                    message_lines.append(git_status_line)
                message_lines.append(t("UPD_GIT_MODE", command="git pull"))
            else:
                message_lines.append(t("UPD_SOURCE_MODE"))
            messagebox.showinfo(t("UPD_TITLE"), "\n\n".join(message_lines))
            return
        prompt = "\n\n".join(
            [
                t("UPD_UPDATE_AVAILABLE", current=current_tag, latest=latest_tag),
                t("UPD_PROMPT_INSTALL"),
            ]
        )
        if not messagebox.askyesno(t("UPD_TITLE"), prompt):
            return
        if not _get_updater_command():
            messagebox.showerror(t("UPD_TITLE"), t("UPD_UPDATER_MISSING"))
            return
        try:
            if not launch_updater(wait_pid=os.getpid()):
                messagebox.showerror(t("UPD_TITLE"), t("UPD_START_FAILED"))
                return
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Failed to start updater: %s", exc, exc_info=True)
            messagebox.showerror(t("UPD_TITLE"), t("UPD_START_FAILED"))
            return
        root.destroy()

    connection_frame = ttk.LabelFrame(
        root, text=t("FRAME_DB_CONNECTION"), padding=(10, 10)
    )
    connection_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
    i18n_widgets["connection_frame"] = connection_frame

    status_label = ttk.Label(
        connection_frame,
        textvariable=connection_status_var,
        justify="left",
        anchor="w",
    )
    status_label.grid(row=0, column=0, sticky="we")
    connection_frame.columnconfigure(0, weight=1)

    def _update_status_wrap(event=None):
        width = connection_frame.winfo_width() or status_label.winfo_width() or 0
        wrap = max(width - 20, 200)
        status_label.config(wraplength=wrap)

    connection_frame.bind("<Configure>", _update_status_wrap)
    status_label.bind("<Configure>", _update_status_wrap)

    connection_controls = ttk.Frame(connection_frame)
    connection_controls.grid(row=1, column=0, sticky="we", pady=(5, 0))
    connection_controls.columnconfigure(1, weight=1)

    lbl_connection = ttk.Label(connection_controls, text=t("LBL_CONNECTION"))
    lbl_connection.grid(row=0, column=0, sticky="w")
    i18n_widgets["lbl_connection"] = lbl_connection
    connection_combo = ttk.Combobox(
        connection_controls,
        textvariable=selected_connection_var,
        state="readonly",
        width=50,
    )
    connection_combo.grid(row=0, column=1, sticky="we", padx=(5, 0))
    connections_state["combobox"] = connection_combo
    connection_combo.bind("<<ComboboxSelected>>", on_connection_change)

    lbl_language = ttk.Label(connection_controls, text=t("LBL_LANGUAGE"))
    lbl_language.grid(row=1, column=0, sticky="w", pady=(6, 0))
    i18n_widgets["lbl_language"] = lbl_language
    lang_combo = ttk.Combobox(
        connection_controls,
        textvariable=lang_var,
        values=["EN", "PL"],
        state="readonly",
        width=6,
    )
    lang_combo.grid(row=1, column=1, sticky="w", padx=(5, 0), pady=(6, 0))
    i18n_widgets["lang_combo"] = lang_combo

    buttons_frame = ttk.Frame(connection_controls)
    buttons_frame.grid(row=0, column=2, sticky="e", padx=(10, 0))

    btn_manage_connections = ttk.Button(
        buttons_frame,
        text=t("BTN_MANAGE_CONNECTIONS"),
        command=open_connections_manager_gui,
    )
    btn_manage_connections.pack(side="left", padx=(0, 8))
    i18n_widgets["btn_manage_connections"] = btn_manage_connections

    btn_test_connection = ttk.Button(
        buttons_frame,
        text=t("BTN_TEST_CONNECTION"),
        command=test_connection_only,
    )
    btn_test_connection.pack(side="left", padx=(0, 8))
    i18n_widgets["btn_test_connection"] = btn_test_connection

    btn_odbc_diagnostics = ttk.Button(
        buttons_frame,
        text=t("BTN_ODBC_DIAGNOSTICS"),
        command=show_odbc_diagnostics_popup,
    )
    btn_odbc_diagnostics.pack(side="left")
    i18n_widgets["btn_odbc_diagnostics"] = btn_odbc_diagnostics

    advanced_frame = ttk.LabelFrame(
        connection_frame, text=t("FRAME_ADVANCED"), padding=(10, 8)
    )
    advanced_frame.grid(row=2, column=0, sticky="we", pady=(8, 0))
    i18n_widgets["advanced_frame"] = advanced_frame

    chk_archive_sql = ttk.Checkbutton(
        advanced_frame,
        text=t("CHK_ARCHIVE_SQL"),
        variable=archive_sql_var,
        command=on_archive_sql_toggle,
    )
    chk_archive_sql.grid(row=0, column=0, columnspan=2, sticky="w")
    i18n_widgets["chk_archive_sql"] = chk_archive_sql

    advanced_frame.columnconfigure(0, weight=1)
    advanced_frame.columnconfigure(1, weight=0)

    lbl_db_timeout = ttk.Label(advanced_frame, text=t("LBL_DB_TIMEOUT_MIN"))
    lbl_db_timeout.grid(row=1, column=0, sticky="w", pady=(6, 0))
    i18n_widgets["lbl_db_timeout"] = lbl_db_timeout

    sp_db_timeout = ttk.Spinbox(
        advanced_frame,
        from_=0,
        to=1440,
        textvariable=db_timeout_min_var,
        width=6,
        command=persist_timeouts_from_ui,
    )
    sp_db_timeout.grid(row=1, column=1, sticky="e", pady=(6, 0))
    sp_db_timeout.bind("<FocusOut>", lambda *_: persist_timeouts_from_ui())
    sp_db_timeout.bind("<Return>", lambda *_: persist_timeouts_from_ui())

    lbl_export_timeout = ttk.Label(advanced_frame, text=t("LBL_EXPORT_TIMEOUT_MIN"))
    lbl_export_timeout.grid(row=2, column=0, sticky="w", pady=(6, 0))
    i18n_widgets["lbl_export_timeout"] = lbl_export_timeout

    sp_export_timeout = ttk.Spinbox(
        advanced_frame,
        from_=0,
        to=1440,
        textvariable=export_timeout_min_var,
        width=6,
        command=persist_timeouts_from_ui,
    )
    sp_export_timeout.grid(row=2, column=1, sticky="e", pady=(6, 0))
    sp_export_timeout.bind("<FocusOut>", lambda *_: persist_timeouts_from_ui())
    sp_export_timeout.bind("<Return>", lambda *_: persist_timeouts_from_ui())

    lbl_timeout_note = ttk.Label(
        advanced_frame, text=t("LBL_TIMEOUT_NOTE"), foreground="gray40"
    )
    lbl_timeout_note.grid(row=3, column=0, columnspan=2, sticky="w", pady=(6, 0))
    i18n_widgets["lbl_timeout_note"] = lbl_timeout_note

    source_frame = ttk.LabelFrame(root, text=t("FRAME_SQL_SOURCE"), padding=(10, 10))
    source_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
    i18n_widgets["source_frame"] = source_frame

    format_frame = ttk.LabelFrame(root, text=t("FRAME_OUTPUT_FORMAT"), padding=(10, 10))
    format_frame.pack(fill=tk.X, padx=10, pady=5)
    i18n_widgets["format_frame"] = format_frame

    template_frame = ttk.LabelFrame(
        root, text=t("FRAME_TEMPLATE_OPTIONS"), padding=(10, 10)
    )
    template_frame.pack(fill=tk.X, padx=10, pady=5)
    i18n_widgets["template_frame"] = template_frame

    result_frame = ttk.LabelFrame(root, text=t("FRAME_RESULTS"), padding=(10, 10))
    result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))
    i18n_widgets["result_frame"] = result_frame

    source_frame.columnconfigure(1, weight=1)
    source_frame.columnconfigure(2, weight=0)
    format_frame.columnconfigure(1, weight=1)
    template_frame.columnconfigure(1, weight=1)
    result_frame.columnconfigure(1, weight=1)
    result_frame.rowconfigure(3, weight=1)

    lbl_selected_sql = ttk.Label(source_frame, text=t("LBL_SELECTED_SQL"))
    lbl_selected_sql.grid(row=0, column=0, sticky="nw")
    i18n_widgets["lbl_selected_sql"] = lbl_selected_sql
    ttk.Label(source_frame, textvariable=sql_label_var, wraplength=600, justify="left").grid(
        row=0, column=1, columnspan=3, sticky="we"
    )

    btn_select_sql = ttk.Button(source_frame, text=t("BTN_SELECT_SQL"), command=choose_sql_file)
    btn_select_sql.grid(
        row=1, column=0, pady=5, sticky="w"
    )
    i18n_widgets["btn_select_sql"] = btn_select_sql
    btn_select_from_list = ttk.Button(
        source_frame, text=t("BTN_SELECT_FROM_LIST"), command=choose_from_list
    )
    btn_select_from_list.grid(
        row=1, column=1, pady=5, sticky="w"
    )
    i18n_widgets["btn_select_from_list"] = btn_select_from_list
    btn_edit_queries = ttk.Button(
        source_frame, text=t("BTN_EDIT_QUERIES"), command=open_queries_manager
    )
    btn_edit_queries.grid(
        row=1, column=2, pady=5, sticky="w"
    )
    i18n_widgets["btn_edit_queries"] = btn_edit_queries
    btn_paste_sql = ttk.Button(
        source_frame, text=t("BTN_PASTE_SQL"), command=open_paste_sql_dialog
    )
    btn_paste_sql.grid(
        row=1, column=3, pady=5, sticky="w"
    )
    i18n_widgets["btn_paste_sql"] = btn_paste_sql

    radio_xlsx = tk.Radiobutton(
        format_frame,
        text=t("FORMAT_XLSX"),
        variable=format_var,
        value="xlsx",
        command=on_format_change,
    )
    radio_xlsx.grid(row=0, column=0, sticky="w")
    i18n_widgets["radio_xlsx"] = radio_xlsx
    radio_csv = tk.Radiobutton(
        format_frame,
        text=t("FORMAT_CSV"),
        variable=format_var,
        value="csv",
        command=on_format_change,
    )
    radio_csv.grid(row=0, column=1, sticky="w")
    i18n_widgets["radio_csv"] = radio_csv

    on_format_change()

    lbl_csv_profile = ttk.Label(format_frame, text=t("LBL_CSV_PROFILE"))
    lbl_csv_profile.grid(
        row=1, column=0, sticky="w", pady=(5, 0)
    )
    i18n_widgets["lbl_csv_profile"] = lbl_csv_profile
    csv_profile_combo = ttk.Combobox(
        format_frame,
        textvariable=selected_csv_profile_var,
        state="readonly",
        width=25,
    )
    csv_profile_combo.grid(row=1, column=1, sticky="w", pady=(5, 0))
    csv_profile_state["combobox"] = csv_profile_combo

    csv_profile_manage_btn = ttk.Button(
        format_frame,
        text=t("BTN_MANAGE_CSV_PROFILES"),
        command=lambda: open_csv_profiles_manager_gui(
            root,
            csv_profile_state,
            selected_csv_profile_var,
            refresh_csv_profile_controls,
        ),
    )
    csv_profile_manage_btn.grid(row=1, column=2, padx=(10, 0), pady=(5, 0), sticky="w")
    csv_profile_state["manage_button"] = csv_profile_manage_btn
    i18n_widgets["csv_profile_manage_btn"] = csv_profile_manage_btn

    ttk.Label(format_frame, textvariable=default_csv_label_var, justify="left", wraplength=600).grid(
        row=2, column=0, columnspan=4, sticky="w", pady=(5, 0)
    )

    lbl_save_as = ttk.Label(format_frame, text=t("LBL_SAVE_AS"))
    lbl_save_as.grid(row=3, column=0, sticky="w", pady=(8, 0))
    i18n_widgets["lbl_save_as"] = lbl_save_as

    ent_save_as = ttk.Entry(format_frame, textvariable=save_as_path_var)
    ent_save_as.grid(row=3, column=1, sticky="we", pady=(8, 0))
    ent_save_as.bind("<FocusOut>", lambda *_: _sync_save_as_extension(show_info=False))
    ent_save_as.bind("<Return>", lambda *_: _sync_save_as_extension(show_info=False))

    btn_save_as = ttk.Button(format_frame, text=t("BTN_SAVE_AS"), command=choose_save_as_path)
    btn_save_as.grid(row=3, column=2, padx=(10, 0), pady=(8, 0), sticky="w")
    i18n_widgets["btn_save_as"] = btn_save_as

    btn_clear_save_as = ttk.Button(format_frame, text=t("BTN_CLEAR"), command=clear_save_as_path)
    btn_clear_save_as.grid(row=3, column=3, padx=(8, 0), pady=(8, 0), sticky="w")
    i18n_widgets["btn_clear_save_as"] = btn_clear_save_as

    lbl_save_as_hint = ttk.Label(
        format_frame,
        text=t("LBL_SAVE_AS_HINT", dir=shorten_path(output_directory, max_len=60)),
        foreground="gray40",
    )
    lbl_save_as_hint.grid(row=4, column=0, columnspan=4, sticky="w", pady=(4, 0))
    i18n_widgets["lbl_save_as_hint"] = lbl_save_as_hint

    refresh_csv_profile_controls(csv_profile_state["config"].get("default_profile"))

    chk_use_template = ttk.Checkbutton(
        template_frame,
        text=t("CHK_USE_TEMPLATE"),
        variable=use_template_var,
        command=on_toggle_template,
    )
    chk_use_template.grid(row=0, column=0, columnspan=2, sticky="w")
    i18n_widgets["chk_use_template"] = chk_use_template

    lbl_template_file = ttk.Label(template_frame, text=t("LBL_TEMPLATE_FILE"))
    lbl_template_file.grid(row=1, column=0, sticky="w", pady=(5, 0))
    i18n_widgets["lbl_template_file"] = lbl_template_file
    choose_template_btn = ttk.Button(
        template_frame, text=t("BTN_SELECT_TEMPLATE"), command=choose_template_file
    )
    choose_template_btn.grid(row=1, column=1, sticky="w", pady=(5, 0))
    template_state["choose_button"] = choose_template_btn
    i18n_widgets["choose_template_btn"] = choose_template_btn
    ttk.Label(
        template_frame,
        textvariable=template_label_var,
        wraplength=600,
        justify="left",
    ).grid(row=2, column=0, columnspan=2, sticky="we")

    lbl_template_sheet = ttk.Label(template_frame, text=t("LBL_TEMPLATE_SHEET"))
    lbl_template_sheet.grid(row=3, column=0, sticky="w", pady=(5, 0))
    i18n_widgets["lbl_template_sheet"] = lbl_template_sheet
    sheet_combobox = ttk.Combobox(
        template_frame,
        textvariable=sheet_name_var,
        state="readonly",
        width=30,
    )
    sheet_combobox.grid(row=3, column=1, sticky="w", pady=(5, 0))
    template_state["sheet_combobox"] = sheet_combobox

    lbl_template_start_cell = ttk.Label(template_frame, text=t("LBL_TEMPLATE_START_CELL"))
    lbl_template_start_cell.grid(row=4, column=0, sticky="w", pady=(5, 0))
    i18n_widgets["lbl_template_start_cell"] = lbl_template_start_cell
    start_cell_entry = ttk.Entry(template_frame, textvariable=start_cell_var, width=10)
    start_cell_entry.grid(row=4, column=1, sticky="w", pady=(5, 0))
    template_state["start_cell_entry"] = start_cell_entry

    include_header_check = ttk.Checkbutton(
        template_frame,
        text=t("CHK_INCLUDE_HEADERS"),
        variable=include_header_var,
    )
    include_header_check.grid(row=5, column=0, columnspan=2, sticky="w", pady=(5, 0))
    template_state["include_header_check"] = include_header_check
    i18n_widgets["include_header_check"] = include_header_check

    update_template_controls_state()
    _refresh_template_ui()
    update_csv_profile_controls_state()

    btn_start = ttk.Button(result_frame, text=t("BTN_START"), command=run_export_gui)
    btn_start.grid(row=0, column=0, pady=(0, 10), sticky="w")
    start_button_holder["widget"] = btn_start
    i18n_widgets["btn_start"] = btn_start
    btn_cancel = ttk.Button(
        result_frame,
        text=t("BTN_CANCEL_RUN"),
        command=request_cancel,
        state=tk.DISABLED,
    )
    btn_cancel.grid(row=0, column=1, padx=(10, 0), pady=(0, 10), sticky="w")
    i18n_widgets["btn_cancel"] = btn_cancel
    btn_report_issue = ttk.Button(
        result_frame,
        text=t("BTN_REPORT_ISSUE"),
        command=lambda: open_github_issue_chooser(parent=root),
    )
    btn_report_issue.grid(row=0, column=2, padx=(10, 0), pady=(0, 10), sticky="w")
    i18n_widgets["btn_report_issue"] = btn_report_issue
    btn_check_updates = ttk.Button(
        result_frame,
        text=t("BTN_CHECK_UPDATES"),
        command=check_updates_gui,
    )
    btn_check_updates.grid(row=0, column=3, padx=(10, 0), pady=(0, 10), sticky="w")
    i18n_widgets["btn_check_updates"] = btn_check_updates
    btn_help = ttk.Button(
        result_frame,
        text=t("BTN_HELP"),
        command=show_help_window,
    )
    btn_help.grid(row=0, column=4, padx=(10, 0), pady=(0, 10), sticky="w")
    i18n_widgets["btn_help"] = btn_help
    btn_open_logs = ttk.Button(
        result_frame, text=t("BTN_OPEN_LOGS"), command=open_logs_folder
    )
    btn_open_logs.grid(row=0, column=5, padx=(10, 0), pady=(0, 10), sticky="w")
    i18n_widgets["btn_open_logs"] = btn_open_logs

    lbl_project_page = tk.Label(
        result_frame, text=t("LBL_PROJECT_PAGE"), fg="blue", cursor="hand2"
    )
    _project_font = tkfont.Font(lbl_project_page, lbl_project_page.cget("font"))
    _project_font.configure(underline=True)
    lbl_project_page.configure(font=_project_font)
    lbl_project_page.bind(
        "<Button-1>", lambda *_: webbrowser.open(REPO_URL)
    )
    lbl_project_page.grid(
        row=0, column=6, padx=(10, 0), pady=(2, 10), sticky="w"
    )
    i18n_widgets["lbl_project_page"] = lbl_project_page

    refresh_connection_combobox()
    refresh_secure_edit_button()
    if selected_connection_var.get():
        apply_selected_connection(show_success=False)
    else:
        set_connection_status(connected=False, key="STATUS_NO_CONNECTION")

    lbl_export_info = ttk.Label(result_frame, text=t("LBL_EXPORT_INFO"))
    lbl_export_info.grid(row=1, column=0, sticky="nw")
    i18n_widgets["lbl_export_info"] = lbl_export_info
    lbl_export_info_value = ttk.Label(
        result_frame,
        textvariable=result_info_var,
        justify="left",
        wraplength=600,
    )
    lbl_export_info_value.grid(row=1, column=1, columnspan=3, sticky="w")

    btn_open_file = ttk.Button(result_frame, text=t("BTN_OPEN_FILE"), command=open_file)
    btn_open_file.grid(row=2, column=0, pady=5, sticky="w")
    i18n_widgets["btn_open_file"] = btn_open_file
    btn_open_folder = ttk.Button(
        result_frame, text=t("BTN_OPEN_FOLDER"), command=open_folder
    )
    btn_open_folder.grid(row=2, column=1, pady=5, sticky="w")
    i18n_widgets["btn_open_folder"] = btn_open_folder

    lbl_errors_short = ttk.Label(result_frame, text=t("LBL_ERRORS_SHORT"))
    lbl_errors_short.grid(row=3, column=0, sticky="nw", pady=(10, 0))
    i18n_widgets["lbl_errors_short"] = lbl_errors_short
    error_frame = ttk.Frame(result_frame)
    error_frame.grid(row=3, column=1, columnspan=3, sticky="nsew", pady=(10, 0))

    error_text = tk.Text(
        error_frame,
        width=90,
        height=6,
        wrap="none",
        state="disabled",
        font=("Consolas", 9),
    )
    error_y_scroll = tk.Scrollbar(error_frame, orient="vertical", command=error_text.yview)
    error_x_scroll = tk.Scrollbar(error_frame, orient="horizontal", command=error_text.xview)
    error_text.configure(yscrollcommand=error_y_scroll.set, xscrollcommand=error_x_scroll.set)

    error_y_scroll.pack(side="right", fill="y")
    error_x_scroll.pack(side="bottom", fill="x")
    error_text.pack(side="left", fill="both", expand=True)

    error_display["widget"] = error_text

    btn_open_file.config(state=tk.DISABLED)
    btn_open_folder.config(state=tk.DISABLED)

    def apply_i18n():
        root.title(f"{t('APP_TITLE_FULL')} {get_app_version_label()}")
        connection_frame.config(text=t("FRAME_DB_CONNECTION"))
        advanced_frame.config(text=t("FRAME_ADVANCED"))
        source_frame.config(text=t("FRAME_SQL_SOURCE"))
        format_frame.config(text=t("FRAME_OUTPUT_FORMAT"))
        template_frame.config(text=t("FRAME_TEMPLATE_OPTIONS"))
        result_frame.config(text=t("FRAME_RESULTS"))
        lbl_connection.config(text=t("LBL_CONNECTION"))
        lbl_language.config(text=t("LBL_LANGUAGE"))
        chk_archive_sql.config(text=t("CHK_ARCHIVE_SQL"))
        btn_manage_connections.config(text=t("BTN_MANAGE_CONNECTIONS"))
        btn_test_connection.config(text=t("BTN_TEST_CONNECTION"))
        lbl_selected_sql.config(text=t("LBL_SELECTED_SQL"))
        btn_select_sql.config(text=t("BTN_SELECT_SQL"))
        btn_select_from_list.config(text=t("BTN_SELECT_FROM_LIST"))
        btn_edit_queries.config(text=t("BTN_EDIT_QUERIES"))
        btn_paste_sql.config(text=t("BTN_PASTE_SQL"))
        radio_xlsx.config(text=t("FORMAT_XLSX"))
        radio_csv.config(text=t("FORMAT_CSV"))
        lbl_csv_profile.config(text=t("LBL_CSV_PROFILE"))
        csv_profile_manage_btn.config(text=t("BTN_MANAGE_CSV_PROFILES"))
        lbl_save_as.config(text=t("LBL_SAVE_AS"))
        btn_save_as.config(text=t("BTN_SAVE_AS"))
        btn_clear_save_as.config(text=t("BTN_CLEAR"))
        lbl_save_as_hint.config(
            text=t("LBL_SAVE_AS_HINT", dir=shorten_path(output_directory, max_len=60))
        )
        chk_use_template.config(text=t("CHK_USE_TEMPLATE"))
        lbl_template_file.config(text=t("LBL_TEMPLATE_FILE"))
        choose_template_btn.config(text=t("BTN_SELECT_TEMPLATE"))
        lbl_template_sheet.config(text=t("LBL_TEMPLATE_SHEET"))
        lbl_template_start_cell.config(text=t("LBL_TEMPLATE_START_CELL"))
        include_header_check.config(text=t("CHK_INCLUDE_HEADERS"))
        btn_start.config(text=t("BTN_START"))
        btn_cancel.config(text=t("BTN_CANCEL_RUN"))
        btn_report_issue.config(text=t("BTN_REPORT_ISSUE"))
        btn_check_updates.config(text=t("BTN_CHECK_UPDATES"))
        btn_help.config(text=t("BTN_HELP"))
        btn_open_logs.config(text=t("BTN_OPEN_LOGS"))
        lbl_project_page.config(text=t("LBL_PROJECT_PAGE"))
        lbl_export_info.config(text=t("LBL_EXPORT_INFO"))
        btn_open_file.config(text=t("BTN_OPEN_FILE"))
        btn_open_folder.config(text=t("BTN_OPEN_FOLDER"))
        lbl_errors_short.config(text=t("LBL_ERRORS_SHORT"))
        btn_odbc_diagnostics.config(text=t("BTN_ODBC_DIAGNOSTICS"))
        refresh_csv_profile_controls(csv_profile_state["config"].get("default_profile"))
        if connection_status_state["key"]:
            is_connected = False
            status_btn = start_button_holder.get("widget")
            if status_btn is not None:
                is_connected = status_btn.cget("state") == tk.NORMAL
            set_connection_status(
                connected=is_connected,
                key=connection_status_state["key"],
                **connection_status_state["params"],
            )
        if pasted_sql_state["report_name"]:
            sql_label_var.set(
                f"{t('LBL_SQL_PASTED')} {pasted_sql_state['report_name']}"
            )

    def on_lang_change(_event=None):  # noqa: ANN001
        selected = (lang_var.get() or "").lower()
        set_lang(selected)
        persist_ui_lang(selected)
        apply_i18n()

    lang_combo.bind("<<ComboboxSelected>>", on_lang_change)

    _center_window(root)
    root.protocol("WM_DELETE_WINDOW", on_close)

    _dbg("run_gui(): entering mainloop()")
    root.mainloop()
    _dbg("run_gui(): mainloop() returned")


# =========================
# SELF-TESTS / DEV CHECKS
# =========================

LOGGER_METHODS = {"debug", "info", "warning", "error", "exception", "critical"}


def _is_logger_call_stmt(stmt) -> bool:
    import ast

    if not isinstance(stmt, ast.Expr):
        return False
    call = stmt.value
    if not isinstance(call, ast.Call):
        return False
    func = call.func
    if not isinstance(func, ast.Attribute):
        return False
    if func.attr not in LOGGER_METHODS:
        return False
    return isinstance(func.value, ast.Name) and func.value.id == "LOGGER"


def _find_logger_try_except_pass_wrappers(source: str) -> list[str]:
    import ast

    tree = ast.parse(source)
    locations: list[str] = []
    for node in ast.walk(tree):
        if not isinstance(node, ast.Try):
            continue
        if node.orelse or node.finalbody:
            continue
        if len(node.handlers) != 1:
            continue
        handler = node.handlers[0]
        if not isinstance(handler.type, ast.Name) or handler.type.id != "Exception":
            continue
        if len(handler.body) != 1 or not isinstance(handler.body[0], ast.Pass):
            continue
        if not node.body or not all(_is_logger_call_stmt(stmt) for stmt in node.body):
            continue
        locations.append(f"line {node.lineno}")
    return locations


def _find_unclosed_openpyxl_workbooks(source: str) -> list[str]:
    import ast

    tree = ast.parse(source)

    def _is_load_workbook_call(node: ast.AST) -> bool:
        if not isinstance(node, ast.Call):
            return False
        func = node.func
        if isinstance(func, ast.Name):
            return func.id == "load_workbook"
        if isinstance(func, ast.Attribute):
            return func.attr == "load_workbook"
        return False

    def _node_to_dotted_path(node: ast.AST) -> str | None:
        if isinstance(node, ast.Name):
            return node.id
        if isinstance(node, ast.Attribute):
            base = _node_to_dotted_path(node.value)
            if base is None:
                return None
            return f"{base}.{node.attr}"
        return None

    def _extract_assigned_names(target: ast.AST) -> list[str]:
        dotted = _node_to_dotted_path(target)
        if dotted is not None:
            return [dotted]
        if isinstance(target, (ast.Tuple, ast.List)):
            out: list[str] = []
            for elt in target.elts:
                out.extend(_extract_assigned_names(elt))
            return out
        return []

    def _expr_contains_load_workbook(expr: ast.AST) -> bool:
        # łapie: load_workbook(...), load_workbook(...).sheetnames, itd.
        return any(_is_load_workbook_call(n) for n in ast.walk(expr))

    def _is_immediate_load_workbook_close(expr: ast.AST) -> bool:
        if not isinstance(expr, ast.Call):
            return False
        func = expr.func
        if not isinstance(func, ast.Attribute) or func.attr != "close":
            return False
        return _is_load_workbook_call(func.value)

    def _collect_closed_names_in_finally(finalbody: list[ast.stmt]) -> set[str]:
        module = ast.Module(body=finalbody, type_ignores=[])
        closed_names: set[str] = set()
        for node in ast.walk(module):
            if not isinstance(node, ast.Call):
                continue
            func = node.func
            if not isinstance(func, ast.Attribute) or func.attr != "close":
                continue
            dotted = _node_to_dotted_path(func.value)
            if dotted is not None:
                closed_names.add(dotted)
        return closed_names

    unclosed_lines: set[int] = set()

    class WorkbookCloseVisitor(ast.NodeVisitor):
        def __init__(self) -> None:
            self.try_finally_closed_stack: list[set[str]] = []

        def _current_closed_names(self) -> set[str]:
            if not self.try_finally_closed_stack:
                return set()
            combined: set[str] = set()
            for names in self.try_finally_closed_stack:
                combined.update(names)
            return combined

        def _check_assignment(self, node: ast.AST, value: ast.AST, targets: list[ast.AST]) -> None:
            if not _is_load_workbook_call(value):
                return
            target_names: list[str] = []
            for target in targets:
                target_names.extend(_extract_assigned_names(target))
            closed_names = self._current_closed_names()
            for name in target_names:
                if name not in closed_names:
                    unclosed_lines.add(node.lineno)

        def visit_Assign(self, node: ast.Assign) -> None:
            self._check_assignment(node, node.value, node.targets)
            self.generic_visit(node)

        def visit_AnnAssign(self, node: ast.AnnAssign) -> None:
            if node.value is not None:
                self._check_assignment(node, node.value, [node.target])
            self.generic_visit(node)

        def visit_Expr(self, node: ast.Expr) -> None:
            if (
                _expr_contains_load_workbook(node.value)
                and not _is_immediate_load_workbook_close(node.value)
            ):
                unclosed_lines.add(node.lineno)
            self.generic_visit(node)

        def visit_Try(self, node: ast.Try) -> None:
            closed_names = _collect_closed_names_in_finally(node.finalbody)
            if node.finalbody:
                self.try_finally_closed_stack.append(closed_names)

            for stmt in node.body:
                self.visit(stmt)
            for handler in node.handlers:
                self.visit(handler)
            for stmt in node.orelse:
                self.visit(stmt)
            for stmt in node.finalbody:
                self.visit(stmt)

            if node.finalbody:
                self.try_finally_closed_stack.pop()

    WorkbookCloseVisitor().visit(tree)

    return [f"line {line}" for line in sorted(unclosed_lines)]


def _selftest_openpyxl_close_scan() -> bool:
    cases = [
        (
            "name_target_closed",
            """
def f(path):
    try:
        wb = load_workbook(path)
    finally:
        wb.close()
""",
            [],
        ),
        (
            "self_attr_target_closed",
            """
def f(self, path):
    try:
        self.wb = load_workbook(path)
    finally:
        self.wb.close()
""",
            [],
        ),
        (
            "obj_attr_target_closed",
            """
def f(obj, path):
    try:
        obj.wb = load_workbook(path)
    finally:
        obj.wb.close()
""",
            [],
        ),
        (
            "different_name_not_closed_by_attr",
            """
def f(self, path):
    try:
        wb = load_workbook(path)
    finally:
        self.wb.close()
""",
            ["line 4"],
        ),
        (
            "immediate_close_allowed",
            """
def f(path):
    load_workbook(path).close()
""",
            [],
        ),
    ]

    ok = True
    for name, source, expected in cases:
        got = _find_unclosed_openpyxl_workbooks(source)
        if got != expected:
            print(
                "[FAIL] _find_unclosed_openpyxl_workbooks() "
                f"case={name} expected={expected!r} got={got!r}"
            )
            ok = False

    return ok


def _selftest_has_multiple_statements() -> bool:
    cases = [
        ("pg_single", "postgresql", "select 1", False),
        ("pg_two", "postgresql", "select 1; select 2", True),
        ("pg_trailing_semicolon", "postgresql", "select 1;", False),
        ("pg_semicolon_in_string", "postgresql", "select ';' as x", False),
        ("pg_comment_semicolon", "postgresql", "-- comment;\nselect 1", False),
        ("pg_comment_between", "postgresql", "select 1; -- comment\nselect 2", True),
        ("pg_block_comment", "postgresql", "/* ; */ select 1", False),
        (
            "pg_do_block_then_select",
            "postgresql",
            "do $$ begin perform 1; perform 2; end $$; select 1",
            True,
        ),
        (
            "pg_dollar_quoted_string",
            "postgresql",
            "select $tag$ a; b $tag$ as x",
            False,
        ),
        ("mysql_two", "mysql", "select 1; select 2", True),
        ("mysql_backticks", "mysql", "select `a;b` from t", False),
        # Python string: "a\\\';b" -> SQL string content: a\';b
        ("mysql_escape", "mysql", "select 'a\\\';b' as x", False),
        ("sqlite_two", "sqlite", "select 1; select 2", True),
        ("sqlite_empty_segments", "sqlite", " ; \n select 1 ;", False),
    ]

    ok = True
    for name, backend, sql, expected in cases:
        try:
            got = has_multiple_statements(sql, backend)
        except Exception as exc:  # noqa: BLE001
            print(f"[FAIL] has_multiple_statements: {name} raised {type(exc).__name__}: {exc}")
            ok = False
            continue

        if got == expected:
            print(f"[OK] has_multiple_statements: {name}")
        else:
            print(f"[FAIL] has_multiple_statements: {name} expected {expected} got {got}")
            ok = False

    return ok



def _selftest_apply_mssql_safe_set() -> bool:
    class _FakeCursor:
        def __init__(self, mode: str):
            self.mode = mode
            self.calls: list[str] = []

        def execute(self, sql):
            s = str(sql)
            self.calls.append(s)
            is_batch = ("\n" in s) or (s.upper().count("SET ") > 1)
            if self.mode == "batch_fails_singles_ok":
                if is_batch:
                    raise RuntimeError("batch not supported")
                return None
            if self.mode == "single_statement_fails":
                if is_batch:
                    raise RuntimeError("batch not supported")
                if s == "SET ANSI_WARNINGS ON":
                    raise RuntimeError("simulated single statement failure")
                return None
            if self.mode == "batch_ok":
                return None
            raise RuntimeError(f"unknown mode: {self.mode}")

    ok = True

    cur_batch_fallback = _FakeCursor("batch_fails_singles_ok")
    result_batch_fallback = _apply_mssql_safe_set(cur_batch_fallback)
    expected_stmt_calls = [
        stmt.strip()
        for stmt in MSSQL_SAFE_SET_SQL.split(";")
        if stmt.strip()
    ]
    if result_batch_fallback is True and cur_batch_fallback.calls == [
        MSSQL_SAFE_SET_SQL,
        *expected_stmt_calls,
    ]:
        print("[OK] _apply_mssql_safe_set(): fallback per statement")
    else:
        print(
            "[FAIL] _apply_mssql_safe_set(): fallback per statement "
            f"result={result_batch_fallback!r} calls={cur_batch_fallback.calls!r}"
        )
        ok = False

    cur_partial_failure = _FakeCursor("single_statement_fails")
    result_partial_failure = _apply_mssql_safe_set(cur_partial_failure)
    if result_partial_failure is False and cur_partial_failure.calls == [
        MSSQL_SAFE_SET_SQL,
        *expected_stmt_calls,
    ]:
        print("[OK] _apply_mssql_safe_set(): keeps trying after statement failure")
    else:
        print(
            "[FAIL] _apply_mssql_safe_set(): partial failure handling "
            f"result={result_partial_failure!r} calls={cur_partial_failure.calls!r}"
        )
        ok = False

    cur_batch_ok = _FakeCursor("batch_ok")
    result_batch_ok = _apply_mssql_safe_set(cur_batch_ok)
    if result_batch_ok is True and cur_batch_ok.calls == [MSSQL_SAFE_SET_SQL]:
        print("[OK] _apply_mssql_safe_set(): batch fast path")
    else:
        print(
            "[FAIL] _apply_mssql_safe_set(): batch fast path "
            f"result={result_batch_ok!r} calls={cur_batch_ok.calls!r}"
        )
        ok = False

    return ok


def _can_write_text_stream(stream) -> bool:
    """Return True when ``stream`` looks writable and is not marked as closed."""
    try:
        return stream is not None and hasattr(stream, "write") and not getattr(stream, "closed", False)
    except Exception:
        return False


def run_self_test() -> bool:
    ok, report = run_self_test_report()

    emitted = False
    stdout = getattr(sys, "stdout", None)
    if _can_write_text_stream(stdout):
        try:
            print(report)
            try:
                stdout.flush()
            except Exception:
                pass
            emitted = True
        except Exception:
            emitted = False

    report_path = _write_self_test_report(report)
    # Popup only as a fallback when we cannot emit to stdout (e.g. pythonw / no console).
    if (not emitted) and report_path:
        _show_self_test_report_popup(report, report_path)

    return ok


def run_self_test_report() -> tuple[bool, str]:
    lines: list[str] = []

    def _append(label: str, value) -> None:
        lines.append(f"{label}: {value}")

    lines.append("=== KKR Query2XLSX self-test report ===")
    _append("app_version", get_app_version_label())
    _append("install_mode", detect_install_mode())
    _append("BASE_DIR", BASE_DIR)
    _append("DATA_DIR", DATA_DIR)
    _append("LOG_FILE_PATH", LOG_FILE_PATH or "")
    _append("sys.executable", sys.executable)
    _append("python version", sys.version.replace("\n", " "))
    _append("platform", platform.platform())
    _append("sqlalchemy.__version__", getattr(sqlalchemy, "__version__", "unknown"))
    _append("openpyxl.__version__", getattr(openpyxl, "__version__", "unknown"))

    try:
        import pyodbc  # type: ignore

        _append("pyodbc", f"available ({getattr(pyodbc, '__version__', 'unknown')})")
        try:
            drivers = pyodbc.drivers()
            _append("pyodbc.drivers()", ", ".join(drivers) if drivers else "<none>")
        except Exception as exc:  # noqa: BLE001
            _append("pyodbc.drivers()", f"<error: {type(exc).__name__}: {exc}>")
    except Exception as exc:  # noqa: BLE001
        _append("pyodbc", f"missing ({type(exc).__name__}: {exc})")
        _append("pyodbc.drivers()", "<unavailable>")

    lines.append("")
    lines.append("=== checks ===")

    ok = True

    is_frozen = bool(getattr(sys, "frozen", False))
    src_path = None
    has_source_file = False

    if not is_frozen:
        file_path = globals().get("__file__")
        if isinstance(file_path, str) and file_path:
            src_path = Path(file_path)
            has_source_file = src_path.exists()

    # 1) Syntax compile smoke test (source only)
    if has_source_file and not is_frozen:
        try:
            import py_compile

            py_compile.compile(str(src_path), doraise=True)  # type: ignore[arg-type]
            lines.append("[OK] py_compile main.pyw")
        except Exception as exc:  # noqa: BLE001
            lines.append(f"[FAIL] py_compile main.pyw: {type(exc).__name__}: {exc}")
            ok = False
    else:
        lines.append("[SKIP] py_compile (no source file / frozen exe)")

    # 2) Code hygiene scan: forbid try: LOGGER.*; except Exception: pass (source only)
    if has_source_file and not is_frozen:
        try:
            source = src_path.read_text(encoding="utf-8")  # type: ignore[union-attr]
            locations = _find_logger_try_except_pass_wrappers(source)
            if locations:
                lines.append(
                    "[FAIL] Found forbidden LOGGER try/except-pass wrappers at: "
                    + ", ".join(locations)
                )
                ok = False
            else:
                lines.append("[OK] no LOGGER try/except-pass wrappers")
        except Exception as exc:  # noqa: BLE001
            lines.append(f"[FAIL] LOGGER wrapper scan: {type(exc).__name__}: {exc}")
            ok = False
    else:
        lines.append("[SKIP] LOGGER wrapper scan (no source file / frozen exe)")

    # 2b) openpyxl load_workbook close scan
    if has_source_file and not is_frozen:
        try:
            source = src_path.read_text(encoding="utf-8")  # type: ignore[union-attr]
            locations = _find_unclosed_openpyxl_workbooks(source)
            if locations:
                formatted_locations = "\n".join(f"- {location}" for location in locations)
                lines.append("[FAIL] Found load_workbook() without close() in finally at:")
                lines.append(formatted_locations)
                lines.append("Wrap load_workbook in try/finally and call wb.close()")
                ok = False
            else:
                lines.append("[OK] openpyxl.load_workbook close handling")
        except Exception as exc:  # noqa: BLE001
            lines.append(f"[FAIL] openpyxl close scan: {type(exc).__name__}: {exc}")
            ok = False
    else:
        lines.append("[SKIP] openpyxl close scan (no source file / frozen exe)")

    # 2c) openpyxl load_workbook close scan self-test (string-only)
    try:
        if _selftest_openpyxl_close_scan():
            lines.append("[OK] _find_unclosed_openpyxl_workbooks() cases")
        else:
            lines.append("[FAIL] _find_unclosed_openpyxl_workbooks() cases")
            ok = False
    except Exception as exc:  # noqa: BLE001
        lines.append(
            "[FAIL] _find_unclosed_openpyxl_workbooks() self-test crashed: "
            f"{type(exc).__name__}: {exc}"
        )
        ok = False

    # 3) Runtime function smoke test (always)
    try:
        diagnostics = odbc_diagnostics_text()
        if isinstance(diagnostics, str):
            lines.append("[OK] odbc_diagnostics_text()")
        else:
            lines.append("[FAIL] odbc_diagnostics_text(): returned non-string value")
            ok = False
    except Exception as exc:  # noqa: BLE001
        lines.append(f"[FAIL] odbc_diagnostics_text(): {type(exc).__name__}: {exc}")
        ok = False

    # 4) has_multiple_statements() self-test (string-only)
    try:
        capture = io.StringIO()
        with contextlib.redirect_stdout(capture):
            has_multi_ok = _selftest_has_multiple_statements()
        emitted = capture.getvalue().strip()
        if emitted:
            lines.extend(emitted.splitlines())

        if has_multi_ok:
            lines.append("[OK] has_multiple_statements() cases")
        else:
            lines.append("[FAIL] has_multiple_statements() cases")
            ok = False
    except Exception as exc:  # noqa: BLE001
        lines.append(f"[FAIL] has_multiple_statements() self-test crashed: {type(exc).__name__}: {exc}")
        ok = False

    # 5) MSSQL session prolog helper self-test (no DB required)
    try:
        capture = io.StringIO()
        with contextlib.redirect_stdout(capture):
            safe_set_ok = _selftest_apply_mssql_safe_set()
        emitted = capture.getvalue().strip()
        if emitted:
            lines.extend(emitted.splitlines())

        if safe_set_ok:
            lines.append("[OK] _apply_mssql_safe_set() cases")
        else:
            lines.append("[FAIL] _apply_mssql_safe_set() cases")
            ok = False
    except Exception as exc:  # noqa: BLE001
        lines.append(f"[FAIL] _apply_mssql_safe_set() self-test crashed: {type(exc).__name__}: {exc}")
        ok = False

    lines.append("")
    lines.append("SELF-TEST: OK" if ok else "SELF-TEST: FAIL")
    report = "\n".join(lines)
    return ok, report


def _write_self_test_report(report: str) -> str | None:
    candidates = [LOG_DIR, DATA_DIR, tempfile.gettempdir()]
    filename = f"kkr-query2xlsx.selftest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    for candidate in candidates:
        if not candidate:
            continue
        try:
            os.makedirs(candidate, exist_ok=True)
            path = os.path.join(candidate, filename)
            with open(path, "w", encoding="utf-8") as fh:
                fh.write(report)
                fh.write("\n")
            return path
        except Exception:
            continue
    return None


def _show_self_test_report_popup(report: str, report_path: str) -> None:
    try:
        popup_root = tk.Tk()
        apply_app_icon(popup_root)
        popup_root.title("Self-test report")
        popup_root.geometry("860x560")
        popup_root.minsize(700, 420)

        popup_root.columnconfigure(0, weight=1)
        popup_root.rowconfigure(1, weight=1)

        info = ttk.Label(
            popup_root,
            text=f"Report saved to:\n{report_path}",
            justify="left",
            anchor="w",
            padding=(12, 10, 12, 6),
        )
        info.grid(row=0, column=0, sticky="we")

        body = ttk.Frame(popup_root, padding=(12, 0, 12, 8))
        body.grid(row=1, column=0, sticky="nsew")
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        text_widget = tk.Text(
            body,
            wrap="none",
            font=("Consolas", 9) if sys.platform == "win32" else None,
            borderwidth=1,
            relief="solid",
        )
        y_scroll = ttk.Scrollbar(body, orient="vertical", command=text_widget.yview)
        x_scroll = ttk.Scrollbar(body, orient="horizontal", command=text_widget.xview)
        text_widget.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        text_widget.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="we")

        text_widget.insert("1.0", report)
        text_widget.config(state="disabled")

        btns = ttk.Frame(popup_root, padding=(12, 0, 12, 12))
        btns.grid(row=2, column=0, sticky="we")

        def copy_all() -> None:
            popup_root.clipboard_clear()
            popup_root.clipboard_append(report or "")

        ttk.Button(btns, text=t("BTN_COPY_ALL"), command=copy_all).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text=t("BTN_CLOSE"), command=popup_root.destroy).pack(side="left")

        popup_root.bind("<Escape>", lambda *_: popup_root.destroy())
        popup_root.mainloop()
    except Exception:
        return


if __name__ == "__main__":
    if os.environ.get("KKR_DEBUG_EXIT") == "1":
        _install_debug_sys_exit_hook()
        _dbg("entered __main__")

    early_parser = argparse.ArgumentParser(add_help=False)
    early_parser.add_argument("--lang", choices=["en", "pl"])
    early_parser.add_argument("--self-test", action="store_true")
    early_args, _ = early_parser.parse_known_args()
    if early_args.lang:
        set_lang(early_args.lang)
    if early_args.self_test:
        sys.exit(0 if run_self_test() else 1)

    parser = argparse.ArgumentParser(description=t("CLI_DESC"))
    parser.add_argument("--self-test", action="store_true", help=t("CLI_SELF_TEST_HELP"))
    parser.add_argument("-c", "--console", action="store_true", help=t("CLI_CONSOLE_HELP"))
    parser.add_argument("-o", "--output", help=t("CLI_OUTPUT_HELP"))
    parser.add_argument("--lang", choices=["en", "pl"], help=t("CLI_LANG_HELP"))
    parser.add_argument("--diag-odbc", action="store_true", help=t("CLI_DIAG_ODBC_HELP"))
    parser.add_argument(
        "--check-update",
        action="store_true",
        help="Check for updates and exit.",
    )
    parser.add_argument(
        "--update",
        action="store_true",
        help="Start updater (if available) and exit.",
    )
    parser.add_argument(
        "--archive-sql",
        action="store_true",
        help="Archive executed SQL (with metadata) to sql_archive/.",
    )
    args = parser.parse_args()

    if args.self_test:
        sys.exit(0 if run_self_test() else 1)

    if args.diag_odbc:
        print(odbc_diagnostics_text())
        sys.exit(0)

    install_mode = detect_install_mode()

    if args.check_update:
        if install_mode == "git":
            release_info = None
            try:
                release_info = get_update_info()
            except Exception:
                release_info = None
            print("Update: repository checkout detected. Use Git to sync with remote (e.g. `git pull`).")
            local_sha = _get_git_full_sha()
            remote_sha = None
            try:
                remote_sha = _fetch_remote_main_sha()
            except Exception:
                remote_sha = None
            if local_sha and remote_sha:
                local_short = local_sha[:7]
                remote_short = remote_sha[:7]
                relation = _classify_git_relation(local_sha, remote_sha)
                if relation == "match":
                    print(
                        f"Update: local matches remote main (local {local_short}, remote {remote_short})"
                    )
                elif relation == "remote_ahead":
                    print(
                        f"Update: remote main is ahead (local {local_short}, remote {remote_short})"
                    )
                elif relation == "local_ahead":
                    print(
                        f"Update: local is ahead of remote main (local {local_short}, remote {remote_short})"
                    )
                elif relation == "diverged":
                    print(
                        f"Update: local and remote main have diverged (local {local_short}, remote {remote_short})"
                    )
                elif relation == "different_unverified":
                    print(
                        "Update: local differs from remote main "
                        f"(local {local_short}, remote {remote_short}) "
                        "(could not compare — no connection or the local commit does not exist on GitHub)"
                    )
                else:
                    print(
                        f"Update: local differs from remote main (local {local_short}, remote {remote_short})"
                    )
            if release_info:
                latest_tag = release_info.get("latest_tag") or ""
                current_tag = release_info.get("current_tag") or f"v{APP_VERSION}"
                if latest_tag:
                    status = (
                        "available"
                        if release_info.get("update_available")
                        else "none"
                    )
                    print(
                        f"Release: {status} (current {current_tag}, latest {latest_tag})"
                    )
            sys.exit(0)
        if install_mode == "source":
            release_info = None
            try:
                release_info = get_update_info()
            except Exception:
                release_info = None
            print(
                "Update: source checkout detected (no Git). Download a new ZIP or clone the repo."
            )
            if release_info:
                latest_tag = release_info.get("latest_tag") or ""
                current_tag = release_info.get("current_tag") or f"v{APP_VERSION}"
                if latest_tag:
                    status = (
                        "available"
                        if release_info.get("update_available")
                        else "none"
                    )
                    print(
                        f"Release: {status} (current {current_tag}, latest {latest_tag})"
                    )
            sys.exit(0)
        try:
            info = get_update_info()
        except Exception as exc:
            LOGGER.warning("Update check failed (CLI): %s", exc, exc_info=True)
            print(f"Update: error ({_build_update_check_message(exc)})")
            sys.exit(0)
        latest_tag = info.get("latest_tag") or ""
        current_tag = info.get("current_tag") or f"v{APP_VERSION}"
        if not latest_tag:
            print(f"Update: error ({t('UPD_ERR_JSON')})")
            sys.exit(0)
        if info.get("update_available"):
            print(f"Update: available (current {current_tag}, latest {latest_tag})")
        else:
            print(f"Update: none (current {current_tag}, latest {latest_tag})")
        sys.exit(0)

    if args.update:
        # Ensure we don't launch an outdated updater if a staged updater exists.
        if install_mode == "exe":
            _apply_pending_updater_update()
        if install_mode != "exe":
            print(
                "Update: not supported for source runs. Use `git pull` or download a new ZIP."
            )
            sys.exit(2)
        if not _get_updater_command():
            print("Update: updater missing (download new ZIP manually)")
            sys.exit(2)
        try:
            if not launch_updater(wait_pid=os.getpid()):
                print("Update: start failed")
                sys.exit(1)
        except Exception:
            print("Update: start failed")
            sys.exit(1)
        sys.exit(0)

    output_directory = bootstrap_data_dir_and_workdirs_or_exit(prefer_gui_prompt=not args.console)

    if args.lang:
        set_lang(args.lang)
        persist_ui_lang(args.lang)
    else:
        stored_lang = load_persisted_ui_lang()
        if stored_lang:
            set_lang(stored_lang)

    if install_mode == "exe":
        _apply_pending_updater_update()

    created_files = bootstrap_local_files()
    if created_files:
        LOGGER.info(
            "Bootstrapped local files from samples: %s",
            ", ".join(created_files),
        )

    connection_store = load_connections()
    connection_store.pop("ui_lang", None)
    selected_name = connection_store.get("last_selected")
    selected_connection = None
    for conn in connection_store.get("connections", []):
        if conn.get("name") == selected_name:
            selected_connection = conn
            break
    if selected_connection is None and connection_store.get("connections"):
        selected_connection = connection_store["connections"][0]

    if args.console:
        if not selected_connection:
            print(t("CLI_NO_CONNECTIONS"))
            sys.exit(1)

        engine_kwargs = {}
        if selected_connection.get("type") == "mssql_odbc":
            engine_kwargs["isolation_level"] = "AUTOCOMMIT"
        try:
            # Console: build runtime URL from details; if password missing -> prompt in console.
            ctype = selected_connection.get("type")
            details = selected_connection.get("details") or {}
            if ctype == "sqlite":
                url = _build_runtime_url(ctype, details, None)
            elif ctype == "mssql_odbc" and bool(details.get("trusted")):
                url = _build_runtime_url(ctype, details, None)
            else:
                pwd = str(details.get("password") or "")
                if not pwd:
                    pwd = getpass.getpass(
                        t(
                            "CONSOLE_PROMPT_PASSWORD",
                            name=selected_connection.get("name"),
                        )
                    )
                    if not str(pwd or "").strip():
                        raise PasswordRequiredError(
                            name=selected_connection.get("name", ""),
                            conn_type=str(ctype or ""),
                        )
                url = _build_runtime_url(ctype, details, pwd)
            engine = create_engine(url, **engine_kwargs)
            if selected_connection.get("type") == "mssql_odbc":
                _ensure_engine_mssql_set_hook(engine)
        except Exception as exc:  # noqa: BLE001
            if isinstance(exc, PasswordRequiredError):
                name = getattr(exc, "name", "") or selected_connection.get("name", "")
                print(t("ERR_PASSWORD_REQUIRED_TITLE"))
                print(t("ERR_PASSWORD_REQUIRED_BODY", name=name))
                sys.exit(1)
            handled = handle_db_driver_error(
                exc,
                selected_connection.get("type"),
                selected_connection.get("name"),
                show_message=lambda *args: print(args[-1] if args else exc),
            )
            if not handled:
                LOGGER.exception(
                    "Failed to initialize console engine for %s (%s)",
                    selected_connection.get("name"),
                    selected_connection.get("type"),
                    exc_info=True,
                )
                title, msg = _format_connection_error(
                    conn_type=str(selected_connection.get("type") or ""),
                    exc=exc,
                    details=selected_connection.get("details") or {},
                )
                print(f"{title}\n{msg}")
            sys.exit(1)

        run_console(
            engine,
            output_directory,
            selected_connection,
            archive_sql=args.archive_sql,
            output_override=args.output,
        )
    else:
        _dbg("__main__: about to call run_gui()")
        run_gui(connection_store, output_directory)
        _dbg("__main__: run_gui() returned")
